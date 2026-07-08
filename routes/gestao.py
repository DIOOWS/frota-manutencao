from flask import Blueprint, render_template, request, redirect, flash, jsonify, send_file
from utils.auth import gestao_required
from models.lancamento_financeiro import LancamentoFinanceiro
from models.fechamento_mensal import FechamentoMensal
from models.conta_pagar_importada import ContaPagarImportada
from models.conta_receber_importada import ContaReceberImportada
from models.conta_recorrente import ContaRecorrente
from models.dashboard_operacional_importado import DashboardOperacionalImportado
from database import db
from datetime import datetime, date, timedelta
from collections import Counter
from decimal import Decimal, InvalidOperation
from sqlalchemy import or_, and_
import calendar
import re
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


gestao_bp = Blueprint("gestao", __name__, url_prefix="/gestao")


# =========================================================
# HELPERS GERAIS
# =========================================================

def dinheiro(valor):
    """
    Converte valores financeiros com segurança.

    Aceita:
    - Decimal, int, float
    - "R$ 1.234,56"
    - "1.234,56"
    - "1234,56"
    - "1234.56"

    Observação: campos com máscara JS serão enviados como "1234.56".
    """
    if valor is None:
        return 0

    if isinstance(valor, Decimal):
        return float(valor)

    if isinstance(valor, (int, float)):
        return float(valor)

    texto_valor = str(valor).strip()

    if not texto_valor:
        return 0

    texto_valor = (
        texto_valor
        .replace("R$", "")
        .replace(" ", "")
    )

    # Formato brasileiro: 1.234,56
    if "," in texto_valor:
        texto_valor = texto_valor.replace(".", "").replace(",", ".")

    try:
        return float(Decimal(texto_valor))
    except (InvalidOperation, ValueError):
        return 0


def normalizar_texto(valor):
    if not valor:
        return None

    return str(valor).strip().upper()


def parse_data(data_str):
    if not data_str:
        return None

    return datetime.strptime(data_str, "%Y-%m-%d").date()


def mes_ano_anterior(mes, ano):
    if mes == 1:
        return 12, ano - 1

    return mes - 1, ano


def obter_fechamento_anterior(mes, ano):
    mes_ant, ano_ant = mes_ano_anterior(mes, ano)

    return FechamentoMensal.query.filter_by(
        mes=mes_ant,
        ano=ano_ant
    ).first()


def obter_saldo_inicial_automatico(mes, ano):
    fechamento_anterior = obter_fechamento_anterior(mes, ano)

    if fechamento_anterior:
        return dinheiro(fechamento_anterior.saldo_final)

    return 0


def obter_saldo_inicial_mes(mes, ano):
    fechamento_atual = FechamentoMensal.query.filter_by(
        mes=mes,
        ano=ano
    ).first()

    if fechamento_atual:
        return dinheiro(fechamento_atual.saldo_inicial)

    return obter_saldo_inicial_automatico(mes, ano)


def status_pago(valor):
    status = normalizar_texto(valor)

    return status in [
        "PAGO",
        "RECEBIDO",
        "OK",
        "QUITADO",
        "BAIXADO"
    ]


def status_cancelado(valor):
    status = normalizar_texto(valor)

    return status in [
        "CANCELADO",
        "CANCELADA"
    ]


def calcular_variacao_percentual(valor_atual, valor_anterior):
    valor_atual = dinheiro(valor_atual)
    valor_anterior = dinheiro(valor_anterior)

    if valor_anterior == 0:
        if valor_atual == 0:
            return 0

        return None

    return ((valor_atual - valor_anterior) / abs(valor_anterior)) * 100


def primeiro_dia_mes(mes, ano):
    return date(ano, mes, 1)


def ultimo_dia_mes(mes, ano):
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    return date(ano, mes, ultimo_dia)


def inicio_mes_datetime(mes, ano):
    return datetime.combine(primeiro_dia_mes(mes, ano), datetime.min.time())


def fim_mes_datetime(mes, ano):
    return datetime.combine(ultimo_dia_mes(mes, ano), datetime.max.time())


def data_para_date(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    return None


def nome_dia_semana(data_ref):
    dias = {
        0: "segunda-feira",
        1: "terça-feira",
        2: "quarta-feira",
        3: "quinta-feira",
        4: "sexta-feira",
        5: "sábado",
        6: "domingo",
    }

    return dias.get(data_ref.weekday(), "")


def conta_esta_paga(conta):
    if getattr(conta, "pago", False):
        return True

    status = normalizar_texto(getattr(conta, "status", None))

    return status in [
        "PAGO",
        "RECEBIDO",
        "OK",
        "QUITADO",
        "BAIXADO"
    ]


def valor_conta_receber(conta):
    return dinheiro(
        getattr(conta, "total", None)
        or getattr(conta, "valor", 0)
    )


def data_movimento_fluxo(conta):
    """
    Fluxo é caixa realizado.
    Portanto a data correta é data_pagamento.
    Se não tiver data_pagamento, não entra no fluxo.
    """

    return data_para_date(getattr(conta, "data_pagamento", None))


# =========================================================
# CÁLCULO FINANCEIRO
# IMPORTADAS + LANÇAMENTOS MANUAIS
# =========================================================

def calcular_totais_financeiros(mes, ano, saldo_inicial=0):

    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)
    fim_dia = ultimo_dia_mes(mes, ano)

    total_entradas = 0
    total_saidas = 0
    total_a_pagar = 0
    total_a_receber = 0

    despesas_counter = Counter()
    despesas_assistencia_counter = Counter()
    despesas_logistica_counter = Counter()

    receitas_counter = Counter()
    clientes_counter = Counter()

    # =====================================================
    # SAÍDAS REALIZADAS
    # Contas pagas pela DATA DE PAGAMENTO dentro do mês
    # =====================================================
    contas_pagar_pagas = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt
    ).all()

    for c in contas_pagar_pagas:
        valor = dinheiro(c.valor)
        total_saidas += valor

        categoria = c.categoria or c.plano_contas or "SEM CATEGORIA"
        setor = c.setor or "ASSISTÊNCIA"

        despesas_counter[categoria] += valor

        if setor == "LOGÍSTICA":
            despesas_logistica_counter[categoria] += valor
        else:
            despesas_assistencia_counter[categoria] += valor

    # =====================================================
    # A PAGAR
    # Contas abertas com vencimento até o fim do mês filtrado
    # Inclui atrasadas de meses anteriores
    # =====================================================
    contas_pagar_abertas = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == False,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_vencimento <= fim_dt
    ).all()

    for c in contas_pagar_abertas:
        valor = dinheiro(c.valor)
        total_a_pagar += valor

    # =====================================================
    # ENTRADAS REALIZADAS
    # Contas recebidas pela DATA DE PAGAMENTO dentro do mês
    # =====================================================
    contas_receber_recebidas = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt
    ).all()

    for c in contas_receber_recebidas:
        valor = valor_conta_receber(c)
        total_entradas += valor

        categoria = c.categoria or c.plano_contas or "SEM CATEGORIA"
        cliente = c.cliente or "SEM CLIENTE"

        receitas_counter[categoria] += valor
        clientes_counter[cliente] += valor

    # =====================================================
    # A RECEBER
    # Contas abertas com vencimento até o fim do mês filtrado
    # Inclui atrasadas de meses anteriores
    # =====================================================
    contas_receber_abertas = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == False,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_vencimento <= fim_dt
    ).all()

    for c in contas_receber_abertas:
        valor = valor_conta_receber(c)
        total_a_receber += valor

    # =====================================================
    # LANÇAMENTOS MANUAIS
    # Mantém por mes/ano porque o lançamento manual já grava
    # o mês com base na data informada no cadastro
    # =====================================================
    lancamentos = LancamentoFinanceiro.query.filter_by(
        mes=mes,
        ano=ano
    ).all()

    for l in lancamentos:

        if status_cancelado(l.status):
            continue

        valor = dinheiro(l.valor)
        tipo = normalizar_texto(l.tipo)
        categoria = l.categoria or "SEM CATEGORIA"
        setor = l.setor or "GERAL"
        cliente = l.cliente or "SEM CLIENTE"

        if tipo == "RECEITA":

            if status_pago(l.status):
                total_entradas += valor

                receitas_counter[categoria] += valor
                clientes_counter[cliente] += valor
            else:
                total_a_receber += valor

        elif tipo == "DESPESA":

            if status_pago(l.status):
                total_saidas += valor

                despesas_counter[categoria] += valor

                if setor == "LOGÍSTICA":
                    despesas_logistica_counter[categoria] += valor
                else:
                    despesas_assistencia_counter[categoria] += valor
            else:
                total_a_pagar += valor

    lucro_mes = total_entradas - total_saidas
    saldo_final = dinheiro(saldo_inicial) + total_entradas - total_saidas

    margem_operacional = 0

    if total_entradas > 0:
        margem_operacional = (lucro_mes / total_entradas) * 100

    return {
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "total_a_pagar": total_a_pagar,
        "total_a_receber": total_a_receber,
        "lucro_mes": lucro_mes,
        "saldo_final": saldo_final,
        "margem_operacional": margem_operacional,

        "ranking_despesas": despesas_counter.most_common(),
        "ranking_despesas_assistencia": despesas_assistencia_counter.most_common(),
        "ranking_despesas_logistica": despesas_logistica_counter.most_common(),
        "ranking_receitas": receitas_counter.most_common(),
        "ranking_clientes": clientes_counter.most_common(),
    }


# =========================================================
# EVOLUÇÃO FINANCEIRA MENSAL
# =========================================================

def calcular_evolucao_mensal(ano):
    labels = []
    entradas = []
    saidas = []
    lucros = []
    saldos_finais = []
    margens = []
    tabela = []

    nomes_meses = [
        "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
        "Jul", "Ago", "Set", "Out", "Nov", "Dez"
    ]

    for mes in range(1, 13):
        saldo_inicial = obter_saldo_inicial_mes(mes, ano)

        totais = calcular_totais_financeiros(
            mes=mes,
            ano=ano,
            saldo_inicial=saldo_inicial
        )

        total_entradas = dinheiro(totais["total_entradas"])
        total_saidas = dinheiro(totais["total_saidas"])
        lucro_mes = dinheiro(totais["lucro_mes"])
        saldo_final = dinheiro(totais["saldo_final"])
        margem_operacional = dinheiro(totais["margem_operacional"])

        labels.append(nomes_meses[mes - 1])
        entradas.append(round(total_entradas, 2))
        saidas.append(round(total_saidas, 2))
        lucros.append(round(lucro_mes, 2))
        saldos_finais.append(round(saldo_final, 2))
        margens.append(round(margem_operacional, 2))

        tabela.append({
            "mes": mes,
            "nome_mes": nomes_meses[mes - 1],
            "ano": ano,
            "saldo_inicial": round(dinheiro(saldo_inicial), 2),
            "entradas": round(total_entradas, 2),
            "saidas": round(total_saidas, 2),
            "lucro": round(lucro_mes, 2),
            "saldo_final": round(saldo_final, 2),
            "margem": round(margem_operacional, 2),
        })

    return {
        "labels": labels,
        "entradas": entradas,
        "saidas": saidas,
        "lucros": lucros,
        "saldos_finais": saldos_finais,
        "margens": margens,
        "tabela": tabela,
    }


# =========================================================
# INTELIGÊNCIA FINANCEIRA / TOMADA DE DECISÃO
# =========================================================

def calcular_inteligencia_financeira(totais, evolucao_mensal, mes, ano, saldo_inicial):
    total_entradas = dinheiro(totais["total_entradas"])
    total_saidas = dinheiro(totais["total_saidas"])
    total_a_pagar = dinheiro(totais["total_a_pagar"])
    total_a_receber = dinheiro(totais["total_a_receber"])
    lucro_mes = dinheiro(totais["lucro_mes"])
    saldo_final = dinheiro(totais["saldo_final"])
    margem = dinheiro(totais["margem_operacional"])

    percentual_saidas = 0
    if total_entradas > 0:
        percentual_saidas = (total_saidas / total_entradas) * 100

    saldo_projetado_pos_obrigacoes = saldo_final + total_a_receber - total_a_pagar

    faltante_equilibrio = 0
    sobra_equilibrio = 0

    if total_saidas > total_entradas:
        faltante_equilibrio = total_saidas - total_entradas
    else:
        sobra_equilibrio = total_entradas - total_saidas

    despesas_top = totais["ranking_despesas"][:5]
    despesas_top3 = totais["ranking_despesas"][:3]

    total_top5 = sum(dinheiro(valor) for _, valor in despesas_top)
    total_top3 = sum(dinheiro(valor) for _, valor in despesas_top3)
    economia_10_top3 = total_top3 * 0.10

    percentual_top5_saidas = 0
    if total_saidas > 0:
        percentual_top5_saidas = (total_top5 / total_saidas) * 100

    maior_despesa_nome = None
    maior_despesa_valor = 0
    maior_despesa_percentual = 0

    if despesas_top:
        maior_despesa_nome = despesas_top[0][0]
        maior_despesa_valor = dinheiro(despesas_top[0][1])

        if total_saidas > 0:
            maior_despesa_percentual = (maior_despesa_valor / total_saidas) * 100

    meses_com_movimento_lista = [
        item for item in evolucao_mensal["tabela"]
        if dinheiro(item["entradas"]) != 0 or dinheiro(item["saidas"]) != 0
    ]

    meses_positivos_lista = [
        item for item in meses_com_movimento_lista
        if dinheiro(item["lucro"]) > 0
    ]

    meses_negativos_lista = [
        item for item in meses_com_movimento_lista
        if dinheiro(item["lucro"]) < 0
    ]

    melhor_mes = None
    pior_mes = None
    media_resultado = 0
    resultado_acumulado_ano = 0
    entradas_acumuladas_ano = 0
    saidas_acumuladas_ano = 0
    margem_media_ano = 0

    if meses_com_movimento_lista:
        melhor_mes = max(meses_com_movimento_lista, key=lambda item: dinheiro(item["lucro"]))
        pior_mes = min(meses_com_movimento_lista, key=lambda item: dinheiro(item["lucro"]))
        resultado_acumulado_ano = sum(dinheiro(item["lucro"]) for item in meses_com_movimento_lista)
        entradas_acumuladas_ano = sum(dinheiro(item["entradas"]) for item in meses_com_movimento_lista)
        saidas_acumuladas_ano = sum(dinheiro(item["saidas"]) for item in meses_com_movimento_lista)
        media_resultado = resultado_acumulado_ano / len(meses_com_movimento_lista)

        if entradas_acumuladas_ano > 0:
            margem_media_ano = (resultado_acumulado_ano / entradas_acumuladas_ano) * 100

    aproveitamento_ano = 0
    if meses_com_movimento_lista:
        aproveitamento_ano = (len(meses_positivos_lista) / len(meses_com_movimento_lista)) * 100

    tendencia_ano = "Sem base suficiente"
    tendencia_ano_classe = "neutral"
    leitura_ano_resumo = "Ainda não há movimento suficiente para leitura estratégica do ano."

    if meses_com_movimento_lista:
        if len(meses_positivos_lista) == 0 and len(meses_negativos_lista) > 0:
            tendencia_ano = "Ano em zona crítica"
            tendencia_ano_classe = "danger"
            leitura_ano_resumo = "Todos os meses com movimento fecharam negativos. A prioridade é recuperar margem e cortar vazamentos recorrentes."
        elif resultado_acumulado_ano < 0:
            tendencia_ano = "Ano pressionado"
            tendencia_ano_classe = "warning"
            leitura_ano_resumo = "O ano ainda está negativo. Existem meses bons, mas eles não compensaram os meses ruins."
        elif margem_media_ano < 10:
            tendencia_ano = "Ano positivo, mas apertado"
            tendencia_ano_classe = "warning"
            leitura_ano_resumo = "O ano está positivo, porém com margem baixa. Qualquer despesa fora do padrão pode virar prejuízo."
        else:
            tendencia_ano = "Ano saudável"
            tendencia_ano_classe = "success"
            leitura_ano_resumo = "A operação acumulada do ano está positiva e com margem operacional favorável."

    diagnostico_ano = []

    if meses_com_movimento_lista:
        diagnostico_ano.append(
            f"Foram analisados {len(meses_com_movimento_lista)} mês(es) com movimento: {len(meses_positivos_lista)} positivo(s) e {len(meses_negativos_lista)} negativo(s)."
        )

        diagnostico_ano.append(
            f"O resultado acumulado do ano está em R$ {resultado_acumulado_ano:,.2f}, com média mensal de R$ {media_resultado:,.2f}."
        )

        if len(meses_positivos_lista) == 0 and len(meses_negativos_lista) > 0:
            diagnostico_ano.append(
                "Nenhum mês analisado fechou positivo. Isso indica que o negócio ainda não encontrou ponto de equilíbrio no ano."
            )
        elif aproveitamento_ano < 50:
            diagnostico_ano.append(
                "Menos da metade dos meses fechou positivo. O foco deve ser padronizar receita e controlar despesas recorrentes."
            )
        else:
            diagnostico_ano.append(
                "A maior parte dos meses analisados fechou positiva. O foco agora é proteger margem e caixa."
            )
    else:
        diagnostico_ano.append(
            "Sem meses com movimento financeiro suficiente para montar diagnóstico anual."
        )

    leitura_vazamentos = "Sem despesas suficientes para leitura dos vazamentos."
    prioridade_corte = "Sem prioridade definida"
    prioridade_corte_classe = "neutral"

    if despesas_top:
        leitura_vazamentos = (
            f"Os 5 maiores grupos consumiram {percentual_top5_saidas:.2f}% das saídas. "
            f"O maior vazamento é {maior_despesa_nome}, com {maior_despesa_percentual:.2f}% do total de saídas."
        )

        if percentual_top5_saidas >= 60 or maior_despesa_percentual >= 25:
            prioridade_corte = "Prioridade alta"
            prioridade_corte_classe = "danger"
        elif percentual_top5_saidas >= 35 or maior_despesa_percentual >= 15:
            prioridade_corte = "Prioridade média"
            prioridade_corte_classe = "warning"
        else:
            prioridade_corte = "Prioridade controlada"
            prioridade_corte_classe = "success"

    hoje = datetime.now()
    projecao = None
    mes_atual = hoje.month == mes and hoje.year == ano

    if mes_atual and hoje.day > 0:
        dias_no_mes = calendar.monthrange(ano, mes)[1]
        dias_restantes = max(dias_no_mes - hoje.day, 0)

        entradas_projetadas = (total_entradas / hoje.day) * dias_no_mes
        saidas_projetadas = (total_saidas / hoje.day) * dias_no_mes
        resultado_projetado = entradas_projetadas - saidas_projetadas
        saldo_final_projetado = dinheiro(saldo_inicial) + resultado_projetado
        caixa_projetado_pos_obrigacoes = saldo_final_projetado + total_a_receber - total_a_pagar

        media_diaria_entradas = total_entradas / hoje.day
        media_diaria_saidas = total_saidas / hoje.day
        media_diaria_resultado = lucro_mes / hoje.day

        necessario_equilibrio_projetado = 0
        if resultado_projetado < 0:
            necessario_equilibrio_projetado = abs(resultado_projetado)

        risco = "Baixo"
        risco_classe = "success"
        leitura = "Mantendo o ritmo atual, o mês tende a fechar positivo."

        if resultado_projetado < 0:
            risco = "Alto"
            risco_classe = "danger"
            leitura = "Mantendo o ritmo atual, o mês tende a fechar negativo. É preciso reduzir saídas ou antecipar recebimentos."
        elif margem < 10:
            risco = "Médio"
            risco_classe = "warning"
            leitura = "O mês tende a fechar positivo, mas com margem apertada. Controle novas despesas."

        projecao = {
            "dia_atual": hoje.day,
            "dias_no_mes": dias_no_mes,
            "dias_restantes": dias_restantes,
            "entradas_projetadas": entradas_projetadas,
            "saidas_projetadas": saidas_projetadas,
            "resultado_projetado": resultado_projetado,
            "saldo_final_projetado": saldo_final_projetado,
            "caixa_projetado_pos_obrigacoes": caixa_projetado_pos_obrigacoes,
            "media_diaria_entradas": media_diaria_entradas,
            "media_diaria_saidas": media_diaria_saidas,
            "media_diaria_resultado": media_diaria_resultado,
            "necessario_equilibrio_projetado": necessario_equilibrio_projetado,
            "risco": risco,
            "risco_classe": risco_classe,
            "leitura": leitura,
        }

    nivel = "success"
    titulo_situacao = "Operação saudável"
    resumo_situacao = "O mês apresenta resultado positivo e margem operacional favorável."

    if lucro_mes < 0:
        nivel = "danger"
        titulo_situacao = "Resultado negativo"
        resumo_situacao = "As saídas superaram as entradas. O foco imediato deve ser reduzir custos e acelerar recebimentos."
    elif margem < 10:
        nivel = "warning"
        titulo_situacao = "Margem apertada"
        resumo_situacao = "O mês está positivo, mas com margem baixa. Qualquer aumento de custo pode comprometer o resultado."
    elif percentual_saidas >= 80:
        nivel = "warning"
        titulo_situacao = "Despesas pressionando"
        resumo_situacao = "As saídas estão consumindo uma parte alta das entradas. Vale revisar os maiores custos."
    elif saldo_projetado_pos_obrigacoes < 0:
        nivel = "danger"
        titulo_situacao = "Risco de caixa"
        resumo_situacao = "Mesmo com o saldo atual, as obrigações em aberto podem deixar o caixa negativo."

    alertas = []

    if total_entradas == 0 and total_saidas == 0:
        alertas.append({
            "nivel": "warning",
            "titulo": "Sem movimento financeiro",
            "descricao": "Não há entradas nem saídas realizadas para o mês filtrado.",
            "acao": "Verifique se as importações possuem data de pagamento nas contas pagas/recebidas."
        })

    if lucro_mes < 0:
        alertas.append({
            "nivel": "danger",
            "titulo": "Prejuízo no mês",
            "descricao": "O mês está fechando com mais saídas realizadas do que entradas realizadas.",
            "acao": "Ataque as maiores despesas e priorize recebimentos em aberto."
        })

    if percentual_saidas >= 80 and total_entradas > 0:
        alertas.append({
            "nivel": "warning",
            "titulo": "Custo consumindo receita",
            "descricao": f"As saídas representam {percentual_saidas:.2f}% das entradas.",
            "acao": "Busque reduzir custos variáveis ou renegociar os maiores pagamentos."
        })

    if total_a_pagar > saldo_final:
        alertas.append({
            "nivel": "danger",
            "titulo": "A pagar maior que saldo final",
            "descricao": "As obrigações em aberto são maiores que o saldo final atual.",
            "acao": "Priorize recebimentos e evite novos compromissos antes de reforçar o caixa."
        })

    if total_a_receber > total_entradas and total_a_receber > 0:
        alertas.append({
            "nivel": "warning",
            "titulo": "Recebíveis importantes em aberto",
            "descricao": "O valor a receber é maior que o que já entrou no mês.",
            "acao": "Acompanhe cobranças e antecipe recebimentos quando possível."
        })

    if maior_despesa_nome:
        alertas.append({
            "nivel": "info",
            "titulo": "Maior vazamento financeiro",
            "descricao": f"{maior_despesa_nome} representa R$ {maior_despesa_valor:,.2f} e {maior_despesa_percentual:.2f}% das saídas.",
            "acao": "Analise se esse custo pode ser reduzido, renegociado ou controlado por limite."
        })

    if not alertas:
        alertas.append({
            "nivel": "success",
            "titulo": "Sem alertas críticos",
            "descricao": "Os principais indicadores do mês estão dentro de um cenário saudável.",
            "acao": "Mantenha o acompanhamento e preserve a margem operacional."
        })

    acoes_recomendadas = []

    if lucro_mes < 0:
        acoes_recomendadas.append("Reduzir imediatamente as 3 maiores despesas realizadas do mês.")
        acoes_recomendadas.append("Priorizar cobrança dos valores em aberto.")
        acoes_recomendadas.append("Revisar se alguma despesa pontual distorceu o mês.")

    if economia_10_top3 > 0:
        acoes_recomendadas.append(
            f"Reduzir 10% nas 3 maiores despesas aumentaria o resultado em aproximadamente R$ {economia_10_top3:,.2f}."
        )

    if total_a_receber > 0:
        acoes_recomendadas.append(
            f"Receber os valores em aberto pode reforçar o caixa em R$ {total_a_receber:,.2f}."
        )

    if total_a_pagar > 0:
        acoes_recomendadas.append(
            f"Planejar os pagamentos em aberto de R$ {total_a_pagar:,.2f} para não pressionar o caixa."
        )

    if mes_atual and projecao and projecao["resultado_projetado"] < 0:
        acoes_recomendadas.append(
            f"Para o mês atual não fechar negativo, será necessário melhorar aproximadamente R$ {projecao['necessario_equilibrio_projetado']:,.2f} até o fim do mês."
        )

    if not acoes_recomendadas:
        acoes_recomendadas.append("Manter controle do orçamento e acompanhar os maiores custos semanalmente.")

    return {
        "nivel": nivel,
        "titulo_situacao": titulo_situacao,
        "resumo_situacao": resumo_situacao,
        "percentual_saidas": percentual_saidas,
        "saldo_projetado_pos_obrigacoes": saldo_projetado_pos_obrigacoes,
        "faltante_equilibrio": faltante_equilibrio,
        "sobra_equilibrio": sobra_equilibrio,
        "despesas_top": despesas_top,
        "despesas_top3": despesas_top3,
        "total_top3": total_top3,
        "total_top5": total_top5,
        "percentual_top5_saidas": percentual_top5_saidas,
        "economia_10_top3": economia_10_top3,
        "maior_despesa_nome": maior_despesa_nome,
        "maior_despesa_valor": maior_despesa_valor,
        "maior_despesa_percentual": maior_despesa_percentual,
        "leitura_vazamentos": leitura_vazamentos,
        "prioridade_corte": prioridade_corte,
        "prioridade_corte_classe": prioridade_corte_classe,
        "meses_com_movimento": len(meses_com_movimento_lista),
        "meses_positivos": len(meses_positivos_lista),
        "meses_negativos": len(meses_negativos_lista),
        "melhor_mes": melhor_mes,
        "pior_mes": pior_mes,
        "media_resultado": media_resultado,
        "resultado_acumulado_ano": resultado_acumulado_ano,
        "entradas_acumuladas_ano": entradas_acumuladas_ano,
        "saidas_acumuladas_ano": saidas_acumuladas_ano,
        "margem_media_ano": margem_media_ano,
        "aproveitamento_ano": aproveitamento_ano,
        "tendencia_ano": tendencia_ano,
        "tendencia_ano_classe": tendencia_ano_classe,
        "leitura_ano_resumo": leitura_ano_resumo,
        "diagnostico_ano": diagnostico_ano,
        "mes_atual": mes_atual,
        "projecao": projecao,
        "alertas": alertas,
        "acoes_recomendadas": acoes_recomendadas,
    }




# =========================================================
# DASHBOARD HÍBRIDO - OBRIGAÇÕES + MODAIS DE KPI
# =========================================================

def formatar_moeda_json(valor):
    return moeda(valor) if "moeda" in globals() else f"R$ {dinheiro(valor):,.2f}"


def status_aberto_importado(conta):
    if getattr(conta, "pago", False):
        return False

    status = normalizar_texto(getattr(conta, "status", None))
    return status not in ["PAGO", "RECEBIDO", "OK", "QUITADO", "BAIXADO", "CANCELADO", "CANCELADA"]


def linha_conta_pagar_dashboard(conta, origem="OBRIGAÇÃO"):
    return {
        "tipo": "DESPESA",
        "origem": origem,
        "data": formatar_data_json(getattr(conta, "data_pagamento", None) or getattr(conta, "data_vencimento", None)),
        "documento": str(getattr(conta, "numero_fatura", None) or "-"),
        "nome": str(getattr(conta, "fornecedor_funcionario", None) or getattr(conta, "fornecedor", None) or "-"),
        "categoria": str(getattr(conta, "categoria", None) or getattr(conta, "plano_contas", None) or "SEM CATEGORIA"),
        "setor": str(getattr(conta, "setor", None) or "-"),
        "vencimento": formatar_data_json(getattr(conta, "data_vencimento", None)),
        "pagamento": formatar_data_json(getattr(conta, "data_pagamento", None)),
        "status": str(getattr(conta, "status", None) or ("PAGO" if getattr(conta, "pago", False) else "PENDENTE")),
        "observacao": str(getattr(conta, "observacoes", None) or "-"),
        "valor": dinheiro(getattr(conta, "valor", 0)),
    }


def linha_conta_receber_dashboard(conta, origem="RECEBIMENTO"):
    return {
        "tipo": "RECEITA",
        "origem": origem,
        "data": formatar_data_json(getattr(conta, "data_pagamento", None) or getattr(conta, "data_vencimento", None)),
        "documento": str(getattr(conta, "numero_fatura", None) or "-"),
        "nome": str(getattr(conta, "cliente", None) or "-"),
        "categoria": str(getattr(conta, "categoria", None) or getattr(conta, "plano_contas", None) or "SEM CATEGORIA"),
        "setor": str(getattr(conta, "setor", None) or "-"),
        "vencimento": formatar_data_json(getattr(conta, "data_vencimento", None)),
        "pagamento": formatar_data_json(getattr(conta, "data_pagamento", None)),
        "status": str(getattr(conta, "status", None) or ("RECEBIDO" if getattr(conta, "pago", False) else "PENDENTE")),
        "observacao": str(getattr(conta, "observacoes", None) or "-"),
        "valor": valor_conta_receber(conta),
    }


def linha_lancamento_dashboard(lancamento):
    tipo = normalizar_texto(getattr(lancamento, "tipo", None)) or "-"
    return {
        "tipo": tipo,
        "origem": str(getattr(lancamento, "origem", None) or "MANUAL"),
        "data": formatar_data_json(getattr(lancamento, "data", None)),
        "documento": "-",
        "nome": str(getattr(lancamento, "cliente", None) or getattr(lancamento, "descricao", None) or "-"),
        "categoria": str(getattr(lancamento, "categoria", None) or "SEM CATEGORIA"),
        "setor": str(getattr(lancamento, "setor", None) or "GERAL"),
        "vencimento": formatar_data_json(getattr(lancamento, "data", None)),
        "pagamento": formatar_data_json(getattr(lancamento, "data", None)),
        "status": str(getattr(lancamento, "status", None) or "-"),
        "observacao": str(getattr(lancamento, "descricao", None) or getattr(lancamento, "observacoes", None) or "-"),
        "valor": dinheiro(getattr(lancamento, "valor", 0)),
    }


def resumo_lista_dashboard(itens):
    valores = [dinheiro(item.get("valor", 0)) for item in itens]
    total = sum(valores)
    quantidade = len(valores)

    return {
        "quantidade": quantidade,
        "total": total,
        "maior": max(valores) if valores else 0,
        "menor": min(valores) if valores else 0,
        "media": (total / quantidade) if quantidade else 0,
    }


def calcular_obrigacoes_dashboard(mes, ano):
    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)
    inicio_dia = primeiro_dia_mes(mes, ano)
    fim_dia = ultimo_dia_mes(mes, ano)
    hoje = date.today()

    # Referência usada para vencer hoje/próximos 7/até fim do mês.
    # Mês atual: usa hoje. Mês passado: considera o fim do mês. Mês futuro: considera o início do mês.
    if hoje.year == ano and hoje.month == mes:
        data_referencia = hoje
    elif date(ano, mes, 1) < date(hoje.year, hoje.month, 1):
        data_referencia = fim_dia
    else:
        data_referencia = inicio_dia

    ref_inicio_dt = datetime.combine(data_referencia, datetime.min.time())
    ref_fim_dt = datetime.combine(data_referencia, datetime.max.time())
    prox_7_dt = datetime.combine(data_referencia + timedelta(days=7), datetime.max.time())

    base_query = ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.pago == False
    )

    # Total do mês = somente contas com vencimento dentro do mês filtrado.
    a_pagar_mes = base_query.filter(
        ContaPagarImportada.data_vencimento >= inicio_dt,
        ContaPagarImportada.data_vencimento <= fim_dt
    ).all()

    # Atrasadas do mês = venceu dentro do mês filtrado e antes da data de referência.
    atrasadas_mes = base_query.filter(
        ContaPagarImportada.data_vencimento >= inicio_dt,
        ContaPagarImportada.data_vencimento < ref_inicio_dt
    ).all()

    vence_hoje = base_query.filter(
        ContaPagarImportada.data_vencimento >= ref_inicio_dt,
        ContaPagarImportada.data_vencimento <= ref_fim_dt
    ).all()

    proximos_7 = base_query.filter(
        ContaPagarImportada.data_vencimento > ref_fim_dt,
        ContaPagarImportada.data_vencimento <= prox_7_dt,
        ContaPagarImportada.data_vencimento <= fim_dt
    ).all()

    # Restante até o fim do mês = depois dos próximos 7 dias.
    # Assim os cartões não se sobrepõem:
    # Total do mês = Atrasadas do mês + Vence hoje + Próximos 7 dias + Restante até fim do mês.
    ate_fim_mes = base_query.filter(
        ContaPagarImportada.data_vencimento > prox_7_dt,
        ContaPagarImportada.data_vencimento <= fim_dt
    ).all()

    pagas_mes = ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.pago == True,
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt
    ).all()

    def total(lista):
        return sum(dinheiro(getattr(item, "valor", 0)) for item in lista)

    return {
        "referencia": data_referencia,
        "a_pagar_total": total(a_pagar_mes),
        "a_pagar_qtd": len(a_pagar_mes),
        "atrasadas": total(atrasadas_mes),
        "atrasadas_qtd": len(atrasadas_mes),
        "vence_hoje": total(vence_hoje),
        "vence_hoje_qtd": len(vence_hoje),
        "proximos_7": total(proximos_7),
        "proximos_7_qtd": len(proximos_7),
        "ate_fim_mes": total(ate_fim_mes),
        "ate_fim_mes_qtd": len(ate_fim_mes),
        "pagas_mes": total(pagas_mes),
        "pagas_mes_qtd": len(pagas_mes),
    }


def montar_itens_dashboard_kpi(tipo, mes, ano):
    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)
    inicio_dia = primeiro_dia_mes(mes, ano)
    fim_dia = ultimo_dia_mes(mes, ano)
    hoje = date.today()

    if hoje.year == ano and hoje.month == mes:
        data_referencia = hoje
    elif date(ano, mes, 1) < date(hoje.year, hoje.month, 1):
        data_referencia = fim_dia
    else:
        data_referencia = inicio_dia

    ref_inicio_dt = datetime.combine(data_referencia, datetime.min.time())
    ref_fim_dt = datetime.combine(data_referencia, datetime.max.time())
    prox_7_dt = datetime.combine(data_referencia + timedelta(days=7), datetime.max.time())

    itens = []
    titulo = "Detalhes"

    if tipo == "saidas":
        titulo = "Saídas realizadas no mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.pago == True,
            ContaPagarImportada.origem_importacao == "PAGAMENTO",
            ContaPagarImportada.data_pagamento >= inicio_dt,
            ContaPagarImportada.data_pagamento <= fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "PAGAMENTO") for c in contas)

        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.mes == mes,
            LancamentoFinanceiro.ano == ano
        ).all()
        for l in lancamentos:
            if status_cancelado(getattr(l, "status", None)) or not status_pago(getattr(l, "status", None)):
                continue
            if normalizar_texto(getattr(l, "tipo", None)) == "DESPESA":
                itens.append(linha_lancamento_dashboard(l))

    elif tipo == "entradas":
        titulo = "Entradas realizadas no mês"
        contas = ContaReceberImportada.query.filter(
            ContaReceberImportada.pago == True,
            ContaReceberImportada.origem_importacao == "RECEBIMENTO",
            ContaReceberImportada.data_pagamento >= inicio_dt,
            ContaReceberImportada.data_pagamento <= fim_dt
        ).all()
        itens.extend(linha_conta_receber_dashboard(c, "RECEBIMENTO") for c in contas)

        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.mes == mes,
            LancamentoFinanceiro.ano == ano
        ).all()
        for l in lancamentos:
            if status_cancelado(getattr(l, "status", None)) or not status_pago(getattr(l, "status", None)):
                continue
            if normalizar_texto(getattr(l, "tipo", None)) == "RECEITA":
                itens.append(linha_lancamento_dashboard(l))

    elif tipo in ["a_pagar", "obrigacoes_total"]:
        titulo = "Total a pagar do mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == False,
            ContaPagarImportada.data_vencimento >= inicio_dt,
            ContaPagarImportada.data_vencimento <= fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    elif tipo == "a_receber":
        titulo = "A receber até o fim do mês"
        contas = ContaReceberImportada.query.filter(
            ContaReceberImportada.pago == False,
            ContaReceberImportada.origem_importacao == "RECEBIMENTO",
            ContaReceberImportada.data_vencimento <= fim_dt
        ).all()
        itens.extend(linha_conta_receber_dashboard(c, "RECEBIMENTO") for c in contas)

        lancamentos = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.mes == mes,
            LancamentoFinanceiro.ano == ano
        ).all()
        for l in lancamentos:
            if status_cancelado(getattr(l, "status", None)):
                continue
            if normalizar_texto(getattr(l, "tipo", None)) == "RECEITA" and not status_pago(getattr(l, "status", None)):
                itens.append(linha_lancamento_dashboard(l))

    elif tipo == "obrigacoes_atrasadas":
        titulo = "Atrasadas do mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == False,
            ContaPagarImportada.data_vencimento >= inicio_dt,
            ContaPagarImportada.data_vencimento < ref_inicio_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    elif tipo == "obrigacoes_hoje":
        titulo = "Obrigações vencendo na referência"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == False,
            ContaPagarImportada.data_vencimento >= ref_inicio_dt,
            ContaPagarImportada.data_vencimento <= ref_fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    elif tipo == "obrigacoes_proximos_7":
        titulo = "Próximos 7 dias do mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == False,
            ContaPagarImportada.data_vencimento > ref_fim_dt,
            ContaPagarImportada.data_vencimento <= prox_7_dt,
            ContaPagarImportada.data_vencimento <= fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    elif tipo == "obrigacoes_ate_fim_mes":
        titulo = "Restante até o fim do mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == False,
            ContaPagarImportada.data_vencimento > prox_7_dt,
            ContaPagarImportada.data_vencimento <= fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    elif tipo == "obrigacoes_pagas_mes":
        titulo = "Obrigações pagas no mês"
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.pago == True,
            ContaPagarImportada.data_pagamento >= inicio_dt,
            ContaPagarImportada.data_pagamento <= fim_dt
        ).all()
        itens.extend(linha_conta_pagar_dashboard(c, "DESPESA_COMPLETA") for c in contas)

    itens = sorted(itens, key=lambda item: (item.get("vencimento") or "99/99/9999", item.get("nome") or ""))
    return titulo, itens


@gestao_bp.route("/api/dashboard-kpi-detalhes")
@gestao_required
def api_dashboard_kpi_detalhes():
    tipo = request.args.get("tipo", "").strip()
    mes = request.args.get("mes", type=int)
    ano = request.args.get("ano", type=int)

    if not tipo or not mes or not ano:
        return jsonify({
            "ok": False,
            "mensagem": "Tipo, mês e ano são obrigatórios.",
            "itens": [],
        }), 400

    titulo, itens = montar_itens_dashboard_kpi(tipo, mes, ano)
    resumo = resumo_lista_dashboard(itens)

    return jsonify({
        "ok": True,
        "tipo": tipo,
        "titulo": titulo,
        "mes": mes,
        "ano": ano,
        "resumo": resumo,
        "itens": itens,
    })


# =========================================================
# DASHBOARD FINANCEIRO
# =========================================================

@gestao_bp.route("/dashboard")
@gestao_required
def dashboard():

    hoje = datetime.now()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year

    fechamento = FechamentoMensal.query.filter_by(
        mes=mes,
        ano=ano
    ).first()

    saldo_inicial = obter_saldo_inicial_mes(mes, ano)

    totais = calcular_totais_financeiros(
        mes=mes,
        ano=ano,
        saldo_inicial=saldo_inicial
    )

    evolucao_mensal = calcular_evolucao_mensal(ano)

    inteligencia_financeira = calcular_inteligencia_financeira(
        totais=totais,
        evolucao_mensal=evolucao_mensal,
        mes=mes,
        ano=ano,
        saldo_inicial=saldo_inicial
    )

    obrigacoes_dashboard = calcular_obrigacoes_dashboard(mes, ano)

    return render_template(
        "gestao/dashboard.html",
        mes=mes,
        ano=ano,
        fechamento=fechamento,

        saldo_inicial=saldo_inicial,
        total_entradas=totais["total_entradas"],
        total_saidas=totais["total_saidas"],
        lucro_mes=totais["lucro_mes"],
        saldo_final=totais["saldo_final"],
        total_a_pagar=totais["total_a_pagar"],
        total_a_receber=totais["total_a_receber"],
        margem_operacional=totais["margem_operacional"],
        obrigacoes_dashboard=obrigacoes_dashboard,

        evolucao_mensal=evolucao_mensal,
        inteligencia_financeira=inteligencia_financeira,

        ranking_despesas=totais["ranking_despesas"],
        ranking_despesas_assistencia=totais["ranking_despesas_assistencia"],
        ranking_despesas_logistica=totais["ranking_despesas_logistica"],
        ranking_receitas=totais["ranking_receitas"],
        ranking_clientes=totais["ranking_clientes"]
    )


# =========================================================
# API - DETALHES DAS CONTAS POR CATEGORIA DE DESPESA
# =========================================================

def pegar_primeiro_valor(objeto, campos, padrao="-"):
    for campo in campos:
        valor = getattr(objeto, campo, None)

        if valor not in [None, ""]:
            return valor

    return padrao


def formatar_data_json(valor):
    if not valor:
        return "-"

    try:
        return valor.strftime("%d/%m/%Y")
    except Exception:
        return str(valor)


def categoria_item_despesa(objeto):
    categoria = getattr(objeto, "categoria", None)
    plano_contas = getattr(objeto, "plano_contas", None)

    return (categoria or plano_contas or "SEM CATEGORIA").strip().upper()


@gestao_bp.route("/api/despesas-categoria")
@gestao_required
def api_despesas_categoria():
    categoria = request.args.get("categoria", "").strip().upper()
    mes = request.args.get("mes", type=int)
    ano = request.args.get("ano", type=int)

    if not categoria or not mes or not ano:
        return jsonify({
            "ok": False,
            "mensagem": "Categoria, mês e ano são obrigatórios.",
            "categoria": categoria,
            "total": 0,
            "total_pago": 0,
            "total_aberto": 0,
            "quantidade": 0,
            "itens": []
        }), 400

    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)

    itens = []
    total = 0
    total_pago = 0
    total_aberto = 0

    contas_pagar = ContaPagarImportada.query.filter(
        or_(
            and_(
                ContaPagarImportada.pago == True,
                ContaPagarImportada.origem_importacao == "PAGAMENTO",
                ContaPagarImportada.data_pagamento >= inicio_dt,
                ContaPagarImportada.data_pagamento <= fim_dt
            ),
            and_(
                ContaPagarImportada.pago == False,
                ContaPagarImportada.origem_importacao == "PAGAMENTO",
                ContaPagarImportada.data_vencimento <= fim_dt
            )
        )
    ).all()

    for conta in contas_pagar:
        categoria_conta = categoria_item_despesa(conta)

        if categoria_conta != categoria:
            continue

        valor = dinheiro(getattr(conta, "valor", 0))
        pago = bool(conta_esta_paga(conta))

        total += valor

        if pago:
            total_pago += valor
            status = "PAGO"
            data = getattr(conta, "data_pagamento", None)
        else:
            total_aberto += valor
            status = "EM ABERTO"
            data = getattr(conta, "data_vencimento", None)

        conta_nome = (
            getattr(conta, "plano_contas", None)
            or getattr(conta, "categoria", None)
            or categoria
        )

        observacao = getattr(conta, "observacoes", None) or "-"
        setor = getattr(conta, "setor", None) or "-"

        itens.append({
            "origem": "IMPORTADA",
            "tipo": "DESPESA",
            "data": formatar_data_json(data),
            "conta": str(conta_nome),
            "fornecedor_funcionario": str(getattr(conta, "fornecedor_funcionario", None) or "-"),
            "numero_fatura": str(getattr(conta, "numero_fatura", None) or "-"),
            "observacao": str(observacao),
            "setor": str(setor),
            "status": status,
            "valor": valor
        })

    lancamentos = LancamentoFinanceiro.query.filter_by(
        mes=mes,
        ano=ano
    ).all()

    for lancamento in lancamentos:
        if status_cancelado(getattr(lancamento, "status", None)):
            continue

        tipo = normalizar_texto(getattr(lancamento, "tipo", None))

        if tipo != "DESPESA":
            continue

        categoria_lancamento = (
            getattr(lancamento, "categoria", None) or "SEM CATEGORIA"
        ).strip().upper()

        if categoria_lancamento != categoria:
            continue

        valor = dinheiro(getattr(lancamento, "valor", 0))
        pago = status_pago(getattr(lancamento, "status", None))

        total += valor

        if pago:
            total_pago += valor
            status = getattr(lancamento, "status", None) or "PAGO"
        else:
            total_aberto += valor
            status = getattr(lancamento, "status", None) or "EM ABERTO"

        conta_nome = (
            getattr(lancamento, "subcategoria", None)
            or getattr(lancamento, "categoria", None)
            or categoria_lancamento
        )

        observacao = (
            getattr(lancamento, "observacoes", None)
            or getattr(lancamento, "observacao", None)
            or getattr(lancamento, "descricao", None)
            or "-"
        )

        itens.append({
            "origem": "MANUAL",
            "tipo": "DESPESA",
            "data": formatar_data_json(getattr(lancamento, "data", None)),
            "conta": str(conta_nome),
            "fornecedor_funcionario": getattr(lancamento, "cliente", None) or "-",
            "numero_fatura": "-",
            "observacao": str(observacao),
            "setor": getattr(lancamento, "setor", None) or "-",
            "status": status,
            "valor": valor
        })

    itens = sorted(
        itens,
        key=lambda item: item["valor"],
        reverse=True
    )

    return jsonify({
        "ok": True,
        "categoria": categoria,
        "mes": mes,
        "ano": ano,
        "total": total,
        "total_pago": total_pago,
        "total_aberto": total_aberto,
        "quantidade": len(itens),
        "itens": itens
    })


# =========================================================
# FECHAMENTO MENSAL
# =========================================================

@gestao_bp.route("/fechamento", methods=["GET", "POST"])
@gestao_required
def fechamento():

    hoje = datetime.now()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year

    fechamento_existente = FechamentoMensal.query.filter_by(
        mes=mes,
        ano=ano
    ).first()

    saldo_inicial = obter_saldo_inicial_mes(mes, ano)
    saldo_inicial_automatico = obter_saldo_inicial_automatico(mes, ano)

    if request.method == "POST":
        mes = request.form.get("mes", type=int)
        ano = request.form.get("ano", type=int)

        saldo_inicial = dinheiro(request.form.get("saldo_inicial"))

        totais = calcular_totais_financeiros(
            mes=mes,
            ano=ano,
            saldo_inicial=saldo_inicial
        )

        fechamento_existente = FechamentoMensal.query.filter_by(
            mes=mes,
            ano=ano
        ).first()

        if fechamento_existente:
            fechamento_existente.saldo_inicial = saldo_inicial
            fechamento_existente.total_entradas = totais["total_entradas"]
            fechamento_existente.total_saidas = totais["total_saidas"]
            fechamento_existente.lucro_mes = totais["lucro_mes"]
            fechamento_existente.saldo_final = totais["saldo_final"]
            fechamento_existente.total_a_pagar = totais["total_a_pagar"]
            fechamento_existente.total_a_receber = totais["total_a_receber"]
            fechamento_existente.margem_operacional = totais["margem_operacional"]
            fechamento_existente.fechado = True
        else:
            fechamento_existente = FechamentoMensal(
                mes=mes,
                ano=ano,
                saldo_inicial=saldo_inicial,
                total_entradas=totais["total_entradas"],
                total_saidas=totais["total_saidas"],
                lucro_mes=totais["lucro_mes"],
                saldo_final=totais["saldo_final"],
                total_a_pagar=totais["total_a_pagar"],
                total_a_receber=totais["total_a_receber"],
                margem_operacional=totais["margem_operacional"],
                fechado=True
            )

            db.session.add(fechamento_existente)

        db.session.commit()

        flash("Fechamento mensal salvo com sucesso!", "success")
        return redirect(f"/gestao/fechamento?mes={mes}&ano={ano}")

    totais = calcular_totais_financeiros(
        mes=mes,
        ano=ano,
        saldo_inicial=saldo_inicial
    )

    return render_template(
        "gestao/fechamento.html",
        mes=mes,
        ano=ano,
        fechamento=fechamento_existente,
        saldo_inicial=saldo_inicial,
        saldo_inicial_automatico=saldo_inicial_automatico,
        total_entradas=totais["total_entradas"],
        total_saidas=totais["total_saidas"],
        lucro_mes=totais["lucro_mes"],
        saldo_final=totais["saldo_final"],
        total_a_pagar=totais["total_a_pagar"],
        total_a_receber=totais["total_a_receber"],
        margem_operacional=totais["margem_operacional"]
    )



@gestao_bp.route("/fechamento/excluir", methods=["POST"])
@gestao_required
def excluir_fechamento():

    mes = request.form.get("mes", type=int)
    ano = request.form.get("ano", type=int)

    if not mes or not ano:
        flash("Informe mês e ano válidos para excluir o fechamento.", "danger")
        return redirect("/gestao/fechamento")

    fechamento_existente = FechamentoMensal.query.filter_by(
        mes=mes,
        ano=ano
    ).first()

    if not fechamento_existente:
        flash("Nenhum fechamento encontrado para excluir neste mês.", "warning")
        return redirect(f"/gestao/fechamento?mes={mes}&ano={ano}")

    try:
        db.session.delete(fechamento_existente)
        db.session.commit()

        flash("Fechamento excluído com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir fechamento: {str(e)}", "danger")

    return redirect(f"/gestao/fechamento?mes={mes}&ano={ano}")


# =========================================================
# LANÇAMENTOS MANUAIS
# =========================================================

@gestao_bp.route("/lancamentos")
@gestao_required
def lancamentos():

    query = LancamentoFinanceiro.query

    mes = request.args.get("mes", type=int)
    ano = request.args.get("ano", type=int)
    tipo = request.args.get("tipo")
    status = request.args.get("status")
    categoria = request.args.get("categoria")

    if mes:
        query = query.filter(LancamentoFinanceiro.mes == mes)

    if ano:
        query = query.filter(LancamentoFinanceiro.ano == ano)

    if tipo:
        query = query.filter(LancamentoFinanceiro.tipo == tipo)

    if status:
        query = query.filter(LancamentoFinanceiro.status == status)

    if categoria:
        query = query.filter(
            LancamentoFinanceiro.categoria.ilike(f"%{categoria}%")
        )

    lancamentos = query.order_by(
        LancamentoFinanceiro.data.desc(),
        LancamentoFinanceiro.id.desc()
    ).all()

    return render_template(
        "gestao/lancamentos.html",
        lancamentos=lancamentos,
        mes=mes,
        ano=ano,
        tipo=tipo,
        status=status,
        categoria=categoria
    )


@gestao_bp.route("/lancamentos/novo", methods=["GET", "POST"])
@gestao_required
def novo_lancamento():

    if request.method == "POST":

        data = parse_data(request.form.get("data"))

        if not data:
            flash("Informe uma data válida.", "danger")
            return redirect("/gestao/lancamentos/novo")

        lancamento = LancamentoFinanceiro(
            data=data,
            tipo=normalizar_texto(request.form.get("tipo")),
            categoria=normalizar_texto(request.form.get("categoria")),
            subcategoria=normalizar_texto(request.form.get("subcategoria")),
            setor=normalizar_texto(request.form.get("setor")),
            cliente=normalizar_texto(request.form.get("cliente")),
            descricao=request.form.get("descricao"),
            valor=request.form.get("valor") or 0,
            status=normalizar_texto(request.form.get("status")),
            origem=normalizar_texto(request.form.get("origem")),
            recorrente=True if request.form.get("recorrente") == "on" else False,
            mes=data.month,
            ano=data.year
        )

        db.session.add(lancamento)
        db.session.commit()

        flash("Lançamento criado com sucesso!", "success")
        return redirect("/gestao/lancamentos")

    return render_template(
        "gestao/form_lancamento.html",
        lancamento=None
    )


@gestao_bp.route("/lancamentos/editar/<int:id>", methods=["GET", "POST"])
@gestao_required
def editar_lancamento(id):

    lancamento = LancamentoFinanceiro.query.get_or_404(id)

    if request.method == "POST":

        data = parse_data(request.form.get("data"))

        if not data:
            flash("Informe uma data válida.", "danger")
            return redirect(f"/gestao/lancamentos/editar/{id}")

        lancamento.data = data
        lancamento.tipo = normalizar_texto(request.form.get("tipo"))
        lancamento.categoria = normalizar_texto(request.form.get("categoria"))
        lancamento.subcategoria = normalizar_texto(request.form.get("subcategoria"))
        lancamento.setor = normalizar_texto(request.form.get("setor"))
        lancamento.cliente = normalizar_texto(request.form.get("cliente"))
        lancamento.descricao = request.form.get("descricao")
        lancamento.valor = request.form.get("valor") or 0
        lancamento.status = normalizar_texto(request.form.get("status"))
        lancamento.origem = normalizar_texto(request.form.get("origem"))
        lancamento.recorrente = True if request.form.get("recorrente") == "on" else False
        lancamento.mes = data.month
        lancamento.ano = data.year

        db.session.commit()

        flash("Lançamento atualizado com sucesso!", "success")
        return redirect("/gestao/lancamentos")

    return render_template(
        "gestao/form_lancamento.html",
        lancamento=lancamento
    )


@gestao_bp.route("/lancamentos/excluir/<int:id>")
@gestao_required
def excluir_lancamento(id):

    lancamento = LancamentoFinanceiro.query.get_or_404(id)

    db.session.delete(lancamento)
    db.session.commit()

    flash("Lançamento excluído com sucesso!", "success")
    return redirect("/gestao/lancamentos")


# =========================================================
# FLUXO DIÁRIO FINANCEIRO
# SOMENTE MOVIMENTAÇÃO REALIZADA
# =========================================================

def calcular_totais_reais_periodo(data_inicio, data_fim):
    total_despesas = 0
    total_receitas = 0

    inicio_dt = datetime.combine(data_inicio, datetime.min.time())
    fim_dt = datetime.combine(data_fim, datetime.max.time())

    contas_pagar = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt
    ).all()

    contas_receber = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt
    ).all()

    for conta in contas_pagar:
        total_despesas += dinheiro(getattr(conta, "valor", 0))

    for conta in contas_receber:
        total_receitas += valor_conta_receber(conta)

    lancamentos = LancamentoFinanceiro.query.filter(
        LancamentoFinanceiro.data >= data_inicio,
        LancamentoFinanceiro.data <= data_fim
    ).all()

    for lancamento in lancamentos:
        if status_cancelado(getattr(lancamento, "status", None)):
            continue

        if not status_pago(getattr(lancamento, "status", None)):
            continue

        tipo = normalizar_texto(getattr(lancamento, "tipo", None))
        valor = dinheiro(getattr(lancamento, "valor", 0))

        if tipo == "DESPESA":
            total_despesas += valor
        elif tipo == "RECEITA":
            total_receitas += valor

    return total_despesas, total_receitas


@gestao_bp.route("/fluxo-diario")
@gestao_required
def fluxo_diario():

    hoje = datetime.now()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year
    visao = request.args.get("visao") or "mes_ano"

    if visao not in ["mes", "ano", "mes_ano"]:
        visao = "mes_ano"

    primeiro_dia = primeiro_dia_mes(mes, ano)
    ultimo_dia = ultimo_dia_mes(mes, ano)
    ultimo_dia_numero = ultimo_dia.day

    inicio_dt = datetime.combine(primeiro_dia, datetime.min.time())
    fim_dt = datetime.combine(ultimo_dia, datetime.max.time())

    # =====================================================
    # TRANSPORTE DOS MESES ANTERIORES
    # Tudo que realmente entrou/saiu de 01/01 até o dia
    # anterior ao mês filtrado
    # =====================================================
    if mes > 1:
        data_inicio_transporte = date(ano, 1, 1)
        data_fim_transporte = primeiro_dia - timedelta(days=1)

        despesas_transporte, receitas_transporte = calcular_totais_reais_periodo(
            data_inicio=data_inicio_transporte,
            data_fim=data_fim_transporte
        )
    else:
        despesas_transporte = 0
        receitas_transporte = 0

    saldo_transporte = dinheiro(receitas_transporte) - dinheiro(despesas_transporte)

    # =====================================================
    # CONTAS IMPORTADAS REALIZADAS NO MÊS
    # Fluxo usa data_pagamento, não data_vencimento
    # =====================================================
    contas_pagar_mes = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt
    ).all()

    contas_receber_mes = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt
    ).all()

    lancamentos_mes = LancamentoFinanceiro.query.filter(
        LancamentoFinanceiro.data >= primeiro_dia,
        LancamentoFinanceiro.data <= ultimo_dia
    ).all()

    despesas_por_dia = {}
    receitas_por_dia = {}

    total_despesas_mes = 0
    total_receitas_mes = 0

    for conta in contas_pagar_mes:
        valor = dinheiro(getattr(conta, "valor", 0))
        data_mov = data_movimento_fluxo(conta)

        if not data_mov:
            continue

        total_despesas_mes += valor
        despesas_por_dia[data_mov] = despesas_por_dia.get(data_mov, 0) + valor

    for conta in contas_receber_mes:
        valor = valor_conta_receber(conta)
        data_mov = data_movimento_fluxo(conta)

        if not data_mov:
            continue

        total_receitas_mes += valor
        receitas_por_dia[data_mov] = receitas_por_dia.get(data_mov, 0) + valor

    for lancamento in lancamentos_mes:
        if status_cancelado(getattr(lancamento, "status", None)):
            continue

        if not status_pago(getattr(lancamento, "status", None)):
            continue

        data_mov = data_para_date(getattr(lancamento, "data", None))

        if not data_mov:
            continue

        tipo = normalizar_texto(getattr(lancamento, "tipo", None))
        valor = dinheiro(getattr(lancamento, "valor", 0))

        if tipo == "DESPESA":
            total_despesas_mes += valor
            despesas_por_dia[data_mov] = despesas_por_dia.get(data_mov, 0) + valor
        elif tipo == "RECEITA":
            total_receitas_mes += valor
            receitas_por_dia[data_mov] = receitas_por_dia.get(data_mov, 0) + valor

    saldo_mes = dinheiro(total_receitas_mes) - dinheiro(total_despesas_mes)

    total_receitas_ano = dinheiro(receitas_transporte) + dinheiro(total_receitas_mes)
    total_despesas_ano = dinheiro(despesas_transporte) + dinheiro(total_despesas_mes)
    saldo_ano = dinheiro(total_receitas_ano) - dinheiro(total_despesas_ano)

    linhas = []
    acumulado = dinheiro(saldo_transporte)

    linhas.append({
        "tipo": "transporte",
        "dia": "** TRANSPORTE MESES ANTERIORES **",
        "data": None,
        "despesas": dinheiro(despesas_transporte),
        "receitas": dinheiro(receitas_transporte),
        "saldo": dinheiro(saldo_transporte),
        "acumulado": dinheiro(saldo_transporte)
    })

    for dia in range(1, ultimo_dia_numero + 1):

        data_ref = date(ano, mes, dia)

        despesas = dinheiro(despesas_por_dia.get(data_ref, 0))
        receitas = dinheiro(receitas_por_dia.get(data_ref, 0))
        saldo_dia = dinheiro(receitas) - dinheiro(despesas)

        acumulado += saldo_dia

        linhas.append({
            "tipo": "dia",
            "dia": f"{dia:02d}/{mes:02d} - {nome_dia_semana(data_ref)}",
            "data": data_ref,
            "despesas": despesas,
            "receitas": receitas,
            "saldo": saldo_dia,
            "acumulado": acumulado
        })

    return render_template(
        "gestao/fluxo_diario.html",
        mes=mes,
        ano=ano,
        visao=visao,
        linhas=linhas,

        total_despesas_mes=total_despesas_mes,
        total_receitas_mes=total_receitas_mes,
        saldo_mes=saldo_mes,

        despesas_transporte=despesas_transporte,
        receitas_transporte=receitas_transporte,
        saldo_transporte=saldo_transporte,

        total_despesas_ano=total_despesas_ano,
        total_receitas_ano=total_receitas_ano,
        saldo_ano=saldo_ano
    )


# =========================================================
# API - DETALHES DO FLUXO POR DIA
# =========================================================

def moeda_json(valor):
    return round(dinheiro(valor), 2)


def texto_json(valor, padrao="-"):
    if valor is None:
        return padrao

    valor = str(valor).strip()

    return valor if valor else padrao


@gestao_bp.route("/api/fluxo-diario/dia")
@gestao_required
def api_fluxo_diario_dia():

    data_str = request.args.get("data", "").strip()

    try:
        data_ref = datetime.strptime(data_str, "%Y-%m-%d").date()
    except Exception:
        return jsonify({
            "ok": False,
            "mensagem": "Data inválida. Use o formato YYYY-MM-DD."
        }), 400

    inicio_dt = datetime.combine(data_ref, datetime.min.time())
    fim_dt = datetime.combine(data_ref, datetime.max.time())

    despesas = []
    receitas = []
    manuais = []

    total_despesas = 0
    total_receitas = 0

    contas_pagar = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt
    ).order_by(
        ContaPagarImportada.valor.desc(),
        ContaPagarImportada.id.desc()
    ).all()

    for conta in contas_pagar:
        valor = dinheiro(getattr(conta, "valor", 0))
        total_despesas += valor

        despesas.append({
            "id": conta.id,
            "origem": "IMPORTADA",
            "tipo": "DESPESA",
            "titulo": texto_json(conta.fornecedor_funcionario, "Fornecedor não informado"),
            "fornecedor_funcionario": texto_json(conta.fornecedor_funcionario),
            "numero_fatura": texto_json(conta.numero_fatura),
            "plano_contas": texto_json(conta.plano_contas),
            "categoria": texto_json(conta.categoria),
            "setor": texto_json(conta.setor),
            "data_documento": formatar_data_json(conta.data_documento),
            "data_vencimento": formatar_data_json(conta.data_vencimento),
            "data_pagamento": formatar_data_json(conta.data_pagamento),
            "valor": moeda_json(valor),
            "status": texto_json(conta.status, "PAGO"),
            "observacoes": texto_json(conta.observacoes)
        })

    contas_receber = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt
    ).order_by(
        ContaReceberImportada.total.desc(),
        ContaReceberImportada.valor.desc(),
        ContaReceberImportada.id.desc()
    ).all()

    for conta in contas_receber:
        valor_base = dinheiro(getattr(conta, "valor", 0))
        juros = dinheiro(getattr(conta, "juros", 0))
        total = valor_conta_receber(conta)
        total_receitas += total

        receitas.append({
            "id": conta.id,
            "origem": "IMPORTADA",
            "tipo": "RECEITA",
            "titulo": texto_json(conta.cliente, "Cliente não informado"),
            "cliente": texto_json(conta.cliente),
            "numero_fatura": texto_json(conta.numero_fatura),
            "plano_contas": texto_json(conta.plano_contas),
            "categoria": texto_json(conta.categoria),
            "setor": texto_json(conta.setor),
            "cobranca": texto_json(conta.cobranca),
            "data_documento": formatar_data_json(conta.data_documento),
            "data_vencimento": formatar_data_json(conta.data_vencimento),
            "data_pagamento": formatar_data_json(conta.data_pagamento),
            "valor": moeda_json(valor_base),
            "juros": moeda_json(juros),
            "total": moeda_json(total),
            "status": texto_json(conta.status, "RECEBIDO"),
            "observacoes": texto_json(conta.observacoes)
        })

    lancamentos = LancamentoFinanceiro.query.filter(
        LancamentoFinanceiro.data == data_ref
    ).order_by(
        LancamentoFinanceiro.valor.desc(),
        LancamentoFinanceiro.id.desc()
    ).all()

    for lancamento in lancamentos:
        if status_cancelado(getattr(lancamento, "status", None)):
            continue

        if not status_pago(getattr(lancamento, "status", None)):
            continue

        tipo = normalizar_texto(getattr(lancamento, "tipo", None))
        valor = dinheiro(getattr(lancamento, "valor", 0))

        if tipo == "RECEITA":
            total_receitas += valor
            tipo_visual = "RECEITA"
        elif tipo == "DESPESA":
            total_despesas += valor
            tipo_visual = "DESPESA"
        else:
            continue

        manuais.append({
            "id": lancamento.id,
            "origem": "MANUAL",
            "tipo": tipo_visual,
            "titulo": texto_json(
                getattr(lancamento, "descricao", None)
                or getattr(lancamento, "subcategoria", None)
                or getattr(lancamento, "categoria", None),
                "Lançamento manual"
            ),
            "categoria": texto_json(getattr(lancamento, "categoria", None)),
            "subcategoria": texto_json(getattr(lancamento, "subcategoria", None)),
            "setor": texto_json(getattr(lancamento, "setor", None)),
            "cliente": texto_json(getattr(lancamento, "cliente", None)),
            "descricao": texto_json(getattr(lancamento, "descricao", None)),
            "origem_lancamento": texto_json(getattr(lancamento, "origem", None)),
            "data": formatar_data_json(getattr(lancamento, "data", None)),
            "valor": moeda_json(valor),
            "status": texto_json(getattr(lancamento, "status", None))
        })

    saldo = dinheiro(total_receitas) - dinheiro(total_despesas)

    return jsonify({
        "ok": True,
        "data": data_ref.strftime("%d/%m/%Y"),
        "data_iso": data_ref.strftime("%Y-%m-%d"),
        "dia_semana": nome_dia_semana(data_ref),
        "resumo": {
            "total_receitas": moeda_json(total_receitas),
            "total_despesas": moeda_json(total_despesas),
            "saldo": moeda_json(saldo),
            "quantidade": len(receitas) + len(despesas) + len(manuais)
        },
        "receitas": receitas,
        "despesas": despesas,
        "manuais": manuais
    })


# =========================================================
# RADAR DE PAGAMENTOS
# SOMENTE DESPESAS - PAGAS E PENDENTES
# =========================================================

def status_visual_conta_pagar(conta, hoje):
    data_vencimento = data_para_date(conta.data_vencimento)

    if conta_esta_paga(conta):
        return {
            "label": "PAGO",
            "classe": "pago",
            "grupo": "pagas"
        }

    if not data_vencimento:
        return {
            "label": "SEM DATA",
            "classe": "sem-data",
            "grupo": "sem_data"
        }

    if data_vencimento < hoje:
        return {
            "label": "ATRASADA",
            "classe": "atrasada",
            "grupo": "atrasadas"
        }

    if data_vencimento == hoje:
        return {
            "label": "VENCE HOJE",
            "classe": "vence-hoje",
            "grupo": "vence_hoje"
        }

    if data_vencimento <= hoje + timedelta(days=7):
        return {
            "label": "PRÓXIMOS 7 DIAS",
            "classe": "proxima",
            "grupo": "proximas"
        }

    return {
        "label": "FUTURA",
        "classe": "futura",
        "grupo": "futuras"
    }


def montar_item_radar_pagar(conta, hoje):
    status_info = status_visual_conta_pagar(conta, hoje)

    data_vencimento = data_para_date(conta.data_vencimento)
    data_pagamento = data_para_date(conta.data_pagamento)

    dias = None

    if data_vencimento and not conta_esta_paga(conta):
        dias = (data_vencimento - hoje).days

    return {
        "id": conta.id,
        "numero_fatura": conta.numero_fatura or "-",
        "fornecedor_funcionario": conta.fornecedor_funcionario or "-",
        "plano_contas": conta.plano_contas or "-",
        "categoria": conta.categoria or "-",
        "setor": conta.setor or "-",
        "data_vencimento": data_vencimento,
        "data_pagamento": data_pagamento,
        "valor": dinheiro(conta.valor),
        "pago": bool(conta_esta_paga(conta)),
        "status": conta.status or "-",
        "observacoes": conta.observacoes or "-",
        "status_label": status_info["label"],
        "status_classe": status_info["classe"],
        "grupo": status_info["grupo"],
        "dias": dias
    }


def somar_itens_radar(lista):
    return sum(dinheiro(item.get("valor")) for item in lista)


@gestao_bp.route("/radar-pagamentos")
@gestao_required
def radar_pagamentos():

    hoje = date.today()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year
    setor = request.args.get("setor", "").strip().upper()
    filtro = request.args.get("filtro", "todos").strip().lower()

    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)

    # =====================================================
    # RADAR USA VENCIMENTO
    # Mostra:
    # - contas pagas com vencimento dentro do mês filtrado
    # - contas abertas vencidas até o fim do mês filtrado
    #   incluindo atrasadas de meses anteriores
    # - contas sem data que foram importadas no mês/ano
    # =====================================================
    query = ContaPagarImportada.query.filter(
        or_(
            and_(
                ContaPagarImportada.pago == True,
                ContaPagarImportada.data_vencimento >= inicio_dt,
                ContaPagarImportada.data_vencimento <= fim_dt
            ),
            and_(
                ContaPagarImportada.pago == False,
                ContaPagarImportada.origem_importacao == "PAGAMENTO",
                ContaPagarImportada.data_vencimento <= fim_dt
            ),
            and_(
                ContaPagarImportada.data_vencimento == None,
                ContaPagarImportada.mes == mes,
                ContaPagarImportada.ano == ano
            )
        )
    )

    if setor:
        query = query.filter(
            ContaPagarImportada.setor.ilike(f"%{setor}%")
        )

    contas = query.order_by(
        ContaPagarImportada.data_vencimento.asc().nullslast(),
        ContaPagarImportada.id.desc()
    ).all()

    itens = [
        montar_item_radar_pagar(conta, hoje)
        for conta in contas
    ]

    atrasadas = [item for item in itens if item["grupo"] == "atrasadas"]
    vence_hoje = [item for item in itens if item["grupo"] == "vence_hoje"]
    proximas = [item for item in itens if item["grupo"] == "proximas"]
    futuras = [item for item in itens if item["grupo"] == "futuras"]
    pagas = [item for item in itens if item["grupo"] == "pagas"]
    sem_data = [item for item in itens if item["grupo"] == "sem_data"]
    abertas = [item for item in itens if not item["pago"]]

    if filtro == "pagas":
        itens_filtrados = pagas
    elif filtro == "abertas":
        itens_filtrados = abertas
    elif filtro == "atrasadas":
        itens_filtrados = atrasadas
    elif filtro == "hoje":
        itens_filtrados = vence_hoje
    elif filtro == "proximas":
        itens_filtrados = proximas
    elif filtro == "futuras":
        itens_filtrados = futuras
    elif filtro == "sem_data":
        itens_filtrados = sem_data
    else:
        itens_filtrados = itens

    total_geral = somar_itens_radar(itens)
    total_pagas = somar_itens_radar(pagas)
    total_abertas = somar_itens_radar(abertas)
    total_atrasadas = somar_itens_radar(atrasadas)
    total_hoje = somar_itens_radar(vence_hoje)
    total_proximas = somar_itens_radar(proximas)
    total_futuras = somar_itens_radar(futuras)

    return render_template(
        "gestao/radar_pagamentos.html",
        hoje=hoje,
        mes=mes,
        ano=ano,
        setor=setor,
        filtro=filtro,

        itens=itens,
        itens_filtrados=itens_filtrados,

        atrasadas=atrasadas,
        vence_hoje=vence_hoje,
        proximas=proximas,
        futuras=futuras,
        pagas=pagas,
        abertas=abertas,
        sem_data=sem_data,

        total_geral=total_geral,
        total_pagas=total_pagas,
        total_abertas=total_abertas,
        total_atrasadas=total_atrasadas,
        total_hoje=total_hoje,
        total_proximas=total_proximas,
        total_futuras=total_futuras
    )


@gestao_bp.route("/radar-pagamentos/pagar/<int:id>", methods=["POST"])
@gestao_required
def radar_marcar_pago(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    data_pagamento_str = request.form.get("data_pagamento")
    observacao_extra = request.form.get("observacao", "").strip()

    data_pagamento = parse_data(data_pagamento_str)

    if not data_pagamento:
        data_pagamento = date.today()

    conta.pago = True
    conta.status = "PAGO"
    conta.data_pagamento = datetime.combine(
        data_pagamento,
        datetime.min.time()
    )

    if observacao_extra:
        observacao_antiga = conta.observacoes or ""
        conta.observacoes = f"{observacao_antiga}\nPagamento: {observacao_extra}".strip()

    db.session.commit()

    flash("Conta marcada como paga com sucesso!", "success")

    return redirect(request.referrer or "/gestao/radar-pagamentos")


@gestao_bp.route("/radar-pagamentos/adiar/<int:id>", methods=["POST"])
@gestao_required
def radar_adiar_pagamento(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    nova_data_str = request.form.get("nova_data_vencimento")
    motivo = request.form.get("motivo", "").strip()

    nova_data = parse_data(nova_data_str)

    if not nova_data:
        flash("Informe uma nova data válida.", "danger")
        return redirect(request.referrer or "/gestao/radar-pagamentos")

    data_antiga = data_para_date(conta.data_vencimento)

    conta.data_vencimento = datetime.combine(
        nova_data,
        datetime.min.time()
    )

    # Mantém mes/ano sincronizado para listagens administrativas,
    # mas as telas financeiras agora usam data_vencimento/data_pagamento.
    conta.mes = nova_data.month
    conta.ano = nova_data.year

    if not conta_esta_paga(conta):
        conta.status = "ADIADO"

    observacao_antiga = conta.observacoes or ""

    texto_adiamento = (
        f"Adiamento: vencimento alterado de "
        f"{data_antiga.strftime('%d/%m/%Y') if data_antiga else '-'} "
        f"para {nova_data.strftime('%d/%m/%Y')}."
    )

    if motivo:
        texto_adiamento += f" Motivo: {motivo}"

    conta.observacoes = f"{observacao_antiga}\n{texto_adiamento}".strip()

    db.session.commit()

    flash("Pagamento adiado com sucesso!", "success")

    return redirect(request.referrer or "/gestao/radar-pagamentos")

# =========================================================
# RADAR FINANCEIRO INTELIGENTE
# DÍVIDA HERDADA AGRUPADA + HISTÓRICO POR MÊS
# =========================================================

def radar_moeda(valor):
    valor = dinheiro(valor)
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def radar_nome_mes(mes):
    nomes = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
    }
    return nomes.get(int(mes or 0), str(mes or ""))


def radar_nome_mes_curto(mes):
    nomes = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }
    return nomes.get(int(mes or 0), str(mes or ""))


def radar_data_input(valor):
    data_ref = data_para_date(valor)
    if not data_ref:
        return ""
    return data_ref.strftime("%Y-%m-%d")


def radar_data_br(valor):
    data_ref = data_para_date(valor)
    if not data_ref:
        return "-"
    return data_ref.strftime("%d/%m/%Y")


def radar_ids_limpos_formulario(nome_campo="contas_ids"):
    ids = request.form.getlist(nome_campo)

    if not ids:
        ids = request.form.getlist(f"{nome_campo}[]")

    ids_limpos = []

    for item in ids:
        try:
            ids_limpos.append(int(item))
        except Exception:
            pass

    return list(dict.fromkeys(ids_limpos))


def radar_intervalo_dia(data_ref):
    return (
        datetime.combine(data_ref, datetime.min.time()),
        datetime.combine(data_ref, datetime.max.time())
    )


def radar_transportada_ja_existe(conta, nova_data, mes_destino, ano_destino):
    """
    Evita duplicidade no transporte do Radar.

    Como o Radar atual usa ContaPagarImportada, a trava compara os campos
    principais da conta transportada no mês destino.
    """
    if not conta or not nova_data:
        return None

    inicio, fim = radar_intervalo_dia(nova_data)

    query = ContaPagarImportada.query.filter(
        ContaPagarImportada.mes == mes_destino,
        ContaPagarImportada.ano == ano_destino,
        ContaPagarImportada.data_vencimento >= inicio,
        ContaPagarImportada.data_vencimento <= fim,
        ContaPagarImportada.valor == getattr(conta, "valor", None),
        ContaPagarImportada.fornecedor_funcionario == getattr(conta, "fornecedor_funcionario", None),
        ContaPagarImportada.plano_contas == getattr(conta, "plano_contas", None),
        ContaPagarImportada.categoria == getattr(conta, "categoria", None),
        ContaPagarImportada.setor == getattr(conta, "setor", None),
        ContaPagarImportada.origem_importacao == "PAGAMENTO"
    )

    if getattr(conta, "id", None):
        query = query.filter(ContaPagarImportada.id != conta.id)

    return query.first()


def radar_competencia_label(data_ref):
    data_ref = data_para_date(data_ref)
    if not data_ref:
        return "-"
    return f"{radar_nome_mes(data_ref.month)}/{data_ref.year}"


def radar_texto(valor, padrao="-"):
    if valor is None:
        return padrao
    valor = str(valor).strip()
    return valor if valor else padrao


def radar_descricao_conta(conta):
    return (
        radar_texto(getattr(conta, "plano_contas", None), "")
        or radar_texto(getattr(conta, "categoria", None), "")
        or radar_texto(getattr(conta, "fornecedor_funcionario", None), "")
        or radar_texto(getattr(conta, "numero_fatura", None), "")
        or f"Conta #{conta.id}"
    )


def radar_fornecedor_conta(conta):
    return radar_texto(getattr(conta, "fornecedor_funcionario", None), "-")


def radar_categoria_conta(conta):
    return radar_texto(getattr(conta, "categoria", None), "-")


def radar_setor_conta(conta):
    return radar_texto(getattr(conta, "setor", None), "GERAL")


def radar_normalizar_chave(valor):
    valor = radar_texto(valor, "")
    return " ".join(valor.upper().split())


def radar_tipo_obrigacao(conta):
    tipo = normalizar_texto(getattr(conta, "tipo_obrigacao", None))

    if tipo in ["RECORRENTE", "PARCELADA", "UNICA"]:
        return tipo

    chave = str(getattr(conta, "chave_conciliacao", "") or "")

    if chave.startswith("RECORRENTE:"):
        return "RECORRENTE"

    if chave.startswith("PARCELADA:"):
        return "PARCELADA"

    return "UNICA"


def radar_conta_parcelada(conta):
    return radar_tipo_obrigacao(conta) == "PARCELADA"


def radar_parcela_label(conta):
    parcela_atual = getattr(conta, "parcela_atual", None)
    total_parcelas = getattr(conta, "total_parcelas", None)

    if parcela_atual and total_parcelas:
        return f"{parcela_atual}/{total_parcelas}"

    if parcela_atual:
        return str(parcela_atual)

    return "-"


def radar_chave_base_obrigacao(descricao, fornecedor, categoria, setor):
    import hashlib

    partes = [
        radar_normalizar_chave(descricao),
        radar_normalizar_chave(fornecedor),
        radar_normalizar_chave(categoria),
        radar_normalizar_chave(setor),
    ]

    base = "|".join(partes)
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def radar_grupo_parcelamento_padrao(descricao, fornecedor, categoria, setor):
    return f"PARCELAMENTO:{radar_chave_base_obrigacao(descricao, fornecedor, categoria, setor)}"


def radar_chave_parcelada(grupo_parcelamento, parcela_atual, total_parcelas):
    return f"PARCELADA:{grupo_parcelamento}:{int(parcela_atual)}:{int(total_parcelas)}"


def radar_somar_meses(data_base, quantidade_meses):
    mes = data_base.month + quantidade_meses
    ano = data_base.year + ((mes - 1) // 12)
    mes = ((mes - 1) % 12) + 1
    dia = min(data_base.day, calendar.monthrange(ano, mes)[1])

    return date(ano, mes, dia)


def radar_chave_grupo(conta):
    import hashlib

    if radar_conta_parcelada(conta):
        grupo_parcelamento = getattr(conta, "grupo_parcelamento", None)

        if grupo_parcelamento:
            return f"PARCELADA:{grupo_parcelamento}"

    partes = [
        radar_normalizar_chave(radar_descricao_conta(conta)),
        radar_normalizar_chave(radar_fornecedor_conta(conta)),
        radar_normalizar_chave(radar_categoria_conta(conta)),
        radar_normalizar_chave(radar_setor_conta(conta)),
    ]

    base = "|".join(partes)
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def radar_mesmo_grupo(item, chave):
    return item.get("grupo_key") == chave


def radar_conta_cancelada(conta):
    return status_cancelado(getattr(conta, "status", None))


def radar_conta_aberta(conta):
    status = normalizar_texto(getattr(conta, "status", None))

    return (
        not conta_esta_paga(conta)
        and not radar_conta_cancelada(conta)
        and status != "TRANSPORTADO"
    )


def radar_valor_conta(conta):
    return dinheiro(getattr(conta, "valor", 0))


def radar_item_conta(conta, hoje):
    data_vencimento = data_para_date(getattr(conta, "data_vencimento", None))
    data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))
    valor = radar_valor_conta(conta)
    pago = conta_esta_paga(conta)
    dias_atraso = 0
    dias_para_vencer = None

    if data_vencimento:
        if data_vencimento < hoje and not pago:
            dias_atraso = (hoje - data_vencimento).days
        elif data_vencimento >= hoje and not pago:
            dias_para_vencer = (data_vencimento - hoje).days

    status_label = "PAGO" if pago else "EM ABERTO"
    status_classe = "pago" if pago else "aberto"

    if not pago and data_vencimento:
        if data_vencimento < hoje:
            status_label = "ATRASADA"
            status_classe = "vencido"
        elif data_vencimento == hoje:
            status_label = "HOJE"
            status_classe = "hoje"
        elif data_vencimento <= hoje + timedelta(days=7):
            status_label = f"{dias_para_vencer} dia(s)"
            status_classe = "proximo"
        else:
            status_label = f"{dias_para_vencer} dia(s)"
            status_classe = "futuro"

    if radar_conta_cancelada(conta):
        status_label = "CANCELADA"
        status_classe = "cancelado"

    return {
        "id": conta.id,
        "conta": conta,
        "grupo_key": radar_chave_grupo(conta),
        "descricao": radar_descricao_conta(conta),
        "fornecedor": radar_fornecedor_conta(conta),
        "categoria": radar_categoria_conta(conta),
        "setor": radar_setor_conta(conta),
        "valor": valor,
        "valor_formatado": radar_moeda(valor),
        "data_vencimento": data_vencimento,
        "data_vencimento_formatada": radar_data_br(data_vencimento),
        "data_vencimento_input": radar_data_input(data_vencimento),
        "data_pagamento": data_pagamento,
        "data_pagamento_formatada": radar_data_br(data_pagamento),
        "data_pagamento_input": radar_data_input(data_pagamento),
        "competencia": radar_competencia_label(data_vencimento),
        "pago": pago,
        "status": radar_texto(getattr(conta, "status", None), "PENDENTE"),
        "status_label": status_label,
        "status_classe": status_classe,
        "dias_atraso": dias_atraso,
        "dias_para_vencer": dias_para_vencer,
        "observacoes": radar_texto(getattr(conta, "observacoes", None), ""),
        "numero_fatura": radar_texto(getattr(conta, "numero_fatura", None), "-"),
        "tipo_obrigacao": radar_tipo_obrigacao(conta),
        "recorrente": radar_tipo_obrigacao(conta) == "RECORRENTE",
        "parcelada": radar_tipo_obrigacao(conta) == "PARCELADA",
        "parcela_atual": getattr(conta, "parcela_atual", None) or "",
        "total_parcelas": getattr(conta, "total_parcelas", None) or "",
        "parcela_label": radar_parcela_label(conta),
        "grupo_parcelamento": radar_texto(getattr(conta, "grupo_parcelamento", None), ""),
    }


def radar_chave_recorrente(recorrencia_id, mes, ano):
    return f"RECORRENTE:{recorrencia_id}:{int(mes):02d}:{int(ano)}"


def radar_mes_ano_iter(inicio, fim):
    mes = inicio.month
    ano = inicio.year

    while (ano, mes) <= (fim.year, fim.month):
        yield mes, ano

        mes += 1

        if mes > 12:
            mes = 1
            ano += 1


def radar_criar_conta_recorrente_mes(recorrencia, mes, ano):
    chave = radar_chave_recorrente(recorrencia.id, mes, ano)

    existente = ContaPagarImportada.query.filter_by(
        chave_conciliacao=chave
    ).first()

    if existente:
        return None

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dia_vencimento = min(int(recorrencia.dia_vencimento or 1), ultimo_dia)
    data_vencimento = date(ano, mes, dia_vencimento)

    observacoes_base = recorrencia.observacoes or ""
    observacao_recorrencia = (
        f"Conta gerada automaticamente pela recorrência #{recorrencia.id}. "
        f"Competência: {radar_competencia_label(data_vencimento)}."
    )

    conta = ContaPagarImportada(
        numero_fatura=f"REC-{recorrencia.id}-{mes:02d}-{ano}",
        fornecedor_funcionario=recorrencia.fornecedor_funcionario,
        plano_contas=recorrencia.plano_contas or recorrencia.descricao,
        categoria=recorrencia.categoria,
        setor=recorrencia.setor or "GERAL",
        data_documento=datetime.combine(date.today(), datetime.min.time()),
        data_vencimento=datetime.combine(data_vencimento, datetime.min.time()),
        valor=recorrencia.valor,
        pago=False,
        status="PENDENTE",
        observacoes=f"{observacao_recorrencia}\n{observacoes_base}".strip(),
        mes=mes,
        ano=ano,
        chave_conciliacao=chave,
        origem_importacao="RECORRENTE",
    )

    db.session.add(conta)

    return conta


def radar_gerar_recorrencias_ate_mes(mes, ano):
    data_limite = ultimo_dia_mes(mes, ano)

    recorrencias = ContaRecorrente.query.filter_by(
        ativo=True
    ).all()

    total_criadas = 0

    for recorrencia in recorrencias:
        data_inicio = recorrencia.data_inicio or date(ano, mes, 1)
        data_fim = recorrencia.data_fim

        if data_inicio > data_limite:
            continue

        fim_geracao = data_limite

        if data_fim and data_fim < fim_geracao:
            fim_geracao = data_fim

        for mes_ref, ano_ref in radar_mes_ano_iter(data_inicio, fim_geracao):
            criada = radar_criar_conta_recorrente_mes(
                recorrencia=recorrencia,
                mes=mes_ref,
                ano=ano_ref
            )

            if criada:
                total_criadas += 1

    if total_criadas:
        db.session.commit()

    return total_criadas



def radar_garantir_parcelas_existentes_ate_mes(mes, ano):
    """
    Garante que contratos parcelados já existentes no banco tenham suas próximas parcelas geradas.

    Exemplo real:
    - Existe Empréstimo Caixa parcela 18/72 com vencimento em março.
    - Ao abrir maio, o sistema cria automaticamente 19/72 abril e 20/72 maio se ainda não existirem.
    - Ao abrir junho, cria 21/72 junho, mantendo as anteriores abertas/pagas conforme a situação.

    Isso é o que faz a coluna Dívida Herdada enxergar Abril/Maio como abertas quando Março já foi paga.
    """

    data_limite = ultimo_dia_mes(mes, ano)
    total_criadas = 0
    total_ajustadas = 0

    contas_parceladas = ContaPagarImportada.query.filter(
        ContaPagarImportada.tipo_obrigacao == "PARCELADA",
        ContaPagarImportada.parcela_atual.isnot(None),
        ContaPagarImportada.total_parcelas.isnot(None),
        ContaPagarImportada.data_vencimento.isnot(None)
    ).order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc()
    ).all()

    for conta_base in contas_parceladas:
        data_base = data_para_date(conta_base.data_vencimento)
        parcela_base = getattr(conta_base, "parcela_atual", None)
        total_parcelas = getattr(conta_base, "total_parcelas", None)

        if not data_base or not parcela_base or not total_parcelas:
            continue

        try:
            parcela_base = int(parcela_base)
            total_parcelas = int(total_parcelas)
        except Exception:
            continue

        if parcela_base < 1 or total_parcelas < parcela_base:
            continue

        descricao = radar_descricao_conta(conta_base)
        fornecedor = radar_texto(getattr(conta_base, "fornecedor_funcionario", None), "")
        categoria = radar_texto(getattr(conta_base, "categoria", None), "")
        setor = radar_texto(getattr(conta_base, "setor", None), "GERAL")

        grupo_parcelamento = getattr(conta_base, "grupo_parcelamento", None)

        if not grupo_parcelamento:
            grupo_parcelamento = radar_grupo_parcelamento_padrao(
                descricao=descricao,
                fornecedor=fornecedor,
                categoria=categoria,
                setor=setor
            )
            conta_base.grupo_parcelamento = grupo_parcelamento
            total_ajustadas += 1

        chave_base = radar_chave_parcelada(
            grupo_parcelamento=grupo_parcelamento,
            parcela_atual=parcela_base,
            total_parcelas=total_parcelas
        )

        if not getattr(conta_base, "chave_conciliacao", None):
            conta_base.chave_conciliacao = chave_base
            total_ajustadas += 1

        if not getattr(conta_base, "dia_vencimento_parcela", None):
            conta_base.dia_vencimento_parcela = data_base.day
            total_ajustadas += 1

        if not getattr(conta_base, "parcela_origem_id", None):
            conta_base.parcela_origem_id = conta_base.id
            total_ajustadas += 1

        # Gera somente até o mês visualizado, para não criar todo o contrato de uma vez.
        for parcela in range(parcela_base + 1, total_parcelas + 1):
            meses_adiante = parcela - parcela_base
            data_vencimento = radar_somar_meses(data_base, meses_adiante)

            if data_vencimento > data_limite:
                break

            chave = radar_chave_parcelada(
                grupo_parcelamento=grupo_parcelamento,
                parcela_atual=parcela,
                total_parcelas=total_parcelas
            )

            existente = ContaPagarImportada.query.filter_by(
                chave_conciliacao=chave
            ).first()

            if existente:
                continue

            observacao_parcela = (
                f"Conta parcelada gerada automaticamente a partir da parcela "
                f"{parcela_base}/{total_parcelas}. "
                f"Parcela {parcela}/{total_parcelas}. "
                f"Grupo: {grupo_parcelamento}. "
                f"Competência: {radar_competencia_label(data_vencimento)}."
            )

            nova_conta = ContaPagarImportada(
                numero_fatura=(
                    f"{radar_texto(getattr(conta_base, 'numero_fatura', None), 'PARC')}-{parcela:03d}"
                ),
                fornecedor_funcionario=getattr(conta_base, "fornecedor_funcionario", None),
                telefone=getattr(conta_base, "telefone", None),
                email=getattr(conta_base, "email", None),
                plano_contas=getattr(conta_base, "plano_contas", None),
                categoria=getattr(conta_base, "categoria", None),
                setor=getattr(conta_base, "setor", None) or "GERAL",
                data_documento=datetime.combine(date.today(), datetime.min.time()),
                data_vencimento=datetime.combine(data_vencimento, datetime.min.time()),
                valor=getattr(conta_base, "valor", 0),
                pago=False,
                status="PENDENTE",
                observacoes=f"{observacao_parcela}\n{radar_texto(getattr(conta_base, 'observacoes', None), '')}".strip(),
                tipo_obrigacao="PARCELADA",
                parcela_atual=parcela,
                total_parcelas=total_parcelas,
                grupo_parcelamento=grupo_parcelamento,
                dia_vencimento_parcela=getattr(conta_base, "dia_vencimento_parcela", None) or data_base.day,
                parcela_origem_id=getattr(conta_base, "parcela_origem_id", None) or conta_base.id,
                mes=data_vencimento.month,
                ano=data_vencimento.year,
                chave_conciliacao=chave,
                origem_importacao="PARCELADA",
            )

            db.session.add(nova_conta)
            total_criadas += 1

    if total_criadas or total_ajustadas:
        db.session.commit()

    return total_criadas


def radar_filtrar_busca(contas, busca):
    busca = radar_normalizar_chave(busca)

    if not busca:
        return contas

    filtradas = []

    for conta in contas:
        texto_busca = " ".join([
            radar_descricao_conta(conta),
            radar_fornecedor_conta(conta),
            radar_categoria_conta(conta),
            radar_setor_conta(conta),
            radar_texto(getattr(conta, "numero_fatura", None), ""),
            radar_texto(getattr(conta, "observacoes", None), ""),
        ])

        if busca in radar_normalizar_chave(texto_busca):
            filtradas.append(conta)

    return filtradas


def radar_agrupar_herdadas(contas_herdadas, hoje):
    grupos_map = {}

    for conta in contas_herdadas:
        item = radar_item_conta(conta, hoje)
        chave = item["grupo_key"]

        if chave not in grupos_map:
            grupos_map[chave] = {
                "grupo_key": chave,
                "descricao": item["descricao"],
                "fornecedor": item["fornecedor"],
                "categoria": item["categoria"],
                "setor": item["setor"],
                "total": 0,
                "total_formatado": radar_moeda(0),
                "qtd": 0,
                "desde_data": item["data_vencimento"],
                "desde_label": item["competencia"],
                "maior_atraso": 0,
                "atraso_medio": 0,
                "historico": [],
                "parcelas_labels": [],
            }

        grupo = grupos_map[chave]
        grupo["total"] += item["valor"]
        grupo["qtd"] += 1
        grupo["historico"].append(item)

        if item.get("parcelada") and item.get("parcela_label") and item.get("parcela_label") != "-":
            grupo["parcelas_labels"].append(item["parcela_label"])

        if item["data_vencimento"]:
            if not grupo["desde_data"] or item["data_vencimento"] < grupo["desde_data"]:
                grupo["desde_data"] = item["data_vencimento"]
                grupo["desde_label"] = item["competencia"]

        grupo["maior_atraso"] = max(grupo["maior_atraso"], item["dias_atraso"] or 0)

    grupos = []

    for grupo in grupos_map.values():
        grupo["historico"] = sorted(
            grupo["historico"],
            key=lambda x: (x["data_vencimento"] or date.max, x["id"])
        )

        if grupo["historico"]:
            grupo["desde_label"] = grupo["historico"][0]["competencia"]

        grupo["total_formatado"] = radar_moeda(grupo["total"])

        valores_historico = [dinheiro(h.get("valor")) for h in grupo["historico"] if dinheiro(h.get("valor")) > 0]
        valores_unicos = sorted(set(round(v, 2) for v in valores_historico))

        if len(valores_unicos) == 1:
            grupo["valor_parcela"] = valores_unicos[0]
            grupo["valor_parcela_formatado"] = radar_moeda(valores_unicos[0])
        elif valores_unicos:
            grupo["valor_parcela"] = valores_unicos[0]
            grupo["valor_parcela_formatado"] = radar_moeda(valores_unicos[0])
        else:
            grupo["valor_parcela"] = 0
            grupo["valor_parcela_formatado"] = radar_moeda(0)

        grupo["parcelas_resumo"] = ", ".join(grupo.get("parcelas_labels", [])[:4])
        if len(grupo.get("parcelas_labels", [])) > 4:
            grupo["parcelas_resumo"] += "..."
        atrasos = [h["dias_atraso"] for h in grupo["historico"] if h["dias_atraso"]]
        grupo["atraso_medio"] = round(sum(atrasos) / len(atrasos)) if atrasos else 0
        grupos.append(grupo)

    return sorted(
        grupos,
        key=lambda g: (g["desde_data"] or date.max, -g["total"])
    )


def radar_somar_itens(lista):
    return sum(dinheiro(item.get("valor")) for item in lista)


def radar_somar_grupos(lista):
    return sum(dinheiro(item.get("total")) for item in lista)


def radar_classificar_risco(pressao, receitas_previstas):
    pressao = dinheiro(pressao)
    receitas_previstas = dinheiro(receitas_previstas)

    if pressao <= 0:
        return "CONTROLADO", "controlado"

    if receitas_previstas <= 0:
        return "ATENÇÃO", "atencao"

    percentual = (pressao / receitas_previstas) * 100

    if percentual <= 50:
        return "CONTROLADO", "controlado"

    if percentual <= 80:
        return "ATENÇÃO", "atencao"

    if percentual <= 100:
        return "ALTO RISCO", "alto-risco"

    return "CRÍTICO", "critico"


def radar_recomendacao(total_divida_herdada, total_proximos, diferenca):
    if total_divida_herdada > 0:
        return "Priorize pagar os meses mais antigos das dívidas herdadas para cortar a bola de neve."

    if diferenca < 0:
        return "A pressão do mês está acima das entradas previstas. Revise despesas e renegocie vencimentos."

    if total_proximos > 0:
        return "Organize o caixa para cobrir os próximos vencimentos sem atrasar novas contas."

    return "Cenário saudável. Continue mantendo as obrigações em dia."


def radar_insight(total_divida_herdada, grupos_herdados, total_pagas_antigas):
    qtd_grupos = len(grupos_herdados)

    if total_divida_herdada > 0:
        return (
            f"Você possui {radar_moeda(total_divida_herdada)} em dívidas herdadas "
            f"agrupadas em {qtd_grupos} tipo(s). Pague os meses em ordem para colocar as contas em dia."
        )

    if total_pagas_antigas > 0:
        return (
            f"Você já pagou {radar_moeda(total_pagas_antigas)} de meses anteriores neste mês. "
            "Isso ajuda a limpar o histórico e reduzir a pressão futura."
        )

    return "Nenhuma dívida herdada aberta encontrada para este período. O fluxo está mais limpo."


def radar_receitas_realizadas_mes(mes, ano):
    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)

    total = 0

    contas_receber = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt
    ).all()

    for conta in contas_receber:
        total += valor_conta_receber(conta)

    lancamentos = LancamentoFinanceiro.query.filter(
        LancamentoFinanceiro.tipo == "RECEITA",
        LancamentoFinanceiro.mes == mes,
        LancamentoFinanceiro.ano == ano
    ).all()

    for lancamento in lancamentos:
        if status_pago(lancamento.status):
            total += dinheiro(lancamento.valor)

    return total


@gestao_bp.route("/radar-financeiro/")
@gestao_required
def radar_financeiro():

    hoje = date.today()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year
    setor = request.args.get("setor", "").strip().upper()
    busca = request.args.get("busca", "").strip()
    visao = request.args.get("visao", "vencimento").strip().lower()
    grupo_key = request.args.get("grupo", "").strip()

    if mes < 1 or mes > 12:
        mes = hoje.month

    inicio_mes = primeiro_dia_mes(mes, ano)
    fim_mes = ultimo_dia_mes(mes, ano)
    inicio_dt = inicio_mes_datetime(mes, ano)
    fim_dt = fim_mes_datetime(mes, ano)

    # Gera automaticamente as obrigações recorrentes até o mês visualizado.
    # Assim, se uma conta recorrente nasceu em março e você abre junho,
    # março, abril, maio e junho serão criados se ainda não existirem.
    radar_gerar_recorrencias_ate_mes(mes, ano)

    # Gera automaticamente parcelas faltantes de contratos parcelados já cadastrados.
    # Exemplo: se já existe 18/72 em março, ao abrir maio ele cria 19/72 abril e 20/72 maio.
    radar_garantir_parcelas_existentes_ate_mes(mes, ano)

    query = ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao.in_(["PAGAMENTO", "MANUAL", "VENCIMENTO", "IMPORTACAO", "RECORRENTE", "PARCELADA"])
    )

    if setor:
        query = query.filter(
            ContaPagarImportada.setor.ilike(f"%{setor}%")
        )

    contas_base = query.order_by(
        ContaPagarImportada.data_vencimento.asc().nullslast(),
        ContaPagarImportada.id.asc()
    ).all()

    contas_base = radar_filtrar_busca(contas_base, busca)

    contas_abertas = [
        c for c in contas_base
        if radar_conta_aberta(c)
    ]

    contas_pagas = [
        c for c in contas_base
        if conta_esta_paga(c)
    ]

    # =====================================================
    # HERDADAS
    # Contas abertas que nasceram/venceram antes do mês filtrado.
    # Cada conta mantém sua própria data, e a coluna agrupa por tipo.
    # =====================================================
    contas_herdadas = [
        c for c in contas_abertas
        if data_para_date(c.data_vencimento)
        and data_para_date(c.data_vencimento) < inicio_mes
    ]

    grupos_herdados = radar_agrupar_herdadas(contas_herdadas, hoje)

    # =====================================================
    # FLUXO NORMAL DO MÊS
    # Aqui não entram as herdadas, para não duplicar no radar.
    # =====================================================
    abertas_mes = [
        c for c in contas_abertas
        if data_para_date(c.data_vencimento)
        and inicio_mes <= data_para_date(c.data_vencimento) <= fim_mes
    ]

    vence_hoje = []
    proximos_7 = []
    ate_fim_mes = []
    atrasadas_mes = []

    for conta in abertas_mes:
        data_venc = data_para_date(conta.data_vencimento)

        if not data_venc:
            continue

        item = radar_item_conta(conta, hoje)

        if data_venc < hoje:
            atrasadas_mes.append(item)
        elif data_venc == hoje:
            vence_hoje.append(item)
        elif data_venc <= hoje + timedelta(days=7):
            proximos_7.append(item)
        else:
            ate_fim_mes.append(item)

    pagas_mes_contas = [
        c for c in contas_pagas
        if data_para_date(c.data_pagamento)
        and inicio_mes <= data_para_date(c.data_pagamento) <= fim_mes
        and (
            not data_para_date(c.data_vencimento)
            or data_para_date(c.data_vencimento) >= inicio_mes
        )
    ]

    pagas_antigas_contas = [
        c for c in contas_pagas
        if data_para_date(c.data_pagamento)
        and inicio_mes <= data_para_date(c.data_pagamento) <= fim_mes
        and data_para_date(c.data_vencimento)
        and data_para_date(c.data_vencimento) < inicio_mes
    ]

    pagas_mes = [
        radar_item_conta(c, hoje)
        for c in sorted(
            pagas_mes_contas,
            key=lambda conta: (data_para_date(conta.data_pagamento) or date.max, conta.id)
        )
    ]

    pagas_meses_anteriores = [
        radar_item_conta(c, hoje)
        for c in sorted(
            pagas_antigas_contas,
            key=lambda conta: (data_para_date(conta.data_pagamento) or date.max, conta.id)
        )
    ]

    total_divida_herdada = radar_somar_grupos(grupos_herdados)
    total_atrasadas_mes = radar_somar_itens(atrasadas_mes)
    total_hoje = radar_somar_itens(vence_hoje)
    total_proximos_7 = radar_somar_itens(proximos_7)
    total_ate_fim_mes = radar_somar_itens(ate_fim_mes)
    total_pagas_mes = radar_somar_itens(pagas_mes)
    total_pagas_meses_anteriores = radar_somar_itens(pagas_meses_anteriores)

    receitas_realizadas = radar_receitas_realizadas_mes(mes, ano)

    # Receita prevista simples: realizado + contas a receber abertas até fim do mês.
    receitas_previstas = receitas_realizadas
    contas_receber_abertas = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == False,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_vencimento <= fim_dt
    ).all()

    for conta_receber in contas_receber_abertas:
        receitas_previstas += valor_conta_receber(conta_receber)

    pressao_imediata = (
        total_divida_herdada
        + total_atrasadas_mes
        + total_hoje
        + total_proximos_7
    )

    diferenca_estimada = receitas_previstas - pressao_imediata
    resultado_parcial = receitas_realizadas - (total_pagas_mes + total_pagas_meses_anteriores)

    risco, risco_classe = radar_classificar_risco(
        pressao=pressao_imediata,
        receitas_previstas=receitas_previstas
    )

    recomendacao = radar_recomendacao(
        total_divida_herdada=total_divida_herdada,
        total_proximos=total_proximos_7,
        diferenca=diferenca_estimada
    )

    insight = radar_insight(
        total_divida_herdada=total_divida_herdada,
        grupos_herdados=grupos_herdados,
        total_pagas_antigas=total_pagas_meses_anteriores
    )

    grupo_selecionado = None

    if grupo_key:
        for grupo in grupos_herdados:
            if grupo["grupo_key"] == grupo_key:
                grupo_selecionado = grupo
                break

    if not grupo_selecionado and grupos_herdados:
        grupo_selecionado = grupos_herdados[0]

    top_despesas = []

    despesas_counter = Counter()

    for item in pagas_mes + pagas_meses_anteriores:
        despesas_counter[item["descricao"]] += item["valor"]

    for nome, valor in despesas_counter.most_common(5):
        percentual = 0
        total_pago_geral = total_pagas_mes + total_pagas_meses_anteriores

        if total_pago_geral:
            percentual = (dinheiro(valor) / total_pago_geral) * 100

        top_despesas.append({
            "nome": nome,
            "valor": valor,
            "valor_formatado": radar_moeda(valor),
            "percentual": percentual,
        })

    colunas = [
        {
            "slug": "herdada",
            "titulo": "DÍVIDA HERDADA",
            "subtitulo": "Contas de meses anteriores não pagas",
            "cor": "danger",
            "total": total_divida_herdada,
            "total_formatado": radar_moeda(total_divida_herdada),
            "badge": sum(g["qtd"] for g in grupos_herdados),
            "grupos": grupos_herdados,
            "vazia": "Sem dívidas herdadas",
        },
        {
            "slug": "hoje",
            "titulo": "VENCE HOJE",
            "subtitulo": f"Vencem hoje ({hoje.strftime('%d/%m/%Y')})",
            "cor": "warning",
            "total": total_hoje,
            "total_formatado": radar_moeda(total_hoje),
            "badge": len(vence_hoje),
            "itens": vence_hoje,
            "vazia": "Sem contas para hoje",
        },
        {
            "slug": "proximos",
            "titulo": "PRÓXIMOS 7 DIAS",
            "subtitulo": "Pressão imediata do caixa",
            "cor": "purple",
            "total": total_proximos_7,
            "total_formatado": radar_moeda(total_proximos_7),
            "badge": len(proximos_7),
            "itens": proximos_7,
            "vazia": "Sem vencimentos próximos",
        },
        {
            "slug": "ate_fim_mes",
            "titulo": "ATÉ FIM DO MÊS",
            "subtitulo": "Contas futuras dentro do mês filtrado",
            "cor": "dark",
            "total": total_ate_fim_mes,
            "total_formatado": radar_moeda(total_ate_fim_mes),
            "badge": len(ate_fim_mes),
            "itens": ate_fim_mes,
            "vazia": "Sem contas futuras no mês",
        },
        {
            "slug": "atrasadas_mes",
            "titulo": "ATRASADAS DO MÊS",
            "subtitulo": "Venceram neste mês e ainda não foram pagas",
            "cor": "blue",
            "total": total_atrasadas_mes,
            "total_formatado": radar_moeda(total_atrasadas_mes),
            "badge": len(atrasadas_mes),
            "itens": atrasadas_mes,
            "vazia": "Sem atrasadas do mês",
        },
        {
            "slug": "pagas_mes",
            "titulo": "PAGAS NO MÊS",
            "subtitulo": f"Pagas em {radar_nome_mes(mes)}",
            "cor": "success",
            "total": total_pagas_mes,
            "total_formatado": radar_moeda(total_pagas_mes),
            "badge": len(pagas_mes),
            "itens": pagas_mes,
            "vazia": "Sem contas pagas no mês",
        },
        {
            "slug": "pagas_antigas",
            "titulo": "PAGAS DE MESES ANT.",
            "subtitulo": "Dívidas antigas pagas agora",
            "cor": "teal",
            "total": total_pagas_meses_anteriores,
            "total_formatado": radar_moeda(total_pagas_meses_anteriores),
            "badge": len(pagas_meses_anteriores),
            "itens": pagas_meses_anteriores,
            "vazia": "Sem dívidas antigas pagas",
        },
    ]

    return render_template(
        "gestao/radar_financeiro.html",
        hoje=hoje,
        mes=mes,
        ano=ano,
        setor=setor,
        busca=busca,
        visao=visao,
        grupo_key=grupo_key,

        moeda=radar_moeda,
        nome_mes=radar_nome_mes,
        nome_mes_curto=radar_nome_mes_curto,

        colunas=colunas,
        grupos_herdados=grupos_herdados,
        grupo_selecionado=grupo_selecionado,

        total_divida_herdada=total_divida_herdada,
        total_atrasadas_mes=total_atrasadas_mes,
        total_hoje=total_hoje,
        total_proximos_7=total_proximos_7,
        total_ate_fim_mes=total_ate_fim_mes,
        total_pagas_mes=total_pagas_mes,
        total_pagas_meses_anteriores=total_pagas_meses_anteriores,

        receitas_realizadas=receitas_realizadas,
        receitas_previstas=receitas_previstas,
        pressao_imediata=pressao_imediata,
        diferenca_estimada=diferenca_estimada,
        resultado_parcial=resultado_parcial,
        risco=risco,
        risco_classe=risco_classe,
        recomendacao=recomendacao,
        insight=insight,
        top_despesas=top_despesas,
    )


def radar_criar_parcelas_automaticas(
    descricao,
    fornecedor,
    categoria,
    setor,
    valor,
    data_vencimento_inicial,
    parcela_atual,
    total_parcelas,
    observacoes,
    numero_fatura=None,
    ja_pago=False,
    data_pagamento=None
):
    grupo_parcelamento = radar_grupo_parcelamento_padrao(
        descricao=descricao,
        fornecedor=fornecedor,
        categoria=categoria,
        setor=setor
    )

    contas_criadas = []

    for parcela in range(parcela_atual, total_parcelas + 1):
        meses_adiante = parcela - parcela_atual
        data_vencimento = radar_somar_meses(data_vencimento_inicial, meses_adiante)
        chave = radar_chave_parcelada(
            grupo_parcelamento=grupo_parcelamento,
            parcela_atual=parcela,
            total_parcelas=total_parcelas
        )

        existente = ContaPagarImportada.query.filter_by(
            chave_conciliacao=chave
        ).first()

        if existente:
            continue

        observacao_parcela = (
            f"Conta parcelada gerada automaticamente. "
            f"Parcela {parcela}/{total_parcelas}. "
            f"Grupo: {grupo_parcelamento}. "
            f"Competência: {radar_competencia_label(data_vencimento)}."
        )

        conta = ContaPagarImportada(
            numero_fatura=(
                f"{numero_fatura}-{parcela:03d}"
                if numero_fatura else
                f"PARC-{grupo_parcelamento[-10:]}-{parcela:03d}-{total_parcelas}"
            ),
            fornecedor_funcionario=fornecedor,
            plano_contas=descricao,
            categoria=categoria,
            setor=setor,
            data_documento=datetime.combine(date.today(), datetime.min.time()),
            data_vencimento=datetime.combine(data_vencimento, datetime.min.time()),
            valor=valor,
            pago=False,
            status="PENDENTE",
            observacoes=f"{observacao_parcela}\n{observacoes or ''}".strip(),
            tipo_obrigacao="PARCELADA",
            parcela_atual=parcela,
            total_parcelas=total_parcelas,
            grupo_parcelamento=grupo_parcelamento,
            dia_vencimento_parcela=data_vencimento_inicial.day,
            parcela_origem_id=None,
            mes=data_vencimento.month,
            ano=data_vencimento.year,
            chave_conciliacao=chave,
            origem_importacao="PARCELADA",
        )

        if ja_pago and parcela == parcela_atual:
            conta.pago = True
            conta.status = "PAGO"
            conta.data_pagamento = datetime.combine(
                data_pagamento or date.today(),
                datetime.min.time()
            )

        db.session.add(conta)
        db.session.flush()

        if not conta.parcela_origem_id:
            conta.parcela_origem_id = conta.id

        contas_criadas.append(conta)

    if contas_criadas:
        origem_id = contas_criadas[0].id

        for conta in contas_criadas:
            conta.parcela_origem_id = origem_id

    return contas_criadas


@gestao_bp.route("/radar-financeiro/novo", methods=["POST"])
@gestao_required
def radar_financeiro_novo():

    descricao = request.form.get("descricao", "").strip()
    fornecedor = request.form.get("fornecedor", "").strip()
    categoria = request.form.get("categoria", "").strip()
    setor = normalizar_texto(request.form.get("setor")) or "GERAL"
    valor = dinheiro(request.form.get("valor"))
    data_vencimento = parse_data(request.form.get("data_vencimento"))
    observacoes = request.form.get("observacoes", "").strip()
    tipo_obrigacao = normalizar_texto(request.form.get("tipo_obrigacao")) or "UNICA"
    parcela_atual = request.form.get("parcela_atual", type=int)
    total_parcelas = request.form.get("total_parcelas", type=int)

    if tipo_obrigacao not in ["UNICA", "RECORRENTE", "PARCELADA"]:
        tipo_obrigacao = "UNICA"

    if tipo_obrigacao == "PARCELADA":
        if not parcela_atual or not total_parcelas or parcela_atual < 1 or total_parcelas < parcela_atual:
            flash("Para conta parcelada, informe parcela atual e total de parcelas corretamente.", "danger")
            return redirect(request.referrer or "/gestao/radar-financeiro/")

    if not descricao:
        flash("Informe a descrição da conta.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    if not data_vencimento:
        flash("Informe a data de vencimento.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    recorrente = tipo_obrigacao == "RECORRENTE" or request.form.get("recorrente") == "on"
    recorrencia = None

    if tipo_obrigacao == "PARCELADA":
        ja_pago = request.form.get("ja_pago") == "on"
        data_pagamento = parse_data(request.form.get("data_pagamento")) or date.today()
        radar_criar_parcelas_automaticas(
            descricao=descricao,
            fornecedor=fornecedor,
            categoria=categoria,
            setor=setor,
            valor=valor,
            data_vencimento_inicial=data_vencimento,
            parcela_atual=parcela_atual,
            total_parcelas=total_parcelas,
            observacoes=observacoes,
            numero_fatura=request.form.get("numero_fatura"),
            ja_pago=ja_pago,
            data_pagamento=data_pagamento
        )
        db.session.commit()
        flash("Conta parcelada criada. As parcelas futuras foram geradas automaticamente.", "success")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    if recorrente:
        recorrencia = ContaRecorrente(
            descricao=descricao,
            fornecedor_funcionario=fornecedor,
            plano_contas=descricao,
            categoria=categoria,
            setor=setor,
            valor=valor,
            dia_vencimento=data_vencimento.day,
            data_inicio=data_vencimento,
            ativo=True,
            observacoes=observacoes,
        )

        db.session.add(recorrencia)
        db.session.flush()

    chave_conciliacao = None
    origem_importacao = "PAGAMENTO"

    if recorrencia:
        chave_conciliacao = radar_chave_recorrente(
            recorrencia_id=recorrencia.id,
            mes=data_vencimento.month,
            ano=data_vencimento.year
        )
        origem_importacao = "RECORRENTE"

    observacoes_conta = observacoes

    if recorrencia:
        observacoes_conta = (
            f"Conta gerada pela recorrência #{recorrencia.id}. "
            f"Competência: {radar_competencia_label(data_vencimento)}.\n"
            f"{observacoes or ''}"
        ).strip()

    conta = ContaPagarImportada(
        numero_fatura=request.form.get("numero_fatura") or (f"REC-{recorrencia.id}-{data_vencimento.month:02d}-{data_vencimento.year}" if recorrencia else None),
        fornecedor_funcionario=fornecedor,
        plano_contas=descricao,
        categoria=categoria,
        setor=setor,
        data_documento=datetime.combine(date.today(), datetime.min.time()),
        data_vencimento=datetime.combine(data_vencimento, datetime.min.time()),
        valor=valor,
        pago=False,
        status="PENDENTE",
        observacoes=observacoes_conta,
        tipo_obrigacao="RECORRENTE" if recorrencia else "UNICA",
        parcela_atual=None,
        total_parcelas=None,
        grupo_parcelamento=None,
        dia_vencimento_parcela=None,
        parcela_origem_id=None,
        mes=data_vencimento.month,
        ano=data_vencimento.year,
        chave_conciliacao=chave_conciliacao,
        origem_importacao=origem_importacao,
    )

    if request.form.get("ja_pago") == "on":
        data_pagamento = parse_data(request.form.get("data_pagamento")) or date.today()
        conta.pago = True
        conta.status = "PAGO"
        conta.data_pagamento = datetime.combine(data_pagamento, datetime.min.time())

    db.session.add(conta)
    db.session.commit()

    if recorrencia:
        flash("Conta cadastrada e recorrência mensal criada. Os próximos meses serão gerados automaticamente.", "success")
    else:
        flash("Conta cadastrada no Radar Financeiro.", "success")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/editar/<int:id>", methods=["POST"])
@gestao_required
def radar_financeiro_editar(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    descricao = request.form.get("descricao", "").strip()
    fornecedor = request.form.get("fornecedor", "").strip()
    categoria = request.form.get("categoria", "").strip()
    setor = normalizar_texto(request.form.get("setor")) or "GERAL"
    status = normalizar_texto(request.form.get("status")) or "PENDENTE"
    valor = dinheiro(request.form.get("valor"))
    data_vencimento = parse_data(request.form.get("data_vencimento"))
    data_pagamento = parse_data(request.form.get("data_pagamento"))
    observacoes = request.form.get("observacoes", "").strip()
    tipo_obrigacao = normalizar_texto(request.form.get("tipo_obrigacao")) or radar_tipo_obrigacao(conta)
    parcela_atual = request.form.get("parcela_atual", type=int)
    total_parcelas = request.form.get("total_parcelas", type=int)

    if tipo_obrigacao not in ["UNICA", "RECORRENTE", "PARCELADA"]:
        tipo_obrigacao = "UNICA"

    if not descricao:
        flash("Informe a descrição da conta.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    conta.plano_contas = descricao
    conta.fornecedor_funcionario = fornecedor
    conta.categoria = categoria
    conta.setor = setor
    conta.valor = valor
    conta.status = status
    conta.observacoes = observacoes
    conta.tipo_obrigacao = tipo_obrigacao

    if tipo_obrigacao == "PARCELADA":
        conta.parcela_atual = parcela_atual
        conta.total_parcelas = total_parcelas
        conta.dia_vencimento_parcela = (data_vencimento.day if data_vencimento else conta.dia_vencimento_parcela)

        if not conta.grupo_parcelamento:
            conta.grupo_parcelamento = radar_grupo_parcelamento_padrao(
                descricao=descricao,
                fornecedor=fornecedor,
                categoria=categoria,
                setor=setor
            )
    else:
        conta.parcela_atual = None
        conta.total_parcelas = None
        conta.grupo_parcelamento = None
        conta.dia_vencimento_parcela = None

    if data_vencimento:
        conta.data_vencimento = datetime.combine(data_vencimento, datetime.min.time())
        conta.mes = data_vencimento.month
        conta.ano = data_vencimento.year

    if status == "PAGO":
        conta.pago = True
        conta.data_pagamento = datetime.combine(
            data_pagamento or date.today(),
            datetime.min.time()
        )
    else:
        conta.pago = False
        if data_pagamento:
            conta.data_pagamento = datetime.combine(data_pagamento, datetime.min.time())
        else:
            conta.data_pagamento = None

    db.session.commit()

    flash("Conta atualizada com sucesso.", "success")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/pagar/<int:id>", methods=["POST"])
@gestao_required
def radar_financeiro_pagar(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    data_pagamento = parse_data(request.form.get("data_pagamento")) or date.today()

    conta.pago = True
    conta.status = "PAGO"
    conta.data_pagamento = datetime.combine(
        data_pagamento,
        datetime.min.time()
    )

    observacao_extra = request.form.get("observacao", "").strip()

    if observacao_extra:
        observacao_antiga = conta.observacoes or ""
        conta.observacoes = f"{observacao_antiga}\nPagamento: {observacao_extra}".strip()

    db.session.commit()

    flash("Conta marcada como paga. Se ela era herdada, agora aparecerá em Pagas de Meses Anteriores.", "success")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/cancelar/<int:id>", methods=["POST"])
@gestao_required
def radar_financeiro_cancelar(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    conta.status = "CANCELADO"
    conta.pago = False

    db.session.commit()

    flash("Conta cancelada no Radar Financeiro.", "success")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/excluir/<int:id>", methods=["POST"])
@gestao_required
def radar_financeiro_excluir(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    db.session.delete(conta)
    db.session.commit()

    flash("Conta excluída definitivamente.", "success")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/excluir-lote", methods=["POST"])
@gestao_required
def radar_financeiro_excluir_lote():

    ids_limpos = radar_ids_limpos_formulario("contas_ids")
    mes_retorno = request.form.get("mes_retorno", type=int)
    ano_retorno = request.form.get("ano_retorno", type=int)

    if not ids_limpos:
        flash("Selecione pelo menos uma conta para excluir.", "warning")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    try:
        contas = ContaPagarImportada.query.filter(
            ContaPagarImportada.id.in_(ids_limpos)
        ).all()

        total = len(contas)

        for conta in contas:
            db.session.delete(conta)

        db.session.commit()

        flash(f"{total} conta(s) excluída(s) com sucesso.", "success")

        if mes_retorno and ano_retorno:
            return redirect(
                f"/gestao/radar-financeiro/?mes={mes_retorno}&ano={ano_retorno}&visao=vencimento"
            )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir contas em lote: {str(e)}", "danger")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/transportar/<int:id>", methods=["POST"])
@gestao_required
def radar_financeiro_transportar(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    mes_destino = request.form.get("mes_destino", type=int)
    ano_destino = request.form.get("ano_destino", type=int)

    if not mes_destino or not ano_destino:
        flash("Informe mês e ano de destino.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    data_vencimento_atual = data_para_date(conta.data_vencimento) or date.today()
    dia = min(data_vencimento_atual.day, calendar.monthrange(ano_destino, mes_destino)[1])
    nova_data = date(ano_destino, mes_destino, dia)

    duplicada = radar_transportada_ja_existe(
        conta=conta,
        nova_data=nova_data,
        mes_destino=mes_destino,
        ano_destino=ano_destino
    )

    if duplicada:
        flash(
            f"Essa conta já existe em {radar_nome_mes(mes_destino)}/{ano_destino}. Nenhuma duplicidade foi criada.",
            "warning"
        )
        return redirect(
            f"/gestao/radar-financeiro/?mes={mes_destino}&ano={ano_destino}&visao=vencimento"
        )

    observacao_origem = (
        f"Conta criada por transporte de {radar_competencia_label(data_vencimento_atual)}. "
        f"Vencimento original: {radar_data_br(data_vencimento_atual)}."
    )

    nova_conta = ContaPagarImportada(
        numero_fatura=conta.numero_fatura,
        fornecedor_funcionario=conta.fornecedor_funcionario,
        plano_contas=conta.plano_contas,
        categoria=conta.categoria,
        setor=conta.setor,
        data_documento=datetime.combine(date.today(), datetime.min.time()),
        data_vencimento=datetime.combine(nova_data, datetime.min.time()),
        valor=conta.valor,
        pago=False,
        status="PENDENTE",
        observacoes=f"{observacao_origem}\n{conta.observacoes or ''}".strip(),
        tipo_obrigacao=radar_tipo_obrigacao(conta),
        parcela_atual=getattr(conta, "parcela_atual", None),
        total_parcelas=getattr(conta, "total_parcelas", None),
        grupo_parcelamento=getattr(conta, "grupo_parcelamento", None),
        dia_vencimento_parcela=getattr(conta, "dia_vencimento_parcela", None),
        parcela_origem_id=getattr(conta, "parcela_origem_id", None),
        mes=mes_destino,
        ano=ano_destino,
        chave_conciliacao=None,
        origem_importacao="PAGAMENTO",
    )

    observacao_antiga = conta.observacoes or ""
    conta.status = "TRANSPORTADO"
    conta.observacoes = (
        f"{observacao_antiga}\n"
        f"Transportada para {radar_nome_mes(mes_destino)}/{ano_destino} em {date.today().strftime('%d/%m/%Y')}."
    ).strip()

    db.session.add(nova_conta)
    db.session.commit()

    flash("Conta transportada mantendo histórico no mês original e criando nova obrigação no destino.", "success")

    return redirect(
        f"/gestao/radar-financeiro/?mes={mes_destino}&ano={ano_destino}&visao=vencimento"
    )


@gestao_bp.route("/radar-financeiro/transportar-lote", methods=["POST"])
@gestao_required
def radar_financeiro_transportar_lote():

    ids = request.form.getlist("contas_ids")

    if not ids:
        ids = request.form.getlist("contas_ids[]")

    mes_destino = request.form.get("mes_destino", type=int)
    ano_destino = request.form.get("ano_destino", type=int)

    if not ids:
        flash("Selecione ao menos uma conta para transportar.", "warning")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    if not mes_destino or not ano_destino or mes_destino < 1 or mes_destino > 12:
        flash("Informe mês e ano de destino.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    total = 0
    ignoradas = 0

    try:
        for id_conta in ids:
            try:
                id_limpo = int(id_conta)
            except Exception:
                ignoradas += 1
                continue

            conta = ContaPagarImportada.query.get(id_limpo)

            if not conta:
                ignoradas += 1
                continue

            status_atual = normalizar_texto(getattr(conta, "status", None))

            if status_atual in ["CANCELADO", "CANCELADA", "TRANSPORTADO"]:
                ignoradas += 1
                continue

            data_vencimento_atual = data_para_date(conta.data_vencimento) or date.today()
            dia = min(data_vencimento_atual.day, calendar.monthrange(ano_destino, mes_destino)[1])
            nova_data = date(ano_destino, mes_destino, dia)

            duplicada = radar_transportada_ja_existe(
                conta=conta,
                nova_data=nova_data,
                mes_destino=mes_destino,
                ano_destino=ano_destino
            )

            if duplicada:
                ignoradas += 1
                continue

            nova_conta = ContaPagarImportada(
                numero_fatura=conta.numero_fatura,
                fornecedor_funcionario=conta.fornecedor_funcionario,
                plano_contas=conta.plano_contas,
                categoria=conta.categoria,
                setor=conta.setor,
                data_documento=datetime.combine(date.today(), datetime.min.time()),
                data_vencimento=datetime.combine(nova_data, datetime.min.time()),
                valor=conta.valor,
                pago=False,
                status="PENDENTE",
                observacoes=(
                    f"Conta transportada em lote a partir de {radar_competencia_label(data_vencimento_atual)}. "
                    f"Vencimento original: {radar_data_br(data_vencimento_atual)}.\n"
                    f"{conta.observacoes or ''}"
                ).strip(),
                tipo_obrigacao=radar_tipo_obrigacao(conta),
                parcela_atual=getattr(conta, "parcela_atual", None),
                total_parcelas=getattr(conta, "total_parcelas", None),
                grupo_parcelamento=getattr(conta, "grupo_parcelamento", None),
                dia_vencimento_parcela=getattr(conta, "dia_vencimento_parcela", None),
                parcela_origem_id=getattr(conta, "parcela_origem_id", None),
                mes=mes_destino,
                ano=ano_destino,
                chave_conciliacao=None,
                origem_importacao="PAGAMENTO",
            )

            db.session.add(nova_conta)

            if not conta_esta_paga(conta):
                observacao_antiga = conta.observacoes or ""
                conta.status = "TRANSPORTADO"
                conta.pago = False
                conta.observacoes = (
                    f"{observacao_antiga}\n"
                    f"Transportada para {radar_nome_mes(mes_destino)}/{ano_destino} em {date.today().strftime('%d/%m/%Y')}."
                ).strip()

            total += 1

        db.session.commit()

        flash(
            f"{total} conta(s) transportada(s) para {radar_nome_mes(mes_destino)}/{ano_destino}. "
            f"Ignoradas: {ignoradas}.",
            "success"
        )

        return redirect(
            f"/gestao/radar-financeiro/?mes={mes_destino}&ano={ano_destino}&visao=vencimento"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao transportar contas em lote: {str(e)}", "danger")

    return redirect(request.referrer or "/gestao/radar-financeiro/")


@gestao_bp.route("/radar-financeiro/transportar-pagas-lote", methods=["POST"])
@gestao_required
def radar_financeiro_transportar_pagas_lote():

    ids = request.form.getlist("contas_ids")

    if not ids:
        ids = request.form.getlist("contas_ids[]")

    mes_destino = request.form.get("mes_destino", type=int)
    ano_destino = request.form.get("ano_destino", type=int)

    if not ids:
        flash("Selecione ao menos uma conta paga para transportar.", "warning")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    if not mes_destino or not ano_destino:
        flash("Informe mês e ano de destino.", "danger")
        return redirect(request.referrer or "/gestao/radar-financeiro/")

    total = 0
    ignoradas = 0

    for id_conta in ids:
        conta = ContaPagarImportada.query.get(id_conta)

        if not conta:
            continue

        data_vencimento_atual = data_para_date(conta.data_vencimento) or date.today()
        dia = min(data_vencimento_atual.day, calendar.monthrange(ano_destino, mes_destino)[1])
        nova_data = date(ano_destino, mes_destino, dia)

        duplicada = radar_transportada_ja_existe(
            conta=conta,
            nova_data=nova_data,
            mes_destino=mes_destino,
            ano_destino=ano_destino
        )

        if duplicada:
            ignoradas += 1
            continue

        nova_conta = ContaPagarImportada(
            numero_fatura=conta.numero_fatura,
            fornecedor_funcionario=conta.fornecedor_funcionario,
            plano_contas=conta.plano_contas,
            categoria=conta.categoria,
            setor=conta.setor,
            data_documento=datetime.combine(date.today(), datetime.min.time()),
            data_vencimento=datetime.combine(nova_data, datetime.min.time()),
            valor=conta.valor,
            pago=False,
            status="PENDENTE",
            observacoes=(
                f"Conta transportada em lote a partir de {radar_competencia_label(data_vencimento_atual)}. "
                f"Vencimento original: {radar_data_br(data_vencimento_atual)}."
            ),
            mes=mes_destino,
            ano=ano_destino,
            chave_conciliacao=None,
            origem_importacao="PAGAMENTO",
        )

        db.session.add(nova_conta)
        total += 1

    db.session.commit()

    flash(
        f"{total} conta(s) transportada(s) para {radar_nome_mes(mes_destino)}/{ano_destino}. "
        f"Duplicadas ignoradas: {ignoradas}.",
        "success"
    )

    return redirect(
        f"/gestao/radar-financeiro/?mes={mes_destino}&ano={ano_destino}&visao=vencimento"
    )


# =========================================================
# CENTRO DE CUSTOS
# Tabela gerencial anual por plano de contas
# Fonte oficial: Contas Pagas importadas
# =========================================================

def formatar_moeda_br(valor):
    valor = dinheiro(valor)
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_numero_br(valor):
    valor = dinheiro(valor)
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def periodo_ano_centro_custos(ano):
    inicio = datetime(ano, 1, 1, 0, 0, 0)
    fim = datetime(ano, 12, 31, 23, 59, 59)
    return inicio, fim


def contas_pagas_centro_custos_query(ano=None):
    """
    Centro de Custos deve bater com o relatório de Contas Pagas.

    Por isso a fonte fica restrita a:
    - pago=True
    - origem_importacao='PAGAMENTO'
    - data_pagamento preenchida

    Não usamos DESPESA_COMPLETA aqui para evitar duplicidade quando os dois
    relatórios forem importados no mesmo período.
    """
    query = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "PAGAMENTO",
        ContaPagarImportada.data_pagamento.isnot(None)
    )

    if ano:
        inicio, fim = periodo_ano_centro_custos(ano)
        query = query.filter(
            ContaPagarImportada.data_pagamento >= inicio,
            ContaPagarImportada.data_pagamento <= fim
        )

    return query


def mes_referencia_centro_custos(conta):
    """No Centro de Custos, a competência é sempre a data real de pagamento."""
    data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))

    if data_pagamento:
        return data_pagamento.month, data_pagamento.year

    return None, None


def nome_plano_centro_custos(conta):
    return (
        getattr(conta, "categoria", None)
        or getattr(conta, "plano_contas", None)
        or "SEM PLANO DE CONTAS"
    ).strip()


def setor_centro_custos(conta):
    setor = normalizar_texto(getattr(conta, "setor", None)) or "ASSISTÊNCIA"

    if setor == "LOGÍSTICA":
        return "LOGÍSTICA"

    return "ASSISTÊNCIA"


def nome_plano_exibicao_centro_custos(plano, setor):
    plano = (plano or "SEM PLANO DE CONTAS").strip()

    if setor == "LOGÍSTICA" and not plano.upper().endswith(" T"):
        return f"{plano} T"

    return plano


def normalizar_plano_comparacao(valor, setor=None):
    plano = normalizar_texto(valor) or "SEM PLANO DE CONTAS"

    if setor == "LOGÍSTICA" and plano.endswith(" T"):
        plano = plano[:-2].strip()

    return plano


def montar_linhas_centro_custos(contas, meses_com_movimento):
    agrupado = {}

    for conta in contas:
        mes, _ = mes_referencia_centro_custos(conta)

        if not mes or mes < 1 or mes > 12:
            continue

        setor = setor_centro_custos(conta)
        plano_base = nome_plano_centro_custos(conta)
        plano_chave = normalizar_plano_comparacao(plano_base, setor)
        plano_exibicao = nome_plano_exibicao_centro_custos(plano_base, setor)
        chave = (setor, plano_chave)

        if chave not in agrupado:
            agrupado[chave] = {
                "setor": setor,
                "plano": plano_exibicao,
                "plano_base": plano_base,
                "meses": {i: 0 for i in range(1, 13)},
                "total": 0,
                "media": 0,
            }

        valor = dinheiro(getattr(conta, "valor", 0))
        agrupado[chave]["meses"][mes] += valor
        agrupado[chave]["total"] += valor

    linhas = list(agrupado.values())

    for linha in linhas:
        linha["media"] = linha["total"] / meses_com_movimento if meses_com_movimento else 0

    return sorted(linhas, key=lambda x: (x["setor"], x["plano"]))


def somar_linhas_centro_custos(linhas, meses_com_movimento):
    total_meses = {i: 0 for i in range(1, 13)}
    total_ano = 0

    for linha in linhas:
        for mes in range(1, 13):
            total_meses[mes] += dinheiro(linha["meses"].get(mes, 0))

        total_ano += dinheiro(linha.get("total", 0))

    return {
        "meses": total_meses,
        "total": total_ano,
        "media": total_ano / meses_com_movimento if meses_com_movimento else 0,
    }


def formatar_data_br_centro_custos(valor):
    data = data_para_date(valor)
    return data.strftime("%d/%m/%Y") if data else "-"


@gestao_bp.route("/centro-custos/detalhes")
@gestao_required
def centro_custos_detalhes():
    ano = request.args.get("ano", type=int) or datetime.now().year
    mes_filtro = request.args.get("mes", type=int)
    setor_filtro = normalizar_texto(request.args.get("setor")) or "ASSISTÊNCIA"
    plano_filtro = request.args.get("plano", "").strip()

    if setor_filtro not in ["ASSISTÊNCIA", "LOGÍSTICA"]:
        setor_filtro = "ASSISTÊNCIA"

    plano_comparacao = normalizar_plano_comparacao(plano_filtro, setor_filtro)

    contas = contas_pagas_centro_custos_query(ano).all()

    despesas = []
    total = 0

    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_custos(conta)

        if ano_ref != ano:
            continue

        if mes_filtro and mes_ref != mes_filtro:
            continue

        setor_conta = setor_centro_custos(conta)

        if setor_conta != setor_filtro:
            continue

        plano_conta = nome_plano_centro_custos(conta)

        if normalizar_plano_comparacao(plano_conta, setor_conta) != plano_comparacao:
            continue

        valor = dinheiro(getattr(conta, "valor", 0))
        total += valor

        data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))
        data_vencimento = data_para_date(getattr(conta, "data_vencimento", None))

        despesas.append({
            "id": conta.id,
            "ordem": data_pagamento or date(1900, 1, 1),
            "fornecedor": getattr(conta, "fornecedor_funcionario", None) or "Sem fornecedor informado",
            "numero_fatura": getattr(conta, "numero_fatura", None) or "-",
            "plano": nome_plano_exibicao_centro_custos(plano_conta, setor_conta),
            "setor": setor_conta,
            "mes": mes_ref,
            "data_pagamento": data_pagamento.strftime("%d/%m/%Y") if data_pagamento else "-",
            "data_vencimento": data_vencimento.strftime("%d/%m/%Y") if data_vencimento else "-",
            "valor": formatar_numero_br(valor),
            "status": getattr(conta, "status", None) or "PAGO",
            "observacoes": getattr(conta, "observacoes", None) or "",
        })

    despesas.sort(key=lambda item: (item["ordem"], item["fornecedor"]))

    for item in despesas:
        item.pop("ordem", None)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    return jsonify({
        "ok": True,
        "ano": ano,
        "mes": mes_filtro,
        "mes_nome": nomes_meses.get(mes_filtro, "Ano completo") if mes_filtro else "Ano completo",
        "setor": setor_filtro,
        "plano": nome_plano_exibicao_centro_custos(plano_filtro, setor_filtro),
        "total": formatar_numero_br(total),
        "quantidade": len(despesas),
        "despesas": despesas,
    })


@gestao_bp.route("/centro-custos/exportar")
@gestao_required
def exportar_centro_custos():
    hoje = datetime.now()
    ano = request.args.get("ano", type=int) or hoje.year

    contas = contas_pagas_centro_custos_query(ano).all()

    meses_ativos = set()
    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_custos(conta)
        if ano_ref == ano and mes_ref:
            meses_ativos.add(mes_ref)

    meses_com_movimento = len(meses_ativos)
    linhas = montar_linhas_centro_custos(contas, meses_com_movimento)

    linhas_assistencia = [l for l in linhas if l["setor"] == "ASSISTÊNCIA"]
    linhas_logistica = [l for l in linhas if l["setor"] == "LOGÍSTICA"]

    total_assistencia = somar_linhas_centro_custos(linhas_assistencia, meses_com_movimento)
    total_logistica = somar_linhas_centro_custos(linhas_logistica, meses_com_movimento)
    total_geral = somar_linhas_centro_custos(linhas, meses_com_movimento)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Centro de Custos"

    azul_escuro = "1E3A8A"
    azul_medio = "2563EB"
    verde = "0F766E"
    cinza_fundo = "F8FAFC"
    cinza_borda = "D9E2EC"
    branco = "FFFFFF"
    preto = "111827"

    fill_titulo = PatternFill("solid", fgColor=azul_escuro)
    fill_header = PatternFill("solid", fgColor=azul_medio)
    fill_assistencia = PatternFill("solid", fgColor="EAF2FF")
    fill_logistica = PatternFill("solid", fgColor="E7F7F3")
    fill_total = PatternFill("solid", fgColor="DBEAFE")
    fill_total_geral = PatternFill("solid", fgColor=azul_escuro)
    fill_zebra = PatternFill("solid", fgColor=cinza_fundo)

    font_titulo = Font(color=branco, bold=True, size=14)
    font_header = Font(color=branco, bold=True)
    font_bold = Font(color=preto, bold=True)
    font_total_geral = Font(color=branco, bold=True)
    border = Border(
        left=Side(style="thin", color=cinza_borda),
        right=Side(style="thin", color=cinza_borda),
        top=Side(style="thin", color=cinza_borda),
        bottom=Side(style="thin", color=cinza_borda),
    )

    colunas = ["Plano de Contas"] + [nomes_meses[m] for m in range(1, 13)] + [f"Total {ano}", "Média Mês"]
    total_colunas = len(colunas)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    cell = ws.cell(row=1, column=1, value=f"Centro de Custos - {ano}")
    cell.fill = fill_titulo
    cell.font = font_titulo
    cell.alignment = Alignment(horizontal="center")

    ws.append([])

    def aplicar_header(row_idx):
        for col_idx, titulo in enumerate(colunas, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=titulo)
            c.fill = fill_header
            c.font = font_header
            c.alignment = Alignment(horizontal="center")
            c.border = border

    def escrever_secao(titulo, linhas, totais, fill_secao, label_total):
        row_secao = ws.max_row + 1
        ws.merge_cells(start_row=row_secao, start_column=1, end_row=row_secao, end_column=total_colunas)
        c = ws.cell(row=row_secao, column=1, value=titulo)
        c.fill = fill_secao
        c.font = font_bold
        c.alignment = Alignment(horizontal="left")

        header_row = ws.max_row + 1
        aplicar_header(header_row)

        for idx, linha in enumerate(linhas, start=1):
            valores = [linha["plano"]]
            valores.extend(dinheiro(linha["meses"].get(m, 0)) for m in range(1, 13))
            valores.append(dinheiro(linha.get("total", 0)))
            valores.append(dinheiro(linha.get("media", 0)))
            ws.append(valores)
            row_idx = ws.max_row

            for col_idx in range(1, total_colunas + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = border
                c.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
                if idx % 2 == 0:
                    c.fill = fill_zebra
                if col_idx > 1:
                    c.number_format = '#,##0.00'

        valores_total = [label_total]
        valores_total.extend(dinheiro(totais["meses"].get(m, 0)) for m in range(1, 13))
        valores_total.append(dinheiro(totais.get("total", 0)))
        valores_total.append(dinheiro(totais.get("media", 0)))
        ws.append(valores_total)
        row_idx = ws.max_row
        for col_idx in range(1, total_colunas + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill_total
            c.font = font_bold
            c.border = border
            c.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
            if col_idx > 1:
                c.number_format = '#,##0.00'

        ws.append([])

    escrever_secao(
        "Assistência",
        linhas_assistencia,
        total_assistencia,
        fill_assistencia,
        "Total Assistência Mês",
    )

    escrever_secao(
        "Logística",
        linhas_logistica,
        total_logistica,
        fill_logistica,
        "Total Logística Mês",
    )

    valores_total_geral = [f"Total despesas {ano}"]
    valores_total_geral.extend(dinheiro(total_geral["meses"].get(m, 0)) for m in range(1, 13))
    valores_total_geral.append(dinheiro(total_geral.get("total", 0)))
    valores_total_geral.append(dinheiro(total_geral.get("media", 0)))
    ws.append(valores_total_geral)
    row_idx = ws.max_row
    for col_idx in range(1, total_colunas + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.fill = fill_total_geral
        c.font = font_total_geral
        c.border = border
        c.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")
        if col_idx > 1:
            c.number_format = '#,##0.00'

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, total_colunas + 1):
        letra = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letra].width = 13

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical="center",
                wrap_text=False,
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"centro_custos_{ano}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@gestao_bp.route("/centro-custos")
@gestao_required
def centro_custos():
    hoje = datetime.now()
    ano = request.args.get("ano", type=int) or hoje.year

    contas = contas_pagas_centro_custos_query(ano).all()

    meses_ativos = set()

    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_custos(conta)

        if ano_ref == ano and mes_ref:
            meses_ativos.add(mes_ref)

    meses_com_movimento = len(meses_ativos)

    linhas = montar_linhas_centro_custos(contas, meses_com_movimento)

    linhas_assistencia = [l for l in linhas if l["setor"] == "ASSISTÊNCIA"]
    linhas_logistica = [l for l in linhas if l["setor"] == "LOGÍSTICA"]

    total_assistencia = somar_linhas_centro_custos(linhas_assistencia, meses_com_movimento)
    total_logistica = somar_linhas_centro_custos(linhas_logistica, meses_com_movimento)
    total_geral = somar_linhas_centro_custos(linhas, meses_com_movimento)

    maior_conta = max(linhas, key=lambda x: x["total"], default=None)

    anos_disponiveis = set()

    for conta in contas_pagas_centro_custos_query().all():
        _, ano_ref = mes_referencia_centro_custos(conta)
        if ano_ref:
            anos_disponiveis.add(ano_ref)

    if ano not in anos_disponiveis:
        anos_disponiveis.add(ano)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    return render_template(
        "gestao/centro_custos.html",
        ano=ano,
        anos_disponiveis=sorted(anos_disponiveis, reverse=True),
        meses=range(1, 13),
        nomes_meses=nomes_meses,
        meses_com_movimento=meses_com_movimento,
        linhas_assistencia=linhas_assistencia,
        linhas_logistica=linhas_logistica,
        total_assistencia=total_assistencia,
        total_logistica=total_logistica,
        total_geral=total_geral,
        maior_conta=maior_conta,
        formatar_moeda_br=formatar_moeda_br,
        formatar_numero_br=formatar_numero_br,
    )


# =========================================================
# CENTRO DE RECEITAS
# Tabela gerencial anual por plano de contas
# Fonte oficial: Contas Recebidas importadas
# =========================================================

def periodo_ano_centro_receitas(ano):
    inicio = datetime(ano, 1, 1, 0, 0, 0)
    fim = datetime(ano, 12, 31, 23, 59, 59)
    return inicio, fim


def contas_recebidas_centro_receitas_query(ano=None):
    """
    Centro de Receitas deve bater com o relatório de Contas Recebidas.

    Fonte restrita a:
    - pago=True
    - origem_importacao='RECEBIMENTO'
    - data_pagamento preenchida
    """
    query = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento.isnot(None)
    )

    if ano:
        inicio, fim = periodo_ano_centro_receitas(ano)
        query = query.filter(
            ContaReceberImportada.data_pagamento >= inicio,
            ContaReceberImportada.data_pagamento <= fim
        )

    return query


def mes_referencia_centro_receitas(conta):
    """No Centro de Receitas, a competência é sempre a data real de recebimento."""
    data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))

    if data_pagamento:
        return data_pagamento.month, data_pagamento.year

    return None, None


def valor_centro_receitas(conta):
    return dinheiro(
        getattr(conta, "total", None)
        or getattr(conta, "valor", 0)
    )


def setor_centro_receitas(conta):
    setor = normalizar_texto(getattr(conta, "setor", None)) or "ASSISTÊNCIA"
    plano = normalizar_texto(getattr(conta, "plano_contas", None)) or ""
    categoria = normalizar_texto(getattr(conta, "categoria", None)) or ""

    if setor == "LOGÍSTICA" or plano.endswith(" T") or categoria.endswith(" T"):
        return "LOGÍSTICA"

    return "ASSISTÊNCIA"


def nome_plano_centro_receitas(conta):
    return (
        getattr(conta, "categoria", None)
        or getattr(conta, "plano_contas", None)
        or "SEM PLANO DE CONTAS"
    ).strip()


def nome_plano_exibicao_centro_receitas(plano, setor):
    plano = (plano or "SEM PLANO DE CONTAS").strip()

    if setor == "LOGÍSTICA" and not plano.upper().endswith(" T"):
        return f"{plano} T"

    return plano


def normalizar_plano_comparacao_receitas(valor, setor=None):
    plano = normalizar_texto(valor) or "SEM PLANO DE CONTAS"

    if setor == "LOGÍSTICA" and plano.endswith(" T"):
        plano = plano[:-2].strip()

    return plano


def montar_linhas_centro_receitas(contas, meses_com_movimento):
    agrupado = {}

    for conta in contas:
        mes, _ = mes_referencia_centro_receitas(conta)

        if not mes or mes < 1 or mes > 12:
            continue

        setor = setor_centro_receitas(conta)
        plano_base = nome_plano_centro_receitas(conta)
        plano_chave = normalizar_plano_comparacao_receitas(plano_base, setor)
        plano_exibicao = nome_plano_exibicao_centro_receitas(plano_base, setor)
        chave = (setor, plano_chave)

        if chave not in agrupado:
            agrupado[chave] = {
                "setor": setor,
                "plano": plano_exibicao,
                "plano_base": plano_base,
                "meses": {i: 0 for i in range(1, 13)},
                "total": 0,
                "media": 0,
            }

        valor = valor_centro_receitas(conta)
        agrupado[chave]["meses"][mes] += valor
        agrupado[chave]["total"] += valor

    linhas = list(agrupado.values())

    for linha in linhas:
        linha["media"] = linha["total"] / meses_com_movimento if meses_com_movimento else 0

    return sorted(linhas, key=lambda x: (x["setor"], x["plano"]))


def somar_linhas_centro_receitas(linhas, meses_com_movimento):
    total_meses = {i: 0 for i in range(1, 13)}
    total_ano = 0

    for linha in linhas:
        for mes in range(1, 13):
            total_meses[mes] += dinheiro(linha["meses"].get(mes, 0))

        total_ano += dinheiro(linha.get("total", 0))

    return {
        "meses": total_meses,
        "total": total_ano,
        "media": total_ano / meses_com_movimento if meses_com_movimento else 0,
    }


@gestao_bp.route("/centro-receitas/detalhes")
@gestao_required
def centro_receitas_detalhes():
    ano = request.args.get("ano", type=int) or datetime.now().year
    mes_filtro = request.args.get("mes", type=int)
    setor_filtro = normalizar_texto(request.args.get("setor")) or "ASSISTÊNCIA"
    plano_filtro = request.args.get("plano", "").strip()

    if setor_filtro not in ["ASSISTÊNCIA", "LOGÍSTICA"]:
        setor_filtro = "ASSISTÊNCIA"

    plano_comparacao = normalizar_plano_comparacao_receitas(plano_filtro, setor_filtro)

    contas = contas_recebidas_centro_receitas_query(ano).all()

    receitas = []
    total = 0

    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_receitas(conta)

        if ano_ref != ano:
            continue

        if mes_filtro and mes_ref != mes_filtro:
            continue

        setor_conta = setor_centro_receitas(conta)

        if setor_conta != setor_filtro:
            continue

        plano_conta = nome_plano_centro_receitas(conta)

        if normalizar_plano_comparacao_receitas(plano_conta, setor_conta) != plano_comparacao:
            continue

        valor = valor_centro_receitas(conta)
        total += valor

        data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))
        data_vencimento = data_para_date(getattr(conta, "data_vencimento", None))

        receitas.append({
            "id": conta.id,
            "ordem": data_pagamento or date(1900, 1, 1),
            "cliente": getattr(conta, "cliente", None) or "Sem cliente informado",
            "numero_fatura": getattr(conta, "numero_fatura", None) or "-",
            "plano": nome_plano_exibicao_centro_receitas(plano_conta, setor_conta),
            "setor": setor_conta,
            "mes": mes_ref,
            "data_pagamento": data_pagamento.strftime("%d/%m/%Y") if data_pagamento else "-",
            "data_vencimento": data_vencimento.strftime("%d/%m/%Y") if data_vencimento else "-",
            "valor": formatar_numero_br(valor),
            "status": getattr(conta, "status", None) or "RECEBIDO",
            "observacoes": getattr(conta, "observacoes", None) or "",
        })

    receitas.sort(key=lambda item: (item["ordem"], item["cliente"]))

    for item in receitas:
        item.pop("ordem", None)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    return jsonify({
        "ok": True,
        "ano": ano,
        "mes": mes_filtro,
        "mes_nome": nomes_meses.get(mes_filtro, "Ano completo") if mes_filtro else "Ano completo",
        "setor": setor_filtro,
        "plano": nome_plano_exibicao_centro_receitas(plano_filtro, setor_filtro),
        "total": formatar_numero_br(total),
        "quantidade": len(receitas),
        "receitas": receitas,
    })


@gestao_bp.route("/centro-receitas")
@gestao_required
def centro_receitas():
    hoje = datetime.now()
    ano = request.args.get("ano", type=int) or hoje.year

    contas = contas_recebidas_centro_receitas_query(ano).all()

    meses_ativos = set()

    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_receitas(conta)

        if ano_ref == ano and mes_ref:
            meses_ativos.add(mes_ref)

    meses_com_movimento = len(meses_ativos)

    linhas = montar_linhas_centro_receitas(contas, meses_com_movimento)

    linhas_assistencia = [l for l in linhas if l["setor"] == "ASSISTÊNCIA"]
    linhas_logistica = [l for l in linhas if l["setor"] == "LOGÍSTICA"]

    total_assistencia = somar_linhas_centro_receitas(linhas_assistencia, meses_com_movimento)
    total_logistica = somar_linhas_centro_receitas(linhas_logistica, meses_com_movimento)
    total_geral = somar_linhas_centro_receitas(linhas, meses_com_movimento)

    maior_conta = max(linhas, key=lambda x: x["total"], default=None)

    anos_disponiveis = set()

    for conta in contas_recebidas_centro_receitas_query().all():
        _, ano_ref = mes_referencia_centro_receitas(conta)
        if ano_ref:
            anos_disponiveis.add(ano_ref)

    if ano not in anos_disponiveis:
        anos_disponiveis.add(ano)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    return render_template(
        "gestao/centro_receitas.html",
        ano=ano,
        anos_disponiveis=sorted(anos_disponiveis, reverse=True),
        meses=range(1, 13),
        nomes_meses=nomes_meses,
        meses_com_movimento=meses_com_movimento,
        linhas_assistencia=linhas_assistencia,
        linhas_logistica=linhas_logistica,
        total_assistencia=total_assistencia,
        total_logistica=total_logistica,
        total_geral=total_geral,
        maior_conta=maior_conta,
        formatar_moeda_br=formatar_moeda_br,
        formatar_numero_br=formatar_numero_br,
    )


@gestao_bp.route("/centro-receitas/exportar")
@gestao_required
def exportar_centro_receitas():
    hoje = datetime.now()
    ano = request.args.get("ano", type=int) or hoje.year

    contas = contas_recebidas_centro_receitas_query(ano).all()

    meses_ativos = set()
    for conta in contas:
        mes_ref, ano_ref = mes_referencia_centro_receitas(conta)
        if ano_ref == ano and mes_ref:
            meses_ativos.add(mes_ref)

    meses_com_movimento = len(meses_ativos)
    linhas = montar_linhas_centro_receitas(contas, meses_com_movimento)
    linhas_assistencia = [l for l in linhas if l["setor"] == "ASSISTÊNCIA"]
    linhas_logistica = [l for l in linhas if l["setor"] == "LOGÍSTICA"]
    total_assistencia = somar_linhas_centro_receitas(linhas_assistencia, meses_com_movimento)
    total_logistica = somar_linhas_centro_receitas(linhas_logistica, meses_com_movimento)
    total_geral = somar_linhas_centro_receitas(linhas, meses_com_movimento)

    nomes_meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Centro de Receitas"

    azul = "003B8E"
    verde = "006B63"
    cinza = "F8FAFC"
    borda_cor = "D9E1EA"
    total_fill = "EAF2FF"
    geral_fill = "003B8E"

    thin = Side(style="thin", color=borda_cor)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Plano de Contas"] + [nomes_meses[m] for m in range(1, 13)] + [f"Total {ano}", "Média Mês"]

    def escrever_titulo(row, titulo, cor):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws.cell(row=row, column=1)
        cell.value = titulo
        cell.fill = PatternFill("solid", fgColor=cor)
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="left")
        return row + 1

    def escrever_cabecalho(row):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=cinza)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        return row + 1

    def escrever_linhas(row, linhas, total, titulo_total):
        for linha in linhas:
            ws.cell(row=row, column=1).value = linha["plano"]
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=1).border = border
            for mes in range(1, 13):
                cell = ws.cell(row=row, column=mes + 1)
                cell.value = dinheiro(linha["meses"].get(mes, 0)) or None
                cell.number_format = '#,##0.00'
                cell.border = border
                cell.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=14).value = dinheiro(linha["total"])
            ws.cell(row=row, column=15).value = dinheiro(linha["media"])
            for col in [14, 15]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            row += 1

        ws.cell(row=row, column=1).value = titulo_total
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=total_fill)
            cell.font = Font(bold=True, color="003B8E")
            cell.border = border
        for mes in range(1, 13):
            ws.cell(row=row, column=mes + 1).value = dinheiro(total["meses"].get(mes, 0)) or None
            ws.cell(row=row, column=mes + 1).number_format = '#,##0.00'
        ws.cell(row=row, column=14).value = dinheiro(total["total"])
        ws.cell(row=row, column=15).value = dinheiro(total["media"])
        ws.cell(row=row, column=14).number_format = '#,##0.00'
        ws.cell(row=row, column=15).number_format = '#,##0.00'
        return row + 2

    row = 1
    row = escrever_titulo(row, "ASSISTÊNCIA", azul)
    row = escrever_cabecalho(row)
    row = escrever_linhas(row, linhas_assistencia, total_assistencia, "TOTAL RECEITA ASSISTÊNCIA MÊS")

    row = escrever_titulo(row, "LOGÍSTICA", verde)
    row = escrever_cabecalho(row)
    row = escrever_linhas(row, linhas_logistica, total_logistica, "TOTAL RECEITA LOGÍSTICA MÊS")

    ws.cell(row=row, column=1).value = f"TOTAL RECEITAS {ano}"
    for col in range(1, 16):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=geral_fill if col == 1 else "EEF4FF")
        cell.font = Font(bold=True, color="FFFFFF" if col == 1 else "0F172A")
        cell.border = border
    for mes in range(1, 13):
        ws.cell(row=row, column=mes + 1).value = dinheiro(total_geral["meses"].get(mes, 0)) or None
        ws.cell(row=row, column=mes + 1).number_format = '#,##0.00'
    ws.cell(row=row, column=14).value = dinheiro(total_geral["total"])
    ws.cell(row=row, column=15).value = dinheiro(total_geral["media"])
    ws.cell(row=row, column=14).number_format = '#,##0.00'
    ws.cell(row=row, column=15).number_format = '#,##0.00'

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 28
    for col in range(2, 16):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"centro_receitas_{ano}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================================
# DASHBOARD OPERACIONAL IMPORTADO
# Cole este bloco em routes/gestao.py
# Requer:
# from models.dashboard_operacional_importado import DashboardOperacionalImportado
# import re
# from collections import Counter, defaultdict
# from openpyxl import load_workbook
# =========================================================

def op_texto(valor, padrao="NÃO INFORMADO"):
    if valor is None:
        return padrao

    texto = str(valor).strip()

    if not texto or texto.upper() in ["NAN", "NONE", "-", "/", "43"]:
        return padrao

    texto = " ".join(texto.split()).upper()
    return texto


def op_data(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor

    try:
        return datetime.strptime(str(valor).strip(), "%d/%m/%Y %H:%M:%S")
    except Exception:
        pass

    try:
        return datetime.strptime(str(valor).strip(), "%d/%m/%Y")
    except Exception:
        return None


def op_frota_3_numeros(valor):
    if valor is None:
        return "NÃO INFORMADO"

    texto = str(valor).strip().upper()
    numeros = re.findall(r"\d+", texto)

    if not numeros:
        return "NÃO INFORMADO"

    digitos = "".join(numeros)
    return digitos[:3].zfill(3)


def op_tipo_servico(defeito):
    texto = op_texto(defeito)

    if texto == "NÃO INFORMADO":
        return "NÃO INFORMADO"

    if "CONTRATO" in texto:
        return "CONTRATO"

    if (
        "REFRIG" in texto
        or "COMPRESSOR" in texto
        or "GÁS" in texto
        or "GAS" in texto
        or "PLACA EUT" in texto
    ):
        return "REFRIGERAÇÃO"

    if (
        "ESTRUT" in texto
        or "BAÚ" in texto
        or "BAU" in texto
        or "LAMINA" in texto
        or "PISO" in texto
        or "PORTA" in texto
        or "PARACHOQUE" in texto
    ):
        return "ESTRUTURAL"

    if "ENCARRO" in texto or "DESENCARRO" in texto:
        return "ENCARROÇAMENTO"

    if "REVISÃO" in texto or "REVISAO" in texto or "KM" in texto:
        return "REVISÃO / KM"

    if "BATERIA" in texto or "ELÉTR" in texto or "ELETR" in texto or "CHICOTE" in texto:
        return "ELÉTRICA"

    if "PINTURA" in texto or "ADESIVO" in texto:
        return "PINTURA / ADESIVO"

    return "OUTROS SERVIÇOS"


def op_encontrar_header(ws):
    """
    Procura a linha do cabeçalho.
    O arquivo enviado normalmente tem cabeçalho na linha 2.
    """
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        valores = [
            op_texto(ws.cell(row=row_idx, column=col_idx).value, "")
            for col_idx in range(1, ws.max_column + 1)
        ]

        if "CLIENTE" in valores and any("ENTRADA" in v for v in valores):
            return row_idx

    return 1


def op_mapa_colunas(ws, header_row):
    mapa = {}

    for col_idx in range(1, ws.max_column + 1):
        nome = op_texto(ws.cell(row=header_row, column=col_idx).value, "")
        if nome:
            mapa[nome] = col_idx

    return mapa


def op_col(mapa, nomes):
    for nome in nomes:
        for chave, indice in mapa.items():
            if nome in chave:
                return indice

    return None


def op_valor(ws, row_idx, col_idx):
    if not col_idx:
        return None

    return ws.cell(row=row_idx, column=col_idx).value


def op_processar_excel_dashboard(arquivo):
    wb = load_workbook(arquivo, data_only=True)
    ws = wb.active

    header_row = op_encontrar_header(ws)
    mapa = op_mapa_colunas(ws, header_row)

    c_cliente = op_col(mapa, ["CLIENTE"])
    c_cidade = op_col(mapa, ["CIDADE"])
    c_entrada = op_col(mapa, ["ENTRADA"])
    c_pronto = op_col(mapa, ["PRONTO"])
    c_saida = op_col(mapa, ["SAÍDA", "SAIDA"])
    c_situacao = op_col(mapa, ["SITUAÇÃO", "SITUACAO"])
    c_veiculo = op_col(mapa, ["VEÍCULO", "VEICULO"])
    c_defeito = op_col(mapa, ["DEFEITO"])
    c_marca = op_col(mapa, ["MARCA"])
    c_frota = op_col(mapa, ["FROTA"])
    c_km = op_col(mapa, ["KM"])

    registros = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        cliente = op_texto(op_valor(ws, row_idx, c_cliente))
        cidade = op_texto(op_valor(ws, row_idx, c_cidade))

        entrada = op_data(op_valor(ws, row_idx, c_entrada))
        pronto = op_data(op_valor(ws, row_idx, c_pronto))
        saida = op_data(op_valor(ws, row_idx, c_saida))

        data_base = entrada or saida or pronto

        # ignora linhas completamente vazias ou sem data útil
        if cliente == "NÃO INFORMADO" and not data_base:
            continue

        if not data_base:
            continue

        defeito = op_texto(op_valor(ws, row_idx, c_defeito))
        frota_original = op_texto(op_valor(ws, row_idx, c_frota))

        registros.append(DashboardOperacionalImportado(
            cliente=cliente,
            cidade_uf=cidade,
            entrada=entrada,
            pronto=pronto,
            saida=saida,
            situacao=op_texto(op_valor(ws, row_idx, c_situacao)),
            veiculo=op_texto(op_valor(ws, row_idx, c_veiculo)),
            defeito=defeito,
            tipo_servico=op_tipo_servico(defeito),
            marca=op_texto(op_valor(ws, row_idx, c_marca)),
            frota_original=frota_original,
            frota_tratada=op_frota_3_numeros(frota_original),
            km=op_texto(op_valor(ws, row_idx, c_km)),
            ano=data_base.year,
            mes=data_base.month,
        ))

    return registros


def op_counter_top(registros, atributo, limite=8, ignorar_nao_informado=True):
    contador = Counter()

    for r in registros:
        valor = getattr(r, atributo, None) or "NÃO INFORMADO"

        if ignorar_nao_informado and valor == "NÃO INFORMADO":
            continue

        contador[valor] += 1

    return contador.most_common(limite)


def op_series_anual(registros):
    contador = Counter()

    for r in registros:
        if r.ano:
            contador[r.ano] += 1

    anos = list(range(min(contador.keys()), max(contador.keys()) + 1)) if contador else []

    return {
        "labels": [str(ano) for ano in anos],
        "valores": [contador.get(ano, 0) for ano in anos],
    }


def op_series_mensal_comparativa(registros, anos_comparar):
    nomes_meses = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
    dados = {}

    for ano in anos_comparar:
        dados[str(ano)] = [0] * 12

    for r in registros:
        if r.ano in anos_comparar and r.mes:
            dados[str(r.ano)][r.mes - 1] += 1

    return {
        "labels": nomes_meses,
        "series": dados,
    }


@gestao_bp.route("/dashboard-operacional", methods=["GET"])
@gestao_required
def dashboard_operacional():
    filtros = {
        "ano": request.args.get("ano", type=int),
        "mes": request.args.get("mes", type=int),
        "cliente": request.args.get("cliente", "").strip(),
        "cidade": request.args.get("cidade", "").strip(),
        "frota": request.args.get("frota", "").strip(),
        "veiculo": request.args.get("veiculo", "").strip(),
        "servico": request.args.get("servico", "").strip(),
        "situacao": request.args.get("situacao", "").strip(),
    }

    query = DashboardOperacionalImportado.query

    if filtros["ano"]:
        query = query.filter(DashboardOperacionalImportado.ano == filtros["ano"])

    if filtros["mes"]:
        query = query.filter(DashboardOperacionalImportado.mes == filtros["mes"])

    if filtros["cliente"]:
        query = query.filter(DashboardOperacionalImportado.cliente == filtros["cliente"])

    if filtros["cidade"]:
        query = query.filter(DashboardOperacionalImportado.cidade_uf == filtros["cidade"])

    if filtros["frota"]:
        query = query.filter(DashboardOperacionalImportado.frota_tratada == filtros["frota"])

    if filtros["veiculo"]:
        query = query.filter(DashboardOperacionalImportado.veiculo == filtros["veiculo"])

    if filtros["servico"]:
        query = query.filter(DashboardOperacionalImportado.tipo_servico == filtros["servico"])

    if filtros["situacao"]:
        query = query.filter(DashboardOperacionalImportado.situacao == filtros["situacao"])

    registros = query.all()
    todos_registros = DashboardOperacionalImportado.query.all()

    def distinct_coluna(coluna):
        return [
            item[0] for item in db.session.query(coluna)
            .filter(coluna.isnot(None))
            .filter(coluna != "NÃO INFORMADO")
            .distinct()
            .order_by(coluna)
            .all()
        ]

    anos_disponiveis = [
        item[0] for item in db.session.query(DashboardOperacionalImportado.ano)
        .filter(DashboardOperacionalImportado.ano.isnot(None))
        .distinct()
        .order_by(DashboardOperacionalImportado.ano)
        .all()
    ]

    meses_disponiveis = [
        {"numero": 1, "nome": "Janeiro"},
        {"numero": 2, "nome": "Fevereiro"},
        {"numero": 3, "nome": "Março"},
        {"numero": 4, "nome": "Abril"},
        {"numero": 5, "nome": "Maio"},
        {"numero": 6, "nome": "Junho"},
        {"numero": 7, "nome": "Julho"},
        {"numero": 8, "nome": "Agosto"},
        {"numero": 9, "nome": "Setembro"},
        {"numero": 10, "nome": "Outubro"},
        {"numero": 11, "nome": "Novembro"},
        {"numero": 12, "nome": "Dezembro"},
    ]

    clientes_disponiveis = distinct_coluna(DashboardOperacionalImportado.cliente)
    cidades_disponiveis = distinct_coluna(DashboardOperacionalImportado.cidade_uf)
    frotas_disponiveis = distinct_coluna(DashboardOperacionalImportado.frota_tratada)
    veiculos_disponiveis = distinct_coluna(DashboardOperacionalImportado.veiculo)
    servicos_disponiveis = distinct_coluna(DashboardOperacionalImportado.tipo_servico)
    situacoes_disponiveis = distinct_coluna(DashboardOperacionalImportado.situacao)

    total_os = len(registros)
    clientes = len(set(r.cliente for r in registros if r.cliente and r.cliente != "NÃO INFORMADO"))
    frotas = len(set(r.frota_tratada for r in registros if r.frota_tratada and r.frota_tratada != "NÃO INFORMADO"))
    cidades = len(set(r.cidade_uf for r in registros if r.cidade_uf and r.cidade_uf != "NÃO INFORMADO"))
    veiculos = len(set(r.veiculo for r in registros if r.veiculo and r.veiculo != "NÃO INFORMADO"))

    anos_para_mensal = []
    if filtros["ano"]:
        anos_para_mensal = [filtros["ano"]]
    elif anos_disponiveis:
        anos_para_mensal = anos_disponiveis[-2:]

    total_base = DashboardOperacionalImportado.query.count()

    return render_template(
        "gestao/dashboard_operacional.html",
        filtros=filtros,
        ano=filtros["ano"],
        mes=filtros["mes"],
        cliente=filtros["cliente"],

        anos_disponiveis=anos_disponiveis,
        meses_disponiveis=meses_disponiveis,
        clientes_disponiveis=clientes_disponiveis,
        cidades_disponiveis=cidades_disponiveis,
        frotas_disponiveis=frotas_disponiveis,
        veiculos_disponiveis=veiculos_disponiveis,
        servicos_disponiveis=servicos_disponiveis,
        situacoes_disponiveis=situacoes_disponiveis,

        total_os=total_os,
        clientes=clientes,
        frotas=frotas,
        cidades=cidades,
        veiculos=veiculos,
        total_base=total_base,

        top_clientes=op_counter_top(registros, "cliente", 8),
        top_cidades=op_counter_top(registros, "cidade_uf", 8),
        top_frotas=op_counter_top(registros, "frota_tratada", 8),
        top_veiculos=op_counter_top(registros, "veiculo", 8),
        top_servicos=op_counter_top(registros, "tipo_servico", 8, ignorar_nao_informado=False),
        top_situacoes=op_counter_top(registros, "situacao", 8, ignorar_nao_informado=False),

        # Agora os gráficos também respeitam os filtros atuais.
        grafico_anual=op_series_anual(registros),
        grafico_mensal=op_series_mensal_comparativa(registros, anos_para_mensal),
    )


@gestao_bp.route("/api/dashboard-operacional-detalhes", methods=["GET"])
@gestao_required
def api_dashboard_operacional_detalhes():
    filtros = {
        "ano": request.args.get("ano", type=int),
        "mes": request.args.get("mes", type=int),
        "cliente": request.args.get("cliente", "").strip(),
        "cidade": request.args.get("cidade", "").strip(),
        "frota": request.args.get("frota", "").strip(),
        "veiculo": request.args.get("veiculo", "").strip(),
        "servico": request.args.get("servico", "").strip(),
        "situacao": request.args.get("situacao", "").strip(),
    }

    query = DashboardOperacionalImportado.query

    if filtros["ano"]:
        query = query.filter(DashboardOperacionalImportado.ano == filtros["ano"])
    if filtros["mes"]:
        query = query.filter(DashboardOperacionalImportado.mes == filtros["mes"])
    if filtros["cliente"]:
        query = query.filter(DashboardOperacionalImportado.cliente == filtros["cliente"])
    if filtros["cidade"]:
        query = query.filter(DashboardOperacionalImportado.cidade_uf == filtros["cidade"])
    if filtros["frota"]:
        query = query.filter(DashboardOperacionalImportado.frota_tratada == filtros["frota"])
    if filtros["veiculo"]:
        query = query.filter(DashboardOperacionalImportado.veiculo == filtros["veiculo"])
    if filtros["servico"]:
        query = query.filter(DashboardOperacionalImportado.tipo_servico == filtros["servico"])
    if filtros["situacao"]:
        query = query.filter(DashboardOperacionalImportado.situacao == filtros["situacao"])

    registros = query.order_by(
        DashboardOperacionalImportado.entrada.desc().nullslast(),
        DashboardOperacionalImportado.id.desc()
    ).limit(500).all()

    def data_fmt(valor):
        if not valor:
            return "-"
        return valor.strftime("%d/%m/%Y")

    itens = []

    for r in registros:
        itens.append({
            "cliente": r.cliente or "-",
            "cidade": r.cidade_uf or "-",
            "entrada": data_fmt(r.entrada),
            "saida": data_fmt(r.saida),
            "situacao": r.situacao or "-",
            "veiculo": r.veiculo or "-",
            "frota": r.frota_tratada or "-",
            "servico": r.tipo_servico or "-",
            "defeito": r.defeito or "-",
        })

    return jsonify({
        "ok": True,
        "total": query.count(),
        "itens": itens,
    })




@gestao_bp.route("/dashboard-operacional/importar", methods=["POST"])
@gestao_required
def dashboard_operacional_importar():
    arquivo = request.files.get("arquivo")

    if not arquivo or not arquivo.filename:
        flash("Selecione um arquivo Excel para importar.", "error")
        return redirect("/gestao/dashboard-operacional")

    substituir = request.form.get("substituir") == "1"

    try:
        registros = op_processar_excel_dashboard(arquivo)

        if substituir:
            DashboardOperacionalImportado.query.delete()
            db.session.commit()

        if registros:
            db.session.bulk_save_objects(registros)
            db.session.commit()

        flash(f"Dashboard operacional importado com sucesso. {len(registros)} registro(s) processado(s).", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao importar arquivo: {str(e)}", "error")

    return redirect("/gestao/dashboard-operacional")


@gestao_bp.route("/dashboard-operacional/limpar", methods=["POST"])
@gestao_required
def dashboard_operacional_limpar():
    try:
        DashboardOperacionalImportado.query.delete()
        db.session.commit()
        flash("Base do Dashboard Operacional foi limpa com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar base: {str(e)}", "error")

    return redirect("/gestao/dashboard-operacional")

# =========================================================
# FATURAMENTO COMERCIAL
# =========================================================

class _PrazoRecebimento:
    def __init__(self, data, cliente, valor):
        self.data = data
        self.cliente = cliente
        self.valor = valor


def _primeiro_vencimento(registro):
    datas = [
        getattr(registro, "vencimento_30", None),
        getattr(registro, "vencimento_60", None),
        getattr(registro, "vencimento_90", None),
    ]
    datas = [d for d in datas if d]
    return min(datas) if datas else None


def _proximo_vencimento(registro, hoje):
    datas = [
        getattr(registro, "vencimento_30", None),
        getattr(registro, "vencimento_60", None),
        getattr(registro, "vencimento_90", None),
    ]
    datas = sorted([d for d in datas if d and d >= hoje])
    return datas[0] if datas else None



@gestao_bp.route("/faturamento-comercial", methods=["GET", "POST"])
@gestao_required
def faturamento_comercial():
    from models.faturamento_comercial import FaturamentoComercial

    hoje = date.today()

    if request.method == "POST":
        registro_id = request.form.get("registro_id")

        if registro_id:
            registro = FaturamentoComercial.query.get_or_404(registro_id)
        else:
            registro = FaturamentoComercial()
            db.session.add(registro)

        registro.cliente = normalizar_texto(request.form.get("cliente"))
        registro.os = normalizar_texto(request.form.get("os"))
        registro.nfse = normalizar_texto(request.form.get("nfse"))
        registro.pedido_compras = normalizar_texto(request.form.get("pedido_compras"))
        registro.valor = dinheiro(request.form.get("valor"))
        registro.comissao = dinheiro(request.form.get("comissao"))
        registro.data_emissao = parse_data(request.form.get("data_emissao"))
        registro.vencimento_30 = parse_data(request.form.get("vencimento_30"))
        registro.vencimento_60 = parse_data(request.form.get("vencimento_60"))
        registro.vencimento_90 = parse_data(request.form.get("vencimento_90"))
        registro.status = normalizar_texto(request.form.get("status")) or "PENDENTE"
        registro.tipo = normalizar_texto(request.form.get("tipo")) or "CONTRATO"
        registro.comercial = normalizar_texto(request.form.get("comercial"))
        registro.observacoes = request.form.get("observacoes")
        registro.pago = registro.status in ["RECEBIDO", "NOTAS PAGAS", "PAGO"]

        data_base = registro.data_emissao or registro.vencimento_30 or hoje
        registro.mes = data_base.month
        registro.ano = data_base.year

        db.session.commit()

        flash("Faturamento comercial salvo com sucesso.", "success")

        mes_retorno = request.form.get("mes_retorno") or registro.mes
        ano_retorno = request.form.get("ano_retorno") or registro.ano

        return redirect(f"/gestao/faturamento-comercial?mes={mes_retorno}&ano={ano_retorno}")

    mes = request.args.get("mes", hoje.month, type=int)
    ano = request.args.get("ano", hoje.year, type=int)
    cliente_filtro = normalizar_texto(request.args.get("cliente"))
    status_filtro = normalizar_texto(request.args.get("status"))
    tipo_filtro = normalizar_texto(request.args.get("tipo"))
    vencimento_filtro = normalizar_texto(request.args.get("vencimento"))

    sort_col = (request.args.get("sort") or "vencimento").strip().lower()
    sort_order = (request.args.get("order") or "asc").strip().lower()

    if sort_col not in ["cliente", "os", "nfse", "valor", "emissao", "vencimento", "status", "tipo"]:
        sort_col = "vencimento"

    if sort_order not in ["asc", "desc"]:
        sort_order = "asc"

    periodo_inicio = primeiro_dia_mes(mes, ano)
    periodo_fim = ultimo_dia_mes(mes, ano)

    # A tela agora trabalha por competência direta: mês + ano.
    # A sincronização continua trazendo todos os registros de Contas a Receber,
    # mas a visualização fica limpa por mês selecionado.
    query = FaturamentoComercial.query.filter(
        FaturamentoComercial.mes == mes,
        FaturamentoComercial.ano == ano,
    )

    if cliente_filtro:
        query = query.filter(FaturamentoComercial.cliente.ilike(f"%{cliente_filtro}%"))

    if status_filtro:
        query = query.filter(FaturamentoComercial.status == status_filtro)

    if tipo_filtro:
        query = query.filter(FaturamentoComercial.tipo == tipo_filtro)

    colunas_ordenacao = {
        "cliente": FaturamentoComercial.cliente,
        "os": FaturamentoComercial.os,
        "nfse": FaturamentoComercial.nfse,
        "valor": FaturamentoComercial.valor,
        "emissao": FaturamentoComercial.data_emissao,
        "vencimento": FaturamentoComercial.vencimento_30,
        "status": FaturamentoComercial.status,
        "tipo": FaturamentoComercial.tipo,
    }

    coluna_ordenacao = colunas_ordenacao.get(sort_col, FaturamentoComercial.vencimento_30)

    if sort_order == "desc":
        ordenacao_principal = coluna_ordenacao.desc().nullslast()
    else:
        ordenacao_principal = coluna_ordenacao.asc().nullslast()

    registros = query.order_by(
        ordenacao_principal,
        FaturamentoComercial.id.desc()
    ).all()

    if vencimento_filtro:
        hoje_ref = hoje
        filtrados = []
        for r in registros:
            datas_venc = [r.vencimento_30, r.vencimento_60, r.vencimento_90]
            datas_venc = [d for d in datas_venc if d]
            if vencimento_filtro == "ATRASADO":
                if (not r.pago) and any(d < hoje_ref for d in datas_venc):
                    filtrados.append(r)
            elif vencimento_filtro == "PROXIMOS_30":
                limite = hoje_ref + timedelta(days=30)
                if (not r.pago) and any(hoje_ref <= d <= limite for d in datas_venc):
                    filtrados.append(r)
            elif vencimento_filtro == "TODOS":
                filtrados.append(r)
        registros = filtrados

    def _is_recebido(registro):
        status = normalizar_texto(getattr(registro, "status", None)) or ""
        return bool(getattr(registro, "pago", False)) or status in ["RECEBIDO", "NOTAS PAGAS", "PAGO", "QUITADO", "BAIXADO"]

    registros_validos = [r for r in registros if (normalizar_texto(r.status) or "") != "CANCELADO"]

    total_notas = sum(dinheiro(r.valor) for r in registros_validos)
    total_recebido = sum(dinheiro(r.valor) for r in registros_validos if _is_recebido(r))
    total_pendente = sum(dinheiro(r.valor) for r in registros_validos if not _is_recebido(r))
    total_comissao = sum(dinheiro(r.comissao) for r in registros_validos)

    percentual_recebido = (total_recebido / total_notas * 100) if total_notas else 0
    percentual_aberto = (total_pendente / total_notas * 100) if total_notas else 0
    percentual_comissao = (total_comissao / total_notas * 100) if total_notas else 0

    tipo_counter = Counter()
    clientes_counter = Counter()

    proximos_recebimentos = []
    atrasadas = []

    for r in registros_validos:
        valor = dinheiro(r.valor)
        tipo_counter[r.tipo or "SEM TIPO"] += valor
        clientes_counter[r.cliente or "SEM CLIENTE"] += valor

        if not _is_recebido(r):
            prox = _proximo_vencimento(r, hoje)
            primeiro = _primeiro_vencimento(r)

            if prox and prox <= hoje + timedelta(days=30):
                proximos_recebimentos.append(_PrazoRecebimento(prox, r.cliente or "SEM CLIENTE", valor))
            elif primeiro and primeiro < hoje:
                atrasadas.append(_PrazoRecebimento(primeiro, r.cliente or "SEM CLIENTE", valor))

    proximos_recebimentos = sorted(proximos_recebimentos, key=lambda x: x.data)[:6]
    atrasadas = sorted(atrasadas, key=lambda x: x.data)[:6]

    total_por_tipo = {
        "CONTRATO": tipo_counter.get("CONTRATO", 0),
        "EXTRA": tipo_counter.get("EXTRA", 0),
        "PARTICULAR": tipo_counter.get("PARTICULAR", 0),
    }

    total_notas_pagas = total_recebido

    return render_template(
        "gestao/faturamento_comercial.html",
        mes=mes,
        ano=ano,
        inicio=periodo_inicio,
        fim=periodo_fim,
        cliente_filtro=cliente_filtro,
        status_filtro=status_filtro,
        tipo_filtro=tipo_filtro,
        vencimento_filtro=vencimento_filtro,
        sort_col=sort_col,
        sort_order=sort_order,
        registros=registros,
        total_notas=total_notas,
        total_recebido=total_recebido,
        total_pendente=total_pendente,
        total_comissao=total_comissao,
        percentual_recebido=percentual_recebido,
        percentual_aberto=percentual_aberto,
        percentual_comissao=percentual_comissao,
        totais_tipo=tipo_counter.most_common(),
        total_por_tipo=total_por_tipo,
        total_notas_pagas=total_notas_pagas,
        ranking_clientes=clientes_counter.most_common(8),
        proximos_recebimentos=proximos_recebimentos,
        atrasadas=atrasadas,
    )




@gestao_bp.route("/faturamento-comercial/exportar")
@gestao_required
def exportar_faturamento_comercial():
    from models.faturamento_comercial import FaturamentoComercial

    hoje = date.today()
    mes = request.args.get("mes", hoje.month, type=int)
    ano = request.args.get("ano", hoje.year, type=int)
    cliente_filtro = normalizar_texto(request.args.get("cliente"))
    status_filtro = normalizar_texto(request.args.get("status"))
    tipo_filtro = normalizar_texto(request.args.get("tipo"))
    vencimento_filtro = normalizar_texto(request.args.get("vencimento"))

    query = FaturamentoComercial.query.filter(
        FaturamentoComercial.mes == mes,
        FaturamentoComercial.ano == ano,
    )

    if cliente_filtro:
        query = query.filter(FaturamentoComercial.cliente.ilike(f"%{cliente_filtro}%"))

    if status_filtro:
        query = query.filter(FaturamentoComercial.status == status_filtro)

    if tipo_filtro:
        query = query.filter(FaturamentoComercial.tipo == tipo_filtro)

    colunas_ordenacao = {
        "cliente": FaturamentoComercial.cliente,
        "os": FaturamentoComercial.os,
        "nfse": FaturamentoComercial.nfse,
        "valor": FaturamentoComercial.valor,
        "emissao": FaturamentoComercial.data_emissao,
        "vencimento": FaturamentoComercial.vencimento_30,
        "status": FaturamentoComercial.status,
        "tipo": FaturamentoComercial.tipo,
    }

    coluna_ordenacao = colunas_ordenacao.get(sort_col, FaturamentoComercial.vencimento_30)

    if sort_order == "desc":
        ordenacao_principal = coluna_ordenacao.desc().nullslast()
    else:
        ordenacao_principal = coluna_ordenacao.asc().nullslast()

    registros = query.order_by(
        ordenacao_principal,
        FaturamentoComercial.id.desc()
    ).all()

    if vencimento_filtro:
        filtrados = []
        limite = hoje + timedelta(days=30)
        for r in registros:
            status = normalizar_texto(getattr(r, "status", None)) or ""
            recebido = bool(getattr(r, "pago", False)) or status in ["RECEBIDO", "NOTAS PAGAS", "PAGO", "QUITADO", "BAIXADO"]
            datas_venc = [r.vencimento_30, r.vencimento_60, r.vencimento_90]
            datas_venc = [d for d in datas_venc if d]

            if vencimento_filtro == "ATRASADO" and (not recebido) and any(d < hoje for d in datas_venc):
                filtrados.append(r)
            elif vencimento_filtro == "PROXIMOS_30" and (not recebido) and any(hoje <= d <= limite for d in datas_venc):
                filtrados.append(r)

        registros = filtrados

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faturamento Comercial"

    cabecalhos = [
        "Cliente", "O.S", "NFSE", "Pedido de Compras", "Valor", "Comissão Comercial",
        "Data de Emissão", "30 Dias", "60 Dias", "90 Dias", "Status", "Tipo", "Comercial", "Observações"
    ]
    ws.append(cabecalhos)

    header_fill = PatternFill("solid", fgColor="002244")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in registros:
        ws.append([
            r.cliente or "",
            r.os or "",
            r.nfse or "",
            r.pedido_compras or "",
            float(dinheiro(r.valor)),
            float(dinheiro(r.comissao)),
            r.data_emissao.strftime("%d/%m/%Y") if r.data_emissao else "",
            r.vencimento_30.strftime("%d/%m/%Y") if r.vencimento_30 else "",
            r.vencimento_60.strftime("%d/%m/%Y") if r.vencimento_60 else "",
            r.vencimento_90.strftime("%d/%m/%Y") if r.vencimento_90 else "",
            r.status or "",
            r.tipo or "",
            r.comercial or "",
            r.observacoes or "",
        ])

    widths = [32, 12, 14, 22, 14, 18, 16, 14, 14, 14, 16, 14, 20, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '"R$" #,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"faturamento_comercial_{mes:02d}_{ano}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@gestao_bp.route("/faturamento-comercial/sincronizar", methods=["POST"])
@gestao_required
def sincronizar_faturamento_comercial():
    from models.faturamento_comercial import FaturamentoComercial

    contas = ContaReceberImportada.query.all()

    criados = 0
    atualizados = 0
    ignorados = 0

    for conta in contas:
        cliente = normalizar_texto(getattr(conta, "cliente", None)) or "SEM CLIENTE"
        nfse = normalizar_texto(getattr(conta, "numero_fatura", None)) or "-"
        valor = dinheiro(getattr(conta, "total", None) or getattr(conta, "valor", 0))

        data_documento = data_para_date(getattr(conta, "data_documento", None))
        data_vencimento = data_para_date(getattr(conta, "data_vencimento", None))
        data_pagamento = data_para_date(getattr(conta, "data_pagamento", None))

        data_base = data_documento or data_vencimento or data_pagamento
        mes_ref = getattr(conta, "mes", None) or (data_base.month if data_base else None)
        ano_ref = getattr(conta, "ano", None) or (data_base.year if data_base else None)

        if not mes_ref or not ano_ref:
            ignorados += 1
            continue

        chave_conciliacao = normalizar_texto(getattr(conta, "chave_conciliacao", None))

        if not chave_conciliacao:
            chave_conciliacao = "|".join([
                str(getattr(conta, "id", "")),
                cliente or "SEM CLIENTE",
                nfse or "-",
                str(mes_ref),
                str(ano_ref),
            ])

        registro = FaturamentoComercial.query.filter(
            FaturamentoComercial.chave_conciliacao == chave_conciliacao
        ).first()

        if not registro:
            # Compatibilidade com registros antigos, criados antes do campo chave_conciliacao.
            # Esse fallback evita criar duplicidade na primeira sincronização após a atualização.
            registro = FaturamentoComercial.query.filter(
                FaturamentoComercial.cliente == cliente,
                FaturamentoComercial.nfse == nfse,
                FaturamentoComercial.mes == mes_ref,
                FaturamentoComercial.ano == ano_ref,
            ).first()

        pago = bool(getattr(conta, "pago", False))
        status_importado = normalizar_texto(getattr(conta, "status", None))
        status = "RECEBIDO" if pago or status_importado in ["RECEBIDO", "PAGO", "QUITADO", "BAIXADO"] else "PENDENTE"

        if registro:
            # Atualiza os campos que vêm da importação, mantendo os campos comerciais manuais.
            registro.cliente = cliente
            registro.nfse = nfse
            registro.valor = valor
            registro.data_emissao = data_documento or registro.data_emissao
            registro.vencimento_30 = data_vencimento or registro.vencimento_30
            registro.observacoes = getattr(conta, "observacoes", None) or registro.observacoes
            # Atualiza pagamento sem deixar a sincronização desfazer baixa manual.
            ja_recebido_manual = bool(registro.pago) or normalizar_texto(registro.status) in [
                "RECEBIDO", "NOTAS PAGAS", "PAGO", "QUITADO", "BAIXADO"
            ]

            if ja_recebido_manual:
                registro.pago = True
                registro.status = "RECEBIDO"
            else:
                registro.pago = pago
                registro.status = status

            # Preserva a competência original do faturamento já criado.
            # Evita a nota sumir do mês atual e voltar para mês anterior como pendente.
            if not registro.mes:
                registro.mes = mes_ref
            if not registro.ano:
                registro.ano = ano_ref

            registro.chave_conciliacao = chave_conciliacao
            registro.origem_registro = "SINCRONIZADO"
            atualizados += 1
            continue

        novo = FaturamentoComercial(
            cliente=cliente,
            nfse=nfse,
            valor=valor,
            comissao=0,
            data_emissao=data_documento,
            vencimento_30=data_vencimento,
            vencimento_60=None,
            vencimento_90=None,
            status=status,
            tipo="CONTRATO",
            observacoes=getattr(conta, "observacoes", None),
            pago=pago,
            mes=mes_ref,
            ano=ano_ref,
            chave_conciliacao=chave_conciliacao,
            origem_registro="SINCRONIZADO",
        )

        db.session.add(novo)
        criados += 1

    db.session.commit()

    mes_retorno = request.form.get("mes", type=int) or date.today().month
    ano_retorno = request.form.get("ano", type=int) or date.today().year

    flash(
        f"Sincronização concluída: {criados} novos, {atualizados} atualizados e {ignorados} ignorados.",
        "success"
    )

    return redirect(f"/gestao/faturamento-comercial?mes={mes_retorno}&ano={ano_retorno}")


@gestao_bp.route("/faturamento-comercial/limpar-sincronizados", methods=["POST"])
@gestao_required
def limpar_faturamento_comercial_sincronizados():
    from models.faturamento_comercial import FaturamentoComercial

    mes = request.form.get("mes", type=int) or date.today().month
    ano = request.form.get("ano", type=int) or date.today().year

    try:
        removidos = FaturamentoComercial.query.filter(
            FaturamentoComercial.mes == mes,
            FaturamentoComercial.ano == ano,
            FaturamentoComercial.origem_registro == "SINCRONIZADO",
        ).delete(synchronize_session=False)

        db.session.commit()

        flash(
            f"{removidos} registro(s) sincronizado(s) removido(s). As notas manuais foram mantidas.",
            "success"
        )
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar registros sincronizados: {str(e)}", "danger")

    return redirect(f"/gestao/faturamento-comercial?mes={mes}&ano={ano}")


@gestao_bp.route("/faturamento-comercial/<int:registro_id>/status", methods=["POST"])
@gestao_required
def faturamento_comercial_status(registro_id):
    from models.faturamento_comercial import FaturamentoComercial

    registro = FaturamentoComercial.query.get_or_404(registro_id)

    mes = request.form.get("mes", type=int) or registro.mes or date.today().month
    ano = request.form.get("ano", type=int) or registro.ano or date.today().year
    status = normalizar_texto(request.form.get("status")) or "PENDENTE"

    status_recebido = [
        "RECEBIDO",
        "NOTAS PAGAS",
        "PAGO",
        "QUITADO",
        "BAIXADO",
    ]

    registro.status = status
    registro.pago = status in status_recebido

    # Não deixa a baixa alterar a competência do faturamento.
    if not registro.mes:
        registro.mes = mes
    if not registro.ano:
        registro.ano = ano

    # Se a nota veio da sincronização, baixa também a conta importada de origem.
    # Assim a próxima sincronização não recria/atualiza como PENDENTE.
    if registro.chave_conciliacao:
        conta = ContaReceberImportada.query.filter(
            ContaReceberImportada.chave_conciliacao == registro.chave_conciliacao
        ).first()

        if conta:
            conta.pago = bool(registro.pago)
            conta.status = "RECEBIDO" if registro.pago else "PENDENTE"

            if registro.pago and not getattr(conta, "data_pagamento", None):
                conta.data_pagamento = datetime.combine(date.today(), datetime.min.time())

    db.session.commit()

    flash("Status atualizado com sucesso.", "success")
    return redirect(f"/gestao/faturamento-comercial?mes={mes}&ano={ano}")


@gestao_bp.route("/faturamento-comercial/<int:registro_id>/excluir", methods=["POST"])
@gestao_required
def faturamento_comercial_excluir(registro_id):
    from models.faturamento_comercial import FaturamentoComercial

    registro = FaturamentoComercial.query.get_or_404(registro_id)

    mes = request.form.get("mes") or registro.mes or date.today().month
    ano = request.form.get("ano") or registro.ano or date.today().year

    db.session.delete(registro)
    db.session.commit()

    flash("Lançamento excluído com sucesso.", "success")
    return redirect(f"/gestao/faturamento-comercial?mes={mes}&ano={ano}")



# =========================================================
# PLANEJAMENTO FINANCEIRO
# =========================================================
# A tela nova de planejamento financeiro fica no blueprint próprio:
# routes/planejamento_financeiro.py
#
# Aqui ficam apenas os redirects para links antigos do sistema.
# Isso evita conflito com o blueprint novo e elimina o erro do campo "status".

@gestao_bp.route("/planejamento_financeiro/")
@gestao_required
def planejamento_financeiro_redirect_underline():
    return redirect("/gestao/planejamento-financeiro/")


@gestao_bp.route("/planejamento_financeiro")
@gestao_required
def planejamento_financeiro_redirect_underline_sem_barra():
    return redirect("/gestao/planejamento-financeiro/")
