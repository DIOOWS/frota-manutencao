from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from flask import Blueprint, render_template, request, jsonify, send_file
from sqlalchemy import or_
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError
import hashlib
import re
import unicodedata
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import db
from utils.auth import gestao_required

from models.conta_pagar_importada import ContaPagarImportada
from models.conta_receber_importada import ContaReceberImportada
from models.planejamento_financeiro import PlanejamentoFinanceiro, PlanejamentoPagamento

from .radar_financeiro import (
    moeda,
    nome_mes,
    primeiro_dia_mes,
    ultimo_dia_mes,
    fim_dia_datetime,
)


planejamento_financeiro_bp = Blueprint(
    "planejamento_financeiro",
    __name__,
    url_prefix="/gestao/planejamento-financeiro",
)


def dinheiro_decimal(valor):
    if valor is None or valor == "":
        return Decimal("0")

    try:
        return Decimal(str(valor))
    except Exception:
        return Decimal("0")


def data_para_date_local(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except Exception:
            pass

    return None


def formatar_data(valor):
    data_ref = data_para_date_local(valor)
    if not data_ref:
        return "-"

    return data_ref.strftime("%d/%m/%Y")


def texto_limpo(valor, padrao="-"):
    if valor is None:
        return padrao

    texto = str(valor).strip()
    return texto if texto else padrao


def buscar_primeiro(conta, campos, padrao="-"):
    for campo in campos:
        valor = getattr(conta, campo, None)
        if valor not in (None, ""):
            return texto_limpo(valor, padrao)

    return padrao


def conta_esta_cancelada(conta):
    status = texto_limpo(getattr(conta, "status", None), "").upper()
    return status in ("CANCELADO", "CANCELADA")


def conta_esta_paga_local(conta):
    if bool(getattr(conta, "pago", False)):
        return True

    status = texto_limpo(getattr(conta, "status", None), "").upper()
    return status in ("PAGO", "RECEBIDO", "OK", "QUITADO", "BAIXADO")




def normalizar_chave(valor):
    """Normaliza texto para gerar uma chave estável entre reimportações."""
    if valor is None:
        return ""

    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto).strip().upper()
    return texto


def valor_chave(valor):
    return str(dinheiro_decimal(valor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def gerar_chave_conta(conta):
    """
    Chave estável da conta. Não depende do ID criado pela importação.

    Usa os campos mais consistentes do relatório para reconhecer a mesma conta
    depois que a importação é apagada e recriada.
    """
    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))
    vencimento_txt = vencimento.isoformat() if vencimento else ""

    fornecedor = buscar_primeiro(
        conta,
        ["fornecedor_funcionario", "fornecedor", "cliente"],
        "",
    )
    conta_nome = buscar_primeiro(
        conta,
        ["plano_contas", "categoria", "descricao", "nome"],
        "",
    )
    documento = buscar_primeiro(
        conta,
        ["numero_fatura", "documento", "numero_documento", "nf", "nota_fiscal"],
        "",
    )
    parcela = buscar_primeiro(
        conta,
        ["parcela_label", "parcela", "numero_parcela"],
        "",
    )

    base = "|".join([
        normalizar_chave(fornecedor),
        normalizar_chave(conta_nome),
        normalizar_chave(documento),
        normalizar_chave(parcela),
        vencimento_txt,
        valor_chave(getattr(conta, "valor", 0)),
    ])

    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def preencher_snapshot_planejamento(registro, conta, chave_conta=None):
    """Atualiza o vínculo atual e preserva dados essenciais do planejamento."""
    registro.conta_id = int(getattr(conta, "id", 0))
    registro.chave_conta = chave_conta or gerar_chave_conta(conta)
    registro.descricao_snapshot = buscar_primeiro(
        conta, ["plano_contas", "categoria", "descricao", "nome"], "SEM CONTA"
    )
    registro.fornecedor_snapshot = buscar_primeiro(
        conta, ["fornecedor_funcionario", "fornecedor", "cliente"], "-"
    )
    valor_atual = dinheiro_decimal(getattr(conta, "valor", 0))
    valor_snapshot_atual = dinheiro_decimal(getattr(registro, "valor_snapshot", None))
    registro.valor_snapshot = max(valor_atual, valor_snapshot_atual)
    registro.vencimento_snapshot = data_para_date_local(getattr(conta, "data_vencimento", None))
    registro.observacao_snapshot = buscar_primeiro(
        conta, ["observacoes", "observacao", "historico", "descricao"], "-"
    )


def localizar_planejamento(conta, mes, ano, por_id=None, por_chave=None):
    conta_id = int(getattr(conta, "id", 0))
    chave = gerar_chave_conta(conta)

    registro = None
    if por_id is not None:
        registro = por_id.get(conta_id)
    if not registro and por_chave is not None:
        registro = por_chave.get(chave)

    if registro:
        mudou = registro.conta_id != conta_id or registro.chave_conta != chave
        preencher_snapshot_planejamento(registro, conta, chave)
        if mudou:
            db.session.flush()
        return registro

    # Durante a montagem da tela, os planejamentos já foram carregados em lote.
    # Não executa consulta individual por conta (evita N+1 queries).
    return None


def montar_item_planejamento(conta, mes, ano, hoje):
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))

    valor = dinheiro_decimal(getattr(conta, "valor", 0))

    conta_nome = buscar_primeiro(
        conta,
        [
            "plano_contas",
            "categoria",
            "descricao",
            "nome",
            "fornecedor_funcionario",
            "fornecedor",
        ],
        "SEM CONTA",
    )

    fornecedor = buscar_primeiro(
        conta,
        ["fornecedor_funcionario", "fornecedor", "cliente"],
        "-",
    )

    categoria = buscar_primeiro(
        conta,
        ["categoria", "plano_contas"],
        "-",
    )

    documento = buscar_primeiro(
        conta,
        ["numero_fatura", "documento", "numero_documento", "nf", "nota_fiscal"],
        "-",
    )

    observacao = buscar_primeiro(
        conta,
        ["observacoes", "observacao", "historico", "descricao"],
        "-",
    )

    parcela_label = buscar_primeiro(
        conta,
        ["parcela_label", "parcela", "numero_parcela"],
        "",
    )

    if parcela_label == "-":
        parcela_label = ""

    status_visual = "SEM DATA"
    is_herdada = False
    is_mes_atual = False
    dias_atraso = 0

    if vencimento:
        is_herdada = vencimento < primeiro_mes
        is_mes_atual = primeiro_mes <= vencimento <= ultimo_mes

        if vencimento < primeiro_mes:
            status_visual = "HERDADA"
        elif vencimento < hoje:
            status_visual = "ATRASADA"
        elif vencimento == hoje:
            status_visual = "VENCE HOJE"
        else:
            status_visual = "MÊS ATUAL"

        if vencimento < hoje:
            dias_atraso = (hoje - vencimento).days

    return {
        "id": int(getattr(conta, "id", 0)),
        "descricao": conta_nome,
        "conta_nome": conta_nome,
        "fornecedor": fornecedor,
        "categoria": categoria,
        "documento": documento,
        "observacao": observacao,
        "parcela_label": parcela_label,
        "vencimento": formatar_data(vencimento),
        "data_vencimento_obj": vencimento,
        "valor": valor,
        "valor_formatado": moeda(valor),
        "is_herdada": is_herdada,
        "is_mes_atual": is_mes_atual,
        "status_planejamento_visual": status_visual,
        "dias_atraso": dias_atraso,
    }


def buscar_contas_do_mes(mes, ano):
    """
    Busca todas as contas em aberto até o fim do mês selecionado.
    Isso inclui herdadas/atrasadas e contas do mês atual.
    Tudo vem em ordem de pagamento: vencimento mais antigo primeiro.
    """
    ultimo_mes = ultimo_dia_mes(mes, ano)

    return ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.pago == False,
        or_(
            ContaPagarImportada.status.is_(None),
            ContaPagarImportada.status != "CANCELADO",
        ),
        ContaPagarImportada.data_vencimento <= fim_dia_datetime(ultimo_mes),
    ).order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc(),
    ).all()




def buscar_contas_exportacao_e_tela(mes, ano):
    """
    Busca contas até o fim da competência, incluindo pagas.
    Necessário para manter histórico: se uma conta planejada em julho foi paga no Radar,
    ela continua aparecendo/exportando em julho como PAGA.
    """
    ultimo_mes = ultimo_dia_mes(mes, ano)

    return ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        or_(
            ContaPagarImportada.status.is_(None),
            ContaPagarImportada.status != "CANCELADO",
        ),
        ContaPagarImportada.data_vencimento <= fim_dia_datetime(ultimo_mes),
    ).order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc(),
    ).all()


def inicio_dia_datetime_local(data_ref):
    return datetime.combine(data_ref, datetime.min.time())




def valor_receber_decimal(conta):
    """Retorna o valor de contas a receber importadas com segurança."""
    return dinheiro_decimal(
        getattr(conta, "total", None)
        or getattr(conta, "valor", None)
        or getattr(conta, "valor_recebido", None)
        or 0
    )


def calcular_resumo_caixa_competencia(mes, ano):
    """
    Resumo do topo da tela.
    - Recebido: entradas confirmadas no Radar/Importações dentro da competência.
    - Pago: despesas confirmadas no Radar dentro da competência.
    - Saldo da conta: recebido - pago.
    """
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    inicio_dt = inicio_dia_datetime_local(primeiro_mes)
    fim_dt = fim_dia_datetime(ultimo_mes)

    contas_recebidas = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt,
    ).all()

    contas_pagas = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt,
    ).all()

    total_recebido = sum(valor_receber_decimal(c) for c in contas_recebidas)
    total_pago = sum(dinheiro_decimal(getattr(c, "valor", 0)) for c in contas_pagas)
    saldo_conta = total_recebido - total_pago

    return total_recebido, total_pago, saldo_conta


def conta_paga_na_competencia(conta, mes, ano):
    """
    Define se a conta paga no Radar pertence à competência do planejamento.
    Regra principal: data_pagamento dentro do mês/ano selecionado.
    Fallback: se não tiver data_pagamento, usa vencimento dentro do mês/ano.
    """
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)

    data_pagamento = data_para_date_local(getattr(conta, "data_pagamento", None))
    if data_pagamento:
        return primeiro_mes <= data_pagamento <= ultimo_mes

    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))
    if vencimento:
        return primeiro_mes <= vencimento <= ultimo_mes

    return False


def pagamento_na_competencia(pagamento, mes, ano, registro=None):
    """Define a competência pelo mês/ano da data prevista do pagamento."""
    data_prevista = data_para_date_local(getattr(pagamento, "data_prevista", None))
    if data_prevista:
        return data_prevista.month == mes and data_prevista.year == ano

    # Compatibilidade com registros antigos sem data prevista.
    if registro is not None:
        return registro.mes == mes and registro.ano == ano

    return False


def buscar_planejamentos(mes, ano, contas=None):
    """
    Carrega em lote os planejamentos ligados às contas exibidas.

    Um pagamento pertence à competência da sua data prevista, mesmo quando o
    registro-pai foi criado em outro mês. Isso permite, por exemplo, planejar
    em julho um pagamento para agosto sem exibi-lo no painel de julho.
    """
    contas = list(contas or [])
    ids = {int(c.id) for c in contas if getattr(c, "id", None) is not None}
    chaves = {gerar_chave_conta(c) for c in contas}
    chaves.discard("")

    consulta = (
        PlanejamentoFinanceiro.query
        .options(selectinload(PlanejamentoFinanceiro.pagamentos_planejados))
        .filter(PlanejamentoFinanceiro.origem == "IMPORTADA")
    )

    vinculos = []
    if ids:
        vinculos.append(PlanejamentoFinanceiro.conta_id.in_(ids))
    if chaves:
        vinculos.append(PlanejamentoFinanceiro.chave_conta.in_(chaves))

    if vinculos:
        consulta = consulta.filter(or_(*vinculos))
    else:
        consulta = consulta.filter(
            PlanejamentoFinanceiro.mes == mes,
            PlanejamentoFinanceiro.ano == ano,
        )

    registros = consulta.order_by(PlanejamentoFinanceiro.id.desc()).all()

    por_id = {}
    por_chave = {}

    def prioridade(registro):
        pagamentos = pagamentos_do_planejamento(registro)
        tem_pagamento_na_competencia = any(
            pagamento_na_competencia(p, mes, ano, registro)
            for p in pagamentos
        )
        pai_na_competencia = registro.mes == mes and registro.ano == ano
        return (2 if tem_pagamento_na_competencia else 0) + (1 if pai_na_competencia else 0)

    melhores_id = {}
    melhores_chave = {}

    for registro in registros:
        prio = prioridade(registro)
        if prio <= 0:
            continue

        if registro.conta_id is not None:
            conta_id = int(registro.conta_id)
            atual = melhores_id.get(conta_id)
            if atual is None or prio > atual[0]:
                melhores_id[conta_id] = (prio, registro)

        if registro.chave_conta:
            atual = melhores_chave.get(registro.chave_conta)
            if atual is None or prio > atual[0]:
                melhores_chave[registro.chave_conta] = (prio, registro)

    por_id = {chave: valor[1] for chave, valor in melhores_id.items()}
    por_chave = {chave: valor[1] for chave, valor in melhores_chave.items()}

    # Atualiza snapshots antigos sem gerar consultas extras.
    alterou = False
    for conta in contas:
        conta_id = int(getattr(conta, "id", 0))
        chave = gerar_chave_conta(conta)
        registro = por_id.get(conta_id) or por_chave.get(chave)
        if registro and not getattr(registro, "chave_conta", None):
            preencher_snapshot_planejamento(registro, conta, chave)
            por_chave[chave] = registro
            alterou = True

    if alterou:
        db.session.commit()

    return por_id, por_chave

def buscar_planejamentos_outros_meses(mes, ano, contas):
    """
    Carrega somente planejamentos de outras competências ligados às contas
    atualmente exibidas. Tudo é buscado em lote, sem consulta dentro de loop.
    """
    contas_validas = [c for c in (contas or []) if not conta_esta_cancelada(c)]
    if not contas_validas:
        return {}, {}

    ids = {int(c.id) for c in contas_validas if getattr(c, "id", None) is not None}
    chaves = {gerar_chave_conta(c) for c in contas_validas}
    chaves.discard("")

    vinculos = []
    if ids:
        vinculos.append(PlanejamentoFinanceiro.conta_id.in_(ids))
    if chaves:
        vinculos.append(PlanejamentoFinanceiro.chave_conta.in_(chaves))
    if not vinculos:
        return {}, {}

    registros = (
        PlanejamentoFinanceiro.query
        .options(selectinload(PlanejamentoFinanceiro.pagamentos_planejados))
        .filter(
            PlanejamentoFinanceiro.origem == "IMPORTADA",
            or_(
                PlanejamentoFinanceiro.mes != mes,
                PlanejamentoFinanceiro.ano != ano,
            ),
            or_(*vinculos),
        )
        .order_by(
            PlanejamentoFinanceiro.ano.asc(),
            PlanejamentoFinanceiro.mes.asc(),
        )
        .all()
    )

    por_id = {}
    por_chave = {}
    for registro in registros:
        # Só bloqueia planejamento que realmente possui valor planejado.
        # Registros históricos/sincronizados sem pagamento não escondem a conta.
        if total_ja_planejado(registro) <= 0:
            continue
        if registro.conta_id is not None:
            por_id.setdefault(int(registro.conta_id), registro)
        if registro.chave_conta:
            por_chave.setdefault(registro.chave_conta, registro)

    return por_id, por_chave


def sincronizar_pagas_do_radar(mes, ano):
    """Sincroniza pagas sem perder vínculos quando os IDs da importação mudarem."""
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    inicio_dt = inicio_dia_datetime_local(primeiro_mes)
    fim_dt = fim_dia_datetime(ultimo_mes)

    contas = ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        or_(
            ContaPagarImportada.data_pagamento >= inicio_dt,
            ContaPagarImportada.data_vencimento >= inicio_dt,
        ),
        or_(
            ContaPagarImportada.data_pagamento <= fim_dt,
            ContaPagarImportada.data_vencimento <= fim_dt,
        ),
    ).all()

    por_id, por_chave = buscar_planejamentos(mes, ano, contas)
    criadas = 0
    atualizadas = 0

    for conta in contas:
        if conta_esta_cancelada(conta) or not conta_esta_paga_local(conta):
            continue
        if not conta_paga_na_competencia(conta, mes, ano):
            continue

        chave = gerar_chave_conta(conta)
        registro = por_id.get(int(conta.id)) or por_chave.get(chave)

        if registro:
            if registro.conta_id != conta.id or registro.chave_conta != chave:
                preencher_snapshot_planejamento(registro, conta, chave)
                atualizadas += 1
            continue

        novo = PlanejamentoFinanceiro(
            conta_id=int(conta.id),
            chave_conta=chave,
            origem="IMPORTADA",
            mes=mes,
            ano=ano,
            status_planejamento="SINCRONIZADA_RADAR",
        )
        preencher_snapshot_planejamento(novo, conta, chave)
        db.session.add(novo)
        por_id[int(conta.id)] = novo
        por_chave[chave] = novo
        criadas += 1

    db.session.commit()
    return criadas, atualizadas


def buscar_registros_por_conta(contas):
    """Carrega todos os planejamentos e pagamentos das contas em poucas consultas."""
    contas = list(contas or [])
    ids = {int(c.id) for c in contas if getattr(c, "id", None) is not None}
    chaves = {gerar_chave_conta(c) for c in contas}
    chaves.discard("")

    vinculos = []
    if ids:
        vinculos.append(PlanejamentoFinanceiro.conta_id.in_(ids))
    if chaves:
        vinculos.append(PlanejamentoFinanceiro.chave_conta.in_(chaves))
    if not vinculos:
        return {}, {}

    registros = (
        PlanejamentoFinanceiro.query
        .options(selectinload(PlanejamentoFinanceiro.pagamentos_planejados))
        .filter(
            PlanejamentoFinanceiro.origem == "IMPORTADA",
            or_(*vinculos),
        )
        .order_by(PlanejamentoFinanceiro.id.desc())
        .all()
    )

    por_id = {}
    por_chave = {}
    for registro in registros:
        if registro.conta_id is not None:
            por_id.setdefault(int(registro.conta_id), []).append(registro)
        if registro.chave_conta:
            por_chave.setdefault(registro.chave_conta, []).append(registro)

    return por_id, por_chave


def registros_da_conta(conta, por_id, por_chave):
    conta_id = int(getattr(conta, "id", 0))
    chave = gerar_chave_conta(conta)
    encontrados = []
    vistos = set()

    for registro in (por_id.get(conta_id, []) + por_chave.get(chave, [])):
        marcador = getattr(registro, "id", None) or id(registro)
        if marcador in vistos:
            continue
        vistos.add(marcador)
        encontrados.append(registro)

    return encontrados


def _pagamentos_unicos_da_conta(registros):
    encontrados = []
    vistos = set()

    for registro in registros:
        for pagamento in pagamentos_do_planejamento(registro):
            pagamento_id = getattr(pagamento, "id", None)
            if pagamento_id is not None:
                marcador = ("ID", int(pagamento_id))
            else:
                marcador = (
                    "LEGADO",
                    int(getattr(registro, "id", 0) or 0),
                    valor_chave(getattr(pagamento, "valor", 0)),
                    str(data_para_date_local(getattr(pagamento, "data_prevista", None)) or ""),
                    texto_limpo(getattr(pagamento, "tipo", None), "TOTAL").upper(),
                )

            if marcador in vistos:
                continue

            vistos.add(marcador)
            encontrados.append((registro, pagamento))

    return encontrados


def _valor_original_da_conta(item_base, registros):
    valores = [dinheiro_decimal(item_base.get("valor", 0))]
    valores.extend(
        dinheiro_decimal(getattr(registro, "valor_snapshot", None))
        for registro in registros
    )
    return max(valores or [Decimal("0")])


def _registro_origem_do_saldo(registros, pagamentos):
    """
    O saldo permanece na competência onde o usuário fez a decisão parcial
    mais recente.

    Não usamos o registro mais antigo, pois podem existir planejamentos
    históricos da mesma conta em outras competências. A origem correta é o
    registro-pai do pagamento filho mais recente.
    """
    if not pagamentos:
        return None

    def chave_recencia(par):
        registro, pagamento = par
        pagamento_id = getattr(pagamento, "id", None)

        # Pagamentos reais da tabela filha sempre têm ID. Eles têm prioridade
        # sobre objetos legados montados em memória.
        if pagamento_id is not None:
            return (
                2,
                int(pagamento_id),
                int(getattr(registro, "ano", 0) or 0),
                int(getattr(registro, "mes", 0) or 0),
                int(getattr(registro, "id", 0) or 0),
            )

        return (
            1,
            0,
            int(getattr(registro, "ano", 0) or 0),
            int(getattr(registro, "mes", 0) or 0),
            int(getattr(registro, "id", 0) or 0),
        )

    registro_origem, _ = max(pagamentos, key=chave_recencia)
    return registro_origem


def _criar_item_saldo_agregado(item_base, registro_origem, valor_original, total_planejado):
    saldo = max(Decimal("0"), valor_original - total_planejado)

    item = dict(item_base)
    item["valor"] = saldo
    item["valor_formatado"] = moeda(saldo)
    item["valor_original"] = valor_original
    item["valor_original_formatado"] = moeda(valor_original)
    item["saldo_restante"] = saldo
    item["saldo_restante_formatado"] = moeda(saldo)
    item["is_saldo_restante"] = True
    item["status_planejamento_visual"] = "SALDO RESTANTE"
    item["descricao"] = f"{item_base.get('descricao', 'CONTA')} (SALDO)"
    item["conta_nome"] = item["descricao"]
    observacao_original = texto_limpo(
        item_base.get("observacao"),
        "",
    )
    detalhe_saldo = (
        f"Saldo ainda não planejado. Valor original: {moeda(valor_original)}. "
        f"Total já planejado: {moeda(total_planejado)}."
    )
    # Mantém a observação original da conta para que a linha continue sendo
    # encontrada pela busca após virar SALDO.
    item["observacao"] = (
        f"{observacao_original} | {detalhe_saldo}"
        if observacao_original
        else detalhe_saldo
    )
    item["status_execucao"] = "AGUARDANDO"
    item["data_pagamento"] = "-"
    item["planejada_outro_mes"] = False
    item["competencia_planejada"] = ""
    item["planejamento_origem_id"] = getattr(registro_origem, "id", None)
    item["texto_busca"] = " ".join([
        texto_limpo(item.get("descricao"), ""),
        texto_limpo(item.get("conta_nome"), ""),
        texto_limpo(item.get("fornecedor"), ""),
        texto_limpo(item.get("categoria"), ""),
        texto_limpo(item.get("documento"), ""),
        texto_limpo(item.get("observacao"), ""),
        texto_limpo(item.get("parcela_label"), ""),
        texto_limpo(item.get("vencimento"), ""),
        texto_limpo(item.get("valor_formatado"), ""),
    ]).strip()
    return item


def _montar_listas_unificadas(mes, ano):
    hoje = date.today()
    contas = buscar_contas_exportacao_e_tela(mes, ano)
    registros_id, registros_chave = buscar_registros_por_conta(contas)

    aguardando = []
    planejadas = []
    pagas = []

    for conta in contas:
        if conta_esta_cancelada(conta):
            continue

        item_base = montar_item_planejamento(conta, mes, ano, hoje)
        item_base["data_pagamento"] = formatar_data(
            getattr(conta, "data_pagamento", None)
        )

        conta_paga = conta_esta_paga_local(conta)
        registros = registros_da_conta(
            conta,
            registros_id,
            registros_chave,
        )
        pagamentos = _pagamentos_unicos_da_conta(registros)

        valor_original = _valor_original_da_conta(item_base, registros)
        total_planejado = sum(
            dinheiro_decimal(getattr(pagamento, "valor", 0))
            for _, pagamento in pagamentos
        )
        total_planejado = min(total_planejado, valor_original)

        # Conta quitada no Radar representa quitação integral.
        # Exibe uma única linha paga com o valor total original da conta,
        # independentemente de existirem previsões parciais anteriores.
        if conta_paga:
            if conta_paga_na_competencia(conta, mes, ano):
                data_pagamento_formatada = formatar_data(
                    getattr(conta, "data_pagamento", None)
                )

                item_pago = dict(item_base)
                item_pago["status_execucao"] = "PAGA"
                item_pago["data_pagamento"] = data_pagamento_formatada
                item_pago["data_prevista"] = data_pagamento_formatada

                item_pago["valor"] = valor_original
                item_pago["valor_formatado"] = moeda(valor_original)
                item_pago["valor_original"] = valor_original
                item_pago["valor_original_formatado"] = moeda(valor_original)
                item_pago["valor_planejado"] = valor_original
                item_pago["valor_planejado_formatado"] = moeda(valor_original)

                item_pago["saldo_restante"] = Decimal("0")
                item_pago["saldo_restante_formatado"] = moeda(Decimal("0"))
                item_pago["tipo_planejamento"] = "TOTAL"
                item_pago["observacao_previsao"] = (
                    "Conta quitada integralmente no Radar."
                )
                item_pago["pagamento_planejado_id"] = None
                item_pago["planejada_outro_mes"] = False
                item_pago["competencia_planejada"] = ""

                pagas.append(item_pago)

            # Conta paga nunca permanece em planejadas ou aguardando.
            continue

        # Enquanto a conta estiver aberta, cada previsão aparece somente
        # na competência da data prevista escolhida.
        for registro, pagamento in pagamentos:
            if not pagamento_na_competencia(pagamento, mes, ano, registro):
                continue

            item_pagamento = criar_item_pagamento(
                item_base,
                registro,
                pagamento,
            )

            valor_pagamento = dinheiro_decimal(
                getattr(pagamento, "valor", 0)
            )
            saldo_agregado = max(
                Decimal("0"),
                valor_original - total_planejado,
            )

            item_pagamento["valor_original"] = valor_original
            item_pagamento["valor_original_formatado"] = moeda(valor_original)
            item_pagamento["valor_planejado"] = valor_pagamento
            item_pagamento["valor_planejado_formatado"] = moeda(valor_pagamento)
            item_pagamento["saldo_restante"] = saldo_agregado
            item_pagamento["saldo_restante_formatado"] = moeda(saldo_agregado)
            item_pagamento["status_execucao"] = "PLANEJADA"
            item_pagamento["data_pagamento"] = "-"

            planejadas.append(item_pagamento)

        if pagamentos:
            registro_origem = _registro_origem_do_saldo(
                registros,
                pagamentos,
            )
            saldo = max(
                Decimal("0"),
                valor_original - total_planejado,
            )

            # O saldo aparece na competência do registro-pai do pagamento
            # parcial mais recente. Como proteção para dados antigos, se houver
            # registro com pagamento na competência aberta, ele prevalece.
            registros_com_pagamento_na_competencia = [
                registro
                for registro, pagamento in pagamentos
                if int(getattr(registro, "mes", 0) or 0) == int(mes)
                and int(getattr(registro, "ano", 0) or 0) == int(ano)
            ]

            if registros_com_pagamento_na_competencia:
                registro_origem = max(
                    registros_com_pagamento_na_competencia,
                    key=lambda registro: int(getattr(registro, "id", 0) or 0),
                )

            if (
                registro_origem is not None
                and int(registro_origem.mes) == int(mes)
                and int(registro_origem.ano) == int(ano)
                and saldo > 0
            ):
                aguardando.append(
                    _criar_item_saldo_agregado(
                        item_base,
                        registro_origem,
                        valor_original,
                        total_planejado,
                    )
                )
        else:
            item = dict(item_base)
            item["status_execucao"] = "AGUARDANDO"
            item["data_pagamento"] = "-"
            item["planejada_outro_mes"] = False
            item["competencia_planejada"] = ""
            aguardando.append(item)

    aguardando.sort(key=lambda item: (
        data_para_date_local(item.get("data_vencimento_obj")) or date.max,
        normalizar_chave(item.get("fornecedor")),
        normalizar_chave(item.get("conta_nome")),
    ))
    planejadas.sort(key=lambda item: (
        data_para_date_local(item.get("data_prevista")) or date.max,
        normalizar_chave(item.get("fornecedor")),
        normalizar_chave(item.get("conta_nome")),
    ))
    pagas.sort(key=lambda item: (
        data_para_date_local(item.get("data_pagamento"))
        or data_para_date_local(item.get("data_prevista"))
        or date.max,
        normalizar_chave(item.get("fornecedor")),
        normalizar_chave(item.get("conta_nome")),
    ))

    return aguardando, planejadas, pagas


@planejamento_financeiro_bp.route("/")
@gestao_required

def index():
    agora = datetime.now()
    mes = request.args.get("mes", type=int) or agora.month
    ano = request.args.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    aguardando, planejadas, executadas = _montar_listas_unificadas(
        mes,
        ano,
    )

    total_aguardando = sum(
        dinheiro_decimal(item.get("valor"))
        for item in aguardando
    )
    total_planejado = sum(
        dinheiro_decimal(
            item.get("valor_planejado", item.get("valor"))
        )
        for item in planejadas
    )
    total_executado = sum(
        dinheiro_decimal(item.get("valor", 0))
        for item in executadas
    )

    (
        total_recebido,
        total_pago_competencia,
        saldo_competencia,
    ) = calcular_resumo_caixa_competencia(mes, ano)

    return render_template(
        "gestao/planejamento_financeiro.html",
        mes=mes,
        ano=ano,
        nome_mes=nome_mes,
        moeda=moeda,
        aguardando=aguardando,
        planejadas=planejadas,
        executadas=executadas,
        total_aguardando=total_aguardando,
        total_planejado=total_planejado,
        total_executado=total_executado,
        total_recebido=total_recebido,
        total_pago_competencia=total_pago_competencia,
        saldo_competencia=saldo_competencia,
    )


def montar_listas_planejamento(mes, ano):
    return _montar_listas_unificadas(mes, ano)

def valor_decimal_item(item):
    return dinheiro_decimal(item.get("valor_planejado", item.get("valor", 0)))


def pagamentos_do_planejamento(registro):
    """Retorna os pagamentos separados; migra registros antigos de forma transparente."""
    pagamentos = list(getattr(registro, "pagamentos_planejados", []) or [])
    if pagamentos:
        return sorted(
            pagamentos,
            key=lambda p: (data_para_date_local(p.data_prevista) or date.max, p.id or 0),
        )

    valor_legado = dinheiro_decimal(getattr(registro, "valor_planejado", None))
    if valor_legado <= 0:
        return []

    # Objeto leve para compatibilidade com planejamentos criados antes da tabela filha.
    class PagamentoLegado:
        pass

    pagamento = PagamentoLegado()
    pagamento.id = None
    pagamento.tipo = texto_limpo(getattr(registro, "tipo_planejamento", None), "TOTAL").upper()
    pagamento.valor = valor_legado
    pagamento.data_prevista = data_para_date_local(getattr(registro, "data_prevista", None))
    pagamento.observacao = getattr(registro, "observacao_previsao", None)
    return [pagamento]


def total_ja_planejado(registro):
    return sum(dinheiro_decimal(p.valor) for p in pagamentos_do_planejamento(registro))


def valor_original_do_planejamento(item_base, registro):
    """
    Usa o snapshot gravado no momento do planejamento como valor original.

    Isso evita perder o saldo quando a linha importada é recriada, alterada ou
    passa a trazer apenas o valor parcial. O valor da conta atual fica apenas
    como fallback para registros antigos sem snapshot.
    """
    valor_snapshot = dinheiro_decimal(getattr(registro, "valor_snapshot", None))
    if valor_snapshot > 0:
        return valor_snapshot

    return dinheiro_decimal(item_base.get("valor", 0))


def criar_item_pagamento(item_base, registro, pagamento):
    item = dict(item_base)
    valor_original = valor_original_do_planejamento(item_base, registro)
    valor_pagamento = dinheiro_decimal(getattr(pagamento, "valor", 0))
    total_planejado = total_ja_planejado(registro)
    saldo_restante = max(Decimal("0"), valor_original - total_planejado)

    item["valor_original"] = valor_original
    item["valor_original_formatado"] = moeda(valor_original)
    item["valor_planejado"] = valor_pagamento
    item["valor_planejado_formatado"] = moeda(valor_pagamento)
    item["saldo_restante"] = saldo_restante
    item["saldo_restante_formatado"] = moeda(saldo_restante)
    item["tipo_planejamento"] = texto_limpo(getattr(pagamento, "tipo", None), "PARCIAL").upper()
    item["data_prevista"] = formatar_data(getattr(pagamento, "data_prevista", None))
    item["observacao_previsao"] = texto_limpo(getattr(pagamento, "observacao", None), "-")
    item["pagamento_planejado_id"] = getattr(pagamento, "id", None)
    return item


def criar_item_saldo_restante(item_base, registro):
    item = dict(item_base)
    valor_original = valor_original_do_planejamento(item_base, registro)
    saldo = max(Decimal("0"), valor_original - total_ja_planejado(registro))
    item["valor"] = saldo
    item["valor_formatado"] = moeda(saldo)
    item["valor_original"] = valor_original
    item["valor_original_formatado"] = moeda(valor_original)
    item["saldo_restante"] = saldo
    item["saldo_restante_formatado"] = moeda(saldo)
    item["is_saldo_restante"] = True
    item["status_planejamento_visual"] = "SALDO RESTANTE"
    item["descricao"] = f"{item_base.get('descricao', 'CONTA')} (SALDO)"
    item["conta_nome"] = item["descricao"]
    item["observacao"] = (
        f"Saldo ainda não planejado. Valor original: {moeda(valor_original)}."
    )
    item["status_execucao"] = "AGUARDANDO"
    item["planejada_outro_mes"] = False
    item["competencia_planejada"] = ""
    return item


def aplicar_dados_previsao_item(item, registro):
    """Acrescenta ao item os dados previstos sem perder o valor original da conta."""
    valor_original = dinheiro_decimal(item.get("valor", 0))
    valor_planejado = dinheiro_decimal(getattr(registro, "valor_planejado", None))
    if valor_planejado <= 0:
        valor_planejado = valor_original

    tipo = texto_limpo(getattr(registro, "tipo_planejamento", None), "TOTAL").upper()
    data_prevista = data_para_date_local(getattr(registro, "data_prevista", None))
    saldo_restante = max(Decimal("0"), valor_original - valor_planejado)

    item["valor_original"] = valor_original
    item["valor_original_formatado"] = moeda(valor_original)
    item["valor_planejado"] = valor_planejado
    item["valor_planejado_formatado"] = moeda(valor_planejado)
    item["saldo_restante"] = saldo_restante
    item["saldo_restante_formatado"] = moeda(saldo_restante)
    item["tipo_planejamento"] = tipo
    item["data_prevista"] = formatar_data(data_prevista)
    item["observacao_previsao"] = texto_limpo(getattr(registro, "observacao_previsao", None), "-")
    return item


def estilos_excel():
    cores = {
        "azul": "0B4EDB",
        "azul_escuro": "061B3D",
        "azul_claro": "EAF2FF",
        "verde": "087539",
        "verde_claro": "EAF8F0",
        "amarelo": "F59E0B",
        "amarelo_claro": "FFF7DD",
        "vermelho": "E11D2E",
        "vermelho_claro": "FFECEE",
        "cinza": "F4F7FB",
        "cinza_texto": "64748B",
        "branco": "FFFFFF",
        "borda": "DCE5F1",
    }
    borda = Border(
        left=Side(style="thin", color=cores["borda"]),
        right=Side(style="thin", color=cores["borda"]),
        top=Side(style="thin", color=cores["borda"]),
        bottom=Side(style="thin", color=cores["borda"]),
    )
    return cores, borda


def preparar_aba(ws, titulo, subtitulo, total_colunas):
    cores, _ = estilos_excel()
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    titulo_cell = ws.cell(1, 1, titulo)
    titulo_cell.font = Font(bold=True, color=cores["branco"], size=16)
    titulo_cell.fill = PatternFill("solid", fgColor=cores["azul_escuro"])
    titulo_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_colunas)
    sub_cell = ws.cell(2, 1, subtitulo)
    sub_cell.font = Font(bold=True, color=cores["azul_escuro"], size=10)
    sub_cell.fill = PatternFill("solid", fgColor=cores["cinza"])
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 23


def escrever_cabecalho(ws, linha, colunas):
    cores, borda = estilos_excel()
    for indice, cabecalho in enumerate(colunas, start=1):
        cell = ws.cell(linha, indice, cabecalho)
        cell.font = Font(bold=True, color=cores["branco"], size=10)
        cell.fill = PatternFill("solid", fgColor=cores["azul"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[linha].height = 28


def estilizar_status(cell, status):
    cores, _ = estilos_excel()
    status = texto_limpo(status, "-").upper()
    if status == "PAGA":
        cell.fill = PatternFill("solid", fgColor=cores["verde_claro"])
        cell.font = Font(bold=True, color=cores["verde"])
    elif status == "PLANEJADA":
        cell.fill = PatternFill("solid", fgColor=cores["azul_claro"])
        cell.font = Font(bold=True, color=cores["azul"])
    elif status == "PARCIAL":
        cell.fill = PatternFill("solid", fgColor=cores["amarelo_claro"])
        cell.font = Font(bold=True, color="A65B00")
    else:
        cell.fill = PatternFill("solid", fgColor=cores["vermelho_claro"])
        cell.font = Font(bold=True, color=cores["vermelho"])


def valor_original_item(item):
    return dinheiro_decimal(item.get("valor_original", item.get("valor", 0)))


def valor_planejado_item(item):
    if item.get("status_execucao") == "AGUARDANDO":
        return Decimal("0")
    return dinheiro_decimal(item.get("valor_planejado", item.get("valor", 0)))


def saldo_item(item):
    if item.get("status_execucao") == "AGUARDANDO":
        return valor_original_item(item)
    return dinheiro_decimal(item.get("saldo_restante", 0))


def data_agenda_item(item):
    return (
        data_para_date_local(item.get("data_prevista"))
        or data_para_date_local(item.get("data_pagamento"))
        or data_para_date_local(item.get("vencimento"))
    )


def escrever_aba_agenda(wb, competencia, planejadas, pagas):
    ws = wb.create_sheet("Agenda de Pagamentos")
    itens = list(planejadas) + list(pagas)
    itens.sort(key=lambda i: (
        data_agenda_item(i) or date.max,
        normalizar_chave(i.get("fornecedor")),
        normalizar_chave(i.get("conta_nome")),
    ))
    total = sum(valor_planejado_item(i) for i in itens)
    preparar_aba(
        ws,
        f"AGENDA DE PAGAMENTOS - {competencia.upper()}",
        f"Exportado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} | "
        f"{len(itens)} conta(s) | Total previsto: {moeda(total)}",
        11,
    )
    colunas = [
        "Data prevista", "Status", "Conta", "Fornecedor", "Parcela",
        "Tipo", "Valor original", "Valor planejado", "Saldo restante",
        "Pago em", "Observação",
    ]
    escrever_cabecalho(ws, 4, colunas)
    _, borda = estilos_excel()
    linha = 5
    for item in itens:
        data_prevista = data_agenda_item(item)
        valores = [
            data_prevista,
            item.get("status_execucao", "-"),
            item.get("conta_nome", "-"),
            item.get("fornecedor", "-"),
            item.get("parcela_label", "-"),
            item.get("tipo_planejamento", "TOTAL"),
            float(valor_original_item(item)),
            float(valor_planejado_item(item)),
            float(saldo_item(item)),
            data_para_date_local(item.get("data_pagamento")),
            item.get("observacao_previsao") or item.get("observacao", "-"),
        ]
        for coluna, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, coluna, valor)
            cell.border = borda
            cell.alignment = Alignment(vertical="center", wrap_text=coluna in (3, 4, 11))
            if coluna in (1, 10) and valor:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif coluna in (7, 8, 9):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        estilizar_status(ws.cell(linha, 2), item.get("status_execucao"))
        if texto_limpo(item.get("tipo_planejamento"), "TOTAL").upper() == "PARCIAL":
            estilizar_status(ws.cell(linha, 6), "PARCIAL")
        linha += 1

    if itens:
        ws.auto_filter.ref = f"A4:K{linha - 1}"
    ws.freeze_panes = "A5"
    larguras = [15, 14, 34, 30, 12, 12, 17, 18, 18, 14, 42]
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    return ws


def escrever_aba_resumo_diario(wb, competencia, planejadas, pagas):
    ws = wb.create_sheet("Resumo Diário")
    itens = list(planejadas) + list(pagas)
    resumo = {}
    for item in itens:
        dia = data_agenda_item(item)
        if not dia:
            continue
        dados = resumo.setdefault(dia, {"quantidade": 0, "planejado": Decimal("0"), "pago": Decimal("0")})
        dados["quantidade"] += 1
        dados["planejado"] += valor_planejado_item(item)
        if item.get("status_execucao") == "PAGA":
            dados["pago"] += dinheiro_decimal(item.get("valor", 0))

    preparar_aba(
        ws,
        f"RESUMO DIÁRIO - {competencia.upper()}",
        "Consolidação do desembolso por data prevista.",
        5,
    )
    escrever_cabecalho(ws, 4, ["Data", "Quantidade", "Valor planejado", "Valor pago", "Falta executar"])
    _, borda = estilos_excel()
    linha = 5
    for dia in sorted(resumo):
        dados = resumo[dia]
        falta = max(Decimal("0"), dados["planejado"] - dados["pago"])
        valores = [dia, dados["quantidade"], float(dados["planejado"]), float(dados["pago"]), float(falta)]
        for coluna, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, coluna, valor)
            cell.border = borda
            cell.alignment = Alignment(horizontal="center" if coluna <= 2 else "right", vertical="center")
            if coluna == 1:
                cell.number_format = "dd/mm/yyyy"
            elif coluna >= 3:
                cell.number_format = 'R$ #,##0.00'
        linha += 1
    if resumo:
        ws.auto_filter.ref = f"A4:E{linha - 1}"
    ws.freeze_panes = "A5"
    for idx, largura in enumerate([15, 14, 20, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    return ws


def escrever_aba_parciais(wb, competencia, planejadas, pagas):
    ws = wb.create_sheet("Pagamentos Parciais")
    itens = [
        i for i in list(planejadas) + list(pagas)
        if texto_limpo(i.get("tipo_planejamento"), "TOTAL").upper() == "PARCIAL"
    ]
    itens.sort(key=lambda i: (data_agenda_item(i) or date.max, normalizar_chave(i.get("conta_nome"))))
    preparar_aba(
        ws,
        f"PAGAMENTOS PARCIAIS - {competencia.upper()}",
        f"{len(itens)} conta(s) com pagamento parcial | Saldo total: {moeda(sum(saldo_item(i) for i in itens))}",
        9,
    )
    escrever_cabecalho(ws, 4, [
        "Data prevista", "Status", "Conta", "Fornecedor", "Valor original",
        "Valor planejado", "Saldo restante", "% planejado", "Observação",
    ])
    _, borda = estilos_excel()
    linha = 5
    for item in itens:
        original = valor_original_item(item)
        planejado = valor_planejado_item(item)
        percentual = float(planejado / original) if original > 0 else 0
        valores = [
            data_agenda_item(item), item.get("status_execucao", "-"),
            item.get("conta_nome", "-"), item.get("fornecedor", "-"),
            float(original), float(planejado), float(saldo_item(item)), percentual,
            item.get("observacao_previsao") or item.get("observacao", "-"),
        ]
        for coluna, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, coluna, valor)
            cell.border = borda
            cell.alignment = Alignment(vertical="center", wrap_text=coluna in (3, 4, 9))
            if coluna == 1 and valor:
                cell.number_format = "dd/mm/yyyy"
            elif coluna in (5, 6, 7):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif coluna == 8:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
        estilizar_status(ws.cell(linha, 2), item.get("status_execucao"))
        linha += 1
    if itens:
        ws.auto_filter.ref = f"A4:I{linha - 1}"
    ws.freeze_panes = "A5"
    for idx, largura in enumerate([15, 14, 34, 30, 18, 18, 18, 15, 42], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    return ws


def escrever_aba_geral(wb, competencia, aguardando, planejadas, pagas):
    ws = wb.create_sheet("Planejamento Geral")
    itens = list(aguardando) + list(planejadas) + list(pagas)
    itens.sort(key=lambda i: (
        0 if i.get("status_execucao") == "PLANEJADA" else 1 if i.get("status_execucao") == "PAGA" else 2,
        data_agenda_item(i) or date.max,
        normalizar_chave(i.get("conta_nome")),
    ))
    preparar_aba(
        ws,
        f"PLANEJAMENTO GERAL - {competencia.upper()}",
        f"{len(itens)} conta(s) entre aguardando, planejadas e pagas.",
        13,
    )
    escrever_cabecalho(ws, 4, [
        "Status", "Vencimento", "Data prevista", "Pago em", "Conta", "Fornecedor",
        "Parcela", "Tipo", "Valor original", "Valor planejado", "Saldo restante",
        "Observação da conta", "Observação do planejamento",
    ])
    _, borda = estilos_excel()
    linha = 5
    for item in itens:
        valores = [
            item.get("status_execucao", "-"),
            data_para_date_local(item.get("vencimento")),
            data_para_date_local(item.get("data_prevista")),
            data_para_date_local(item.get("data_pagamento")),
            item.get("conta_nome", "-"), item.get("fornecedor", "-"),
            item.get("parcela_label", "-"), item.get("tipo_planejamento", "-"),
            float(valor_original_item(item)), float(valor_planejado_item(item)),
            float(saldo_item(item)), item.get("observacao", "-"),
            item.get("observacao_previsao", "-"),
        ]
        for coluna, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, coluna, valor)
            cell.border = borda
            cell.alignment = Alignment(vertical="center", wrap_text=coluna in (5, 6, 12, 13))
            if coluna in (2, 3, 4) and valor:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif coluna in (9, 10, 11):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        estilizar_status(ws.cell(linha, 1), item.get("status_execucao"))
        linha += 1
    if itens:
        ws.auto_filter.ref = f"A4:M{linha - 1}"
    ws.freeze_panes = "A5"
    larguras = [14, 14, 15, 14, 34, 30, 12, 12, 18, 18, 18, 38, 40]
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    return ws


def escrever_aba_executivo(wb, competencia, aguardando, planejadas, pagas, mes, ano):
    ws = wb.create_sheet("Resumo Executivo")
    preparar_aba(
        ws,
        f"RESUMO EXECUTIVO - {competencia.upper()}",
        f"Posição gerencial exportada em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        6,
    )
    cores, borda = estilos_excel()
    total_recebido, total_pago_caixa, saldo_caixa = calcular_resumo_caixa_competencia(mes, ano)
    total_aguardando = sum(valor_original_item(i) for i in aguardando)
    total_planejado = sum(valor_planejado_item(i) for i in planejadas)
    total_executado = sum(
        dinheiro_decimal(i.get("valor", 0))
        for i in pagas
    )
    saldo_parcial = sum(
        saldo_item(i)
        for i in planejadas
    )
    quantidade_parciais = sum(
        1
        for i in planejadas
        if texto_limpo(i.get("tipo_planejamento"), "TOTAL").upper() == "PARCIAL"
    )
    quantidade_integrais = len(planejadas) + len(pagas) - quantidade_parciais

    indicadores = [
        ("Recebido na competência", total_recebido, "moeda", cores["azul_claro"]),
        ("Pago na competência", total_pago_caixa, "moeda", cores["verde_claro"]),
        ("Saldo da conta", saldo_caixa, "moeda", cores["amarelo_claro"]),
        ("Valor aguardando decisão", total_aguardando, "moeda", cores["vermelho_claro"]),
        ("Valor planejado", total_planejado, "moeda", cores["azul_claro"]),
        ("Valor executado", total_executado, "moeda", cores["verde_claro"]),
        ("Saldo de pagamentos parciais", saldo_parcial, "moeda", cores["amarelo_claro"]),
        ("Contas aguardando", len(aguardando), "numero", cores["cinza"]),
        ("Contas planejadas", len(planejadas), "numero", cores["azul_claro"]),
        ("Contas pagas", len(pagas), "numero", cores["verde_claro"]),
        ("Planejamentos parciais", quantidade_parciais, "numero", cores["amarelo_claro"]),
        ("Planejamentos integrais", quantidade_integrais, "numero", cores["cinza"]),
    ]

    linha = 4
    for indice, (rotulo, valor, tipo, fundo) in enumerate(indicadores):
        coluna_base = 1 if indice % 2 == 0 else 4
        if indice % 2 == 0 and indice > 0:
            linha += 3
        ws.merge_cells(start_row=linha, start_column=coluna_base, end_row=linha, end_column=coluna_base + 2)
        rotulo_cell = ws.cell(linha, coluna_base, rotulo)
        rotulo_cell.font = Font(bold=True, color=cores["cinza_texto"], size=10)
        rotulo_cell.fill = PatternFill("solid", fgColor=fundo)
        rotulo_cell.border = borda
        rotulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=linha + 1, start_column=coluna_base, end_row=linha + 1, end_column=coluna_base + 2)
        valor_cell = ws.cell(linha + 1, coluna_base, float(valor) if tipo == "moeda" else int(valor))
        valor_cell.font = Font(bold=True, color=cores["azul_escuro"], size=15)
        valor_cell.fill = PatternFill("solid", fgColor=fundo)
        valor_cell.border = borda
        valor_cell.alignment = Alignment(horizontal="center", vertical="center")
        if tipo == "moeda":
            valor_cell.number_format = 'R$ #,##0.00'
        ws.row_dimensions[linha].height = 22
        ws.row_dimensions[linha + 1].height = 30

    for coluna in range(1, 7):
        ws.column_dimensions[get_column_letter(coluna)].width = 18
    return ws


@planejamento_financeiro_bp.route("/sincronizar-radar", methods=["POST"])
@gestao_required
def sincronizar_radar():
    agora = datetime.now()
    mes = request.form.get("mes", type=int) or agora.month
    ano = request.form.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    try:
        criadas, atualizadas = sincronizar_pagas_do_radar(mes, ano)
        return jsonify({
            "ok": True,
            "message": f"Sincronização concluída. {criadas} conta(s) adicionada(s) e {atualizadas} vínculo(s) recuperado(s).",
            "criadas": criadas,
            "atualizadas": atualizadas,
        })
    except Exception as erro:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "message": f"Erro ao sincronizar Radar: {str(erro)}",
        }), 500


@planejamento_financeiro_bp.route("/exportar")
@gestao_required
def exportar_planejamento():
    agora = datetime.now()
    mes = request.args.get("mes", type=int) or agora.month
    ano = request.args.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    aguardando, planejadas, pagas = montar_listas_planejamento(mes, ano)
    competencia = f"{nome_mes(mes)}/{ano}"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    escrever_aba_agenda(wb, competencia, planejadas, pagas)
    escrever_aba_resumo_diario(wb, competencia, planejadas, pagas)
    escrever_aba_parciais(wb, competencia, planejadas, pagas)
    escrever_aba_geral(wb, competencia, aguardando, planejadas, pagas)
    escrever_aba_executivo(wb, competencia, aguardando, planejadas, pagas, mes, ano)

    wb.active = 0
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"planejamento_financeiro_{nome_mes(mes).lower()}_{ano}.xlsx".replace(" ", "_")
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@planejamento_financeiro_bp.route("/planejar/<int:conta_id>", methods=["POST"])
@gestao_required

def planejar(conta_id):
    mes = request.form.get("mes", type=int)
    ano = request.form.get("ano", type=int)

    if not mes or not ano or not 1 <= mes <= 12:
        return jsonify({"ok": False, "message": "Mês e ano inválidos."}), 400

    conta = db.session.get(ContaPagarImportada, conta_id)
    if not conta:
        return jsonify({
            "ok": False,
            "message": "Conta não encontrada. Sincronize a importação e tente novamente.",
        }), 404

    if conta_esta_cancelada(conta):
        return jsonify({
            "ok": False,
            "message": "Não é possível planejar uma conta cancelada.",
        }), 409

    if conta_esta_paga_local(conta):
        return jsonify({
            "ok": False,
            "message": "Esta conta já está paga no Radar.",
        }), 409

    chave = gerar_chave_conta(conta)
    tipo = texto_limpo(
        request.form.get("tipo_planejamento"),
        "TOTAL",
    ).upper()

    if tipo not in ("TOTAL", "PARCIAL"):
        return jsonify({
            "ok": False,
            "message": "Tipo de planejamento inválido.",
        }), 400

    data_prevista = data_para_date_local(
        request.form.get("data_prevista")
    )
    observacao = texto_limpo(
        request.form.get("observacao_previsao"),
        "",
    )

    registro = PlanejamentoFinanceiro.query.filter_by(
        conta_id=conta_id,
        origem="IMPORTADA",
        mes=mes,
        ano=ano,
    ).first()

    if registro is None and chave:
        registro = PlanejamentoFinanceiro.query.filter_by(
            chave_conta=chave,
            origem="IMPORTADA",
            mes=mes,
            ano=ano,
        ).first()
        if registro is not None:
            registro.conta_id = conta_id

    if registro is None:
        registro = PlanejamentoFinanceiro(
            conta_id=conta_id,
            chave_conta=chave,
            origem="IMPORTADA",
            mes=mes,
            ano=ano,
            status_planejamento="PLANEJADA",
        )
        db.session.add(registro)
        preencher_snapshot_planejamento(registro, conta, chave)
        db.session.flush()
    else:
        preencher_snapshot_planejamento(registro, conta, chave)
        registro.status_planejamento = "PLANEJADA"

    pagamentos_filhos = list(
        getattr(registro, "pagamentos_planejados", []) or []
    )
    if not pagamentos_filhos:
        valor_legado = dinheiro_decimal(
            getattr(registro, "valor_planejado", None)
        )
        if valor_legado > 0:
            pagamento_legado = PlanejamentoPagamento(
                planejamento_id=registro.id,
                tipo=texto_limpo(
                    getattr(registro, "tipo_planejamento", None),
                    "TOTAL",
                ).upper(),
                valor=valor_legado,
                data_prevista=data_para_date_local(
                    getattr(registro, "data_prevista", None)
                ),
                observacao=getattr(
                    registro,
                    "observacao_previsao",
                    None,
                ),
            )
            db.session.add(pagamento_legado)
            db.session.flush()
            pagamentos_filhos.append(pagamento_legado)

    registros_mesma_conta = (
        PlanejamentoFinanceiro.query
        .options(
            selectinload(
                PlanejamentoFinanceiro.pagamentos_planejados
            )
        )
        .filter(
            PlanejamentoFinanceiro.origem == "IMPORTADA",
            or_(
                PlanejamentoFinanceiro.conta_id == conta_id,
                PlanejamentoFinanceiro.chave_conta == chave,
            ),
        )
        .all()
    )

    item_base = {
        "valor": dinheiro_decimal(getattr(conta, "valor", 0))
    }
    valor_original = _valor_original_da_conta(
        item_base,
        registros_mesma_conta,
    )

    pagamentos_existentes = _pagamentos_unicos_da_conta(
        registros_mesma_conta
    )
    total_anterior = sum(
        dinheiro_decimal(getattr(pagamento, "valor", 0))
        for _, pagamento in pagamentos_existentes
    )
    saldo_disponivel = max(
        Decimal("0"),
        valor_original - total_anterior,
    )

    if saldo_disponivel <= 0:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "message": "Esta conta já está totalmente planejada.",
        }), 409

    if tipo == "PARCIAL":
        valor_planejado = dinheiro_decimal(
            request.form.get("valor_planejado")
        )
        if valor_planejado <= 0:
            db.session.rollback()
            return jsonify({
                "ok": False,
                "message": "Informe um valor previsto maior que zero.",
            }), 400
        if not data_prevista:
            db.session.rollback()
            return jsonify({
                "ok": False,
                "message": "Informe a data prevista para o pagamento.",
            }), 400
        if valor_planejado > saldo_disponivel:
            db.session.rollback()
            return jsonify({
                "ok": False,
                "message": (
                    "O valor previsto não pode ultrapassar o saldo "
                    f"restante de {moeda(saldo_disponivel)}."
                ),
            }), 400
    else:
        valor_planejado = saldo_disponivel

        # O botão "Vou pagar" planeja o valor total na competência
        # atualmente aberta. Não podemos usar automaticamente um vencimento
        # herdado de outro mês, pois isso faria a conta desaparecer de
        # Aguardando sem aparecer em Planejadas nesta tela.
        if not data_prevista:
            vencimento_conta = data_para_date_local(
                getattr(conta, "data_vencimento", None)
            )

            if (
                vencimento_conta
                and vencimento_conta.month == mes
                and vencimento_conta.year == ano
            ):
                data_prevista = vencimento_conta
            else:
                hoje = date.today()
                if hoje.month == mes and hoje.year == ano:
                    data_prevista = hoje
                else:
                    data_prevista = ultimo_dia_mes(mes, ano)

    pagamento = PlanejamentoPagamento(
        planejamento_id=registro.id,
        tipo=tipo,
        valor=valor_planejado,
        data_prevista=data_prevista,
        observacao=observacao or None,
    )
    db.session.add(pagamento)

    total_atualizado = total_anterior + valor_planejado
    registro.valor_snapshot = valor_original
    registro.valor_planejado = total_atualizado
    registro.tipo_planejamento = (
        "TOTAL"
        if total_atualizado >= valor_original
        else "PARCIAL"
    )
    registro.data_prevista = data_prevista
    registro.observacao_previsao = observacao or None
    registro.status_planejamento = "PLANEJADA"

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "message": (
                "Já existe um planejamento desta conta nesta competência. "
                "Atualize a página e tente novamente."
            ),
        }), 409
    except Exception as erro:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "message": f"Erro ao salvar previsão: {erro}",
        }), 500

    saldo_final = max(
        Decimal("0"),
        valor_original - total_atualizado,
    )

    return jsonify({
        "ok": True,
        "message": (
            f"Valor parcial previsto. O saldo de "
            f"{moeda(saldo_final)} permanece em Aguardando Decisão."
            if saldo_final > 0
            else "Conta totalmente planejada."
        ),
        "saldo_restante": float(saldo_final),
        "saldo_restante_formatado": moeda(saldo_final),
        "valor_planejado": float(valor_planejado),
        "valor_planejado_formatado": moeda(valor_planejado),
        "tipo_planejamento": tipo,
        "data_prevista": formatar_data(data_prevista),
        "pagamento_planejado_id": pagamento.id,
    })

@planejamento_financeiro_bp.route("/remover/<int:conta_id>", methods=["POST"])
@gestao_required
def remover(conta_id):
    mes = request.form.get("mes", type=int)
    ano = request.form.get("ano", type=int)
    pagamento_id = request.form.get("pagamento_planejado_id", type=int)

    if not mes or not ano:
        return jsonify({"ok": False, "message": "Mês e ano inválidos."}), 400

    # Quando a linha representa uma parcela planejada, remove somente aquela
    # previsão, inclusive quando a data prevista pertence a outro mês.
    if pagamento_id:
        pagamento = PlanejamentoPagamento.query.get(pagamento_id)
        if not pagamento:
            return jsonify({"ok": False, "message": "Previsão não encontrada."}), 404

        registro = PlanejamentoFinanceiro.query.get(pagamento.planejamento_id)
        db.session.delete(pagamento)
        db.session.flush()

        if registro:
            pagamentos_restantes = list(registro.pagamentos_planejados or [])
            total_restante = sum(
                dinheiro_decimal(p.valor)
                for p in pagamentos_restantes
                if p.id != pagamento_id
            )

            registro.valor_planejado = total_restante if total_restante > 0 else None
            registro.tipo_planejamento = (
                "TOTAL"
                if total_restante >= dinheiro_decimal(registro.valor_snapshot)
                else "PARCIAL"
            )

            if pagamentos_restantes:
                ultimo = sorted(
                    [p for p in pagamentos_restantes if p.id != pagamento_id],
                    key=lambda p: (data_para_date_local(p.data_prevista) or date.min, p.id or 0),
                )[-1] if any(p.id != pagamento_id for p in pagamentos_restantes) else None
                registro.data_prevista = getattr(ultimo, "data_prevista", None) if ultimo else None
                registro.observacao_previsao = getattr(ultimo, "observacao", None) if ultimo else None
            else:
                registro.data_prevista = None
                registro.observacao_previsao = None
                registro.status_planejamento = "AGUARDANDO"

        db.session.commit()
        return jsonify({"ok": True, "message": "Previsão removida do planejamento."})

    conta = ContaPagarImportada.query.get(conta_id)
    chave = gerar_chave_conta(conta) if conta else None

    filtros = [PlanejamentoFinanceiro.origem == "IMPORTADA"]
    if chave:
        filtros.append(or_(
            PlanejamentoFinanceiro.conta_id == conta_id,
            PlanejamentoFinanceiro.chave_conta == chave,
        ))
    else:
        filtros.append(PlanejamentoFinanceiro.conta_id == conta_id)

    registro = (
        PlanejamentoFinanceiro.query
        .options(selectinload(PlanejamentoFinanceiro.pagamentos_planejados))
        .filter(*filtros)
        .order_by(PlanejamentoFinanceiro.id.desc())
        .first()
    )

    if registro:
        db.session.delete(registro)
        db.session.commit()

    return jsonify({"ok": True, "message": "Conta removida do planejamento."})