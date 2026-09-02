import os
import re
import zipfile
import urllib.request
import secrets
from io import BytesIO
from datetime import datetime, timedelta

import cloudinary.uploader
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    session,
    flash,
    abort,
    send_file,
    current_app,
    jsonify,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from sqlalchemy import or_, and_

from database import db
from models.cliente import Cliente
from models.usuario import Usuario
from models.evidencia_frota import (
    EvidenciaRegistro,
    EvidenciaCampoPai,
    EvidenciaCampoFilho,
    EvidenciaTipoFoto,
    EvidenciaImagem,
    EvidenciaTabelaControle,
    EvidenciaTabelaColuna,
    EvidenciaTabelaLinha,
    EvidenciaTabelaCelula,
    EvidenciaLinkPublico,
)


evidencias_frota_bp = Blueprint(
    "evidencias_frota",
    __name__,
    url_prefix="/gestao/evidencias",
)

TIPOS_PADRAO = [
    "Antes",
    "Depois",
    "Durante",
    "Finalizado",
    "Avaria",
    "Componente removido",
    "Componente instalado",
    "Etiqueta",
    "Comprovante",
    "Outro",
]


# =========================================================
# HELPERS
# =========================================================

def usuario_logado():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return Usuario.query.get(user_id)


def usuario_eh_admin_ou_gestao():
    return session.get("user_role") in ["admin", "gestao", "gestor"]


def exigir_login():
    if not session.get("user_id"):
        return redirect("/login")
    return None


def requisicao_ajax():
    return request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json"


def responder_ok(mensagem="Operação realizada", **extra):
    payload = {"ok": True, "mensagem": mensagem}
    payload.update(extra)
    return jsonify(payload)


def responder_erro(mensagem="Não foi possível concluir", status=400, **extra):
    payload = {"ok": False, "erro": mensagem}
    payload.update(extra)
    return jsonify(payload), status


def voltar_ou_json(url, mensagem="Operação realizada", **extra):
    if requisicao_ajax():
        return responder_ok(mensagem, redirect_url=url, **extra)
    flash(mensagem, "success")
    return redirect(url)


def texto(valor):
    return "" if valor is None else str(valor).strip()


def normalizar_texto(valor):
    valor = texto(valor)
    if not valor:
        return ""
    return " ".join(valor.split()).upper()


def limpar_nome_arquivo(valor):
    valor = normalizar_texto(valor) or "SEM_NOME"
    valor = valor.replace("/", "-").replace("\\", "-")
    valor = re.sub(r"[^A-Z0-9_\-\. ]+", "", valor)
    valor = valor.replace(" ", "_")
    return valor[:90] or "SEM_NOME"


def cliente_permitido(cliente_id=None, cliente_nome=None):
    if usuario_eh_admin_ou_gestao():
        return True

    usuario = usuario_logado()
    if not usuario or not getattr(usuario, "cliente", None):
        return False

    if cliente_id and usuario.cliente_id == int(cliente_id):
        return True

    return normalizar_texto(usuario.cliente.nome) == normalizar_texto(cliente_nome)


def aplicar_permissao_registros(query):
    if usuario_eh_admin_ou_gestao():
        return query

    usuario = usuario_logado()
    if not usuario or not getattr(usuario, "cliente", None):
        return query.filter(EvidenciaRegistro.id == 0)

    return query.filter(
        or_(
            EvidenciaRegistro.cliente_id == usuario.cliente_id,
            db.func.upper(EvidenciaRegistro.cliente_nome) == normalizar_texto(usuario.cliente.nome),
        )
    )


def obter_registro_ou_404(registro_id):
    registro = EvidenciaRegistro.query.get_or_404(registro_id)
    if not cliente_permitido(registro.cliente_id, registro.cliente_nome):
        abort(403)
    return registro


def obter_pai_ou_404(pai_id):
    pai = EvidenciaCampoPai.query.get_or_404(pai_id)
    obter_registro_ou_404(pai.registro_id)
    return pai


def obter_tabela_ou_404(tabela_id):
    tabela = EvidenciaTabelaControle.query.get_or_404(tabela_id)
    obter_pai_ou_404(tabela.campo_pai_id)
    return tabela


def obter_coluna_ou_404(coluna_id):
    coluna = EvidenciaTabelaColuna.query.get_or_404(coluna_id)
    obter_tabela_ou_404(coluna.tabela_id)
    return coluna


def obter_linha_ou_404(linha_id):
    linha = EvidenciaTabelaLinha.query.get_or_404(linha_id)
    obter_tabela_ou_404(linha.tabela_id)
    return linha


def montar_mapa_celulas(tabela):
    mapa = {}
    if not tabela:
        return mapa
    celulas = EvidenciaTabelaCelula.query.join(EvidenciaTabelaLinha).filter(
        EvidenciaTabelaLinha.tabela_id == tabela.id
    ).all()
    for celula in celulas:
        mapa[f"{celula.linha_id}_{celula.coluna_id}"] = celula.valor or ""
    return mapa


def montar_grupos_cabecalho(tabela):
    """
    Monta cabeçalhos agrupados por sequência de colunas com o mesmo grupo.
    Não muda o banco; usa o campo coluna.grupo existente.
    """
    grupos = []
    if not tabela:
        return grupos

    ultimo_grupo = None
    for coluna in tabela.colunas:
        grupo_nome = texto(getattr(coluna, "grupo", None))
        cor = texto(getattr(coluna, "cor", None)) or "padrao"

        # Coluna sem grupo fica isolada para não grudar indevidamente em outra.
        chave = f"grupo:{normalizar_texto(grupo_nome)}" if grupo_nome else f"coluna:{coluna.id}"
        titulo = grupo_nome if grupo_nome else (coluna.nome or "")

        if ultimo_grupo and ultimo_grupo["chave"] == chave:
            ultimo_grupo["colspan"] += 1
            ultimo_grupo["colunas"].append(coluna)
        else:
            ultimo_grupo = {
                "chave": chave,
                "titulo": titulo,
                "colspan": 1,
                "cor": cor,
                "colunas": [coluna],
            }
            grupos.append(ultimo_grupo)

    return grupos


def salvar_estado_tabela_do_request(tabela):
    """Salva o que está digitado antes de adicionar/remover linha ou coluna."""
    if not tabela:
        return

    novo_titulo = texto(request.form.get("tabela_titulo")) or texto(request.form.get("titulo"))
    if novo_titulo:
        tabela.titulo = novo_titulo

    existentes = {
        (celula.linha_id, celula.coluna_id): celula
        for celula in EvidenciaTabelaCelula.query.join(EvidenciaTabelaLinha).filter(
            EvidenciaTabelaLinha.tabela_id == tabela.id
        ).all()
    }

    grupos_bulk = {}
    for nome_campo, valor_campo in request.form.items():
        if not nome_campo.startswith("grupo_bulk_"):
            continue
        ids = nome_campo.replace("grupo_bulk_", "").split("_")
        for id_coluna in ids:
            if str(id_coluna).isdigit():
                grupos_bulk[int(id_coluna)] = texto(valor_campo)

    for coluna in tabela.colunas:
        novo_nome = texto(request.form.get(f"coluna_nome_{coluna.id}"))
        novo_grupo = grupos_bulk.get(coluna.id, texto(request.form.get(f"coluna_grupo_{coluna.id}")))
        nova_cor = texto(request.form.get(f"coluna_cor_{coluna.id}"))
        if novo_nome:
            coluna.nome = novo_nome
        coluna.grupo = novo_grupo or None
        coluna.cor = nova_cor or getattr(coluna, "cor", None) or "padrao"
        nova_ordem = texto(request.form.get(f"coluna_ordem_{coluna.id}"))
        if nova_ordem.isdigit():
            coluna.ordem = int(nova_ordem)

    for linha in tabela.linhas:
        novo_rotulo = texto(request.form.get(f"linha_rotulo_{linha.id}"))
        if novo_rotulo:
            linha.rotulo = novo_rotulo

        for coluna in tabela.colunas:
            campo = f"celula_{linha.id}_{coluna.id}"
            valor = texto(request.form.get(campo))
            celula = existentes.get((linha.id, coluna.id))
            if celula:
                celula.valor = valor or None
                celula.atualizado_em = datetime.utcnow()
            elif valor:
                db.session.add(EvidenciaTabelaCelula(
                    linha_id=linha.id,
                    coluna_id=coluna.id,
                    valor=valor,
                ))

    tabela.atualizado_em = datetime.utcnow()


def obter_filho_ou_404(filho_id):
    filho = EvidenciaCampoFilho.query.get_or_404(filho_id)
    obter_registro_ou_404(filho.campo_pai.registro_id)
    return filho


def registro_id_da_imagem(imagem):
    if getattr(imagem, "campo_pai", None):
        return imagem.campo_pai.registro_id
    if getattr(imagem, "campo_filho", None) and imagem.campo_filho.campo_pai:
        return imagem.campo_filho.campo_pai.registro_id
    abort(404)


def obter_imagem_ou_404(imagem_id):
    imagem = EvidenciaImagem.query.get_or_404(imagem_id)
    registro_id = registro_id_da_imagem(imagem)
    obter_registro_ou_404(registro_id)
    return imagem


def origem_imagem(img):
    if getattr(img, "tipo_foto", None):
        return img.tipo_foto
    if getattr(img, "campo_filho", None):
        return img.campo_filho.nome
    return "Sem tipo"


def pai_da_imagem(img):
    if getattr(img, "campo_pai", None):
        return img.campo_pai
    if getattr(img, "campo_filho", None):
        return img.campo_filho.campo_pai
    return None


def registro_da_imagem(img):
    pai = pai_da_imagem(img)
    return pai.registro if pai else None


def garantir_tipos_padrao():
    if EvidenciaTipoFoto.query.first():
        return
    for ordem, nome in enumerate(TIPOS_PADRAO, start=1):
        db.session.add(EvidenciaTipoFoto(nome=nome, ordem=ordem, ativo=True))
    db.session.commit()


def tipos_foto_para_select():
    garantir_tipos_padrao()
    return EvidenciaTipoFoto.query.filter_by(ativo=True).order_by(
        EvidenciaTipoFoto.ordem.asc(),
        EvidenciaTipoFoto.nome.asc(),
    ).all()


def salvar_tipo_foto_se_novo(nome):
    nome = texto(nome)
    if not nome:
        return ""

    existente = EvidenciaTipoFoto.query.filter(
        db.func.upper(EvidenciaTipoFoto.nome) == normalizar_texto(nome)
    ).first()
    if existente:
        if not existente.ativo:
            existente.ativo = True
            db.session.commit()
        return existente.nome

    maior_ordem = db.session.query(db.func.max(EvidenciaTipoFoto.ordem)).scalar() or 0
    tipo = EvidenciaTipoFoto(nome=nome, ordem=maior_ordem + 1, ativo=True)
    db.session.add(tipo)
    db.session.commit()
    return tipo.nome


def imagens_do_pai(pai_id):
    return EvidenciaImagem.query.outerjoin(
        EvidenciaCampoFilho,
        EvidenciaImagem.campo_filho_id == EvidenciaCampoFilho.id,
    ).filter(
        or_(
            EvidenciaImagem.campo_pai_id == int(pai_id),
            EvidenciaCampoFilho.campo_pai_id == int(pai_id),
        )
    ).order_by(
        EvidenciaImagem.tipo_foto.asc().nullslast(),
        EvidenciaImagem.ordem.asc(),
        EvidenciaImagem.id.asc(),
    ).all()


def preparar_registro_para_tela(registro):
    for pai in registro.campos_pai:
        pai.imagens_tela = imagens_do_pai(pai.id)
        pai.total_imagens = len(pai.imagens_tela)
        if getattr(pai, "tabela_controle", None):
            pai.tabela_controle.mapa_celulas = montar_mapa_celulas(pai.tabela_controle)
            pai.tabela_controle.grupos_cabecalho = montar_grupos_cabecalho(pai.tabela_controle)
    return registro


def migrar_imagens_antigas_para_pai():
    """
    Após aplicar a alteração no banco, este helper preenche campo_pai_id e tipo_foto
    nas imagens antigas que ainda estão presas ao campo filho.
    Não apaga filho e não remove nenhuma foto.
    """
    imagens = EvidenciaImagem.query.filter(
        EvidenciaImagem.campo_pai_id.is_(None),
        EvidenciaImagem.campo_filho_id.isnot(None),
    ).all()

    alteradas = 0
    for img in imagens:
        if not img.campo_filho or not img.campo_filho.campo_pai:
            continue
        img.campo_pai_id = img.campo_filho.campo_pai_id
        if not img.tipo_foto:
            img.tipo_foto = img.campo_filho.nome
        alteradas += 1

    if alteradas:
        db.session.commit()
    return alteradas


def cloudinary_configurado():
    return all([
        os.getenv("CLOUD_NAME"),
        os.getenv("API_KEY"),
        os.getenv("API_SECRET"),
    ])


def salvar_imagem_evidencia(arquivo, registro, pai, tipo_foto=None, filho=None):
    if not arquivo or not arquivo.filename:
        return None

    nome_original = secure_filename(arquivo.filename) or "imagem.jpg"
    origem = tipo_foto or (filho.nome if filho else "SEM_TIPO")

    if cloudinary_configurado():
        folder = "/".join([
            "easy_control",
            "evidencias",
            limpar_nome_arquivo(registro.cliente_nome),
            limpar_nome_arquivo(registro.frota or registro.placa or "veiculo"),
            limpar_nome_arquivo(pai.nome),
            limpar_nome_arquivo(origem),
        ])

        upload = cloudinary.uploader.upload(
            arquivo,
            folder=folder,
            resource_type="image",
        )

        return {
            "url": upload.get("secure_url"),
            "public_id": upload.get("public_id"),
            "caminho_local": None,
            "nome_original": nome_original,
        }

    base_upload = current_app.config.get("UPLOAD_FOLDER") or os.path.join(
        current_app.root_path,
        "static",
        "uploads",
    )

    pasta_relativa = os.path.join(
        "evidencias",
        limpar_nome_arquivo(registro.cliente_nome),
        limpar_nome_arquivo(registro.frota or registro.placa or "veiculo"),
        limpar_nome_arquivo(pai.nome),
        limpar_nome_arquivo(origem),
    )

    pasta_destino = os.path.join(base_upload, pasta_relativa)
    os.makedirs(pasta_destino, exist_ok=True)

    nome_final = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{nome_original}"
    caminho_final = os.path.join(pasta_destino, nome_final)
    arquivo.save(caminho_final)

    url = f"/static/uploads/{pasta_relativa.replace(os.sep, '/')}/{nome_final}"

    return {
        "url": url,
        "public_id": None,
        "caminho_local": caminho_final,
        "nome_original": nome_original,
    }


def clientes_para_select():
    if usuario_eh_admin_ou_gestao():
        return Cliente.query.order_by(Cliente.nome.asc()).all()

    usuario = usuario_logado()
    if usuario and getattr(usuario, "cliente", None):
        return [usuario.cliente]

    return []


def contar_imagens_registro(registro):
    total = 0
    for pai in registro.campos_pai:
        total += len(imagens_do_pai(pai.id))
    return total


def registros_filtrados():
    cliente_id = request.args.get("cliente_id", type=int)
    busca = texto(request.args.get("busca"))
    campo = texto(request.args.get("campo"))

    query = aplicar_permissao_registros(EvidenciaRegistro.query)

    if cliente_id:
        query = query.filter(EvidenciaRegistro.cliente_id == cliente_id)

    if busca:
        like = f"%{busca}%"
        query = query.filter(or_(
            EvidenciaRegistro.cliente_nome.ilike(like),
            EvidenciaRegistro.frota.ilike(like),
            EvidenciaRegistro.placa.ilike(like),
        ))

    if campo:
        query = query.join(EvidenciaCampoPai).filter(
            EvidenciaCampoPai.nome.ilike(f"%{campo}%")
        )

    return query.order_by(
        EvidenciaRegistro.atualizado_em.desc(),
        EvidenciaRegistro.id.desc(),
    ).all()


# =========================================================
# TELAS
# =========================================================

@evidencias_frota_bp.route("/")
def index():
    resp = exigir_login()
    if resp:
        return resp

    migrar_imagens_antigas_para_pai()
    registros = registros_filtrados()

    for r in registros:
        r.total_campos_pai = len(r.campos_pai)
        r.total_filhos = 0
        r.total_imagens = contar_imagens_registro(r)

    return render_template(
        "gestao/evidencias_frota/index.html",
        registros=registros,
        clientes=clientes_para_select(),
        cliente_id=request.args.get("cliente_id", type=int),
        busca=texto(request.args.get("busca")),
        campo=texto(request.args.get("campo")),
    )


@evidencias_frota_bp.route("/novo", methods=["GET", "POST"])
def novo_registro():
    resp = exigir_login()
    if resp:
        return resp

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id", type=int)
        frota = texto(request.form.get("frota"))
        placa = texto(request.form.get("placa"))

        if not cliente_id:
            flash("Selecione um cliente.", "danger")
            return redirect(request.url)

        cliente = Cliente.query.get_or_404(cliente_id)

        if not cliente_permitido(cliente.id, cliente.nome):
            abort(403)

        if not frota and not placa:
            flash("Informe a frota ou a placa.", "danger")
            return redirect(request.url)

        registro = EvidenciaRegistro(
            cliente_id=cliente.id,
            cliente_nome=cliente.nome,
            frota=frota or None,
            placa=normalizar_texto(placa) or None,
        )

        db.session.add(registro)
        db.session.commit()

        flash("Painel de evidências criado com sucesso!", "success")
        return redirect(f"/gestao/evidencias/{registro.id}")

    return render_template(
        "gestao/evidencias_frota/form.html",
        clientes=clientes_para_select(),
    )


@evidencias_frota_bp.route("/<int:registro_id>")
def detalhe(registro_id):
    resp = exigir_login()
    if resp:
        return resp

    migrar_imagens_antigas_para_pai()
    registro = preparar_registro_para_tela(obter_registro_ou_404(registro_id))

    links_publicos = EvidenciaLinkPublico.query.filter_by(registro_id=registro.id).order_by(
        EvidenciaLinkPublico.criado_em.desc(),
        EvidenciaLinkPublico.id.desc(),
    ).all()

    return render_template(
        "gestao/evidencias_frota/detalhe.html",
        registro=registro,
        links_publicos=links_publicos,
        pode_gerar_link=usuario_eh_admin_ou_gestao(),
        tipos_foto=tipos_foto_para_select(),
    )


@evidencias_frota_bp.route("/<int:registro_id>/excluir", methods=["POST"])
def excluir_registro(registro_id):
    resp = exigir_login()
    if resp:
        return resp

    registro = obter_registro_ou_404(registro_id)
    db.session.delete(registro)
    db.session.commit()

    flash("Painel de evidências excluído.", "success")
    return redirect("/gestao/evidencias/")


# =========================================================
# CAMPO PAI / TIPOS / ROTAS ANTIGAS DE FILHO
# =========================================================

@evidencias_frota_bp.route("/<int:registro_id>/pai/novo", methods=["POST"])
def criar_pai(registro_id):
    resp = exigir_login()
    if resp:
        return resp

    registro = obter_registro_ou_404(registro_id)
    nome = texto(request.form.get("nome"))

    if not nome:
        flash("Informe o nome do campo pai.", "danger")
        return redirect(f"/gestao/evidencias/{registro.id}")

    pai = EvidenciaCampoPai(
        registro_id=registro.id,
        nome=nome,
        ordem=len(registro.campos_pai) + 1,
    )

    db.session.add(pai)
    db.session.commit()

    flash("Campo pai criado.", "success")
    return redirect(f"/gestao/evidencias/{registro.id}")


@evidencias_frota_bp.route("/pai/<int:pai_id>/editar", methods=["POST"])
def editar_pai(pai_id):
    resp = exigir_login()
    if resp:
        return resp

    pai = obter_pai_ou_404(pai_id)
    nome = texto(request.form.get("nome"))

    if nome:
        pai.nome = nome
        db.session.commit()
        flash("Campo pai atualizado.", "success")

    return redirect(f"/gestao/evidencias/{pai.registro_id}")


@evidencias_frota_bp.route("/pai/<int:pai_id>/excluir", methods=["POST"])
def excluir_pai(pai_id):
    resp = exigir_login()
    if resp:
        return resp

    pai = obter_pai_ou_404(pai_id)
    registro_id = pai.registro_id

    db.session.delete(pai)
    db.session.commit()

    flash("Campo pai excluído.", "success")
    return redirect(f"/gestao/evidencias/{registro_id}")


@evidencias_frota_bp.route("/tipos/novo", methods=["POST"])
def criar_tipo_foto():
    resp = exigir_login()
    if resp:
        return resp

    if not usuario_eh_admin_ou_gestao():
        abort(403)

    registro_id = request.form.get("registro_id", type=int)
    nome = salvar_tipo_foto_se_novo(request.form.get("nome"))

    if nome:
        flash("Tipo de foto cadastrado.", "success")
    else:
        flash("Informe o nome do tipo.", "danger")

    return redirect(f"/gestao/evidencias/{registro_id}" if registro_id else "/gestao/evidencias/")


@evidencias_frota_bp.route("/tipos/<int:tipo_id>/desativar", methods=["POST"])
def desativar_tipo_foto(tipo_id):
    resp = exigir_login()
    if resp:
        return resp

    if not usuario_eh_admin_ou_gestao():
        abort(403)

    tipo = EvidenciaTipoFoto.query.get_or_404(tipo_id)
    tipo.ativo = False
    db.session.commit()

    registro_id = request.form.get("registro_id", type=int)
    flash("Tipo de foto desativado.", "success")
    return redirect(f"/gestao/evidencias/{registro_id}" if registro_id else "/gestao/evidencias/")


# Estas rotas antigas ficam ativas somente para não quebrar dados ou telas antigas.
@evidencias_frota_bp.route("/pai/<int:pai_id>/filho/novo", methods=["POST"])
def criar_filho(pai_id):
    resp = exigir_login()
    if resp:
        return resp

    pai = obter_pai_ou_404(pai_id)
    nome = texto(request.form.get("nome"))

    if not nome:
        flash("Informe o nome do campo filho.", "danger")
        return redirect(f"/gestao/evidencias/{pai.registro_id}")

    filho = EvidenciaCampoFilho(campo_pai_id=pai.id, nome=nome, ordem=len(pai.filhos) + 1)
    db.session.add(filho)
    db.session.commit()

    flash("Campo filho criado.", "success")
    return redirect(f"/gestao/evidencias/{pai.registro_id}")


@evidencias_frota_bp.route("/filho/<int:filho_id>/editar", methods=["POST"])
def editar_filho(filho_id):
    resp = exigir_login()
    if resp:
        return resp

    filho = obter_filho_ou_404(filho_id)
    nome = texto(request.form.get("nome"))

    if nome:
        filho.nome = nome
        db.session.commit()
        flash("Campo filho atualizado.", "success")

    return redirect(f"/gestao/evidencias/{filho.campo_pai.registro_id}")


@evidencias_frota_bp.route("/filho/<int:filho_id>/excluir", methods=["POST"])
def excluir_filho(filho_id):
    resp = exigir_login()
    if resp:
        return resp

    filho = obter_filho_ou_404(filho_id)
    registro_id = filho.campo_pai.registro_id

    # Mantém as fotos: antes de excluir o filho, transfere para o pai.
    for img in list(filho.imagens):
        img.campo_pai_id = filho.campo_pai_id
        img.tipo_foto = img.tipo_foto or filho.nome
        img.campo_filho_id = None

    db.session.delete(filho)
    db.session.commit()

    flash("Campo filho excluído e imagens preservadas no campo pai.", "success")
    return redirect(f"/gestao/evidencias/{registro_id}")


# =========================================================
# TABELA DE CONTROLE POR CAMPO PAI
# =========================================================


def inteiro_limitado(valor, padrao=1, minimo=0, maximo=100):
    try:
        numero = int(valor)
    except (TypeError, ValueError):
        numero = padrao
    return max(minimo, min(maximo, numero))


def reorganizar_ordem_colunas(tabela):
    colunas = sorted(tabela.colunas, key=lambda c: (getattr(c, "ordem", 0) or 0, c.id or 0))
    for indice, coluna in enumerate(colunas, start=1):
        coluna.ordem = indice


def abrir_espaco_coluna(tabela, ordem_alvo):
    for coluna in tabela.colunas:
        if (coluna.ordem or 0) >= ordem_alvo:
            coluna.ordem = (coluna.ordem or 0) + 1


def criar_estrutura_tabela(tabela, qtd_colunas, qtd_linhas):
    """Cria colunas e linhas iniciais para a tabela, sem apagar nada existente."""
    inicio_coluna = len(tabela.colunas) + 1
    inicio_linha = len(tabela.linhas) + 1

    for i in range(qtd_colunas):
        numero = inicio_coluna + i
        db.session.add(EvidenciaTabelaColuna(
            tabela_id=tabela.id,
            grupo=None,
            nome=f"Coluna {numero}",
            tipo="texto",
            cor="padrao",
            ordem=numero,
        ))

    for i in range(qtd_linhas):
        numero = inicio_linha + i
        db.session.add(EvidenciaTabelaLinha(
            tabela_id=tabela.id,
            rotulo=f"Linha {numero}",
            ordem=numero,
        ))

    tabela.atualizado_em = datetime.utcnow()


@evidencias_frota_bp.route("/pai/<int:pai_id>/tabela/criar", methods=["POST"])
def criar_tabela_controle(pai_id):
    resp = exigir_login()
    if resp:
        return resp

    pai = obter_pai_ou_404(pai_id)
    titulo = texto(request.form.get("titulo")) or pai.nome
    qtd_colunas = inteiro_limitado(request.form.get("qtd_colunas"), padrao=5, minimo=1, maximo=40)
    qtd_linhas = inteiro_limitado(request.form.get("qtd_linhas"), padrao=10, minimo=1, maximo=300)

    if pai.tabela_controle:
        flash("Este campo pai já possui uma tabela de controle.", "warning")
        return redirect(f"/gestao/evidencias/{pai.registro_id}#tabelaPai{pai.id}")

    tabela = EvidenciaTabelaControle(
        campo_pai_id=pai.id,
        titulo=titulo,
        ativa=True,
        ordem=1,
    )
    db.session.add(tabela)
    db.session.flush()
    criar_estrutura_tabela(tabela, qtd_colunas, qtd_linhas)
    db.session.commit()

    flash(f"Tabela criada com {qtd_colunas} coluna(s) e {qtd_linhas} linha(s).", "success")
    return redirect(f"/gestao/evidencias/{pai.registro_id}#tabelaPai{pai.id}")


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/editar", methods=["POST"])
def editar_tabela_controle(tabela_id):
    resp = exigir_login()
    if resp:
        return resp

    tabela = obter_tabela_ou_404(tabela_id)
    titulo = texto(request.form.get("titulo"))
    if titulo:
        tabela.titulo = titulo
        tabela.atualizado_em = datetime.utcnow()
        db.session.commit()
        flash("Tabela atualizada.", "success")
    return redirect(f"/gestao/evidencias/{tabela.campo_pai.registro_id}#tabelaPai{tabela.campo_pai_id}")


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/excluir", methods=["POST"])
def excluir_tabela_controle(tabela_id):
    resp = exigir_login()
    if resp:
        return resp

    tabela = obter_tabela_ou_404(tabela_id)
    registro_id = tabela.campo_pai.registro_id
    db.session.delete(tabela)
    db.session.commit()
    url = f"/gestao/evidencias/{registro_id}"
    return voltar_ou_json(url, "Tabela de controle excluída.", tabela_removida=True)


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/coluna/nova", methods=["POST"])
def criar_coluna_tabela(tabela_id):
    resp = exigir_login()
    if resp:
        if requisicao_ajax():
            return responder_erro("login", 401)
        return resp

    tabela = obter_tabela_ou_404(tabela_id)
    salvar_estado_tabela_do_request(tabela)

    nome = texto(request.form.get("nome"))
    grupo = texto(request.form.get("grupo")) or None
    tipo = texto(request.form.get("tipo")) or "texto"
    cor = texto(request.form.get("cor")) or "padrao"
    quantidade = inteiro_limitado(request.form.get("quantidade"), padrao=1, minimo=1, maximo=40)

    inserir_apos = texto(request.form.get("inserir_apos_coluna_id"))
    if inserir_apos.isdigit():
        coluna_base = EvidenciaTabelaColuna.query.filter_by(id=int(inserir_apos), tabela_id=tabela.id).first()
    else:
        coluna_base = None

    if coluna_base:
        ordem_inicial = (coluna_base.ordem or 0) + 1
    elif grupo:
        colunas_do_grupo = [c for c in tabela.colunas if normalizar_texto(c.grupo) == normalizar_texto(grupo)]
        ordem_inicial = (max([(c.ordem or 0) for c in colunas_do_grupo]) + 1) if colunas_do_grupo else (len(tabela.colunas) + 1)
    else:
        ordem_inicial = len(tabela.colunas) + 1

    abrir_espaco_coluna(tabela, ordem_inicial)
    for i in range(quantidade):
        numero = ordem_inicial + i
        nome_final = nome if quantidade == 1 and nome else f"Coluna {numero}"
        db.session.add(EvidenciaTabelaColuna(
            tabela_id=tabela.id,
            grupo=grupo,
            nome=nome_final,
            tipo=tipo,
            cor=cor,
            ordem=numero,
        ))

    reorganizar_ordem_colunas(tabela)
    tabela.atualizado_em = datetime.utcnow()
    db.session.commit()

    url = f"/gestao/evidencias/{tabela.campo_pai.registro_id}#tabelaPai{tabela.campo_pai_id}"
    return voltar_ou_json(url, f"{quantidade} coluna(s) adicionada(s).", tabela_id=tabela.id, campo_pai_id=tabela.campo_pai_id)



@evidencias_frota_bp.route("/tabela/coluna/<int:coluna_id>/excluir", methods=["POST"])
def excluir_coluna_tabela(coluna_id):
    resp = exigir_login()
    if resp:
        return resp

    coluna = obter_coluna_ou_404(coluna_id)
    tabela = coluna.tabela
    salvar_estado_tabela_do_request(tabela)
    registro_id = tabela.campo_pai.registro_id
    campo_pai_id = tabela.campo_pai_id
    db.session.delete(coluna)
    reorganizar_ordem_colunas(tabela)
    tabela.atualizado_em = datetime.utcnow()
    db.session.commit()
    url = f"/gestao/evidencias/{registro_id}#tabelaPai{campo_pai_id}"
    return voltar_ou_json(url, "Coluna excluída.", tabela_id=tabela.id, campo_pai_id=campo_pai_id)


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/linha/nova", methods=["POST"])
def criar_linha_tabela(tabela_id):
    resp = exigir_login()
    if resp:
        return resp

    tabela = obter_tabela_ou_404(tabela_id)
    salvar_estado_tabela_do_request(tabela)
    rotulo = texto(request.form.get("rotulo"))
    quantidade = inteiro_limitado(request.form.get("quantidade"), padrao=1, minimo=1, maximo=300)

    inicio = len(tabela.linhas) + 1
    for i in range(quantidade):
        numero = inicio + i
        rotulo_final = rotulo if quantidade == 1 and rotulo else f"Linha {numero}"
        db.session.add(EvidenciaTabelaLinha(
            tabela_id=tabela.id,
            rotulo=rotulo_final,
            ordem=numero,
        ))

    tabela.atualizado_em = datetime.utcnow()
    db.session.commit()

    flash(f"{quantidade} linha(s) adicionada(s).", "success")
    return redirect(f"/gestao/evidencias/{tabela.campo_pai.registro_id}#tabelaPai{tabela.campo_pai_id}")


@evidencias_frota_bp.route("/tabela/linha/<int:linha_id>/excluir", methods=["POST"])
def excluir_linha_tabela(linha_id):
    resp = exigir_login()
    if resp:
        return resp

    linha = obter_linha_ou_404(linha_id)
    tabela = linha.tabela
    salvar_estado_tabela_do_request(tabela)
    registro_id = tabela.campo_pai.registro_id
    campo_pai_id = tabela.campo_pai_id
    db.session.delete(linha)
    tabela.atualizado_em = datetime.utcnow()
    db.session.commit()
    url = f"/gestao/evidencias/{registro_id}#tabelaPai{campo_pai_id}"
    return voltar_ou_json(url, "Linha excluída.", tabela_id=tabela.id, campo_pai_id=campo_pai_id, linha_id=linha_id)


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/celulas/salvar", methods=["POST"])
def salvar_celulas_tabela(tabela_id):
    resp = exigir_login()
    if resp:
        return resp

    tabela = obter_tabela_ou_404(tabela_id)

    novo_titulo = texto(request.form.get("tabela_titulo"))
    if novo_titulo:
        tabela.titulo = novo_titulo

    existentes = {
        (celula.linha_id, celula.coluna_id): celula
        for celula in EvidenciaTabelaCelula.query.join(EvidenciaTabelaLinha).filter(
            EvidenciaTabelaLinha.tabela_id == tabela.id
        ).all()
    }

    grupos_bulk = {}
    for nome_campo, valor_campo in request.form.items():
        if not nome_campo.startswith("grupo_bulk_"):
            continue
        ids = nome_campo.replace("grupo_bulk_", "").split("_")
        for id_coluna in ids:
            if str(id_coluna).isdigit():
                grupos_bulk[int(id_coluna)] = texto(valor_campo)

    for coluna in tabela.colunas:
        novo_nome = texto(request.form.get(f"coluna_nome_{coluna.id}"))
        novo_grupo = grupos_bulk.get(coluna.id, texto(request.form.get(f"coluna_grupo_{coluna.id}")))
        nova_cor = texto(request.form.get(f"coluna_cor_{coluna.id}"))
        if novo_nome:
            coluna.nome = novo_nome
        coluna.grupo = novo_grupo or None
        if nova_cor:
            coluna.cor = nova_cor
        nova_ordem = texto(request.form.get(f"coluna_ordem_{coluna.id}"))
        if nova_ordem.isdigit():
            coluna.ordem = int(nova_ordem)

    for linha in tabela.linhas:
        novo_rotulo = texto(request.form.get(f"linha_rotulo_{linha.id}"))
        if novo_rotulo:
            linha.rotulo = novo_rotulo

        for coluna in tabela.colunas:
            chave = (linha.id, coluna.id)
            campo = f"celula_{linha.id}_{coluna.id}"
            valor = texto(request.form.get(campo))
            celula = existentes.get(chave)
            if celula:
                celula.valor = valor or None
                celula.atualizado_em = datetime.utcnow()
            elif valor:
                db.session.add(EvidenciaTabelaCelula(
                    linha_id=linha.id,
                    coluna_id=coluna.id,
                    valor=valor,
                ))

    tabela.atualizado_em = datetime.utcnow()
    db.session.commit()
    url = f"/gestao/evidencias/{tabela.campo_pai.registro_id}#tabelaPai{tabela.campo_pai_id}"
    return voltar_ou_json(url, "Tabela salva.", tabela_id=tabela.id, campo_pai_id=tabela.campo_pai_id)


@evidencias_frota_bp.route("/tabela/<int:tabela_id>/autosave", methods=["POST"])
def autosalvar_tabela(tabela_id):
    resp = exigir_login()
    if resp:
        return jsonify({"ok": False, "erro": "login"}), 401

    tabela = obter_tabela_ou_404(tabela_id)
    salvar_estado_tabela_do_request(tabela)
    db.session.commit()

    return jsonify({
        "ok": True,
        "mensagem": "Salvo automaticamente",
        "tabela_id": tabela.id,
        "atualizado_em": datetime.utcnow().strftime("%d/%m/%Y %H:%M:%S"),
    })


# =========================================================
# IMAGENS
# =========================================================

@evidencias_frota_bp.route("/pai/<int:pai_id>/imagens", methods=["POST"])
def upload_imagens_pai(pai_id):
    resp = exigir_login()
    if resp:
        return resp

    pai = obter_pai_ou_404(pai_id)
    registro = pai.registro
    arquivos = request.files.getlist("imagens")
    legenda_padrao = texto(request.form.get("legenda"))
    tipo_foto = texto(request.form.get("tipo_foto"))
    tipo_novo = texto(request.form.get("tipo_foto_novo"))

    if tipo_novo:
        tipo_foto = salvar_tipo_foto_se_novo(tipo_novo)

    if not tipo_foto:
        tipo_foto = "Outro"

    adicionadas = 0
    for arquivo in arquivos:
        if not arquivo or not arquivo.filename:
            continue

        dados = salvar_imagem_evidencia(arquivo, registro, pai, tipo_foto=tipo_foto)
        if not dados or not dados.get("url"):
            continue

        imagem = EvidenciaImagem(
            campo_pai_id=pai.id,
            campo_filho_id=None,
            tipo_foto=tipo_foto,
            imagem_url=dados["url"],
            public_id=dados.get("public_id"),
            caminho_local=dados.get("caminho_local"),
            nome_original=dados.get("nome_original"),
            legenda=legenda_padrao or None,
            ordem=len(imagens_do_pai(pai.id)) + adicionadas + 1,
        )

        db.session.add(imagem)
        adicionadas += 1

    db.session.commit()

    mensagem = f"{adicionadas} imagem(ns) enviada(s)." if adicionadas else "Nenhuma imagem válida foi enviada."
    if requisicao_ajax():
        return responder_ok(mensagem, registro_id=registro.id, campo_pai_id=pai.id)
    flash(mensagem, "success" if adicionadas else "warning")
    return redirect(f"/gestao/evidencias/{registro.id}")


@evidencias_frota_bp.route("/filho/<int:filho_id>/imagens", methods=["POST"])
def upload_imagens(filho_id):
    # Compatibilidade: quem ainda chamar a rota antiga terá a foto salva no pai.
    resp = exigir_login()
    if resp:
        return resp

    filho = obter_filho_ou_404(filho_id)
    pai = filho.campo_pai
    registro = pai.registro
    arquivos = request.files.getlist("imagens")
    legenda_padrao = texto(request.form.get("legenda"))
    tipo_foto = texto(request.form.get("tipo_foto")) or filho.nome or "Outro"

    adicionadas = 0
    for arquivo in arquivos:
        if not arquivo or not arquivo.filename:
            continue
        dados = salvar_imagem_evidencia(arquivo, registro, pai, tipo_foto=tipo_foto, filho=filho)
        if not dados or not dados.get("url"):
            continue
        imagem = EvidenciaImagem(
            campo_pai_id=pai.id,
            campo_filho_id=None,
            tipo_foto=tipo_foto,
            imagem_url=dados["url"],
            public_id=dados.get("public_id"),
            caminho_local=dados.get("caminho_local"),
            nome_original=dados.get("nome_original"),
            legenda=legenda_padrao or None,
            ordem=len(imagens_do_pai(pai.id)) + adicionadas + 1,
        )
        db.session.add(imagem)
        adicionadas += 1

    db.session.commit()
    mensagem = f"{adicionadas} imagem(ns) enviada(s)." if adicionadas else "Nenhuma imagem válida foi enviada."
    if requisicao_ajax():
        return responder_ok(mensagem, registro_id=registro.id, campo_pai_id=pai.id)
    flash(mensagem, "success" if adicionadas else "warning")
    return redirect(f"/gestao/evidencias/{registro.id}")


@evidencias_frota_bp.route("/imagem/<int:imagem_id>/editar", methods=["POST"])
def editar_imagem(imagem_id):
    resp = exigir_login()
    if resp:
        return resp

    imagem = obter_imagem_ou_404(imagem_id)
    registro_id = registro_id_da_imagem(imagem)

    tipo_foto = texto(request.form.get("tipo_foto"))
    tipo_novo = texto(request.form.get("tipo_foto_novo"))
    if tipo_novo:
        tipo_foto = salvar_tipo_foto_se_novo(tipo_novo)

    if tipo_foto:
        imagem.tipo_foto = tipo_foto

    imagem.legenda = texto(request.form.get("legenda")) or None
    db.session.commit()

    url = f"/gestao/evidencias/{registro_id}"
    return voltar_ou_json(url, "Imagem atualizada.", registro_id=registro_id, imagem_id=imagem.id)


@evidencias_frota_bp.route("/imagem/<int:imagem_id>/excluir", methods=["POST"])
def excluir_imagem(imagem_id):
    resp = exigir_login()
    if resp:
        return resp

    imagem = obter_imagem_ou_404(imagem_id)
    registro_id = registro_id_da_imagem(imagem)

    if imagem.public_id and cloudinary_configurado():
        try:
            cloudinary.uploader.destroy(imagem.public_id, resource_type="image")
        except Exception:
            pass

    if imagem.caminho_local and os.path.exists(imagem.caminho_local):
        try:
            os.remove(imagem.caminho_local)
        except Exception:
            pass

    db.session.delete(imagem)
    db.session.commit()

    url = f"/gestao/evidencias/{registro_id}"
    return voltar_ou_json(url, "Imagem excluída.", registro_id=registro_id, imagem_id=imagem_id)


# =========================================================
# LINK PÚBLICO DO CLIENTE
# =========================================================

def exigir_gestao_links():
    resp = exigir_login()
    if resp:
        return resp

    if not usuario_eh_admin_ou_gestao():
        abort(403)

    return None


def gerar_token_link_publico():
    while True:
        token = secrets.token_urlsafe(32)
        existente = EvidenciaLinkPublico.query.filter_by(token=token).first()
        if not existente:
            return token


def obter_link_publico_ou_404(token):
    link = EvidenciaLinkPublico.query.filter_by(token=token).first_or_404()

    if not link.esta_disponivel():
        abort(404)

    link.ultimo_acesso_em = datetime.utcnow()
    db.session.commit()
    return link


def imagens_do_registro(registro_id):
    registro = EvidenciaRegistro.query.get_or_404(int(registro_id))
    imagens = []
    for pai in registro.campos_pai:
        imagens.extend(imagens_do_pai(pai.id))
    return sorted(
        imagens,
        key=lambda img: (
            pai_da_imagem(img).ordem if pai_da_imagem(img) else 0,
            origem_imagem(img),
            img.ordem or 0,
            img.id or 0,
        )
    )


@evidencias_frota_bp.route("/<int:registro_id>/links/criar", methods=["POST"])
def criar_link_publico(registro_id):
    resp = exigir_gestao_links()
    if resp:
        return resp

    registro = obter_registro_ou_404(registro_id)

    validade = request.form.get("validade_dias", type=int)
    expira_em = None
    if validade and validade > 0:
        expira_em = datetime.utcnow() + timedelta(days=validade)

    link = EvidenciaLinkPublico(
        registro_id=registro.id,
        token=gerar_token_link_publico(),
        permitir_excel=True if request.form.get("permitir_excel") == "on" else False,
        permitir_zip=True if request.form.get("permitir_zip") == "on" else False,
        criado_por=session.get("user_id"),
        expira_em=expira_em,
        ativo=True,
    )

    db.session.add(link)
    db.session.commit()

    flash("Link público gerado com sucesso.", "success")
    return redirect(f"/gestao/evidencias/{registro.id}")


@evidencias_frota_bp.route("/links/<int:link_id>/desativar", methods=["POST"])
def desativar_link_publico(link_id):
    resp = exigir_gestao_links()
    if resp:
        return resp

    link = EvidenciaLinkPublico.query.get_or_404(link_id)
    obter_registro_ou_404(link.registro_id)

    link.ativo = False
    db.session.commit()

    flash("Link público desativado.", "success")
    return redirect(f"/gestao/evidencias/{link.registro_id}")


@evidencias_frota_bp.route("/publico/<token>")
def visualizacao_publica(token):
    link = obter_link_publico_ou_404(token)
    registro = preparar_registro_para_tela(link.registro)

    return render_template(
        "gestao/evidencias_frota/publico.html",
        link=link,
        registro=registro,
    )


@evidencias_frota_bp.route("/publico/<token>/excel")
def exportar_excel_publico(token):
    link = obter_link_publico_ou_404(token)

    if not link.permitir_excel:
        abort(403)

    return gerar_excel_registro_publico(link.registro_id)


@evidencias_frota_bp.route("/publico/<token>/imagens.zip")
def exportar_zip_publico(token):
    link = obter_link_publico_ou_404(token)

    if not link.permitir_zip:
        abort(403)

    imagens = imagens_do_registro(link.registro_id)

    memoria = BytesIO()
    with zipfile.ZipFile(memoria, "w", zipfile.ZIP_DEFLATED) as zf:
        if not imagens:
            zf.writestr("SEM_IMAGENS.txt", "Nenhuma imagem encontrada neste painel.")
        for img in imagens:
            adicionar_imagem_ao_zip(zf, img)

    memoria.seek(0)
    return send_file(
        memoria,
        as_attachment=True,
        download_name=f"evidencias_cliente_{link.registro_id}_imagens.zip",
        mimetype="application/zip",
    )


# =========================================================
# EXPORTAÇÕES
# =========================================================

def imagens_por_filtros(registro_id=None):
    query = EvidenciaImagem.query.outerjoin(
        EvidenciaCampoPai,
        EvidenciaImagem.campo_pai_id == EvidenciaCampoPai.id,
    ).outerjoin(
        EvidenciaCampoFilho,
        EvidenciaImagem.campo_filho_id == EvidenciaCampoFilho.id,
    )

    # Filtro de permissão é aplicado pelo registro vinculado ao campo pai novo ou ao filho antigo.
    if usuario_eh_admin_ou_gestao():
        pass
    else:
        usuario = usuario_logado()
        if not usuario or not getattr(usuario, "cliente", None):
            return []

    if registro_id:
        query = query.filter(or_(
            EvidenciaCampoPai.registro_id == int(registro_id),
            EvidenciaCampoFilho.campo_pai.has(EvidenciaCampoPai.registro_id == int(registro_id)),
        ))

    cliente_id = request.args.get("cliente_id", type=int)
    campo = texto(request.args.get("campo"))
    busca = texto(request.args.get("busca"))

    imagens = query.all()
    filtradas = []
    for img in imagens:
        pai = pai_da_imagem(img)
        registro = pai.registro if pai else None
        if not registro:
            continue
        if not cliente_permitido(registro.cliente_id, registro.cliente_nome):
            continue
        if cliente_id and registro.cliente_id != cliente_id:
            continue
        if campo and campo.upper() not in (pai.nome or "").upper():
            continue
        if busca:
            base = " ".join([
                registro.cliente_nome or "",
                registro.frota or "",
                registro.placa or "",
                pai.nome or "",
                origem_imagem(img),
                img.legenda or "",
            ]).upper()
            if busca.upper() not in base:
                continue
        filtradas.append(img)

    return sorted(
        filtradas,
        key=lambda img: (
            (registro_da_imagem(img).cliente_nome if registro_da_imagem(img) else ""),
            (registro_da_imagem(img).frota if registro_da_imagem(img) else "") or "",
            (registro_da_imagem(img).placa if registro_da_imagem(img) else "") or "",
            pai_da_imagem(img).ordem if pai_da_imagem(img) else 0,
            origem_imagem(img),
            img.ordem or 0,
            img.id or 0,
        )
    )


def ajustar_excel(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 45)


def estilizar_cabecalho(ws):
    fill = PatternFill("solid", fgColor="1f2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


@evidencias_frota_bp.route("/exportar/excel")
def exportar_excel():
    resp = exigir_login()
    if resp:
        return resp

    imagens = imagens_por_filtros()

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.append(["Cliente", "Frota", "Placa", "Campo pai", "Tipo da foto", "Qtd. imagens"])

    resumo = {}
    for img in imagens:
        pai = pai_da_imagem(img)
        registro = pai.registro if pai else None
        if not registro:
            continue
        chave = (
            registro.cliente_nome,
            registro.frota or "",
            registro.placa or "",
            pai.nome,
            origem_imagem(img),
        )
        resumo[chave] = resumo.get(chave, 0) + 1

    for chave, qtd in sorted(resumo.items()):
        ws_resumo.append(list(chave) + [qtd])
    estilizar_cabecalho(ws_resumo)
    ajustar_excel(ws_resumo)

    ws_imagens = wb.create_sheet("Imagens")
    ws_imagens.append([
        "Cliente", "Frota", "Placa", "Campo pai", "Tipo da foto",
        "Legenda", "Nome original", "URL/arquivo", "Enviado em",
    ])

    for img in imagens:
        pai = pai_da_imagem(img)
        registro = pai.registro if pai else None
        if not registro:
            continue
        ws_imagens.append([
            registro.cliente_nome,
            registro.frota or "",
            registro.placa or "",
            pai.nome,
            origem_imagem(img),
            img.legenda or "",
            img.nome_original or "",
            img.imagem_url,
            img.enviado_em.strftime("%d/%m/%Y %H:%M") if img.enviado_em else "",
        ])
    estilizar_cabecalho(ws_imagens)
    ajustar_excel(ws_imagens)

    nomes_usados = {"Resumo", "Imagens"}
    campos = {}
    for img in imagens:
        pai = pai_da_imagem(img)
        nome = pai.nome if pai else "SEM CAMPO"
        campos.setdefault(nome, []).append(img)

    for nome_campo, imgs in sorted(campos.items()):
        titulo = re.sub(r"[\\/*?:\[\]]", "-", nome_campo)[:28] or "Campo"
        base_titulo = titulo
        indice = 2
        while titulo in nomes_usados:
            titulo = f"{base_titulo[:25]} {indice}"
            indice += 1
        nomes_usados.add(titulo)

        ws = wb.create_sheet(titulo)
        ws.append(["Cliente", "Frota", "Placa", "Tipo da foto", "Imagem", "Legenda", "URL/arquivo"])
        for img in imgs:
            registro = registro_da_imagem(img)
            ws.append([
                registro.cliente_nome if registro else "",
                registro.frota if registro else "",
                registro.placa if registro else "",
                origem_imagem(img),
                img.nome_original or f"Imagem {img.id}",
                img.legenda or "",
                img.imagem_url,
            ])
        estilizar_cabecalho(ws)
        ajustar_excel(ws)

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)

    return send_file(
        saida,
        as_attachment=True,
        download_name=f"evidencias_frotas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@evidencias_frota_bp.route("/<int:registro_id>/exportar/excel")
def exportar_excel_registro(registro_id):
    return exportar_excel_por_registro(registro_id)


def nome_aba_seguro(nome, usados):
    titulo = re.sub(r"[\\/*?:\[\]]", "-", nome or "Tabela")[:28] or "Tabela"
    base = titulo
    indice = 2
    while titulo in usados:
        titulo = f"{base[:25]} {indice}"
        indice += 1
    usados.add(titulo)
    return titulo


def estilizar_tabela_controle(ws):
    fill = PatternFill("solid", fgColor="E5E7EB")
    font = Font(color="111827", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2)):
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def adicionar_aba_tabela_controle(wb, tabela, usados):
    if not tabela:
        return
    ws = wb.create_sheet(nome_aba_seguro(tabela.titulo or "Tabela", usados))

    grupos = montar_grupos_cabecalho(tabela)
    linha_grupos = []
    for grupo in grupos:
        linha_grupos.extend([grupo.get("titulo") or ""] * int(grupo.get("colspan") or 1))
    ws.append(linha_grupos)
    ws.append([coluna.nome for coluna in tabela.colunas])

    coluna_inicio = 1
    for grupo in grupos:
        colspan = int(grupo.get("colspan") or 1)
        if colspan > 1:
            ws.merge_cells(start_row=1, start_column=coluna_inicio, end_row=1, end_column=coluna_inicio + colspan - 1)
        coluna_inicio += colspan

    mapa = montar_mapa_celulas(tabela)
    for linha in tabela.linhas:
        row = []
        for idx, coluna in enumerate(tabela.colunas):
            valor = mapa.get(f"{linha.id}_{coluna.id}", "")
            if idx == 0 and not valor and linha.rotulo and not str(linha.rotulo).startswith("Linha "):
                valor = linha.rotulo
            row.append(valor)
        ws.append(row)

    estilizar_tabela_controle(ws)
    ajustar_excel(ws)


def montar_excel_registro(imagens, registro_ids_extra=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidências"
    ws.append(["Cliente", "Frota", "Placa", "Campo pai", "Tipo da foto", "Legenda", "Imagem"])

    for img in imagens:
        pai = pai_da_imagem(img)
        registro = pai.registro if pai else None
        if not registro:
            continue
        ws.append([
            registro.cliente_nome,
            registro.frota or "",
            registro.placa or "",
            pai.nome,
            origem_imagem(img),
            img.legenda or "",
            img.imagem_url,
        ])

    estilizar_cabecalho(ws)
    ajustar_excel(ws)

    # Abas das tabelas de controle vinculadas aos campos pai deste registro.
    usados = {ws.title}
    registro_ids = set(registro_ids_extra or [])
    for img in imagens:
        registro = registro_da_imagem(img)
        if registro:
            registro_ids.add(registro.id)
    for registro_id in sorted(registro_ids):
        registro = EvidenciaRegistro.query.get(registro_id)
        if not registro:
            continue
        for pai in registro.campos_pai:
            if getattr(pai, "tabela_controle", None):
                adicionar_aba_tabela_controle(wb, pai.tabela_controle, usados)

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


def exportar_excel_por_registro(registro_id):
    obter_registro_ou_404(registro_id)
    imagens = imagens_por_filtros(registro_id=registro_id)
    saida = montar_excel_registro(imagens, registro_ids_extra=[int(registro_id)])

    return send_file(
        saida,
        as_attachment=True,
        download_name=f"evidencias_registro_{registro_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def gerar_excel_registro_publico(registro_id):
    imagens = imagens_do_registro(registro_id)
    saida = montar_excel_registro(imagens, registro_ids_extra=[int(registro_id)])

    return send_file(
        saida,
        as_attachment=True,
        download_name=f"evidencias_cliente_{registro_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def caminho_local_por_url_estatica(url):
    url = texto(url)
    if not url.startswith("/static/"):
        return None
    caminho_relativo = url.lstrip("/").replace("/", os.sep)
    caminho_absoluto = os.path.join(current_app.root_path, caminho_relativo)
    return caminho_absoluto if os.path.exists(caminho_absoluto) else None


def adicionar_imagem_ao_zip(zf, img):
    """
    V7.2: ZIP sem pasta de campo filho.

    Estrutura final:
    CLIENTE/FROTA_OU_PLACA/CAMPO_PAI/[TIPO] imagem.jpg

    O tipo da foto fica no nome do arquivo, não vira uma pasta.
    Assim o ZIP não cria mais subpastas de filho dentro do pai.
    """
    pai = pai_da_imagem(img)
    registro = pai.registro if pai else None

    extensao = os.path.splitext(img.nome_original or "imagem.jpg")[1] or ".jpg"
    nome_base = limpar_nome_arquivo(img.nome_original or f"imagem_{img.id}{extensao}")
    tipo = limpar_nome_arquivo(origem_imagem(img) or "SEM_TIPO")

    nome_final = f"{img.id}_{tipo}_{nome_base}"

    caminho_zip = "/".join([
        limpar_nome_arquivo(registro.cliente_nome if registro else "SEM_CLIENTE"),
        limpar_nome_arquivo((registro.frota or registro.placa) if registro else "VEICULO"),
        limpar_nome_arquivo(pai.nome if pai else "SEM_CAMPO"),
        nome_final,
    ])

    if img.caminho_local and os.path.exists(img.caminho_local):
        zf.write(img.caminho_local, caminho_zip)
        return True

    caminho_estatico = caminho_local_por_url_estatica(img.imagem_url)
    if caminho_estatico:
        zf.write(caminho_estatico, caminho_zip)
        return True

    if img.imagem_url and str(img.imagem_url).startswith(("http://", "https://")):
        try:
            req = urllib.request.Request(img.imagem_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                zf.writestr(caminho_zip, resp.read())
                return True
        except Exception as e:
            zf.writestr(
                caminho_zip + ".txt",
                "Não foi possível baixar a imagem automaticamente.\n"
                f"URL: {img.imagem_url}\n"
                f"Erro: {str(e)}"
            )
            return False

    zf.writestr(caminho_zip + ".txt", f"Imagem sem caminho válido. ID: {getattr(img, 'id', '-')}." )
    return False


@evidencias_frota_bp.route("/exportar/imagens.zip")
def exportar_zip():
    resp = exigir_login()
    if resp:
        return resp

    imagens = imagens_por_filtros()

    memoria = BytesIO()
    with zipfile.ZipFile(memoria, "w", zipfile.ZIP_DEFLATED) as zf:
        if not imagens:
            zf.writestr("SEM_IMAGENS.txt", "Nenhuma imagem encontrada para os filtros informados.")
        for img in imagens:
            adicionar_imagem_ao_zip(zf, img)

    memoria.seek(0)
    return send_file(
        memoria,
        as_attachment=True,
        download_name=f"evidencias_imagens_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
        mimetype="application/zip",
    )


@evidencias_frota_bp.route("/<int:registro_id>/exportar/imagens.zip")
def exportar_zip_registro(registro_id):
    resp = exigir_login()
    if resp:
        return resp

    obter_registro_ou_404(registro_id)
    imagens = imagens_por_filtros(registro_id=registro_id)

    memoria = BytesIO()
    with zipfile.ZipFile(memoria, "w", zipfile.ZIP_DEFLATED) as zf:
        if not imagens:
            zf.writestr("SEM_IMAGENS.txt", "Nenhuma imagem encontrada neste painel.")
        for img in imagens:
            adicionar_imagem_ao_zip(zf, img)

    memoria.seek(0)
    return send_file(
        memoria,
        as_attachment=True,
        download_name=f"evidencias_registro_{registro_id}_imagens.zip",
        mimetype="application/zip",
    )
