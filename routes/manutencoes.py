from flask import Blueprint, render_template, request, redirect, session, flash, jsonify, send_file
from models.manutencao import Manutencao
from models.cliente import Cliente
from models.afericao_termometro import AfericaoTermometro
from models.causa_manutencao import CausaManutencao
from models.problema_manutencao import ProblemaManutencao
from models.usuario import Usuario
from database import db
from datetime import datetime
import json
import os
import cloudinary
import cloudinary.uploader
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO


manutencao_bp = Blueprint("manutencao", __name__, url_prefix="/manutencoes")


# ==========================================
# 🔐 HELPERS MULTI-CLIENTE
# ==========================================
def usuario_logado():
    user_id = session.get("user_id")

    if not user_id:
        return None

    return Usuario.query.get(user_id)


def usuario_eh_admin_ou_gestao():
    return session.get("user_role") in ["admin", "gestao", "gestor"]


def nome_cliente_usuario_logado():
    usuario = usuario_logado()

    if not usuario:
        return None

    if not usuario.cliente:
        return None

    return normalizar_texto(usuario.cliente.nome)


def aplicar_filtro_cliente(query):
    """
    Admin/Gestão/Gestor visualizam tudo.
    Usuário cliente visualiza somente manutenções do cliente vinculado no cadastro do usuário.

    IMPORTANTE:
    As manutenções continuam usando o campo Manutencao.cliente já existente.
    O vínculo profissional fica em Usuario.cliente_id -> Cliente.nome.
    """
    if usuario_eh_admin_ou_gestao():
        return query

    cliente_nome = nome_cliente_usuario_logado()

    if not cliente_nome:
        return query.filter(Manutencao.id == 0)

    return query.filter(
        db.func.upper(Manutencao.cliente) == cliente_nome.upper()
    )


def manutencao_pertence_ao_usuario(manutencao):
    if usuario_eh_admin_ou_gestao():
        return True

    cliente_nome = nome_cliente_usuario_logado()

    if not cliente_nome:
        return False

    return normalizar_texto(manutencao.cliente) == cliente_nome


# ==========================================
# 🔥 PERMISSÃO OPERACIONAL
# ==========================================
def usuario_tem_permissao_operacional():
    return session.get("user_role") in ["admin", "gestao", "gestor"]


# ==========================================
# 🔥 HELPERS CLOUDINARY
# ==========================================
def cloudinary_esta_configurado():
    return all([
        os.getenv("CLOUD_NAME"),
        os.getenv("API_KEY"),
        os.getenv("API_SECRET"),
    ])


# ==========================================
# 🔥 HELPERS IMAGENS
# ==========================================
def salvar_imagens():
    arquivos = request.files.getlist("imagens")
    return salvar_imagens_arquivos(arquivos)


def salvar_imagens_arquivos(arquivos):
    caminhos = []

    if not cloudinary_esta_configurado():
        print("⚠️ Cloudinary não configurado. Upload ignorado no ambiente atual.")
        return caminhos

    for arquivo in arquivos:
        if arquivo and arquivo.filename:
            try:
                upload = cloudinary.uploader.upload(arquivo)
                caminhos.append(upload["secure_url"])
            except Exception as e:
                print("❌ Erro ao enviar imagem para o Cloudinary:", e)

    return caminhos


def carregar_lista_imagens(valor):
    try:
        lista = json.loads(valor) if valor else []
        return lista if isinstance(lista, list) else []
    except Exception:
        return []


def extrair_public_id_cloudinary(url):
    try:
        path = urlparse(url).path
        partes = path.split("/upload/")
        if len(partes) < 2:
            return None

        resto = partes[1]

        if resto.startswith("v") and "/" in resto:
            primeira, restante = resto.split("/", 1)
            if primeira[1:].isdigit():
                resto = restante

        if "." in resto:
            resto = resto.rsplit(".", 1)[0]

        return resto
    except Exception:
        return None


# ==========================================
# 🔥 HELPERS DADOS
# ==========================================
def parse_data(valor):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date() if valor else None
    except Exception:
        return None


def normalizar_texto(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    if not valor:
        return None

    return " ".join(valor.split()).upper()


def normalizar_simples(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    if not valor:
        return None

    return " ".join(valor.split())




def identificador_veiculo(numero_frota=None, placa=None):
    """Usa a frota quando existir; se não existir, usa a placa como identificador do veículo."""
    frota = normalizar_simples(numero_frota)
    placa_norm = normalizar_texto(placa)

    if frota:
        return frota

    return placa_norm

def calcular_dtm(data_entrada, data_saida):
    if not data_entrada or not data_saida:
        return None

    if data_saida < data_entrada:
        raise ValueError("A data de saída não pode ser menor que a data de entrada.")

    dias = (data_saida - data_entrada).days

    return max(dias, 1)


# ==========================================
# 🔥 HELPERS CAUSAS / PROBLEMAS
# ==========================================
def listar_causas_json():
    causas = CausaManutencao.query.filter_by(ativo=True).order_by(CausaManutencao.nome).all()

    return [
        {
            "id": c.id,
            "nome": c.nome,
            "problemas": [
                {
                    "id": p.id,
                    "nome": p.nome
                }
                for p in sorted(
                    [p for p in c.problemas if p.ativo],
                    key=lambda item: item.nome or ""
                )
            ]
        }
        for c in causas
    ]


# ==========================================
# 🔥 HELPERS AFERIÇÃO TERMÔMETRO
# ==========================================
def carregar_afericao(numero_frota, os, tipo_termometro):
    if not numero_frota or not os:
        return {
            "id": None,
            "afericao": "",
            "data_afericao": "",
            "status": "",
            "imagens": []
        }

    reg = AfericaoTermometro.query.filter_by(
        numero_frota=str(numero_frota).strip(),
        os=str(os).strip(),
        tipo_termometro=tipo_termometro
    ).first()

    if not reg:
        return {
            "id": None,
            "afericao": "",
            "data_afericao": "",
            "status": "",
            "imagens": []
        }

    return {
        "id": reg.id,
        "afericao": reg.afericao or "",
        "data_afericao": reg.data_afericao.strftime("%Y-%m-%d") if reg.data_afericao else "",
        "status": reg.status or "",
        "imagens": carregar_lista_imagens(reg.imagens)
    }


def salvar_ou_atualizar_afericao(
    numero_frota,
    os,
    tipo_termometro,
    afericao,
    data_afericao,
    status,
    novas_imagens=None
):
    if not numero_frota or not os:
        return

    numero_frota = str(numero_frota).strip()
    os = str(os).strip()
    novas_imagens = novas_imagens or []

    reg = AfericaoTermometro.query.filter_by(
        numero_frota=numero_frota,
        os=os,
        tipo_termometro=tipo_termometro
    ).first()

    imagens_atuais = carregar_lista_imagens(reg.imagens) if reg else []

    afericao = normalizar_texto(afericao)
    status = normalizar_texto(status)

    if not (
        str(afericao or "").strip()
        or data_afericao
        or str(status or "").strip()
        or imagens_atuais
        or novas_imagens
    ):
        if reg:
            db.session.delete(reg)
        return

    total_imagens = len(imagens_atuais) + len(novas_imagens)
    if total_imagens > 4:
        raise ValueError(
            f"O termômetro de {tipo_termometro.lower()} permite no máximo 4 imagens."
        )

    if not reg:
        reg = AfericaoTermometro(
            numero_frota=numero_frota,
            os=os,
            tipo_termometro=tipo_termometro
        )
        db.session.add(reg)

    reg.numero_frota = numero_frota
    reg.os = os
    reg.tipo_termometro = tipo_termometro
    reg.afericao = afericao
    reg.data_afericao = data_afericao
    reg.status = status
    reg.imagens = json.dumps(imagens_atuais + novas_imagens)


def mover_afericoes_se_trocar_frota_ou_os(numero_frota_antiga, os_antiga, numero_frota_nova, os_nova):
    if not numero_frota_antiga or not os_antiga:
        return

    regs = AfericaoTermometro.query.filter_by(
        numero_frota=str(numero_frota_antiga).strip(),
        os=str(os_antiga).strip()
    ).all()

    for reg in regs:
        reg.numero_frota = str(numero_frota_nova).strip() if numero_frota_nova else reg.numero_frota
        reg.os = str(os_nova).strip() if os_nova else reg.os


# ==========================================
# 🔄 MIGRAR CAUSAS/PROBLEMAS ANTIGOS
# ==========================================
@manutencao_bp.route("/admin/migrar-causas-problemas", methods=["GET", "POST"])
def migrar_causas_problemas():
    if not session.get("user_id"):
        return redirect("/login")

    if not usuario_tem_permissao_operacional():
        return jsonify({
            "ok": False,
            "message": "Sem permissão."
        }), 403

    registros = Manutencao.query.all()

    causas_criadas = 0
    problemas_criados = 0
    vinculos_existentes = 0
    ignorados = 0

    try:
        for m in registros:
            causa_nome = normalizar_texto(m.causa)
            problema_nome = normalizar_texto(m.problema)

            if not causa_nome or not problema_nome:
                ignorados += 1
                continue

            causa = CausaManutencao.query.filter(
                db.func.upper(CausaManutencao.nome) == causa_nome.upper()
            ).first()

            if not causa:
                causa = CausaManutencao(
                    nome=causa_nome,
                    ativo=True
                )
                db.session.add(causa)
                db.session.flush()
                causas_criadas += 1
            else:
                causa.ativo = True

            problema = ProblemaManutencao.query.filter(
                ProblemaManutencao.causa_id == causa.id,
                db.func.upper(ProblemaManutencao.nome) == problema_nome.upper()
            ).first()

            if not problema:
                problema = ProblemaManutencao(
                    causa_id=causa.id,
                    nome=problema_nome,
                    ativo=True
                )
                db.session.add(problema)
                problemas_criados += 1
            else:
                problema.ativo = True
                vinculos_existentes += 1

        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "Migração concluída com sucesso.",
            "total_manutencoes_lidas": len(registros),
            "causas_criadas": causas_criadas,
            "problemas_criados": problemas_criados,
            "vinculos_existentes": vinculos_existentes,
            "ignorados_sem_causa_ou_problema": ignorados
        })

    except Exception as e:
        db.session.rollback()

        return jsonify({
            "ok": False,
            "message": f"Erro ao migrar causas e problemas: {str(e)}"
        }), 500


# ==========================================
# API CAUSAS / PROBLEMAS
# ==========================================
@manutencao_bp.route("/api/causas", methods=["GET"])
def api_listar_causas():
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    return jsonify({
        "ok": True,
        "causas": listar_causas_json()
    })


@manutencao_bp.route("/api/causas", methods=["POST"])
def api_criar_causa():
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    nome = normalizar_texto(request.form.get("nome"))

    if not nome:
        return jsonify({"ok": False, "message": "Informe o nome da causa."}), 400

    existente = CausaManutencao.query.filter(
        db.func.upper(CausaManutencao.nome) == nome.upper()
    ).first()

    if existente:
        existente.ativo = True
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "Causa já existia e foi ativada.",
            "causa": {
                "id": existente.id,
                "nome": existente.nome
            }
        })

    causa = CausaManutencao(nome=nome, ativo=True)
    db.session.add(causa)
    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Causa criada com sucesso.",
        "causa": {
            "id": causa.id,
            "nome": causa.nome
        }
    })


@manutencao_bp.route("/api/causas/<int:causa_id>", methods=["POST"])
def api_editar_causa(causa_id):
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    causa = CausaManutencao.query.get_or_404(causa_id)
    nome = normalizar_texto(request.form.get("nome"))

    if not nome:
        return jsonify({"ok": False, "message": "Informe o nome da causa."}), 400

    causa.nome = nome
    causa.ativo = True
    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Causa atualizada com sucesso.",
        "causa": {
            "id": causa.id,
            "nome": causa.nome
        }
    })


@manutencao_bp.route("/api/problemas", methods=["POST"])
def api_criar_problema():
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    causa_id = request.form.get("causa_id")
    nome = normalizar_texto(request.form.get("nome"))

    if not causa_id:
        return jsonify({"ok": False, "message": "Selecione uma causa."}), 400

    if not nome:
        return jsonify({"ok": False, "message": "Informe o nome do problema."}), 400

    causa = CausaManutencao.query.get(causa_id)

    if not causa:
        return jsonify({"ok": False, "message": "Causa não encontrada."}), 404

    existente = ProblemaManutencao.query.filter(
        ProblemaManutencao.causa_id == causa.id,
        db.func.upper(ProblemaManutencao.nome) == nome.upper()
    ).first()

    if existente:
        existente.ativo = True
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "Problema já existia e foi ativado.",
            "problema": {
                "id": existente.id,
                "nome": existente.nome,
                "causa_id": causa.id
            }
        })

    problema = ProblemaManutencao(
        causa_id=causa.id,
        nome=nome,
        ativo=True
    )
    db.session.add(problema)
    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Problema criado com sucesso.",
        "problema": {
            "id": problema.id,
            "nome": problema.nome,
            "causa_id": causa.id
        }
    })


@manutencao_bp.route("/api/problemas/<int:problema_id>", methods=["POST"])
def api_editar_problema(problema_id):
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    problema = ProblemaManutencao.query.get_or_404(problema_id)

    causa_id = request.form.get("causa_id")
    nome = normalizar_texto(request.form.get("nome"))

    if not causa_id:
        return jsonify({"ok": False, "message": "Selecione uma causa."}), 400

    if not nome:
        return jsonify({"ok": False, "message": "Informe o nome do problema."}), 400

    causa = CausaManutencao.query.get(causa_id)

    if not causa:
        return jsonify({"ok": False, "message": "Causa não encontrada."}), 404

    problema.causa_id = causa.id
    problema.nome = nome
    problema.ativo = True
    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Problema atualizado com sucesso.",
        "problema": {
            "id": problema.id,
            "nome": problema.nome,
            "causa_id": causa.id
        }
    })


# ==========================================
# 🖼️ EXCLUIR IMAGEM MANUTENÇÃO
# ==========================================
@manutencao_bp.route("/<int:id>/excluir-imagem", methods=["POST"])
def excluir_imagem(id):
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    m = Manutencao.query.get_or_404(id)

    if not manutencao_pertence_ao_usuario(m):
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    imagem_url = request.form.get("imagem_url")
    if not imagem_url:
        return jsonify({"ok": False, "message": "Imagem não informada."}), 400

    imagens_atuais = carregar_lista_imagens(m.imagens)

    if imagem_url not in imagens_atuais:
        return jsonify({"ok": False, "message": "Imagem não encontrada nesse registro."}), 404

    imagens_atuais.remove(imagem_url)
    m.imagens = json.dumps(imagens_atuais)

    public_id = extrair_public_id_cloudinary(imagem_url)
    if public_id and cloudinary_esta_configurado():
        try:
            cloudinary.uploader.destroy(public_id)
        except Exception as e:
            print("Erro ao apagar imagem no Cloudinary:", e)

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Imagem excluída com sucesso!",
        "restantes": len(imagens_atuais)
    })


# ==========================================
# 🖼️ EXCLUIR IMAGEM AFERIÇÃO
# ==========================================
@manutencao_bp.route("/afericao/<int:afericao_id>/excluir-imagem", methods=["POST"])
def excluir_imagem_afericao(afericao_id):
    if not session.get("user_id"):
        return jsonify({"ok": False, "message": "Sessão expirada."}), 401

    if not usuario_tem_permissao_operacional():
        return jsonify({"ok": False, "message": "Sem permissão."}), 403

    afericao = AfericaoTermometro.query.get_or_404(afericao_id)

    imagem_url = request.form.get("imagem_url")
    if not imagem_url:
        return jsonify({"ok": False, "message": "Imagem não informada."}), 400

    imagens_atuais = carregar_lista_imagens(afericao.imagens)

    if imagem_url not in imagens_atuais:
        return jsonify({"ok": False, "message": "Imagem não encontrada nessa aferição."}), 404

    imagens_atuais.remove(imagem_url)
    afericao.imagens = json.dumps(imagens_atuais)

    public_id = extrair_public_id_cloudinary(imagem_url)
    if public_id and cloudinary_esta_configurado():
        try:
            cloudinary.uploader.destroy(public_id)
        except Exception as e:
            print("Erro ao apagar imagem da aferição:", e)

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Imagem da aferição excluída com sucesso!",
        "restantes": len(imagens_atuais)
    })


# ==========================================
# ➕ NOVA MANUTENÇÃO
# ==========================================
@manutencao_bp.route("/", methods=["GET", "POST"])
def nova():
    if not session.get("user_id"):
        return redirect("/login")

    if not usuario_tem_permissao_operacional():
        return redirect("/")

    if request.method == "POST":
        data_convertida = parse_data(request.form.get("data"))
        data_saida_convertida = parse_data(request.form.get("data_saida"))
        dtm_calculado = calcular_dtm(data_convertida, data_saida_convertida)
        caminhos_imagens = salvar_imagens()

        numero_frota = normalizar_simples(request.form.get("numero_frota"))
        placa = normalizar_texto(request.form.get("placa"))
        os_numero = normalizar_simples(request.form.get("os"))
        identificador_afericao = identificador_veiculo(numero_frota, placa)

        placa_novas_imagens = salvar_imagens_arquivos(request.files.getlist("placa_imagens"))
        ambiente_novas_imagens = salvar_imagens_arquivos(request.files.getlist("ambiente_imagens"))

        nova_manutencao = Manutencao(
            data=data_convertida,
            data_saida=data_saida_convertida,
            dtm=dtm_calculado,
            numero_frota=numero_frota,
            placa=placa,
            bau=normalizar_texto(request.form.get("bau")),
            tipo_veiculo=normalizar_texto(request.form.get("tipo_veiculo")),
            tipo_servico=normalizar_texto(request.form.get("tipo_servico")),
            tipo_atendimento=normalizar_texto(request.form.get("tipo_atendimento")),
            tipo_manutencao=normalizar_texto(request.form.get("tipo_manutencao")),
            status=normalizar_texto(request.form.get("status")),
            observacao=normalizar_texto(request.form.get("observacao")),
            cliente=normalizar_texto(request.form.get("cliente")),
            os=os_numero,
            causa=normalizar_texto(request.form.get("causa")),
            problema=normalizar_texto(request.form.get("problema")),
            imagens=json.dumps(caminhos_imagens)
        )

        try:
            db.session.add(nova_manutencao)

            salvar_ou_atualizar_afericao(
                numero_frota=identificador_afericao,
                os=os_numero,
                tipo_termometro="PLACA",
                afericao=request.form.get("placa_afericao"),
                data_afericao=parse_data(request.form.get("placa_data_afericao")),
                status=request.form.get("placa_status"),
                novas_imagens=placa_novas_imagens
            )

            salvar_ou_atualizar_afericao(
                numero_frota=identificador_afericao,
                os=os_numero,
                tipo_termometro="AMBIENTE",
                afericao=request.form.get("ambiente_afericao"),
                data_afericao=parse_data(request.form.get("ambiente_data_afericao")),
                status=request.form.get("ambiente_status"),
                novas_imagens=ambiente_novas_imagens
            )

            db.session.commit()

            flash("Manutenção salva com sucesso!", "success")
            return redirect("/manutencoes/lista")

        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")
            return redirect(request.url)

    clientes = Cliente.query.order_by(Cliente.nome).all()
    causas = CausaManutencao.query.filter_by(ativo=True).order_by(CausaManutencao.nome).all()

    return render_template(
        "manutencoes/form.html",
        clientes=clientes,
        causas=causas,
        causas_json=listar_causas_json(),
        m=None,
        imagens_lista=[],
        afericao_placa={"id": None, "afericao": "", "data_afericao": "", "status": "", "imagens": []},
        afericao_ambiente={"id": None, "afericao": "", "data_afericao": "", "status": "", "imagens": []}
    )


# ==========================================
# 📋 LISTA
# ==========================================
@manutencao_bp.route("/lista")
def lista():
    if not session.get("user_id"):
        return redirect("/login")

    filtro = request.args.get("filtro")
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    query = aplicar_filtro_cliente(Manutencao.query)

    if filtro == "andamento":
        query = query.filter(Manutencao.status.ilike("%ANDAMENTO%"))
    elif filtro == "corretiva":
        query = query.filter(Manutencao.tipo_servico.ilike("%CORRETIVA%"))
    elif filtro == "preventiva":
        query = query.filter(Manutencao.tipo_servico.ilike("%PREVENTIVA%"))

    if data_inicio and data_fim:
        try:
            inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
            fim = datetime.strptime(data_fim, "%Y-%m-%d")
            query = query.filter(Manutencao.data.between(inicio, fim))
        except Exception:
            pass

    registros = query.order_by(
        Manutencao.data.desc().nullslast()
    ).all()

    for r in registros:
        identificador_afericao = identificador_veiculo(r.numero_frota, getattr(r, "placa", None))

        afericao_placa = AfericaoTermometro.query.filter_by(
            numero_frota=str(identificador_afericao).strip() if identificador_afericao else "",
            os=str(r.os).strip() if r.os else "",
            tipo_termometro="PLACA"
        ).first()

        afericao_ambiente = AfericaoTermometro.query.filter_by(
            numero_frota=str(identificador_afericao).strip() if identificador_afericao else "",
            os=str(r.os).strip() if r.os else "",
            tipo_termometro="AMBIENTE"
        ).first()

        r.placa_afericao = afericao_placa.afericao if afericao_placa else None
        r.placa_data_afericao = (
            afericao_placa.data_afericao.strftime("%d/%m/%Y")
            if afericao_placa and afericao_placa.data_afericao else None
        )
        r.placa_status = afericao_placa.status if afericao_placa else None
        r.placa_imagens = carregar_lista_imagens(afericao_placa.imagens) if afericao_placa else []

        r.ambiente_afericao = afericao_ambiente.afericao if afericao_ambiente else None
        r.ambiente_data_afericao = (
            afericao_ambiente.data_afericao.strftime("%d/%m/%Y")
            if afericao_ambiente and afericao_ambiente.data_afericao else None
        )
        r.ambiente_status = afericao_ambiente.status if afericao_ambiente else None
        r.ambiente_imagens = carregar_lista_imagens(afericao_ambiente.imagens) if afericao_ambiente else []

        r.imagens_lista = carregar_lista_imagens(r.imagens)

    return render_template(
        "manutencoes/lista.html",
        registros=registros
    )


# ==========================================
# ✏️ EDITAR
# ==========================================
@manutencao_bp.route("/editar/<int:id>", methods=["GET", "POST"])
def editar(id):
    if not session.get("user_id"):
        return redirect("/login")

    if not usuario_tem_permissao_operacional():
        return redirect("/")

    m = Manutencao.query.get_or_404(id)

    if not manutencao_pertence_ao_usuario(m):
        flash("Você não tem permissão para acessar esta manutenção.", "danger")
        return redirect("/manutencoes/lista")

    if request.method == "POST":
        numero_frota_antiga = identificador_veiculo(m.numero_frota, getattr(m, "placa", None))
        os_antiga = m.os

        m.data = parse_data(request.form.get("data"))
        m.data_saida = parse_data(request.form.get("data_saida"))
        m.dtm = calcular_dtm(m.data, m.data_saida)

        novas_imagens = salvar_imagens()

        if novas_imagens:
            imagens_atuais = carregar_lista_imagens(m.imagens)
            m.imagens = json.dumps(imagens_atuais + novas_imagens)

        m.numero_frota = normalizar_simples(request.form.get("numero_frota"))
        m.placa = normalizar_texto(request.form.get("placa"))
        identificador_afericao_novo = identificador_veiculo(m.numero_frota, m.placa)
        m.bau = normalizar_texto(request.form.get("bau"))
        m.tipo_veiculo = normalizar_texto(request.form.get("tipo_veiculo"))
        m.tipo_servico = normalizar_texto(request.form.get("tipo_servico"))
        m.tipo_atendimento = normalizar_texto(request.form.get("tipo_atendimento"))
        m.tipo_manutencao = normalizar_texto(request.form.get("tipo_manutencao"))
        m.status = normalizar_texto(request.form.get("status"))
        m.observacao = normalizar_texto(request.form.get("observacao"))
        m.cliente = normalizar_texto(request.form.get("cliente"))
        m.os = normalizar_simples(request.form.get("os"))
        m.causa = normalizar_texto(request.form.get("causa"))
        m.problema = normalizar_texto(request.form.get("problema"))

        placa_novas_imagens = salvar_imagens_arquivos(request.files.getlist("placa_imagens"))
        ambiente_novas_imagens = salvar_imagens_arquivos(request.files.getlist("ambiente_imagens"))

        try:
            mover_afericoes_se_trocar_frota_ou_os(
                numero_frota_antiga=numero_frota_antiga,
                os_antiga=os_antiga,
                numero_frota_nova=identificador_afericao_novo,
                os_nova=m.os
            )

            salvar_ou_atualizar_afericao(
                numero_frota=identificador_afericao_novo,
                os=m.os,
                tipo_termometro="PLACA",
                afericao=request.form.get("placa_afericao"),
                data_afericao=parse_data(request.form.get("placa_data_afericao")),
                status=request.form.get("placa_status"),
                novas_imagens=placa_novas_imagens
            )

            salvar_ou_atualizar_afericao(
                numero_frota=identificador_afericao_novo,
                os=m.os,
                tipo_termometro="AMBIENTE",
                afericao=request.form.get("ambiente_afericao"),
                data_afericao=parse_data(request.form.get("ambiente_data_afericao")),
                status=request.form.get("ambiente_status"),
                novas_imagens=ambiente_novas_imagens
            )

            db.session.commit()

            flash("Manutenção atualizada com sucesso!", "success")
            return redirect("/frotas/" + str(identificador_veiculo(m.numero_frota, m.placa)))

        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")
            return redirect(request.url)

    clientes = Cliente.query.order_by(Cliente.nome).all()
    imagens_lista = carregar_lista_imagens(m.imagens)
    causas = CausaManutencao.query.filter_by(ativo=True).order_by(CausaManutencao.nome).all()

    identificador_afericao = identificador_veiculo(m.numero_frota, getattr(m, "placa", None))
    afericao_placa = carregar_afericao(identificador_afericao, m.os, "PLACA")
    afericao_ambiente = carregar_afericao(identificador_afericao, m.os, "AMBIENTE")

    return render_template(
        "manutencoes/form.html",
        m=m,
        clientes=clientes,
        causas=causas,
        causas_json=listar_causas_json(),
        imagens_lista=imagens_lista,
        afericao_placa=afericao_placa,
        afericao_ambiente=afericao_ambiente
    )


# ==========================================
# 🗑️ EXCLUIR
# ==========================================
@manutencao_bp.route("/excluir/<int:id>")
def excluir(id):
    if not session.get("user_id"):
        return redirect("/login")

    if not usuario_tem_permissao_operacional():
        return redirect("/")

    m = Manutencao.query.get_or_404(id)

    if not manutencao_pertence_ao_usuario(m):
        flash("Você não tem permissão para excluir esta manutenção.", "danger")
        return redirect("/manutencoes/lista")

    identificador_afericao = identificador_veiculo(m.numero_frota, getattr(m, "placa", None))

    if identificador_afericao and m.os:
        AfericaoTermometro.query.filter_by(
            numero_frota=str(identificador_afericao).strip(),
            os=str(m.os).strip()
        ).delete()

    db.session.delete(m)
    db.session.commit()

    flash("Manutenção excluída com sucesso!", "success")
    return redirect("/manutencoes/lista")


# ==========================================
# 📊 EXPORTAR EXCEL
# ==========================================
@manutencao_bp.route("/exportar-excel")
def exportar_excel():
    if not session.get("user_id"):
        return redirect("/login")

    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    query = aplicar_filtro_cliente(Manutencao.query)

    if data_inicio and data_fim:
        try:
            inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
            fim = datetime.strptime(data_fim, "%Y-%m-%d")

            query = query.filter(
                Manutencao.data.between(inicio, fim)
            )

        except Exception:
            pass

    registros = query.order_by(
        Manutencao.data.desc().nullslast()
    ).all()

    wb = Workbook()
    ws = wb.active

    ws.title = "Manutenções"

    headers = [
        "DATA ENTRADA",
        "DATA SAÍDA",
        "ATM",
        "FROTA",
        "PLACA",
        "OS",
        "BAÚ",
        "TIPO VEÍCULO",
        "TIPO SERVIÇO",
        "TIPO ATENDIMENTO",
        "TIPO MANUTENÇÃO",
        "STATUS",
        "CLIENTE",
        "CAUSA",
        "PROBLEMA",
        "PLACA AFERIÇÃO",
        "PLACA STATUS",
        "AMBIENTE AFERIÇÃO",
        "AMBIENTE STATUS",
        "OBSERVAÇÃO",
    ]

    ws.append(headers)

    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for r in registros:
        identificador_afericao = identificador_veiculo(r.numero_frota, getattr(r, "placa", None))

        afericao_placa = AfericaoTermometro.query.filter_by(
            numero_frota=str(identificador_afericao).strip() if identificador_afericao else "",
            os=str(r.os).strip() if r.os else "",
            tipo_termometro="PLACA"
        ).first()

        afericao_ambiente = AfericaoTermometro.query.filter_by(
            numero_frota=str(identificador_afericao).strip() if identificador_afericao else "",
            os=str(r.os).strip() if r.os else "",
            tipo_termometro="AMBIENTE"
        ).first()

        ws.append([
            r.data.strftime("%d/%m/%Y") if r.data else "",
            r.data_saida.strftime("%d/%m/%Y") if r.data_saida else "",
            r.dtm if r.dtm is not None else "",
            r.numero_frota or "",
            getattr(r, "placa", None) or "",
            r.os or "",
            r.bau or "",
            r.tipo_veiculo or "",
            r.tipo_servico or "",
            r.tipo_atendimento or "",
            r.tipo_manutencao or "",
            r.status or "",
            r.cliente or "",
            r.causa or "",
            r.problema or "",
            afericao_placa.afericao if afericao_placa else "",
            afericao_placa.status if afericao_placa else "",
            afericao_ambiente.afericao if afericao_ambiente else "",
            afericao_ambiente.status if afericao_ambiente else "",
            r.observacao or "",
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except Exception:
                pass

        adjusted_width = max_length + 5
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="relatorio_manutencoes.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )