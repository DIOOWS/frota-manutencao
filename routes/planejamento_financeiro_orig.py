from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from flask import Blueprint, render_template, request, jsonify, send_file
from sqlalchemy import or_
import hashlib
import re
import unicodedata
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import db
from utils.auth import gestao_required

from models.conta_pagar_importada import ContaPagarImportada
from models.conta_receber_importada import ContaReceberImportada
from models.planejamento_financeiro import PlanejamentoFinanceiro

from .radar_financeiro import (
    moeda,
    nome_mes,
    primeiro_dia_mes,
    ultimo_dia_mes,
    fim_dia_datetime,
)


planejamento_financeiro_bp = Blueprint(
    "planejamento_financeiro",
    __name__,
    url_prefix="/gestao/planejamento-financeiro",
)


def dinheiro_decimal(valor):
    if valor is None or valor == "":
        return Decimal("0")

    try:
        return Decimal(str(valor))
    except Exception:
        return Decimal("0")


def data_para_date_local(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except Exception:
            pass

    return None


def formatar_data(valor):
    data_ref = data_para_date_local(valor)
    if not data_ref:
        return "-"

    return data_ref.strftime("%d/%m/%Y")


def texto_limpo(valor, padrao="-"):
    if valor is None:
        return padrao

    texto = str(valor).strip()
    return texto if texto else padrao


def buscar_primeiro(conta, campos, padrao="-"):
    for campo in campos:
        valor = getattr(conta, campo, None)
        if valor not in (None, ""):
            return texto_limpo(valor, padrao)

    return padrao


def conta_esta_cancelada(conta):
    status = texto_limpo(getattr(conta, "status", None), "").upper()
    return status in ("CANCELADO", "CANCELADA")


def conta_esta_paga_local(conta):
    if bool(getattr(conta, "pago", False)):
        return True

    status = texto_limpo(getattr(conta, "status", None), "").upper()
    return status in ("PAGO", "RECEBIDO", "OK", "QUITADO", "BAIXADO")




def normalizar_chave(valor):
    """Normaliza texto para gerar uma chave estável entre reimportações."""
    if valor is None:
        return ""

    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto).strip().upper()
    return texto


def valor_chave(valor):
    return str(dinheiro_decimal(valor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def gerar_chave_conta(conta):
    """
    Chave estável da conta. Não depende do ID criado pela importação.

    Usa os campos mais consistentes do relatório para reconhecer a mesma conta
    depois que a importação é apagada e recriada.
    """
    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))
    vencimento_txt = vencimento.isoformat() if vencimento else ""

    fornecedor = buscar_primeiro(
        conta,
        ["fornecedor_funcionario", "fornecedor", "cliente"],
        "",
    )
    conta_nome = buscar_primeiro(
        conta,
        ["plano_contas", "categoria", "descricao", "nome"],
        "",
    )
    documento = buscar_primeiro(
        conta,
        ["numero_fatura", "documento", "numero_documento", "nf", "nota_fiscal"],
        "",
    )
    parcela = buscar_primeiro(
        conta,
        ["parcela_label", "parcela", "numero_parcela"],
        "",
    )

    base = "|".join([
        normalizar_chave(fornecedor),
        normalizar_chave(conta_nome),
        normalizar_chave(documento),
        normalizar_chave(parcela),
        vencimento_txt,
        valor_chave(getattr(conta, "valor", 0)),
    ])

    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def preencher_snapshot_planejamento(registro, conta, chave_conta=None):
    """Atualiza o vínculo atual e preserva dados essenciais do planejamento."""
    registro.conta_id = int(getattr(conta, "id", 0))
    registro.chave_conta = chave_conta or gerar_chave_conta(conta)
    registro.descricao_snapshot = buscar_primeiro(
        conta, ["plano_contas", "categoria", "descricao", "nome"], "SEM CONTA"
    )
    registro.fornecedor_snapshot = buscar_primeiro(
        conta, ["fornecedor_funcionario", "fornecedor", "cliente"], "-"
    )
    registro.valor_snapshot = dinheiro_decimal(getattr(conta, "valor", 0))
    registro.vencimento_snapshot = data_para_date_local(getattr(conta, "data_vencimento", None))
    registro.observacao_snapshot = buscar_primeiro(
        conta, ["observacoes", "observacao", "historico", "descricao"], "-"
    )


def localizar_planejamento(conta, mes, ano, por_id=None, por_chave=None):
    conta_id = int(getattr(conta, "id", 0))
    chave = gerar_chave_conta(conta)

    registro = None
    if por_id is not None:
        registro = por_id.get(conta_id)
    if not registro and por_chave is not None:
        registro = por_chave.get(chave)

    if registro:
        mudou = registro.conta_id != conta_id or registro.chave_conta != chave
        preencher_snapshot_planejamento(registro, conta, chave)
        if mudou:
            db.session.flush()
        return registro

    return PlanejamentoFinanceiro.query.filter(
        PlanejamentoFinanceiro.origem == "IMPORTADA",
        PlanejamentoFinanceiro.mes == mes,
        PlanejamentoFinanceiro.ano == ano,
        or_(
            PlanejamentoFinanceiro.conta_id == conta_id,
            PlanejamentoFinanceiro.chave_conta == chave,
        ),
    ).first()


def montar_item_planejamento(conta, mes, ano, hoje):
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))

    valor = dinheiro_decimal(getattr(conta, "valor", 0))

    conta_nome = buscar_primeiro(
        conta,
        [
            "plano_contas",
            "categoria",
            "descricao",
            "nome",
            "fornecedor_funcionario",
            "fornecedor",
        ],
        "SEM CONTA",
    )

    fornecedor = buscar_primeiro(
        conta,
        ["fornecedor_funcionario", "fornecedor", "cliente"],
        "-",
    )

    categoria = buscar_primeiro(
        conta,
        ["categoria", "plano_contas"],
        "-",
    )

    documento = buscar_primeiro(
        conta,
        ["numero_fatura", "documento", "numero_documento", "nf", "nota_fiscal"],
        "-",
    )

    observacao = buscar_primeiro(
        conta,
        ["observacoes", "observacao", "historico", "descricao"],
        "-",
    )

    parcela_label = buscar_primeiro(
        conta,
        ["parcela_label", "parcela", "numero_parcela"],
        "",
    )

    if parcela_label == "-":
        parcela_label = ""

    status_visual = "SEM DATA"
    is_herdada = False
    is_mes_atual = False
    dias_atraso = 0

    if vencimento:
        is_herdada = vencimento < primeiro_mes
        is_mes_atual = primeiro_mes <= vencimento <= ultimo_mes

        if vencimento < primeiro_mes:
            status_visual = "HERDADA"
        elif vencimento < hoje:
            status_visual = "ATRASADA"
        elif vencimento == hoje:
            status_visual = "VENCE HOJE"
        else:
            status_visual = "MÊS ATUAL"

        if vencimento < hoje:
            dias_atraso = (hoje - vencimento).days

    return {
        "id": int(getattr(conta, "id", 0)),
        "descricao": conta_nome,
        "conta_nome": conta_nome,
        "fornecedor": fornecedor,
        "categoria": categoria,
        "documento": documento,
        "observacao": observacao,
        "parcela_label": parcela_label,
        "vencimento": formatar_data(vencimento),
        "data_vencimento_obj": vencimento,
        "valor": valor,
        "valor_formatado": moeda(valor),
        "is_herdada": is_herdada,
        "is_mes_atual": is_mes_atual,
        "status_planejamento_visual": status_visual,
        "dias_atraso": dias_atraso,
    }


def buscar_contas_do_mes(mes, ano):
    """
    Busca todas as contas em aberto até o fim do mês selecionado.
    Isso inclui herdadas/atrasadas e contas do mês atual.
    Tudo vem em ordem de pagamento: vencimento mais antigo primeiro.
    """
    ultimo_mes = ultimo_dia_mes(mes, ano)

    return ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.pago == False,
        or_(
            ContaPagarImportada.status.is_(None),
            ContaPagarImportada.status != "CANCELADO",
        ),
        ContaPagarImportada.data_vencimento <= fim_dia_datetime(ultimo_mes),
    ).order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc(),
    ).all()




def buscar_contas_exportacao_e_tela(mes, ano):
    """
    Busca contas até o fim da competência, incluindo pagas.
    Necessário para manter histórico: se uma conta planejada em julho foi paga no Radar,
    ela continua aparecendo/exportando em julho como PAGA.
    """
    ultimo_mes = ultimo_dia_mes(mes, ano)

    return ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        or_(
            ContaPagarImportada.status.is_(None),
            ContaPagarImportada.status != "CANCELADO",
        ),
        ContaPagarImportada.data_vencimento <= fim_dia_datetime(ultimo_mes),
    ).order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc(),
    ).all()


def inicio_dia_datetime_local(data_ref):
    return datetime.combine(data_ref, datetime.min.time())




def valor_receber_decimal(conta):
    """Retorna o valor de contas a receber importadas com segurança."""
    return dinheiro_decimal(
        getattr(conta, "total", None)
        or getattr(conta, "valor", None)
        or getattr(conta, "valor_recebido", None)
        or 0
    )


def calcular_resumo_caixa_competencia(mes, ano):
    """
    Resumo do topo da tela.
    - Recebido: entradas confirmadas no Radar/Importações dentro da competência.
    - Pago: despesas confirmadas no Radar dentro da competência.
    - Saldo da conta: recebido - pago.
    """
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    inicio_dt = inicio_dia_datetime_local(primeiro_mes)
    fim_dt = fim_dia_datetime(ultimo_mes)

    contas_recebidas = ContaReceberImportada.query.filter(
        ContaReceberImportada.pago == True,
        ContaReceberImportada.origem_importacao == "RECEBIMENTO",
        ContaReceberImportada.data_pagamento >= inicio_dt,
        ContaReceberImportada.data_pagamento <= fim_dt,
    ).all()

    contas_pagas = ContaPagarImportada.query.filter(
        ContaPagarImportada.pago == True,
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        ContaPagarImportada.data_pagamento >= inicio_dt,
        ContaPagarImportada.data_pagamento <= fim_dt,
    ).all()

    total_recebido = sum(valor_receber_decimal(c) for c in contas_recebidas)
    total_pago = sum(dinheiro_decimal(getattr(c, "valor", 0)) for c in contas_pagas)
    saldo_conta = total_recebido - total_pago

    return total_recebido, total_pago, saldo_conta


def conta_paga_na_competencia(conta, mes, ano):
    """
    Define se a conta paga no Radar pertence à competência do planejamento.
    Regra principal: data_pagamento dentro do mês/ano selecionado.
    Fallback: se não tiver data_pagamento, usa vencimento dentro do mês/ano.
    """
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)

    data_pagamento = data_para_date_local(getattr(conta, "data_pagamento", None))
    if data_pagamento:
        return primeiro_mes <= data_pagamento <= ultimo_mes

    vencimento = data_para_date_local(getattr(conta, "data_vencimento", None))
    if vencimento:
        return primeiro_mes <= vencimento <= ultimo_mes

    return False


def buscar_planejamentos(mes, ano, contas=None):
    registros = PlanejamentoFinanceiro.query.filter_by(
        origem="IMPORTADA",
        mes=mes,
        ano=ano,
    ).all()

    por_id = {int(r.conta_id): r for r in registros if r.conta_id is not None}
    por_chave = {r.chave_conta: r for r in registros if getattr(r, "chave_conta", None)}

    # Migra automaticamente os registros antigos ainda ligados ao ID atual.
    alterou = False
    for conta in contas or []:
        conta_id = int(getattr(conta, "id", 0))
        registro = por_id.get(conta_id)
        if registro and not getattr(registro, "chave_conta", None):
            chave = gerar_chave_conta(conta)
            preencher_snapshot_planejamento(registro, conta, chave)
            por_chave[chave] = registro
            alterou = True

    if alterou:
        db.session.commit()

    return por_id, por_chave


def sincronizar_pagas_do_radar(mes, ano):
    """Sincroniza pagas sem perder vínculos quando os IDs da importação mudarem."""
    primeiro_mes = primeiro_dia_mes(mes, ano)
    ultimo_mes = ultimo_dia_mes(mes, ano)
    inicio_dt = inicio_dia_datetime_local(primeiro_mes)
    fim_dt = fim_dia_datetime(ultimo_mes)

    contas = ContaPagarImportada.query.filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
        or_(
            ContaPagarImportada.data_pagamento >= inicio_dt,
            ContaPagarImportada.data_vencimento >= inicio_dt,
        ),
        or_(
            ContaPagarImportada.data_pagamento <= fim_dt,
            ContaPagarImportada.data_vencimento <= fim_dt,
        ),
    ).all()

    por_id, por_chave = buscar_planejamentos(mes, ano, contas)
    criadas = 0
    atualizadas = 0

    for conta in contas:
        if conta_esta_cancelada(conta) or not conta_esta_paga_local(conta):
            continue
        if not conta_paga_na_competencia(conta, mes, ano):
            continue

        chave = gerar_chave_conta(conta)
        registro = por_id.get(int(conta.id)) or por_chave.get(chave)

        if registro:
            if registro.conta_id != conta.id or registro.chave_conta != chave:
                preencher_snapshot_planejamento(registro, conta, chave)
                atualizadas += 1
            continue

        novo = PlanejamentoFinanceiro(
            conta_id=int(conta.id),
            chave_conta=chave,
            origem="IMPORTADA",
            mes=mes,
            ano=ano,
            status_planejamento="SINCRONIZADA_RADAR",
        )
        preencher_snapshot_planejamento(novo, conta, chave)
        db.session.add(novo)
        por_id[int(conta.id)] = novo
        por_chave[chave] = novo
        criadas += 1

    db.session.commit()
    return criadas, atualizadas


@planejamento_financeiro_bp.route("/")
@gestao_required
def index():
    hoje = date.today()
    agora = datetime.now()

    mes = request.args.get("mes", type=int) or agora.month
    ano = request.args.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    contas = buscar_contas_exportacao_e_tela(mes, ano)
    planejamentos_id, planejamentos_chave = buscar_planejamentos(mes, ano, contas)

    # Planejamentos ativos em outras competências. Eles continuam aparecendo
    # em Aguardando Decisão, porém marcados como indisponíveis na interface.
    outros_planejamentos = PlanejamentoFinanceiro.query.filter(
        PlanejamentoFinanceiro.origem == "IMPORTADA",
        or_(
            PlanejamentoFinanceiro.mes != mes,
            PlanejamentoFinanceiro.ano != ano,
        ),
    ).order_by(
        PlanejamentoFinanceiro.ano.asc(),
        PlanejamentoFinanceiro.mes.asc(),
    ).all()

    outros_por_id = {}
    outros_por_chave = {}

    for planejamento_antigo in outros_planejamentos:
        conta_antiga = None
        if planejamento_antigo.conta_id is not None:
            conta_antiga = ContaPagarImportada.query.get(int(planejamento_antigo.conta_id))

        # Planejamentos de contas já pagas ou canceladas não bloqueiam.
        if conta_antiga and (conta_esta_paga_local(conta_antiga) or conta_esta_cancelada(conta_antiga)):
            continue

        if planejamento_antigo.conta_id is not None:
            outros_por_id.setdefault(int(planejamento_antigo.conta_id), planejamento_antigo)

        if planejamento_antigo.chave_conta:
            outros_por_chave.setdefault(planejamento_antigo.chave_conta, planejamento_antigo)

    aguardando = []
    planejadas = []
    executadas = []

    for conta in contas:
        if conta_esta_cancelada(conta):
            continue

        item = montar_item_planejamento(conta, mes, ano, hoje)
        registro = localizar_planejamento(conta, mes, ano, planejamentos_id, planejamentos_chave)

        if registro:
            if conta_esta_paga_local(conta):
                item["status_execucao"] = "PAGA"
                item["data_pagamento"] = formatar_data(getattr(conta, "data_pagamento", None))
                executadas.append(item)
            else:
                item["status_execucao"] = "PLANEJADA"
                item["data_pagamento"] = "-"
                planejadas.append(item)
        elif not conta_esta_paga_local(conta):
            item["status_execucao"] = "AGUARDANDO"
            item["data_pagamento"] = "-"

            chave_atual = gerar_chave_conta(conta)
            planejamento_outro_mes = (
                outros_por_id.get(int(conta.id))
                or outros_por_chave.get(chave_atual)
            )

            if planejamento_outro_mes:
                item["planejada_outro_mes"] = True
                item["competencia_planejada"] = (
                    f"{nome_mes(planejamento_outro_mes.mes).upper()}/"
                    f"{planejamento_outro_mes.ano}"
                )
            else:
                item["planejada_outro_mes"] = False
                item["competencia_planejada"] = ""

            aguardando.append(item)

    total_aguardando = sum(dinheiro_decimal(i.get("valor")) for i in aguardando)
    total_planejado = sum(dinheiro_decimal(i.get("valor")) for i in planejadas)
    total_executado = sum(dinheiro_decimal(i.get("valor")) for i in executadas)
    total_recebido, total_pago_competencia, saldo_competencia = calcular_resumo_caixa_competencia(mes, ano)

    return render_template(
        "gestao/planejamento_financeiro.html",
        mes=mes,
        ano=ano,
        nome_mes=nome_mes,
        moeda=moeda,
        aguardando=aguardando,
        planejadas=planejadas,
        executadas=executadas,
        total_aguardando=total_aguardando,
        total_planejado=total_planejado,
        total_executado=total_executado,
        total_recebido=total_recebido,
        total_pago_competencia=total_pago_competencia,
        saldo_competencia=saldo_competencia,
    )




def montar_listas_planejamento(mes, ano):
    hoje = date.today()
    contas = buscar_contas_exportacao_e_tela(mes, ano)
    planejamentos_id, planejamentos_chave = buscar_planejamentos(mes, ano, contas)

    aguardando = []
    planejadas = []
    pagas = []

    for conta in contas:
        if conta_esta_cancelada(conta):
            continue

        item = montar_item_planejamento(conta, mes, ano, hoje)
        item["data_pagamento"] = formatar_data(getattr(conta, "data_pagamento", None))
        registro = localizar_planejamento(conta, mes, ano, planejamentos_id, planejamentos_chave)

        if registro:
            if conta_esta_paga_local(conta):
                item["status_execucao"] = "PAGA"
                pagas.append(item)
            else:
                item["status_execucao"] = "PLANEJADA"
                planejadas.append(item)
        elif not conta_esta_paga_local(conta):
            item["status_execucao"] = "AGUARDANDO"
            aguardando.append(item)

    return aguardando, planejadas, pagas


def valor_decimal_item(item):
    return dinheiro_decimal(item.get("valor", 0))


def escrever_aba_planejamento(wb, titulo, competencia, itens, colunas):
    ws = wb.create_sheet(titulo)

    azul = "0B4EDB"
    azul_escuro = "061B3D"
    cinza = "F4F7FB"
    branco = "FFFFFF"
    borda_cor = "DCE5F1"

    border = Border(
        left=Side(style="thin", color=borda_cor),
        right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor),
        bottom=Side(style="thin", color=borda_cor),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    ws.cell(row=1, column=1).value = f"{titulo.upper()} - {competencia}"
    ws.cell(row=1, column=1).font = Font(bold=True, color=branco, size=14)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=azul_escuro)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(colunas))
    ws.cell(row=2, column=1).value = f"Exportado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')} | Total: {len(itens)} conta(s) | Valor: {moeda(sum(valor_decimal_item(i) for i in itens))}"
    ws.cell(row=2, column=1).font = Font(bold=True, color=azul_escuro, size=10)
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=cinza)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, (cabecalho, _) in enumerate(colunas, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = cabecalho
        cell.font = Font(bold=True, color=branco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    linha = 5
    for item in itens:
        for col_idx, (_, chave) in enumerate(colunas, start=1):
            valor = item.get(chave, "-")
            if chave == "valor":
                valor = float(valor_decimal_item(item))
            cell = ws.cell(row=linha, column=col_idx)
            cell.value = valor
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if chave == "valor":
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        linha += 1

    linha_total = linha + 1
    ws.cell(row=linha_total, column=1).value = "TOTAL"
    ws.cell(row=linha_total, column=1).font = Font(bold=True, color=azul_escuro)
    ws.cell(row=linha_total, column=len(colunas)).value = float(sum(valor_decimal_item(i) for i in itens))
    ws.cell(row=linha_total, column=len(colunas)).number_format = 'R$ #,##0.00'
    ws.cell(row=linha_total, column=len(colunas)).font = Font(bold=True, color=azul_escuro)

    larguras = {
        "Status": 18,
        "Vencimento": 14,
        "Pago em": 14,
        "Conta": 38,
        "Fornecedor": 34,
        "Observação": 45,
        "Valor": 16,
        "Origem": 16,
        "Parcela": 12,
    }
    for col_idx, (cabecalho, _) in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(cabecalho, 18)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(colunas))}{max(4, linha - 1)}"
    return ws



@planejamento_financeiro_bp.route("/sincronizar-radar", methods=["POST"])
@gestao_required
def sincronizar_radar():
    agora = datetime.now()
    mes = request.form.get("mes", type=int) or agora.month
    ano = request.form.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    try:
        criadas, atualizadas = sincronizar_pagas_do_radar(mes, ano)
        return jsonify({
            "ok": True,
            "message": f"Sincronização concluída. {criadas} conta(s) adicionada(s) e {atualizadas} vínculo(s) recuperado(s).",
            "criadas": criadas,
            "atualizadas": atualizadas,
        })
    except Exception as erro:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "message": f"Erro ao sincronizar Radar: {str(erro)}",
        }), 500

@planejamento_financeiro_bp.route("/exportar")
@gestao_required
def exportar_planejamento():
    agora = datetime.now()
    mes = request.args.get("mes", type=int) or agora.month
    ano = request.args.get("ano", type=int) or agora.year

    if not 1 <= mes <= 12:
        mes = agora.month

    aguardando, planejadas, pagas = montar_listas_planejamento(mes, ano)
    competencia = f"{nome_mes(mes)}/{ano}"

    wb = openpyxl.Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    colunas_base = [
        ("Status", "status_execucao"),
        ("Vencimento", "vencimento"),
        ("Conta", "conta_nome"),
        ("Fornecedor", "fornecedor"),
        ("Observação", "observacao"),
        ("Parcela", "parcela_label"),
        ("Valor", "valor"),
    ]

    colunas_pagas = [
        ("Status", "status_execucao"),
        ("Pago em", "data_pagamento"),
        ("Vencimento", "vencimento"),
        ("Conta", "conta_nome"),
        ("Fornecedor", "fornecedor"),
        ("Observação", "observacao"),
        ("Parcela", "parcela_label"),
        ("Valor", "valor"),
    ]

    escrever_aba_planejamento(wb, "Aguardando Decisão", competencia, aguardando, colunas_base)
    escrever_aba_planejamento(wb, "Planejadas", competencia, planejadas, colunas_base)
    escrever_aba_planejamento(wb, "Pagas", competencia, pagas, colunas_pagas)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"planejamento_financeiro_{nome_mes(mes).lower()}_{ano}.xlsx".replace(" ", "_")

    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@planejamento_financeiro_bp.route("/planejar/<int:conta_id>", methods=["POST"])
@gestao_required
def planejar(conta_id):
    mes = request.form.get("mes", type=int)
    ano = request.form.get("ano", type=int)

    if not mes or not ano:
        return jsonify({"ok": False, "message": "Mês e ano inválidos."}), 400

    conta = ContaPagarImportada.query.get(conta_id)
    if not conta:
        return jsonify({"ok": False, "message": "Conta não encontrada. Sincronize a importação e tente novamente."}), 404

    chave = gerar_chave_conta(conta)

    # Bloqueia a mesma conta/parcela enquanto existir planejamento ativo
    # em outra competência.
    #
    # A verificação usa conta_id OU chave_conta:
    # - conta_id cobre a mesma linha ainda existente na importação;
    # - chave_conta cobre reimportações em que o ID foi recriado.
    #
    # Não limitamos apenas ao status "PLANEJADA", pois registros antigos
    # podem ter status vazio ou outro texto. O que define se ainda está
    # ativo é a conta vinculada continuar em aberto.
    planejamentos_outros_meses = PlanejamentoFinanceiro.query.filter(
        PlanejamentoFinanceiro.origem == "IMPORTADA",
        or_(
            PlanejamentoFinanceiro.conta_id == conta_id,
            PlanejamentoFinanceiro.chave_conta == chave,
        ),
        or_(
            PlanejamentoFinanceiro.mes != mes,
            PlanejamentoFinanceiro.ano != ano,
        ),
    ).order_by(
        PlanejamentoFinanceiro.ano.asc(),
        PlanejamentoFinanceiro.mes.asc(),
    ).all()

    planejamento_ativo = None

    for planejamento_anterior in planejamentos_outros_meses:
        conta_vinculada = None

        if planejamento_anterior.conta_id is not None:
            conta_vinculada = ContaPagarImportada.query.get(
                int(planejamento_anterior.conta_id)
            )

        # Se a conta vinculada existe e já foi paga ou cancelada,
        # o planejamento antigo não deve bloquear uma nova competência.
        if conta_vinculada:
            if conta_esta_paga_local(conta_vinculada):
                continue
            if conta_esta_cancelada(conta_vinculada):
                continue

        # Se o vínculo antigo não existe mais, mas o planejamento permanece
        # no banco, ele continua sendo considerado ativo até ser removido.
        planejamento_ativo = planejamento_anterior
        break

    if planejamento_ativo:
        competencia_existente = (
            f"{nome_mes(planejamento_ativo.mes)}/"
            f"{planejamento_ativo.ano}"
        )

        return jsonify({
            "ok": False,
            "message": (
                "Esta mesma conta/parcela já está planejada em "
                f"{competencia_existente} e ainda está em aberto. "
                "Remova o planejamento anterior ou registre o pagamento."
            ),
            "competencia": competencia_existente,
        }), 409

    existe = PlanejamentoFinanceiro.query.filter(
        PlanejamentoFinanceiro.origem == "IMPORTADA",
        PlanejamentoFinanceiro.mes == mes,
        PlanejamentoFinanceiro.ano == ano,
        or_(
            PlanejamentoFinanceiro.conta_id == conta_id,
            PlanejamentoFinanceiro.chave_conta == chave,
        ),
    ).first()

    if existe:
        preencher_snapshot_planejamento(existe, conta, chave)
        existe.status_planejamento = "PLANEJADA"
    else:
        existe = PlanejamentoFinanceiro(
            conta_id=conta_id,
            chave_conta=chave,
            origem="IMPORTADA",
            mes=mes,
            ano=ano,
            status_planejamento="PLANEJADA",
        )
        preencher_snapshot_planejamento(existe, conta, chave)
        db.session.add(existe)

    db.session.commit()
    return jsonify({"ok": True, "message": "Conta planejada para pagamento."})


@planejamento_financeiro_bp.route("/remover/<int:conta_id>", methods=["POST"])
@gestao_required
def remover(conta_id):
    mes = request.form.get("mes", type=int)
    ano = request.form.get("ano", type=int)

    if not mes or not ano:
        return jsonify({"ok": False, "message": "Mês e ano inválidos."}), 400

    conta = ContaPagarImportada.query.get(conta_id)
    chave = gerar_chave_conta(conta) if conta else None

    filtros = [
        PlanejamentoFinanceiro.origem == "IMPORTADA",
        PlanejamentoFinanceiro.mes == mes,
        PlanejamentoFinanceiro.ano == ano,
    ]

    if chave:
        filtros.append(or_(
            PlanejamentoFinanceiro.conta_id == conta_id,
            PlanejamentoFinanceiro.chave_conta == chave,
        ))
    else:
        filtros.append(PlanejamentoFinanceiro.conta_id == conta_id)

    registro = PlanejamentoFinanceiro.query.filter(*filtros).first()

    if registro:
        db.session.delete(registro)
        db.session.commit()

    return jsonify({"ok": True, "message": "Conta removida do planejamento."})