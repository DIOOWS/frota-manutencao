from flask import Blueprint, render_template, request, redirect, flash
from utils.auth import gestao_required
from database import db
from models.conta_pagar_importada import ContaPagarImportada
from models.conta_receber_importada import ContaReceberImportada
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO, StringIO
from sqlalchemy import or_
import openpyxl
import unicodedata

try:
    import pandas as pd
except Exception:
    pd = None


importacoes_bp = Blueprint(
    "importacoes",
    __name__,
    url_prefix="/gestao/importacoes"
)


# =========================================================
# CONFIGURAÇÃO DE ARQUIVOS
# =========================================================

EXTENSOES_PERMITIDAS = {"xlsx", "xlsm", "xls", "xlsb"}


# =========================================================
# HELPERS GERAIS
# =========================================================

def texto(valor):
    if valor is None:
        return ""

    try:
        if pd is not None and pd.isna(valor):
            return ""
    except Exception:
        pass

    return str(valor).strip()


def normalizar_nome_coluna(valor):
    valor = texto(valor).lower()

    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(
        caractere for caractere in valor
        if not unicodedata.combining(caractere)
    )

    valor = valor.replace("\n", " ")
    valor = valor.replace("\r", " ")
    valor = " ".join(valor.split())

    return valor


def normalizar_texto_chave(valor):
    valor = texto(valor).upper()

    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(
        caractere for caractere in valor
        if not unicodedata.combining(caractere)
    )

    valor = valor.replace("\n", " ")
    valor = valor.replace("\r", " ")
    valor = " ".join(valor.split())

    return valor


def normalizar_data_chave(valor):
    if not valor:
        return ""

    try:
        return valor.strftime("%Y-%m-%d")
    except Exception:
        return texto(valor)


def normalizar_decimal_chave(valor):
    try:
        return f"{Decimal(valor or 0).quantize(Decimal('0.01'))}"
    except Exception:
        return "0.00"


def extensao_arquivo(filename):
    if not filename or "." not in filename:
        return ""

    return filename.rsplit(".", 1)[1].lower().strip()


def arquivo_excel_valido(filename):
    return extensao_arquivo(filename) in EXTENSOES_PERMITIDAS


def detectar_tipo_real_excel(conteudo):
    """
    Detecta o tipo real do arquivo pelo conteúdo interno.

    PK = xlsx/xlsm/xlsb
    D0 CF = xls antigo real
    HTML/XML = exportações antigas que o Excel abre como .xls
    """
    inicio = conteudo[:1000].lstrip().lower()

    if conteudo.startswith(b"PK"):
        return "zip_excel"

    if conteudo.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls_antigo"

    if (
        inicio.startswith(b"<html")
        or b"<table" in inicio
        or b"<body" in inicio
        or b"<!doctype html" in inicio
    ):
        return "html_excel"

    if (
        inicio.startswith(b"<?xml")
        or b"<workbook" in inicio
        or b"urn:schemas-microsoft-com:office:spreadsheet" in inicio
    ):
        return "xml_excel"

    return "desconhecido"


def normalizar_bool(valor):
    if valor is True:
        return True

    valor_txt = texto(valor).upper()

    return valor_txt in [
        "TRUE",
        "VERDADEIRO",
        "SIM",
        "S",
        "1",
        "PAGO",
        "OK",
        "RECEBIDO"
    ]


def normalizar_decimal(valor):
    if valor is None or texto(valor) == "":
        return Decimal("0.00")

    if isinstance(valor, (int, float, Decimal)):
        try:
            return Decimal(str(valor)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal("0.00")

    valor_txt = str(valor).strip()
    valor_txt = valor_txt.replace("R$", "")
    valor_txt = valor_txt.replace(" ", "")

    # Padrão brasileiro: 1.234,56
    if "," in valor_txt:
        valor_txt = valor_txt.replace(".", "")
        valor_txt = valor_txt.replace(",", ".")

    try:
        return Decimal(valor_txt).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def normalizar_data(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor

    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())

    formatos = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]

    valor_txt = texto(valor)

    for formato in formatos:
        try:
            return datetime.strptime(valor_txt, formato)
        except Exception:
            pass

    return None


def data_para_input(valor):
    if not valor:
        return ""

    try:
        return valor.strftime("%Y-%m-%d")
    except Exception:
        return ""


def identificar_setor(plano_contas, receita=False):
    plano = texto(plano_contas).upper()

    if plano.endswith(" T"):
        return "LOGÍSTICA"

    if receita:
        return "GERAL"

    return "ASSISTÊNCIA"


def limpar_categoria(plano_contas):
    plano = texto(plano_contas)

    if plano.upper().endswith(" T"):
        return plano[:-2].strip()

    return plano


def valor_por_coluna(row_dict, nome):
    """
    Busca valor da coluna de forma flexível:
    - ignora acentos
    - ignora maiúsculas/minúsculas
    - ignora espaços extras
    - aceita pequenas variações de cabeçalho
    """
    if not row_dict:
        return ""

    nome_normalizado = normalizar_nome_coluna(nome)

    for chave, valor in row_dict.items():
        if normalizar_nome_coluna(chave) == nome_normalizado:
            return valor

    return ""


def valor_por_colunas(row_dict, nomes):
    for nome in nomes:
        valor = valor_por_coluna(row_dict, nome)

        if texto(valor) != "":
            return valor

    return ""


def redirect_importacoes(mes, ano, extra=""):
    url = f"/gestao/importacoes/?mes={mes}&ano={ano}"

    if extra:
        url += extra

    return redirect(url)


def validar_mes_ano(mes, ano):
    if not mes or not ano:
        return False

    if mes < 1 or mes > 12:
        return False

    if ano < 2000:
        return False

    return True


# =========================================================
# HELPERS DE CONCILIAÇÃO
# =========================================================

def gerar_chave_conta_pagar(
    numero_fatura,
    fornecedor_funcionario,
    valor,
    data_vencimento,
    plano_contas
):
    partes = [
        normalizar_texto_chave(numero_fatura),
        normalizar_texto_chave(fornecedor_funcionario),
        normalizar_decimal_chave(valor),
        normalizar_data_chave(data_vencimento),
        normalizar_texto_chave(plano_contas),
    ]

    return "|".join(partes)


def gerar_chave_conta_receber(
    numero_fatura,
    cliente,
    valor,
    data_vencimento,
    plano_contas
):
    partes = [
        normalizar_texto_chave(numero_fatura),
        normalizar_texto_chave(cliente),
        normalizar_decimal_chave(valor),
        normalizar_data_chave(data_vencimento),
        normalizar_texto_chave(plano_contas),
    ]

    return "|".join(partes)


def buscar_conta_pagar_existente(
    chave_conciliacao,
    numero_fatura,
    fornecedor_funcionario,
    valor,
    data_vencimento,
    plano_contas
):
    conta = None

    if chave_conciliacao:
        conta = ContaPagarImportada.query.filter_by(
            chave_conciliacao=chave_conciliacao
        ).first()

    if conta:
        return conta

    query = ContaPagarImportada.query.filter(
        ContaPagarImportada.numero_fatura == numero_fatura,
        ContaPagarImportada.fornecedor_funcionario == fornecedor_funcionario,
        ContaPagarImportada.valor == valor,
        ContaPagarImportada.plano_contas == plano_contas
    )

    if data_vencimento:
        query = query.filter(
            ContaPagarImportada.data_vencimento == data_vencimento
        )

    return query.first()




def buscar_despesa_completa_existente(
    chave_conciliacao,
    numero_fatura,
    fornecedor_funcionario,
    valor,
    data_vencimento,
    plano_contas
):
    """
    Busca somente registros da nova importação completa de despesas.

    Protegido com no_autoflush para evitar que o SQLAlchemy tente salvar
    objetos incompletos no meio da importação quando uma consulta .first()
    é executada dentro do loop.
    """
    with db.session.no_autoflush:
        conta = None

        if chave_conciliacao:
            conta = ContaPagarImportada.query.filter(
                ContaPagarImportada.chave_conciliacao == chave_conciliacao,
                ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA"
            ).first()

        if conta:
            return conta

        query = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
            ContaPagarImportada.numero_fatura == numero_fatura,
            ContaPagarImportada.fornecedor_funcionario == fornecedor_funcionario,
            ContaPagarImportada.valor == valor,
            ContaPagarImportada.plano_contas == plano_contas
        )

        if data_vencimento:
            query = query.filter(
                ContaPagarImportada.data_vencimento == data_vencimento
            )

        return query.first()


def buscar_conta_receber_existente(
    chave_conciliacao,
    numero_fatura,
    cliente,
    valor,
    data_vencimento,
    plano_contas
):
    conta = None

    if chave_conciliacao:
        conta = ContaReceberImportada.query.filter_by(
            chave_conciliacao=chave_conciliacao
        ).first()

    if conta:
        return conta

    query = ContaReceberImportada.query.filter(
        ContaReceberImportada.numero_fatura == numero_fatura,
        ContaReceberImportada.cliente == cliente,
        ContaReceberImportada.plano_contas == plano_contas
    )

    if data_vencimento:
        query = query.filter(
            ContaReceberImportada.data_vencimento == data_vencimento
        )

    query = query.filter(
        or_(
            ContaReceberImportada.total == valor,
            ContaReceberImportada.valor == valor
        )
    )

    return query.first()


# =========================================================
# LEITURA OPENPYXL - XLSX / XLSM
# =========================================================

def achar_linha_cabecalho_openpyxl(sheet, primeira_coluna):
    primeira_coluna = primeira_coluna.upper()

    for row in sheet.iter_rows():
        valor = texto(row[0].value).upper()

        if valor == primeira_coluna:
            return row[0].row

    return None


def montar_linhas_openpyxl(sheet, header_row):
    headers = []

    for cell in sheet[header_row]:
        headers.append(texto(cell.value))

    linhas = []

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None or texto(v) == "" for v in row):
            continue

        item = {}

        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None

        linhas.append(item)

    return linhas


# =========================================================
# LEITURA PANDAS - XLS / XLSB / HTML / XML
# =========================================================

def achar_linha_cabecalho_dataframe(df_raw, primeira_coluna):
    primeira_coluna = primeira_coluna.upper()

    for idx, row in df_raw.iterrows():
        if len(row) == 0:
            continue

        primeiro_valor = texto(row.iloc[0]).upper()

        if primeiro_valor == primeira_coluna:
            return idx

    return None


def montar_linhas_dataframe(df_raw, primeira_coluna):
    header_index = achar_linha_cabecalho_dataframe(df_raw, primeira_coluna)

    if header_index is None:
        raise ValueError(
            f"Cabeçalho não encontrado. A primeira coluna precisa conter '{primeira_coluna}'."
        )

    headers = []

    for valor in df_raw.iloc[header_index].tolist():
        headers.append(texto(valor))

    linhas = []

    for _, row in df_raw.iloc[header_index + 1:].iterrows():
        valores = row.tolist()

        if all(texto(v) == "" for v in valores):
            continue

        item = {}

        for idx, header in enumerate(headers):
            if header:
                item[header] = valores[idx] if idx < len(valores) else None

        linhas.append(item)

    return linhas


def montar_linhas_pandas_excel(conteudo, extensao, primeira_coluna):
    if pd is None:
        raise ImportError(
            "Para importar arquivos .xls ou .xlsb, instale pandas, xlrd e pyxlsb."
        )

    engine = None

    if extensao == "xls":
        engine = "xlrd"

    if extensao == "xlsb":
        engine = "pyxlsb"

    try:
        df_raw = pd.read_excel(
            BytesIO(conteudo),
            header=None,
            engine=engine
        )

        return montar_linhas_dataframe(df_raw, primeira_coluna)

    except Exception as e:
        raise ValueError(
            f"Não foi possível ler a planilha .{extensao}. "
            f"Verifique se o arquivo não está corrompido. Erro: {str(e)}"
        )


def montar_linhas_pandas_html(conteudo, primeira_coluna):
    if pd is None:
        raise ImportError(
            "Para importar arquivos HTML com extensão .xls, instale pandas, lxml, html5lib e beautifulsoup4."
        )

    encodings = ["utf-8", "latin1", "cp1252"]
    flavors = ["lxml", "bs4", "html5lib"]

    erros = []

    for enc in encodings:
        html = conteudo.decode(enc, errors="ignore")

        for flavor in flavors:
            try:
                tabelas = pd.read_html(
                    StringIO(html),
                    header=None,
                    flavor=flavor
                )

                if not tabelas:
                    continue

                for df_raw in tabelas:
                    try:
                        return montar_linhas_dataframe(df_raw, primeira_coluna)
                    except Exception as e:
                        erros.append(f"{flavor}/{enc}: {str(e)}")
                        continue

            except Exception as e:
                erros.append(f"{flavor}/{enc}: {str(e)}")

    raise ValueError(
        "Não foi possível ler o arquivo .xls como HTML/XML. "
        "Abra no Excel e salve novamente como .xlsx. "
        "Detalhes: " + " | ".join(erros[:5])
    )


def montar_linhas_pandas_tentativas(conteudo, extensao, primeira_coluna):
    erros = []

    try:
        return montar_linhas_pandas_excel(conteudo, "xls", primeira_coluna)
    except Exception as e:
        erros.append(f"XLS: {str(e)}")

    try:
        return montar_linhas_pandas_html(conteudo, primeira_coluna)
    except Exception as e:
        erros.append(f"HTML/XML: {str(e)}")

    raise ValueError(
        "O arquivo foi reconhecido como .xls, mas não consegui ler a estrutura interna. "
        "Abra no Excel e use: Arquivo > Salvar Como > Pasta de Trabalho do Excel (*.xlsx). "
        "Detalhes: " + " | ".join(erros)
    )


def ler_linhas_upload(file_storage, primeira_coluna="Nº Fatura"):
    filename = file_storage.filename
    extensao = extensao_arquivo(filename)

    if not arquivo_excel_valido(filename):
        raise ValueError(
            "Formato não aceito. Envie uma planilha Excel nos formatos .xlsx, .xlsm, .xls ou .xlsb."
        )

    conteudo = file_storage.read()

    if not conteudo:
        raise ValueError("O arquivo enviado está vazio.")

    tipo_real = detectar_tipo_real_excel(conteudo)

    if tipo_real == "xls_antigo":
        return montar_linhas_pandas_excel(conteudo, "xls", primeira_coluna)

    if tipo_real in ["html_excel", "xml_excel"]:
        return montar_linhas_pandas_html(conteudo, primeira_coluna)

    if tipo_real == "zip_excel":

        if extensao == "xlsb":
            return montar_linhas_pandas_excel(conteudo, "xlsb", primeira_coluna)

        try:
            workbook = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise ValueError(
                "Não foi possível abrir a planilha. "
                "O arquivo parece ser Excel moderno, mas pode estar corrompido. "
                "Abra no Excel e salve novamente como .xlsx. "
                f"Erro: {str(e)}"
            )

        header_row = achar_linha_cabecalho_openpyxl(sheet, primeira_coluna)

        if not header_row:
            raise ValueError(
                f"Cabeçalho não encontrado. A primeira coluna precisa conter '{primeira_coluna}'."
            )

        return montar_linhas_openpyxl(sheet, header_row)

    if extensao == "xls":
        return montar_linhas_pandas_tentativas(conteudo, extensao, primeira_coluna)

    raise ValueError(
        "O arquivo enviado não parece ser uma planilha Excel válida. "
        "Abra o arquivo no Excel e salve novamente como 'Pasta de Trabalho do Excel (*.xlsx)'."
    )


# =========================================================
# TELA ÚNICA DE IMPORTAÇÕES
# =========================================================

@importacoes_bp.route("", strict_slashes=False)
@importacoes_bp.route("/", strict_slashes=False)
@gestao_required
def index():

    hoje = datetime.now()

    mes = request.args.get("mes", type=int) or hoje.month
    ano = request.args.get("ano", type=int) or hoje.year

    busca_pagar = texto(request.args.get("busca_pagar"))
    status_pagar = texto(request.args.get("status_pagar")).upper()
    setor_pagar = texto(request.args.get("setor_pagar")).upper()

    busca_receber = texto(request.args.get("busca_receber"))
    status_receber = texto(request.args.get("status_receber")).upper()

    busca_despesas_completo = texto(request.args.get("busca_despesas_completo"))
    status_despesas_completo = texto(request.args.get("status_despesas_completo")).upper()
    setor_despesas_completo = texto(request.args.get("setor_despesas_completo")).upper()

    # Agora a tela trabalha com o cenário novo:
    # 1. Contas Pagas = saídas já realizadas
    # 2. Contas a Receber = entradas/recebíveis importados
    contas_pagar_query = ContaPagarImportada.query.filter_by(
        mes=mes,
        ano=ano
    ).filter(
        ContaPagarImportada.origem_importacao == "PAGAMENTO"
    )

    if busca_pagar:
        like = f"%{busca_pagar}%"
        contas_pagar_query = contas_pagar_query.filter(
            or_(
                ContaPagarImportada.numero_fatura.ilike(like),
                ContaPagarImportada.fornecedor_funcionario.ilike(like),
                ContaPagarImportada.telefone.ilike(like),
                ContaPagarImportada.email.ilike(like),
                ContaPagarImportada.plano_contas.ilike(like),
                ContaPagarImportada.categoria.ilike(like),
                ContaPagarImportada.setor.ilike(like),
                ContaPagarImportada.status.ilike(like),
                ContaPagarImportada.observacoes.ilike(like),
            )
        )

    if status_pagar in ["PAGO", "PENDENTE"]:
        contas_pagar_query = contas_pagar_query.filter(
            ContaPagarImportada.status == status_pagar
        )

    if setor_pagar in ["ASSISTÊNCIA", "LOGÍSTICA"]:
        contas_pagar_query = contas_pagar_query.filter(
            ContaPagarImportada.setor == setor_pagar
        )

    contas_pagar = contas_pagar_query.order_by(
        ContaPagarImportada.data_pagamento.asc(),
        ContaPagarImportada.id.asc()
    ).all()

    contas_receber_query = ContaReceberImportada.query.filter_by(
        mes=mes,
        ano=ano
    ).filter(
        or_(
            ContaReceberImportada.origem_importacao.is_(None),
            ContaReceberImportada.origem_importacao != "MANUAL"
        )
    )

    if busca_receber:
        like = f"%{busca_receber}%"
        contas_receber_query = contas_receber_query.filter(
            or_(
                ContaReceberImportada.numero_fatura.ilike(like),
                ContaReceberImportada.cliente.ilike(like),
                ContaReceberImportada.telefone.ilike(like),
                ContaReceberImportada.email.ilike(like),
                ContaReceberImportada.plano_contas.ilike(like),
                ContaReceberImportada.categoria.ilike(like),
                ContaReceberImportada.setor.ilike(like),
                ContaReceberImportada.cobranca.ilike(like),
                ContaReceberImportada.status.ilike(like),
                ContaReceberImportada.observacoes.ilike(like),
            )
        )

    if status_receber in ["RECEBIDO", "PENDENTE"]:
        contas_receber_query = contas_receber_query.filter(
            ContaReceberImportada.status == status_receber
        )

    contas_receber = contas_receber_query.order_by(
        ContaReceberImportada.data_vencimento.asc(),
        ContaReceberImportada.id.asc()
    ).all()

    despesas_completo_query = ContaPagarImportada.query.filter_by(
        mes=mes,
        ano=ano
    ).filter(
        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA"
    )

    if busca_despesas_completo:
        like = f"%{busca_despesas_completo}%"
        despesas_completo_query = despesas_completo_query.filter(
            or_(
                ContaPagarImportada.numero_fatura.ilike(like),
                ContaPagarImportada.fornecedor_funcionario.ilike(like),
                ContaPagarImportada.telefone.ilike(like),
                ContaPagarImportada.email.ilike(like),
                ContaPagarImportada.plano_contas.ilike(like),
                ContaPagarImportada.categoria.ilike(like),
                ContaPagarImportada.setor.ilike(like),
                ContaPagarImportada.status.ilike(like),
                ContaPagarImportada.observacoes.ilike(like),
            )
        )

    if status_despesas_completo in ["PAGO", "PENDENTE"]:
        despesas_completo_query = despesas_completo_query.filter(
            ContaPagarImportada.status == status_despesas_completo
        )

    if setor_despesas_completo in ["ASSISTÊNCIA", "LOGÍSTICA"]:
        despesas_completo_query = despesas_completo_query.filter(
            ContaPagarImportada.setor == setor_despesas_completo
        )

    despesas_completo = despesas_completo_query.order_by(
        ContaPagarImportada.data_vencimento.asc(),
        ContaPagarImportada.id.asc()
    ).all()

    total_pagar = sum([float(c.valor or 0) for c in contas_pagar])
    total_pagar_pago = sum([float(c.valor or 0) for c in contas_pagar if c.pago])
    total_pagar_pendente = sum([float(c.valor or 0) for c in contas_pagar if not c.pago])

    total_pagar_assistencia = sum([
        float(c.valor or 0) for c in contas_pagar
        if c.setor == "ASSISTÊNCIA"
    ])

    total_pagar_logistica = sum([
        float(c.valor or 0) for c in contas_pagar
        if c.setor == "LOGÍSTICA"
    ])

    total_receber = sum([float(c.total or c.valor or 0) for c in contas_receber])

    total_receber_pago = sum([
        float(c.total or c.valor or 0) for c in contas_receber
        if c.pago
    ])

    total_receber_pendente = sum([
        float(c.total or c.valor or 0) for c in contas_receber
        if not c.pago
    ])

    total_despesas_completo = sum([float(c.valor or 0) for c in despesas_completo])
    total_despesas_completo_pago = sum([
        float(c.valor or 0) for c in despesas_completo
        if c.pago
    ])
    total_despesas_completo_pendente = sum([
        float(c.valor or 0) for c in despesas_completo
        if not c.pago
    ])
    total_despesas_completo_assistencia = sum([
        float(c.valor or 0) for c in despesas_completo
        if c.setor == "ASSISTÊNCIA"
    ])
    total_despesas_completo_logistica = sum([
        float(c.valor or 0) for c in despesas_completo
        if c.setor == "LOGÍSTICA"
    ])

    return render_template(
        "gestao/importacoes.html",

        mes=mes,
        ano=ano,

        contas_pagar=contas_pagar,
        contas_receber=contas_receber,
        despesas_completo=despesas_completo,

        total_pagar=total_pagar,
        total_pagar_pago=total_pagar_pago,
        total_pagar_pendente=total_pagar_pendente,
        total_pagar_assistencia=total_pagar_assistencia,
        total_pagar_logistica=total_pagar_logistica,

        total_receber=total_receber,
        total_receber_pago=total_receber_pago,
        total_receber_pendente=total_receber_pendente,

        total_despesas_completo=total_despesas_completo,
        total_despesas_completo_pago=total_despesas_completo_pago,
        total_despesas_completo_pendente=total_despesas_completo_pendente,
        total_despesas_completo_assistencia=total_despesas_completo_assistencia,
        total_despesas_completo_logistica=total_despesas_completo_logistica,

        busca_pagar=busca_pagar,
        status_pagar=status_pagar,
        setor_pagar=setor_pagar,
        busca_receber=busca_receber,
        status_receber=status_receber,
        busca_despesas_completo=busca_despesas_completo,
        status_despesas_completo=status_despesas_completo,
        setor_despesas_completo=setor_despesas_completo,

        data_para_input=data_para_input
    )


# =========================================================
# IMPORTAR CONTAS PAGAS
# Relatório de pagamentos realizados
# Agora aceita arquivo com vários meses/ano inteiro.
# O mês/ano são definidos automaticamente pela Dt. Pgto de cada linha.
# =========================================================

@importacoes_bp.route("/contas-pagas", methods=["POST"])
@gestao_required
def importar_contas_pagas():

    arquivo = request.files.get("arquivo_pagas")

    # Usado apenas para voltar para a competência que o usuário estava visualizando.
    hoje = datetime.now()
    mes_retorno = request.form.get("mes_retorno", type=int) or hoje.month
    ano_retorno = request.form.get("ano_retorno", type=int) or hoje.year

    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha de Contas Pagas.", "danger")
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}")

    if not arquivo_excel_valido(arquivo.filename):
        flash(
            "Formato não aceito. Envie uma planilha Excel nos formatos .xlsx, .xlsm, .xls ou .xlsb.",
            "danger"
        )
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}")

    try:
        linhas = ler_linhas_upload(arquivo, "Nº Fatura")

        total_criado = 0
        total_atualizado = 0
        total_ignorado = 0
        total_linhas = 0
        competencias_importadas = set()

        for linha in linhas:
            plano_contas = texto(valor_por_coluna(linha, "Pl. Contas"))

            numero_fatura = texto(valor_por_coluna(linha, "Nº Fatura"))
            fornecedor_funcionario = texto(valor_por_coluna(linha, "Fornecedor/Funcionário"))
            telefone = texto(valor_por_coluna(linha, "Telefone"))
            email = texto(valor_por_coluna(linha, "Email"))

            data_pagamento = normalizar_data(valor_por_coluna(linha, "Dt. Pgto"))
            data_vencimento = normalizar_data(valor_por_coluna(linha, "Dt. Vencto"))
            data_documento = normalizar_data(valor_por_coluna(linha, "Dt. Docto"))

            valor = normalizar_decimal(valor_por_coluna(linha, "Valor"))

            # Contas Pagas só entram quando existe Dt. Pgto.
            # É essa data que define automaticamente mês/ano no banco.
            if not data_pagamento:
                total_ignorado += 1
                continue

            observacoes = texto(
                valor_por_colunas(
                    linha,
                    [
                        "Observações",
                        "Observacoes",
                        "Observação",
                        "Observacao",
                        "Obs",
                        "OBS",
                        "Histórico",
                        "Historico",
                        "Descrição",
                        "Descricao"
                    ]
                )
            )

            chave_conciliacao = gerar_chave_conta_pagar(
                numero_fatura=numero_fatura,
                fornecedor_funcionario=fornecedor_funcionario,
                valor=valor,
                data_vencimento=data_vencimento,
                plano_contas=plano_contas
            )

            conta = buscar_conta_pagar_existente(
                chave_conciliacao=chave_conciliacao,
                numero_fatura=numero_fatura,
                fornecedor_funcionario=fornecedor_funcionario,
                valor=valor,
                data_vencimento=data_vencimento,
                plano_contas=plano_contas
            )

            if conta:
                total_atualizado += 1
            else:
                conta = ContaPagarImportada()
                db.session.add(conta)
                total_criado += 1

            conta.numero_fatura = numero_fatura
            conta.fornecedor_funcionario = fornecedor_funcionario
            conta.telefone = telefone
            conta.email = email

            conta.plano_contas = plano_contas
            conta.categoria = limpar_categoria(plano_contas)
            conta.setor = identificar_setor(plano_contas, receita=False)

            conta.data_documento = data_documento
            conta.data_vencimento = data_vencimento
            conta.data_pagamento = data_pagamento

            conta.valor = valor

            conta.pago = True
            conta.status = "PAGO"

            conta.observacoes = observacoes

            # NOVO CENÁRIO:
            # Não força mais tudo para o mês escolhido na tela.
            # A competência vem da data real de pagamento da linha.
            conta.mes = data_pagamento.month
            conta.ano = data_pagamento.year
            competencias_importadas.add(f"{data_pagamento.month:02d}/{data_pagamento.year}")

            conta.chave_conciliacao = chave_conciliacao
            conta.origem_importacao = "PAGAMENTO"
            conta.importado_em = datetime.utcnow()

            total_linhas += 1

        db.session.commit()

        competencias_txt = ", ".join(sorted(competencias_importadas)) or "nenhuma competência"

        flash(
            (
                f"Contas Pagas importadas com sucesso! "
                f"{total_linhas} linha(s) processada(s). "
                f"Competências identificadas: {competencias_txt}. "
                f"Criadas: {total_criado}. "
                f"Atualizadas: {total_atualizado}. "
                f"Ignoradas sem Dt. Pgto: {total_ignorado}."
            ),
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao importar Contas Pagas: {str(e)}", "danger")

    return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}")


# =========================================================
# IMPORTAR DESPESAS COMPLETO
# Relatório completo de contas a pagar: pagas + pendentes
# =========================================================

@importacoes_bp.route("/despesas-completo", methods=["POST"])
@gestao_required
def importar_despesas_completo():

    arquivo = request.files.get("arquivo_despesas_completo")
    mes_retorno = request.form.get("mes_despesas_completo", type=int)
    ano_retorno = request.form.get("ano_despesas_completo", type=int)

    if not validar_mes_ano(mes_retorno, ano_retorno):
        flash("Informe mês e ano válidos para Despesas Completo.", "danger")
        return redirect("/gestao/importacoes/")

    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha de Despesas Completo.", "danger")
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#despesas-completo")

    if not arquivo_excel_valido(arquivo.filename):
        flash(
            "Formato não aceito. Envie uma planilha Excel nos formatos .xlsx, .xlsm, .xls ou .xlsb.",
            "danger"
        )
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#despesas-completo")

    try:
        linhas = ler_linhas_upload(arquivo, "Nº Fatura")

        total_criado = 0
        total_atualizado = 0
        total_ignorado = 0
        total_linhas = 0
        total_pagas = 0
        total_pendentes = 0

        # =====================================================
        # OTIMIZAÇÃO PRODUÇÃO / RENDER
        # =====================================================
        # Antes o sistema fazia 1 SELECT no banco para cada linha do Excel.
        # Com 1.500+ linhas isso pode estourar memória/tempo no Render.
        # Agora primeiro montamos os dados em memória, coletamos as chaves
        # e buscamos as contas já existentes em UMA consulta só.
        registros_processados = []
        chaves_conciliacao = set()

        for linha in linhas:
            plano_contas = texto(valor_por_coluna(linha, "Pl. Contas"))

            numero_fatura = texto(valor_por_coluna(linha, "Nº Fatura"))
            fornecedor_funcionario = texto(valor_por_coluna(linha, "Fornecedor/Funcionário"))
            telefone = texto(valor_por_coluna(linha, "Telefone"))
            email = texto(valor_por_coluna(linha, "Email"))

            data_pagamento = normalizar_data(valor_por_coluna(linha, "Dt. Pgto"))
            data_vencimento = normalizar_data(valor_por_coluna(linha, "Dt. Vencto"))
            data_documento = normalizar_data(valor_por_coluna(linha, "Dt. Docto"))

            valor = normalizar_decimal(valor_por_coluna(linha, "Valor"))
            pago = normalizar_bool(valor_por_coluna(linha, "Pg?"))

            if not data_vencimento and not data_pagamento:
                total_ignorado += 1
                continue

            observacoes = texto(
                valor_por_colunas(
                    linha,
                    [
                        "Observações",
                        "Observacoes",
                        "Observação",
                        "Observacao",
                        "Obs",
                        "OBS",
                        "Histórico",
                        "Historico",
                        "Descrição",
                        "Descricao"
                    ]
                )
            )

            chave_conciliacao = gerar_chave_conta_pagar(
                numero_fatura=numero_fatura,
                fornecedor_funcionario=fornecedor_funcionario,
                valor=valor,
                data_vencimento=data_vencimento,
                plano_contas=plano_contas
            )

            data_base_competencia = data_vencimento or data_pagamento

            registros_processados.append({
                "chave_conciliacao": chave_conciliacao,
                "numero_fatura": numero_fatura,
                "fornecedor_funcionario": fornecedor_funcionario,
                "telefone": telefone,
                "email": email,
                "plano_contas": plano_contas,
                "categoria": limpar_categoria(plano_contas),
                "setor": identificar_setor(plano_contas, receita=False),
                "data_documento": data_documento,
                "data_vencimento": data_vencimento,
                "data_pagamento": data_pagamento if pago else None,
                "valor": valor,
                "pago": bool(pago),
                "status": "PAGO" if pago else "PENDENTE",
                "observacoes": observacoes,
                "mes": data_base_competencia.month if data_base_competencia else mes_retorno,
                "ano": data_base_competencia.year if data_base_competencia else ano_retorno,
            })

            if chave_conciliacao:
                chaves_conciliacao.add(chave_conciliacao)

        contas_existentes = {}

        if chaves_conciliacao:
            chaves_lista = list(chaves_conciliacao)
            tamanho_lote = 500

            with db.session.no_autoflush:
                for inicio in range(0, len(chaves_lista), tamanho_lote):
                    lote = chaves_lista[inicio:inicio + tamanho_lote]

                    existentes_lote = ContaPagarImportada.query.filter(
                        ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA",
                        ContaPagarImportada.chave_conciliacao.in_(lote)
                    ).all()

                    for conta_existente in existentes_lote:
                        if conta_existente.chave_conciliacao:
                            contas_existentes[conta_existente.chave_conciliacao] = conta_existente

        # Evita processar a mesma chave duas vezes na mesma importação.
        # Se a planilha vier duplicada, prevalece a última ocorrência.
        contas_criadas_nesta_importacao = {}

        for dados in registros_processados:
            chave_conciliacao = dados["chave_conciliacao"]

            conta = contas_criadas_nesta_importacao.get(chave_conciliacao)

            if not conta:
                conta = contas_existentes.get(chave_conciliacao)

            if conta:
                total_atualizado += 1
            else:
                conta = ContaPagarImportada()
                db.session.add(conta)
                total_criado += 1

                if chave_conciliacao:
                    contas_criadas_nesta_importacao[chave_conciliacao] = conta

            conta.numero_fatura = dados["numero_fatura"]
            conta.fornecedor_funcionario = dados["fornecedor_funcionario"]
            conta.telefone = dados["telefone"]
            conta.email = dados["email"]

            conta.plano_contas = dados["plano_contas"]
            conta.categoria = dados["categoria"]
            conta.setor = dados["setor"]

            conta.data_documento = dados["data_documento"]
            conta.data_vencimento = dados["data_vencimento"]
            conta.data_pagamento = dados["data_pagamento"]

            conta.valor = dados["valor"]
            conta.pago = dados["pago"]
            conta.status = dados["status"]
            conta.observacoes = dados["observacoes"]

            conta.mes = dados["mes"]
            conta.ano = dados["ano"]

            conta.chave_conciliacao = chave_conciliacao
            conta.origem_importacao = "DESPESA_COMPLETA"
            conta.importado_em = datetime.utcnow()

            if dados["pago"]:
                total_pagas += 1
            else:
                total_pendentes += 1

            total_linhas += 1

        db.session.commit()

        flash(
            (
                f"Despesas Completo importado com sucesso! "
                f"{total_linhas} linha(s) processada(s). "
                f"Pagas: {total_pagas}. "
                f"Pendentes: {total_pendentes}. "
                f"Criadas: {total_criado}. "
                f"Atualizadas: {total_atualizado}. "
                f"Ignoradas: {total_ignorado}."
            ),
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao importar Despesas Completo: {str(e)}", "danger")

    return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#despesas-completo")


# =========================================================
# IMPORTAR CONTAS A RECEBER
# Agora aceita arquivo com vários meses/ano inteiro.
# A competência é definida automaticamente pela data da linha:
# - recebido com Dt. Pgto: usa Dt. Pgto
# - pendente: usa Dt. Vencto
# =========================================================

@importacoes_bp.route("/contas-a-receber", methods=["POST"])
@gestao_required
def importar_contas_receber():

    arquivo = request.files.get("arquivo_receber")

    # Usado apenas para voltar para a competência que o usuário estava visualizando.
    hoje = datetime.now()
    mes_retorno = request.form.get("mes_retorno", type=int) or hoje.month
    ano_retorno = request.form.get("ano_retorno", type=int) or hoje.year

    if not arquivo or not arquivo.filename:
        flash("Selecione uma planilha de Contas a Receber.", "danger")
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#contas-receber")

    if not arquivo_excel_valido(arquivo.filename):
        flash(
            "Formato não aceito. Envie uma planilha Excel nos formatos .xlsx, .xlsm, .xls ou .xlsb.",
            "danger"
        )
        return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#contas-receber")

    try:
        linhas = ler_linhas_upload(arquivo, "Nº Fatura")

        total_criado = 0
        total_atualizado = 0
        total_ignorado = 0
        total_linhas = 0
        total_recebidas = 0
        total_pendentes = 0
        competencias_importadas = set()

        for linha in linhas:
            plano_contas = texto(valor_por_coluna(linha, "Pl. Contas"))

            numero_fatura = texto(valor_por_coluna(linha, "Nº Fatura"))
            cliente = texto(valor_por_coluna(linha, "Cliente"))
            telefone = texto(valor_por_coluna(linha, "Telefone"))
            email = texto(valor_por_coluna(linha, "Email"))

            data_pagamento = normalizar_data(valor_por_coluna(linha, "Dt. Pgto"))
            data_vencimento = normalizar_data(valor_por_coluna(linha, "Dt. Vencto"))
            data_documento = normalizar_data(valor_por_coluna(linha, "Dt. Docto"))

            pago = normalizar_bool(valor_por_coluna(linha, "Pg?"))
            valor = normalizar_decimal(valor_por_coluna(linha, "Valor"))
            juros = normalizar_decimal(valor_por_coluna(linha, "Juros"))
            total = normalizar_decimal(valor_por_coluna(linha, "Total"))

            if total == Decimal("0.00"):
                total = valor + juros

            # Define a competência automaticamente.
            # Se já recebeu e existe Dt. Pgto, entra no mês do recebimento real.
            # Se ainda está pendente, entra no mês de vencimento.
            data_base_competencia = data_pagamento if data_pagamento else data_vencimento

            if not data_base_competencia:
                total_ignorado += 1
                continue

            cobranca = texto(valor_por_coluna(linha, "Cobrança"))

            observacoes = texto(
                valor_por_colunas(
                    linha,
                    [
                        "Observações",
                        "Observacoes",
                        "Observação",
                        "Observacao",
                        "Obs",
                        "OBS",
                        "Histórico",
                        "Historico",
                        "Descrição",
                        "Descricao"
                    ]
                )
            )

            valor_chave = total if total != Decimal("0.00") else valor

            chave_conciliacao = gerar_chave_conta_receber(
                numero_fatura=numero_fatura,
                cliente=cliente,
                valor=valor_chave,
                data_vencimento=data_vencimento,
                plano_contas=plano_contas
            )

            conta = buscar_conta_receber_existente(
                chave_conciliacao=chave_conciliacao,
                numero_fatura=numero_fatura,
                cliente=cliente,
                valor=valor_chave,
                data_vencimento=data_vencimento,
                plano_contas=plano_contas
            )

            if conta:
                total_atualizado += 1
            else:
                conta = ContaReceberImportada()
                db.session.add(conta)
                total_criado += 1

            conta.numero_fatura = numero_fatura
            conta.cliente = cliente
            conta.telefone = telefone
            conta.email = email

            conta.plano_contas = plano_contas
            conta.categoria = limpar_categoria(plano_contas)
            conta.setor = identificar_setor(plano_contas, receita=True)

            conta.cobranca = cobranca

            conta.data_documento = data_documento
            conta.data_vencimento = data_vencimento

            if data_pagamento:
                conta.data_pagamento = data_pagamento
            elif not pago:
                conta.data_pagamento = None

            conta.valor = valor
            conta.juros = juros
            conta.total = total

            conta.pago = pago
            conta.status = "RECEBIDO" if pago else "PENDENTE"

            if pago:
                total_recebidas += 1
            else:
                total_pendentes += 1

            conta.observacoes = observacoes

            # NOVO CENÁRIO:
            # Não força mais tudo para o mês escolhido na tela.
            # A competência vem da data real da linha.
            conta.mes = data_base_competencia.month
            conta.ano = data_base_competencia.year
            competencias_importadas.add(f"{data_base_competencia.month:02d}/{data_base_competencia.year}")

            conta.chave_conciliacao = chave_conciliacao
            conta.origem_importacao = "RECEBIMENTO"
            conta.importado_em = datetime.utcnow()

            total_linhas += 1

        db.session.commit()

        competencias_txt = ", ".join(sorted(competencias_importadas)) or "nenhuma competência"

        flash(
            (
                f"Contas a Receber importadas com sucesso! "
                f"{total_linhas} linha(s) processada(s). "
                f"Competências identificadas: {competencias_txt}. "
                f"Recebidas: {total_recebidas}. "
                f"Pendentes: {total_pendentes}. "
                f"Criadas: {total_criado}. "
                f"Atualizadas: {total_atualizado}. "
                f"Ignoradas sem data: {total_ignorado}."
            ),
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao importar Contas a Receber: {str(e)}", "danger")

    return redirect(f"/gestao/importacoes/?mes={mes_retorno}&ano={ano_retorno}#contas-receber")


# =========================================================
# EDITAR / EXCLUIR CONTAS PAGAS
# =========================================================

@importacoes_bp.route("/contas-pagas/editar/<int:id>", methods=["POST"])
@gestao_required
def editar_conta_paga(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    mes_redirect = request.form.get("mes_redirect", type=int) or conta.mes
    ano_redirect = request.form.get("ano_redirect", type=int) or conta.ano

    try:
        plano_contas = texto(request.form.get("plano_contas"))

        conta.numero_fatura = texto(request.form.get("numero_fatura"))
        conta.fornecedor_funcionario = texto(request.form.get("fornecedor_funcionario"))
        conta.telefone = texto(request.form.get("telefone"))
        conta.email = texto(request.form.get("email"))

        conta.plano_contas = plano_contas
        conta.categoria = limpar_categoria(plano_contas)
        conta.setor = texto(request.form.get("setor")) or identificar_setor(plano_contas, receita=False)

        conta.data_documento = normalizar_data(request.form.get("data_documento"))
        conta.data_vencimento = normalizar_data(request.form.get("data_vencimento"))
        conta.data_pagamento = normalizar_data(request.form.get("data_pagamento"))

        conta.valor = normalizar_decimal(request.form.get("valor"))

        conta.pago = True
        conta.status = "PAGO"

        conta.observacoes = texto(request.form.get("observacoes"))

        conta.chave_conciliacao = gerar_chave_conta_pagar(
            numero_fatura=conta.numero_fatura,
            fornecedor_funcionario=conta.fornecedor_funcionario,
            valor=conta.valor,
            data_vencimento=conta.data_vencimento,
            plano_contas=conta.plano_contas
        )

        if not conta.origem_importacao:
            conta.origem_importacao = "PAGAMENTO"

        db.session.commit()

        flash("Conta paga atualizada com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao editar conta paga: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect)


@importacoes_bp.route("/contas-pagas/excluir/<int:id>", methods=["POST"])
@gestao_required
def excluir_conta_paga(id):

    conta = ContaPagarImportada.query.get_or_404(id)

    mes_redirect = request.form.get("mes_redirect", type=int) or conta.mes
    ano_redirect = request.form.get("ano_redirect", type=int) or conta.ano

    try:
        db.session.delete(conta)
        db.session.commit()

        flash("Conta paga excluída com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir conta paga: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect)


# =========================================================
# EDITAR / EXCLUIR CONTAS A RECEBER IMPORTADAS
# =========================================================

@importacoes_bp.route("/contas-a-receber/editar/<int:id>", methods=["POST"])
@gestao_required
def editar_conta_receber(id):

    conta = ContaReceberImportada.query.get_or_404(id)

    mes_redirect = request.form.get("mes_redirect", type=int) or conta.mes
    ano_redirect = request.form.get("ano_redirect", type=int) or conta.ano

    try:
        plano_contas = texto(request.form.get("plano_contas"))

        conta.numero_fatura = texto(request.form.get("numero_fatura"))
        conta.cliente = texto(request.form.get("cliente"))
        conta.telefone = texto(request.form.get("telefone"))
        conta.email = texto(request.form.get("email"))

        conta.plano_contas = plano_contas
        conta.categoria = limpar_categoria(plano_contas)
        conta.setor = texto(request.form.get("setor")) or identificar_setor(plano_contas, receita=True)

        conta.cobranca = texto(request.form.get("cobranca"))

        conta.data_documento = normalizar_data(request.form.get("data_documento"))
        conta.data_vencimento = normalizar_data(request.form.get("data_vencimento"))
        conta.data_pagamento = normalizar_data(request.form.get("data_pagamento"))

        conta.valor = normalizar_decimal(request.form.get("valor"))
        conta.juros = normalizar_decimal(request.form.get("juros"))
        conta.total = normalizar_decimal(request.form.get("total"))

        if conta.total == Decimal("0.00"):
            conta.total = normalizar_decimal(conta.valor) + normalizar_decimal(conta.juros)

        conta.pago = True if request.form.get("pago") == "on" else False
        conta.status = "RECEBIDO" if conta.pago else "PENDENTE"

        conta.observacoes = texto(request.form.get("observacoes"))

        valor_chave = conta.total if conta.total != Decimal("0.00") else conta.valor

        conta.chave_conciliacao = gerar_chave_conta_receber(
            numero_fatura=conta.numero_fatura,
            cliente=conta.cliente,
            valor=valor_chave,
            data_vencimento=conta.data_vencimento,
            plano_contas=conta.plano_contas
        )

        if not conta.origem_importacao:
            conta.origem_importacao = "RECEBIMENTO"

        db.session.commit()

        flash("Conta a receber atualizada com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao editar conta a receber: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect)


@importacoes_bp.route("/contas-a-receber/excluir/<int:id>", methods=["POST"])
@gestao_required
def excluir_conta_receber(id):

    conta = ContaReceberImportada.query.get_or_404(id)

    mes_redirect = request.form.get("mes_redirect", type=int) or conta.mes
    ano_redirect = request.form.get("ano_redirect", type=int) or conta.ano

    try:
        db.session.delete(conta)
        db.session.commit()

        flash("Conta a receber excluída com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir conta a receber: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect)


# =========================================================
# LIMPAR IMPORTAÇÕES DO MÊS
# =========================================================

@importacoes_bp.route("/contas-pagas/limpar", methods=["POST"])
@gestao_required
def limpar_contas_pagas_mes():

    hoje = datetime.now()
    mes_redirect = request.form.get("mes", type=int) or hoje.month
    ano_redirect = request.form.get("ano", type=int) or hoje.year

    try:
        registros = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "PAGAMENTO"
        ).all()

        total_removido = len(registros)

        for registro in registros:
            db.session.delete(registro)

        db.session.commit()

        flash(
            f"Toda a importação de Contas Pagas foi limpa com sucesso. "
            f"{total_removido} registro(s) importado(s) removido(s) de todos os meses.",
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar toda a importação de Contas Pagas: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect)


@importacoes_bp.route("/contas-a-receber/limpar", methods=["POST"])
@gestao_required
def limpar_contas_receber_mes():

    hoje = datetime.now()
    mes_redirect = request.form.get("mes", type=int) or hoje.month
    ano_redirect = request.form.get("ano", type=int) or hoje.year

    try:
        registros = ContaReceberImportada.query.filter(
            or_(
                ContaReceberImportada.origem_importacao == "RECEBIMENTO",
                ContaReceberImportada.origem_importacao == "VENCIMENTO",
                ContaReceberImportada.origem_importacao.is_(None),
            )
        ).all()

        total_removido = len(registros)

        for registro in registros:
            db.session.delete(registro)

        db.session.commit()

        flash(
            f"Toda a importação de Contas a Receber foi limpa com sucesso. "
            f"{total_removido} registro(s) importado(s) removido(s) de todos os meses.",
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar toda a importação de Contas a Receber: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect, "#contas-receber")


@importacoes_bp.route("/despesas-completo/limpar", methods=["POST"])
@gestao_required
def limpar_despesas_completo_mes():

    hoje = datetime.now()
    mes_redirect = request.form.get("mes", type=int) or hoje.month
    ano_redirect = request.form.get("ano", type=int) or hoje.year

    try:
        registros = ContaPagarImportada.query.filter(
            ContaPagarImportada.origem_importacao == "DESPESA_COMPLETA"
        ).all()

        total_removido = len(registros)

        for registro in registros:
            db.session.delete(registro)

        db.session.commit()

        flash(
            f"Toda a importação de Despesas Completo/Radar foi limpa com sucesso. "
            f"{total_removido} registro(s) removido(s) de todos os meses.",
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar toda a importação de Despesas Completo: {str(e)}", "danger")

    return redirect_importacoes(mes_redirect, ano_redirect, "#despesas-completo")


# =========================================================
# ROTAS ANTIGAS MANTIDAS PARA NÃO QUEBRAR LINKS EXISTENTES
# =========================================================

@importacoes_bp.route("/contas-a-pagar", methods=["POST"])
@gestao_required
def importar_contas_pagar():
    """
    Rota antiga mantida por segurança.
    No novo cenário, a tela usa /contas-pagas.
    """
    return importar_contas_pagas()


@importacoes_bp.route("/contas-a-pagar/editar/<int:id>", methods=["POST"])
@gestao_required
def editar_conta_pagar(id):
    """
    Rota antiga mantida por segurança.
    """
    return editar_conta_paga(id)


@importacoes_bp.route("/contas-a-pagar/excluir/<int:id>", methods=["POST"])
@gestao_required
def excluir_conta_pagar(id):
    """
    Rota antiga mantida por segurança.
    """
    return excluir_conta_paga(id)